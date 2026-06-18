"""Tkinter GUI for the IATA bulk validator.

Layout:
  ┌──────────────────────────────────────────────────────┐
  │ IATA CheckACode — Bulk Validator                     │
  ├──────────────────────────────────────────────────────┤
  │ Input file:  [path____________________] [Browse]     │
  │ Sheet:       [dropdown]                              │
  │ IATA column: [dropdown]                              │
  │ Row range:   [start] to [end]   (blank = all)        │
  │ Output dir:  [path____________________] [Browse]     │
  ├──────────────────────────────────────────────────────┤
  │ [Start]  [Pause]  [Resume]  [Stop]                   │
  ├──────────────────────────────────────────────────────┤
  │ Progress: ████████░░░  234 / 3000 (7.8%) — 4h 12m    │
  │                                                      │
  │ Log:                                                 │
  │ ┌──────────────────────────────────────────────────┐ │
  │ │ 14:02:11  32302491  VALID  TRAVEL POINT ...      │ │
  │ │ 14:02:14  32302492  INVALID                      │ │
  │ │ ...                                              │ │
  │ └──────────────────────────────────────────────────┘ │
  └──────────────────────────────────────────────────────┘
"""

from __future__ import annotations

import logging
import queue
import threading
import time
import tkinter as tk
import winsound
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from . import (
    __version__, auth, config, excel_io, graph_mailer, mailer_client, mailer_io,
    oep_client, oep_presets, traffic_client, updater, zenith_client,
    zenith_history_analyzer, zenith_history_downloader, zenith_history_parser,
    zenith_loads_index, zenith_pnr_client, zenith_pnr_history_analyzer,
    zenith_pnr_history_downloader,
)
from .mailer_log import MailerLog
from .traffic_sources import SOURCES as TRAFFIC_SOURCES
from .zenith_pnr_cache import ZenithPNRCache
from .zenith_pnr_history_cache import ZenithPNRHistoryCache
from .bd_agency_client import Agency, fetch_all_agencies, filter_status
from .bd_cache import BDAgencyCache
from .zenith_cache import ZenithCache
from .bd_matcher import (
    FIELD_ADDRESS,
    FIELD_LICENSE,
    FIELD_NAME,
    AgencyIndex,
)
from .cache import Cache
from .parser import LookupResult, now_iso
from .validator import (
    CaptchaChallenge,
    IATAValidator,
    ValidatorStopped,
    make_validator,
)

log = logging.getLogger(__name__)


# IATA worker → GUI message types
MSG_LOG = "log"
MSG_PROGRESS = "progress"
MSG_RESULT = "result"
MSG_CAPTCHA = "captcha"
MSG_DONE = "done"
MSG_ERROR = "error"

# BD Agency worker → GUI message types
MSG_BD_LOG = "bd_log"
MSG_BD_PROGRESS = "bd_progress"
MSG_BD_DONE = "bd_done"
MSG_BD_ERROR = "bd_error"
MSG_BD_REFRESHED = "bd_refreshed"   # payload: (count:int, last_refresh:str)

# OEP worker → GUI message types
MSG_OEP_LOG = "oep_log"
MSG_OEP_DONE = "oep_done"         # payload: dict (see _oep_worker_run)
MSG_OEP_ERROR = "oep_error"
MSG_OEP_BUSY = "oep_busy"         # payload: str (status text)

# Traffic tab worker → GUI message types
MSG_TRAFFIC_LOG = "traffic_log"
MSG_TRAFFIC_BUSY = "traffic_busy"   # payload: str (status text)
MSG_TRAFFIC_DONE = "traffic_done"   # payload: dict (rows, view, source_label, dates)
MSG_TRAFFIC_ERROR = "traffic_error"

# Zenith worker → GUI message types
MSG_ZENITH_LOG = "zenith_log"
MSG_ZENITH_PROGRESS = "zenith_progress"   # payload: (done, total, ok, nf, err)
MSG_ZENITH_RESULT = "zenith_result"        # payload: LookupResult
MSG_ZENITH_DONE = "zenith_done"            # payload: str (output path)
MSG_ZENITH_ERROR = "zenith_error"
MSG_ZENITH_LOGGED_IN = "zenith_logged_in"  # payload: dict(state_values)
MSG_ZENITH_LOGIN_FAILED = "zenith_login_failed"
# Zenith Flight Loads sub-tab
MSG_ZENITH_FL_LOG = "zenith_fl_log"
MSG_ZENITH_FL_PROGRESS = "zenith_fl_progress"  # (chunk_label, done_chunks, total_chunks, rows)
MSG_ZENITH_PAX_PROGRESS = "zenith_pax_progress"  # (done_legs, total_legs, label, pax_so_far)
MSG_ZENITH_PAX_DONE = "zenith_pax_done"          # payload: dict(path, legs, pax)
MSG_ZENITH_PAX_ERROR = "zenith_pax_error"
MSG_ZENITH_FL_DONE = "zenith_fl_done"           # payload: str(output path)
MSG_ZENITH_FL_ERROR = "zenith_fl_error"
# Zenith Flight History Analyzer sub-tab
MSG_ZENITH_FH_PROGRESS = "zenith_fh_progress"  # (i, total, current_file)
MSG_ZENITH_FH_PARSED = "zenith_fh_parsed"      # payload: int (event count)
MSG_ZENITH_FH_DONE = "zenith_fh_done"          # payload: HistoryAuditReport
MSG_ZENITH_FH_ERROR = "zenith_fh_error"
MSG_ZENITH_PNRMISUSE_DONE = "zenith_pnrmisuse_done"    # payload: summary dict
MSG_ZENITH_PNRMISUSE_ERROR = "zenith_pnrmisuse_error"  # payload: str
MSG_ZENITH_DOSSIER_DONE = "zenith_dossier_done"        # payload: summary dict
MSG_ZENITH_DOSSIER_ERROR = "zenith_dossier_error"      # payload: str
MSG_ZENITH_DOSSIER_PROGRESS = "zenith_dossier_progress"  # payload: (i, n, pnr)
# Zenith Flight History downloader (Phase 3)
MSG_ZENITH_DL_STATUS = "zenith_dl_status"      # payload: str (status line)
MSG_ZENITH_DL_PROGRESS = "zenith_dl_progress"  # (i, total, label, status_code)
MSG_ZENITH_DL_DONE = "zenith_dl_done"          # payload: dict (counts + folder)
MSG_ZENITH_DL_ERROR = "zenith_dl_error"
# Zenith PNR enrichment (Phase A — Phase 2)
MSG_ZENITH_PNR_PROGRESS = "zenith_pnr_progress"  # (i, total, code, status)
MSG_ZENITH_PNR_DONE = "zenith_pnr_done"          # payload: HistoryAuditReport (enriched)
MSG_ZENITH_PNR_ERROR = "zenith_pnr_error"
# Zenith bulk PNR lookup (standalone)
MSG_ZENITH_BULK_PROGRESS = "zenith_bulk_progress"  # (i, total, code, status)
MSG_ZENITH_BULK_DONE = "zenith_bulk_done"          # payload: dict (path, counts)
MSG_ZENITH_BULK_ERROR = "zenith_bulk_error"
# Bulk Mailer
MSG_MAIL_PROGRESS = "mail_progress"   # (i, total, to, status)
MSG_MAIL_DONE = "mail_done"           # payload: dict (drafted/sent/failed/skipped)
MSG_MAIL_ERROR = "mail_error"

# Updater worker → GUI message types
MSG_UPDATE_LOG = "update_log"
MSG_UPDATE_PROGRESS = "update_progress"  # payload: (downloaded:int, total:int)
MSG_UPDATE_FOUND = "update_found"        # payload: UpdateInfo
MSG_UPDATE_NONE = "update_none"          # payload: UpdateInfo (or None)
MSG_UPDATE_DOWNLOADED = "update_downloaded"  # payload: Path
MSG_UPDATE_ERROR = "update_error"


# ---------------------------------------------------------------------------
# Central usage-telemetry registry (future-proof)
# ---------------------------------------------------------------------------
# Any completion/"done" message listed here is reported to the usage backend
# from ONE place in `_handle_msg`, so a newly added feature is tracked the
# moment its DONE message is added here — no per-feature wiring. The value is
# (action_label, count_key) where count_key, when set, is read from a dict
# payload to record the batch size.
_USAGE_EVENTS: "dict[str, tuple[str, str | None]]" = {
    MSG_ZENITH_LOGGED_IN: ("zenith_login", None),
    MSG_ZENITH_DONE: ("zenith_customer", None),
    MSG_ZENITH_FL_DONE: ("zenith_flight_loads", None),
    MSG_ZENITH_PAX_DONE: ("zenith_passenger", "pax"),
    MSG_ZENITH_FH_DONE: ("zenith_history_analyze", None),
    MSG_ZENITH_DL_DONE: ("zenith_history_download", None),
    MSG_ZENITH_PNR_DONE: ("zenith_pnr_enrich", None),
    MSG_ZENITH_BULK_DONE: ("zenith_pnr_bulk", None),
    MSG_OEP_DONE: ("oep_movement", None),
    MSG_TRAFFIC_DONE: ("traffic_movement", None),
    MSG_MAIL_DONE: ("mailer_send", "sent"),
    MSG_UPDATE_FOUND: ("update_available", None),
    MSG_UPDATE_DOWNLOADED: ("update_downloaded", None),
}

# Failure events — tracked the same way so the Usage view shows attempts that
# errored, not only successes. Payload is usually the error string; we attach a
# short, capped note for diagnostics (never record contents). Adding a new
# feature's error message here is the only step to track its failures.
_USAGE_ERRORS: "dict[str, str]" = {
    MSG_ERROR: "iata_error",
    MSG_BD_ERROR: "bd_error",
    MSG_OEP_ERROR: "oep_error",
    MSG_TRAFFIC_ERROR: "traffic_error",
    MSG_ZENITH_ERROR: "zenith_customer_error",
    MSG_ZENITH_LOGIN_FAILED: "zenith_login_failed",
    MSG_ZENITH_FL_ERROR: "zenith_flight_loads_error",
    MSG_ZENITH_PAX_ERROR: "zenith_passenger_error",
    MSG_ZENITH_FH_ERROR: "zenith_history_analyze_error",
    MSG_ZENITH_DL_ERROR: "zenith_history_download_error",
    MSG_ZENITH_PNR_ERROR: "zenith_pnr_enrich_error",
    MSG_ZENITH_BULK_ERROR: "zenith_pnr_bulk_error",
    MSG_MAIL_ERROR: "mailer_error",
    MSG_UPDATE_ERROR: "update_error",
}


class App:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Travel Ops Console")
        self.root.geometry("1080x820")
        self.root.minsize(900, 640)

        # ----- IATA tab state -----
        self.input_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.column_name = tk.StringVar()
        self.start_row = tk.StringVar(value="2")
        self.end_row = tk.StringVar(value="")
        self.output_dir = tk.StringVar(value=str(Path.home() / "Documents"))

        self._sheet_columns: list[str] = []
        self._worker: threading.Thread | None = None
        self._validator: IATAValidator | None = None
        self._pause_event = threading.Event()
        self._pause_event.set()
        self._stop_flag = threading.Event()
        self._captcha_clear_flag = threading.Event()

        # ----- BD tab state -----
        self.bd_mode = tk.StringVar(value="full")  # "full" or "lookup"
        self.bd_input_path = tk.StringVar()
        self.bd_sheet_name = tk.StringVar()
        self.bd_column_name = tk.StringVar()
        self.bd_output_dir = tk.StringVar(value=str(Path.home() / "Documents"))
        self.bd_include_expired = tk.BooleanVar(value=True)
        self.bd_match_name = tk.BooleanVar(value=True)       # default ON
        self.bd_match_license = tk.BooleanVar(value=True)    # default ON
        self.bd_match_address = tk.BooleanVar(value=False)   # default OFF

        self._bd_sheet_columns: list[str] = []
        self._bd_worker: threading.Thread | None = None
        self._bd_cache = BDAgencyCache(config.BD_CACHE_DB)

        # ----- OEP tab state -----
        from datetime import date, timedelta
        today = date.today()
        year_ago = today - timedelta(days=365)
        self.oep_date_from = tk.StringVar(value=year_ago.isoformat())
        self.oep_date_to = tk.StringVar(value=today.isoformat())
        self.oep_mode = tk.StringVar(value="country")
        # country | division | category | gender
        self.oep_gender = tk.StringVar(value="")        # "", "1", "2", "3"
        self.oep_country_filter = tk.StringVar(value="")  # country option label
        self.oep_output_dir = tk.StringVar(value=str(Path.home() / "Documents"))
        self.oep_top_n = tk.StringVar(value="5")
        # ----- Traffic tab state -----
        _traffic_labels = [s.label for s in TRAFFIC_SOURCES.values()]
        self.traffic_source = tk.StringVar(value=_traffic_labels[0] if _traffic_labels else "")
        self.traffic_date_from = tk.StringVar(value="")
        self.traffic_date_to = tk.StringVar(value="")
        self.traffic_view = tk.StringVar(value="country")
        self.traffic_csv_path = tk.StringVar(value="")
        self.traffic_output_dir = tk.StringVar(value=str(Path.home() / "Documents"))
        self._traffic_worker = None
        self._traffic_last_result = None
        self.oep_preset_name = tk.StringVar(value="")

        self._oep_worker: threading.Thread | None = None
        self._oep_last_result: dict | None = None
        self._oep_country_options: list[oep_client.Option] = []
        # Set by the background country-list thread, drained by the main
        # thread's `_oep_poll_country_load`. Tuple of (labels, error_msg).
        self._oep_pending_country_payload: tuple[list[str], str | None] | None = None
        self._oep_preset_store = oep_presets.PresetStore(config.OEP_PRESET_FILE)
        # Embedded matplotlib chart; instantiated lazily on first time-series view.
        self._oep_chart_canvas = None
        self._oep_chart_figure = None

        # ----- Zenith tab state -----
        self.zenith_input_path = tk.StringVar()
        self.zenith_sheet_name = tk.StringVar()
        self.zenith_column_name = tk.StringVar()
        self.zenith_output_dir = tk.StringVar(value=str(Path.home() / "Documents"))
        self.zenith_username = tk.StringVar()
        self.zenith_company = tk.StringVar(value="usba")
        self.zenith_concurrency = tk.IntVar(value=3)
        self.zenith_delay_s = tk.DoubleVar(value=0.8)
        self.zenith_skip_cached = tk.BooleanVar(value=True)
        self.zenith_test_mode = tk.BooleanVar(value=False)

        self._zenith_sheet_columns: list[str] = []
        self._zenith_session: zenith_client.ZenithSession | None = None
        self._zenith_worker: threading.Thread | None = None
        self._zenith_stop_flag = threading.Event()
        self._zenith_pause_flag = threading.Event()
        self._zenith_cache = ZenithCache(config.ZENITH_CACHE_DB)

        # ----- Bulk Mailer tab state -----
        self.mail_mapping_path = tk.StringVar()
        self.mail_attach_dir = tk.StringVar()
        self.mail_subject = tk.StringVar()
        self.mail_mode = tk.StringVar(value="draft")   # "draft" | "send"
        self.mail_delay_s = tk.DoubleVar(value=1.0)
        self.mail_skip_sent = tk.BooleanVar(value=True)
        # Transport: "outlook" (local desktop), "graph" (M365 sign-in),
        # or "smtp" (any provider). Outlook is the default: it sends from
        # whatever account is already added to desktop Outlook, needing
        # neither SMTP basic-auth nor admin Graph consent — both of which
        # locked-down corporate M365 tenants commonly block.
        self.mail_transport = tk.StringVar(value="outlook")
        self.mail_smtp_preset = tk.StringVar(value="Gmail / Google Workspace")
        self.mail_smtp_host = tk.StringVar(value="smtp.gmail.com")
        self.mail_smtp_port = tk.IntVar(value=587)
        self.mail_smtp_sender = tk.StringVar()
        self.mail_smtp_password = tk.StringVar()
        self.mail_smtp_remember = tk.BooleanVar(value=True)
        self.mail_outlook_account = tk.StringVar()   # chosen Outlook sender
        self._graph_session = None                    # graph_mailer.GraphSession
        self.mail_graph_status = tk.StringVar(value="Not signed in")
        self._mail_rows: list = []                     # list[mailer_io.MailRow]
        self._mail_worker: threading.Thread | None = None
        self._mail_stop_flag = threading.Event()
        self._mail_log = MailerLog(config.MAILER_LOG_DB)

        # ----- Shared message queue -----
        self._msg_queue: "queue.Queue[tuple[str, object]]" = queue.Queue()

        # ----- Auth + updater state -----
        self._signed_in_user: dict | None = None
        self._update_worker: threading.Thread | None = None

        self._setup_styles()
        self._build_ui()
        self._refresh_bd_status_label()
        # Show the correct auth button (Sign in or Sign out) from launch,
        # before `ensure_signed_in` runs.
        self._refresh_user_label()
        self.root.after(100, self._poll_queue)
        # Live tab-label refresh — appends "●" while a worker is running
        # and "(N cached)" once an idle tab has cache to show off.
        self.root.after(500, self._refresh_tab_labels)
        # Usage telemetry: report every feature use — including tabs added in
        # the future — through one central, fire-and-forget path.
        self._telemetry_ready = False
        self._last_view: dict[str, str] = {}
        self._install_usage_tracking()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _refresh_tab_labels(self) -> None:
        """Annotate outer-notebook tabs with live worker / cache state.

        Re-runs on a 2-second tick. Keeps the labels short — a single
        glyph plus an optional count, so the user can see "Zenith ●"
        means a worker is busy without leaving their current tab.
        """
        try:
            running = {
                "iata": self._worker is not None and self._worker.is_alive(),
                "bd": self._bd_worker is not None and self._bd_worker.is_alive(),
                "oep": self._oep_worker is not None and self._oep_worker.is_alive(),
                "zenith": any(
                    w is not None and w.is_alive() for w in (
                        getattr(self, "_zenith_worker", None),
                        getattr(self, "_zenith_fl_worker", None),
                        getattr(self, "_zenith_fh_worker", None),
                        getattr(self, "_zenith_pnr_worker", None),
                        getattr(self, "_zenith_bulk_worker", None),
                        getattr(self, "_zenith_dl_worker", None),
                    )
                ),
            }
            cache_counts = {
                "iata": self._cache.count() if hasattr(self, "_cache") and self._cache else 0,
                "bd": (
                    self._bd_cache.count()
                    if hasattr(self, "_bd_cache") and self._bd_cache else 0
                ),
                "zenith": (
                    self._zenith_pnr_cache.count()
                    if hasattr(self, "_zenith_pnr_cache") and self._zenith_pnr_cache else 0
                ),
            }
            for key, widget in self._tab_widgets.items():
                base = self._tab_base_labels[widget]
                pieces = [base]
                if running.get(key):
                    pieces.append("●")
                count = cache_counts.get(key)
                if count:
                    pieces.append(f"({count:,} cached)")
                label = " ".join(pieces)
                try:
                    self.outer_notebook.tab(widget, text=label)
                except tk.TclError:
                    pass
        except Exception:  # noqa: BLE001 — never let the timer kill the app
            log.exception("tab-label refresh failed")
        finally:
            try:
                self.root.after(2000, self._refresh_tab_labels)
            except tk.TclError:
                return

    # ------------------------------------------------------------------
    # Usage telemetry — one central, future-proof path
    # ------------------------------------------------------------------
    def _track(
        self,
        action: str,
        *,
        target: str | None = None,
        count: int = 0,
        notes: str | None = None,
    ) -> None:
        """Fire-and-forget usage event. Safe from any thread, never blocks the
        UI, never raises. Every feature reports through here so the backend
        Usage view reflects what users actually do."""
        def _send() -> None:
            try:
                auth.log_lookup_event(
                    action=action, target=target,
                    count=int(count or 0), notes=notes,
                )
            except Exception as exc:  # noqa: BLE001
                log.debug("usage telemetry %r failed: %s", action, exc)
        try:
            threading.Thread(target=_send, daemon=True).start()
        except Exception:  # noqa: BLE001
            pass

    def _install_usage_tracking(self) -> None:
        """Bind tab-change telemetry on EVERY notebook in the widget tree.

        Walking the whole tree means tabs and sub-tabs added in the future are
        tracked automatically — no per-feature wiring. Completion events are
        handled centrally via `_USAGE_EVENTS` in `_handle_msg`.
        """
        def _walk(widget) -> None:
            for child in widget.winfo_children():
                if isinstance(child, ttk.Notebook):
                    child.bind(
                        "<<NotebookTabChanged>>", self._on_tab_changed, add="+",
                    )
                _walk(child)
        try:
            _walk(self.root)
        except Exception:  # noqa: BLE001
            log.debug("usage tab-tracking install failed", exc_info=True)
        # Ignore the tab events that fire while the window first maps; only
        # count real navigation once the UI has settled.
        self.root.after(1500, lambda: setattr(self, "_telemetry_ready", True))

    @staticmethod
    def _clean_tab_label(text: str) -> str:
        """Strip the live "●" / "(N cached)" annotations from a tab label."""
        t = (text or "").replace("●", "").strip()
        if t.endswith("cached)") and "(" in t:
            t = t[: t.rfind("(")].strip()
        return t

    def _on_tab_changed(self, event) -> None:
        """Report which tab the user opened. Deduped per notebook so only real
        navigation (not repeated refreshes) is logged."""
        if not getattr(self, "_telemetry_ready", False):
            return
        try:
            nb = event.widget
            if not hasattr(nb, "select"):
                nb = self.root.nametowidget(nb)
            sel = nb.select()
            if not sel:
                return
            base = ""
            try:
                widget = self.root.nametowidget(sel)
                base = self._tab_base_labels.get(widget, "")
            except (tk.TclError, AttributeError, KeyError):
                base = ""
            if not base:
                base = self._clean_tab_label(nb.tab(sel, "text"))
        except tk.TclError:
            return
        if not base:
            return
        key = str(nb)
        if self._last_view.get(key) == base:
            return
        self._last_view[key] = base
        self._track(action="view", target=base)

    # ------------------------------------------------------------------
    # Theme + custom widget styles
    # ------------------------------------------------------------------

    # Semantic palette — single source of truth for every coloured
    # widget. If a future user requests dark mode this is the spot to
    # branch on `sv_ttk.get_theme()` and emit alternate hexes.
    _COLOR_PRIMARY = "#0078D4"   # Microsoft accent blue
    _COLOR_PRIMARY_HOVER = "#106EBE"
    _COLOR_DANGER = "#C42B1C"    # WinUI red for Stop / errors
    _COLOR_SUCCESS = "#107C10"   # WinUI green for "Signed in"
    _COLOR_WARNING = "#9D5D00"   # warm amber for partial states
    _COLOR_MUTED = "#64748b"     # slate-500 for hints / metadata
    _COLOR_SECTION = "#0F6CBD"   # brand blue for section titles (warmer than slate)
    _COLOR_ROW_BAD = "#FDE7E9"   # soft red row tint (e.g. QUESTIONABLE verdicts)
    _COLOR_ROW_GOOD = "#DFF6DD"  # soft green row tint
    _COLOR_ROW_WARN = "#FFF4CE"  # soft amber row tint

    def _setup_styles(self) -> None:
        """Apply the Sun Valley (Windows-11) base theme + semantic styles.

        sv_ttk takes over the ttk look entirely; we layer our named
        styles on top *after* `set_theme` so they win. If sv_ttk fails
        to import (e.g. unbundled dev environment) we fall back to the
        previous vista/winnative chain — the app still works, just
        looks plainer.
        """
        style = ttk.Style()
        try:
            import sv_ttk
            # Respect the in-memory dark/light preference so a toggle
            # → _setup_styles round-trip doesn't snap back to light.
            mode = "dark" if getattr(self, "_theme_is_dark", False) else "light"
            sv_ttk.set_theme(mode)
        except Exception:  # noqa: BLE001 — degrade gracefully
            log.warning("sv_ttk unavailable; using fallback ttk theme")
            for name in ("vista", "winnative", "clam"):
                if name in style.theme_names():
                    style.theme_use(name)
                    break

        # Typography hierarchy
        style.configure(
            "Section.TLabel",
            font=("Segoe UI Semibold", 11),
            foreground=self._COLOR_SECTION,
        )
        style.configure(
            "SectionLg.TLabel",
            font=("Segoe UI Semibold", 13),
            foreground=self._COLOR_SECTION,
        )
        style.configure(
            "Hint.TLabel",
            foreground=self._COLOR_MUTED,
            font=("Segoe UI", 9),
        )

        # Primary action — REAL accent blue. sv_ttk renders its blue via a
        # custom layout element tied to the style name "Accent.TButton";
        # plain background config on a TButton is ignored by the theme. So
        # we clone Accent's layout onto our historical "Primary.TButton"
        # name — every existing `style="Primary.TButton"` button now goes
        # blue with zero call-site churn.
        try:
            style.layout("Primary.TButton", style.layout("Accent.TButton"))
            style.configure("Primary.TButton", font=("Segoe UI Semibold", 10))
        except tk.TclError:
            # Fallback theme (no Accent): approximate with colour config.
            style.configure(
                "Primary.TButton", font=("Segoe UI Semibold", 10),
                foreground="white", background=self._COLOR_PRIMARY,
            )
            try:
                style.map("Primary.TButton",
                    background=[("active", self._COLOR_PRIMARY_HOVER)])
            except tk.TclError:
                pass

        # Danger — Stop / destructive controls. sv_ttk has no red-fill
        # button and ignores TButton background, so we signal danger with
        # bold RED TEXT on the standard button chrome (always visible,
        # unlike white-on-grey when a bg override silently fails).
        style.configure(
            "Danger.TButton",
            font=("Segoe UI Semibold", 10),
            foreground=self._COLOR_DANGER,
        )
        try:
            style.map("Danger.TButton",
                foreground=[
                    ("active", "#A82A1C"),
                    ("disabled", "#C9A8A4"),
                ])
        except tk.TclError:
            pass

        # Status chips
        style.configure(
            "Success.TLabel",
            foreground=self._COLOR_SUCCESS,
            font=("Segoe UI Semibold", 10),
        )
        style.configure(
            "Warning.TLabel",
            foreground=self._COLOR_WARNING,
            font=("Segoe UI Semibold", 10),
        )
        style.configure(
            "Error.TLabel",
            foreground=self._COLOR_DANGER,
            font=("Segoe UI Semibold", 10),
        )

    # ------------------------------------------------------------------
    # Layout helpers
    # ------------------------------------------------------------------

    class _Tooltip:
        """Lightweight hover tooltip for any widget.

        Shows after a short delay so accidental cursor passes don't
        flash the popup; hides on leave or click. The popup is a
        borderless `Toplevel` so it floats above the main window
        without altering layout.
        """

        DELAY_MS = 350
        WRAP_WIDTH = 320

        def __init__(
            self,
            widget: tk.Widget,
            text: str,
            *,
            background: str = "#1f2937",   # slate-800
            foreground: str = "#f8fafc",   # slate-50
        ) -> None:
            self.widget = widget
            self.text = text
            self.background = background
            self.foreground = foreground
            self._tip: tk.Toplevel | None = None
            self._after_id: str | None = None
            widget.bind("<Enter>", self._on_enter, add="+")
            widget.bind("<Leave>", self._on_leave, add="+")
            widget.bind("<ButtonPress>", self._on_leave, add="+")

        def _on_enter(self, _event=None) -> None:
            self._cancel_pending()
            try:
                self._after_id = self.widget.after(self.DELAY_MS, self._show)
            except tk.TclError:
                pass

        def _on_leave(self, _event=None) -> None:
            self._cancel_pending()
            self._hide()

        def _cancel_pending(self) -> None:
            if self._after_id is not None:
                try:
                    self.widget.after_cancel(self._after_id)
                except tk.TclError:
                    pass
                self._after_id = None

        def _show(self) -> None:
            if self._tip is not None:
                return
            try:
                x = self.widget.winfo_rootx() + 18
                y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
            except tk.TclError:
                return
            tip = tk.Toplevel(self.widget)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{x}+{y}")
            tip.configure(background=self.background)
            frame = tk.Frame(
                tip, background=self.background,
                padx=10, pady=6, borderwidth=0,
            )
            frame.pack()
            label = tk.Label(
                frame, text=self.text, justify="left",
                background=self.background, foreground=self.foreground,
                font=("Segoe UI", 9), wraplength=self.WRAP_WIDTH,
            )
            label.pack()
            self._tip = tip

        def _hide(self) -> None:
            if self._tip is not None:
                try:
                    self._tip.destroy()
                except tk.TclError:
                    pass
                self._tip = None

    def _attach_tooltip(self, widget: tk.Widget, text: str) -> None:
        """Public sugar — attach a tooltip without managing the instance."""
        App._Tooltip(widget, text)

    def _attach_log_placeholder(self, text_widget: tk.Text, hint: str) -> None:
        """Render ghost text in an empty log widget; vanish on first real line.

        Tk's `Text` has no native placeholder, so we paint the hint with
        a muted tag at insert-position 1.0 and listen for any user-driven
        insertion to clear it. The widget's normal callers (which all
        unlock state="normal" before .insert("end", ...)) blow this away
        the moment the run actually starts.
        """
        text_widget.tag_configure(
            "placeholder", foreground=self._COLOR_MUTED,
            font=("Segoe UI", 9, "italic"),
        )
        text_widget.configure(state="normal")
        text_widget.insert("1.0", hint, ("placeholder",))
        text_widget.configure(state="disabled")
        # Inserting the placeholder above flips Tk's modified flag to true.
        # Reset it so the FIRST real write produces a fresh false→true
        # <<Modified>> transition — otherwise the listener never fires.
        try:
            text_widget.edit_modified(False)
        except tk.TclError:
            pass

        # Holds the <<Modified>> binding id so we can detach once the
        # placeholder is cleared. Tk's `Text` has no `event_remove`; the
        # correct way to stop listening is to unbind the handler.
        modified_funcid: list[str] = []

        def _clear_on_first_real_write(*_args) -> None:
            content = text_widget.get("1.0", "end-1c").strip()
            if content == hint.strip():
                return
            # Remove the placeholder line if it's still the first line.
            first_line = text_widget.get("1.0", "1.end")
            if first_line.strip() == hint.split("\n", 1)[0].strip():
                text_widget.configure(state="normal")
                text_widget.delete("1.0", "2.0")
                text_widget.configure(state="disabled")
            # Placeholder handled — stop reacting to every subsequent edit.
            if modified_funcid:
                try:
                    text_widget.unbind("<<Modified>>", modified_funcid[0])
                except tk.TclError:
                    pass

        # The standard `<<Modified>>` virtual event fires on every edit.
        def _on_modified(_event) -> None:
            try:
                text_widget.edit_modified(False)
            except tk.TclError:
                return
            _clear_on_first_real_write()

        modified_funcid.append(text_widget.bind("<<Modified>>", _on_modified))

    @staticmethod
    def _make_scrollable(parent: ttk.Frame) -> ttk.Frame:
        """Wrap `parent` in a vertical scrollable Canvas and return the inner Frame.

        Widgets added to the returned frame scroll vertically when content
        overflows the visible area. Mouse-wheel events are captured so
        users don't have to grab the scrollbar.
        """
        outer = ttk.Frame(parent)
        outer.pack(fill="both", expand=True)
        canvas = tk.Canvas(outer, borderwidth=0, highlightthickness=0)
        scroll = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        inner = ttk.Frame(canvas)
        window_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_configure(_event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(event):
            # Make the inner frame at least as wide as the canvas viewport,
            # so child widgets that use `fill="x"` actually stretch.
            canvas.itemconfigure(window_id, width=event.width)

        inner.bind("<Configure>", _on_inner_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        # Mouse-wheel scroll while pointer is over the canvas.
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-event.delta / 120), "units")

        def _bind_wheel(_e):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_wheel(_e):
            canvas.unbind_all("<MouseWheel>")

        canvas.bind("<Enter>", _bind_wheel)
        canvas.bind("<Leave>", _unbind_wheel)
        inner.bind("<Enter>", _bind_wheel)
        inner.bind("<Leave>", _unbind_wheel)

        return inner

    def _section(
        self,
        parent: tk.Widget,
        title: str,
        help_text: str = "",
    ) -> ttk.Frame:
        """Section heading + separator + inner content frame.

        When `help_text` is non-empty, an ⓘ icon is rendered next to the
        title; hovering it shows the help text as a tooltip. This keeps
        the visual surface clean while preserving the long-form
        explanation for users who want it.

        Returns the inner frame. Caller should `.pack(...)` widgets into it.
        """
        wrapper = ttk.Frame(parent)
        wrapper.pack(fill="x", pady=(8, 4), padx=4)
        header = ttk.Frame(wrapper)
        header.pack(fill="x")
        ttk.Label(header, text=title, style="Section.TLabel").pack(side="left")
        if help_text:
            info = ttk.Label(
                header, text="ⓘ",
                foreground=self._COLOR_MUTED,
                font=("Segoe UI", 10),
                cursor="question_arrow",
            )
            info.pack(side="left", padx=(6, 0))
            self._attach_tooltip(info, help_text)
        ttk.Separator(wrapper, orient="horizontal").pack(fill="x", pady=(2, 6))
        body = ttk.Frame(wrapper)
        body.pack(fill="x")
        return body

    @staticmethod
    def _form_row(
        parent: ttk.Frame,
        row: int,
        label: str,
        widget: tk.Widget,
        *,
        suffix: tk.Widget | None = None,
        label_width: int = 16,
    ) -> None:
        """Add a `Label: [widget]` row to a 2- or 3-column grid in `parent`."""
        ttk.Label(parent, text=label, width=label_width, anchor="w").grid(
            row=row, column=0, sticky="w", padx=(2, 8), pady=4
        )
        widget.grid(row=row, column=1, sticky="ew", padx=(0, 4), pady=4)
        if suffix is not None:
            suffix.grid(row=row, column=2, padx=(4, 2), pady=4)
        parent.columnconfigure(1, weight=1)

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _build_ui(self) -> None:
        # Bottom status bar — packed first so the Notebook fills the rest.
        status_frm = ttk.Frame(self.root, relief="sunken", borderwidth=1)
        status_frm.pack(side="bottom", fill="x")

        self.status_user_label = ttk.Label(status_frm, text="Not signed in")
        self.status_user_label.pack(side="left", padx=8, pady=2)

        self.status_version_label = ttk.Label(
            status_frm, text=f"v{__version__}", foreground="#475569"
        )
        self.status_version_label.pack(side="right", padx=(0, 8), pady=2)

        self.btn_check_updates = ttk.Button(
            status_frm,
            text="Check for updates",
            command=self._on_check_for_updates,
        )
        self.btn_check_updates.pack(side="right", padx=4, pady=2)

        # Theme toggle — flips sv_ttk between light/dark with one click.
        # Defaults to light, matches what `_setup_styles` applied.
        self._theme_is_dark = False
        self.btn_theme = ttk.Button(
            status_frm, text="☾  Dark", command=self._toggle_theme, width=10,
        )
        self.btn_theme.pack(side="right", padx=4, pady=2)

        # Two mutually-exclusive auth buttons sitting in the same slot —
        # whichever is appropriate gets packed in `_refresh_user_label`.
        self.btn_sign_out = ttk.Button(
            status_frm, text="Sign out", command=self._on_sign_out,
        )
        self.btn_sign_in = ttk.Button(
            status_frm, text="Sign in", command=self._on_sign_in,
        )

        # Wrap everything in a Notebook so each tool gets its own tab.
        notebook = ttk.Notebook(self.root)
        self.outer_notebook = notebook
        notebook.pack(fill="both", expand=True, padx=8, pady=8)

        iata_tab = ttk.Frame(notebook)
        bd_tab = ttk.Frame(notebook)
        traffic_tab = ttk.Frame(notebook)
        zenith_tab = ttk.Frame(notebook)
        mailer_tab = ttk.Frame(notebook)
        # Tab text is a base name; `_refresh_tab_labels` appends running
        # state ("●") and cache counts so users see status at a glance.
        # BD Overseas Movement (OEP) now lives as a sub-tab under Traffic
        # Movement, alongside the air-traffic sources.
        self._tab_base_labels = {
            iata_tab: "IATA Code Validator",
            bd_tab: "BD Travel Agency Lookup",
            traffic_tab: "Traffic Movement",
            zenith_tab: "Zenith",
            mailer_tab: "Bulk Mailer",
        }
        notebook.add(iata_tab, text=self._tab_base_labels[iata_tab])
        notebook.add(bd_tab, text=self._tab_base_labels[bd_tab])
        notebook.add(traffic_tab, text=self._tab_base_labels[traffic_tab])
        notebook.add(zenith_tab, text=self._tab_base_labels[zenith_tab])
        notebook.add(mailer_tab, text=self._tab_base_labels[mailer_tab])
        self._tab_widgets = {
            "iata": iata_tab, "bd": bd_tab, "traffic": traffic_tab,
            "zenith": zenith_tab, "mailer": mailer_tab,
        }

        self._build_iata_tab(iata_tab)
        self._build_bd_tab(bd_tab)
        self._build_traffic_tab(traffic_tab)  # builds Air Traffic + OEP sub-tabs
        self._build_zenith_tab(zenith_tab)
        self._build_mailer_tab(mailer_tab)

    # ==================================================================
    # Traffic Movement tab — air traffic (multi-source) + BD overseas labour
    # ==================================================================

    def _build_traffic_tab(self, parent: ttk.Frame) -> None:
        """Parent 'Traffic Movement' tab. Two sub-tabs, both movement data:
        'Air Traffic' (multi-source air passenger/movement) and 'BD Overseas
        Movement' (the OEP labour-clearance view), unified under one tab."""
        inner_nb = ttk.Notebook(parent)
        inner_nb.pack(fill="both", expand=True)
        air_inner = ttk.Frame(inner_nb)
        oep_inner = ttk.Frame(inner_nb)
        inner_nb.add(air_inner, text="Air Traffic")
        inner_nb.add(oep_inner, text="BD Overseas Movement")
        self._build_air_traffic_subtab(air_inner)
        self._build_oep_tab(oep_inner)

    def _build_air_traffic_subtab(self, parent: ttk.Frame) -> None:
        parent = self._make_scrollable(parent)

        self._section(
            parent,
            "Air traffic & passenger movement  ·  multi-source",
            help_text=(
                "Pull passenger / movement data from open aviation sources into "
                "one normalized table. Pick a source, optionally a date range, a "
                "view, then Run. Export to Excel. Each source plugs in behind one "
                "interface — more get added over time."
            ),
        )

        body = self._section(parent, "Source & filters")
        self._traffic_label_to_id = {s.label: s.id for s in TRAFFIC_SOURCES.values()}
        self.traffic_source_combo = ttk.Combobox(
            body, textvariable=self.traffic_source, state="readonly",
            values=list(self._traffic_label_to_id.keys()), width=46,
        )
        self.traffic_source_combo.bind(
            "<<ComboboxSelected>>", lambda _e: self._traffic_on_source_change()
        )
        self._form_row(body, 0, "Source:", self.traffic_source_combo)

        date_row = ttk.Frame(body)
        ttk.Label(date_row, text="From:", width=6, anchor="w").pack(side="left")
        ttk.Entry(date_row, textvariable=self.traffic_date_from, width=12).pack(
            side="left", padx=(0, 12))
        ttk.Label(date_row, text="To:", width=4, anchor="w").pack(side="left")
        ttk.Entry(date_row, textvariable=self.traffic_date_to, width=12).pack(
            side="left", padx=(0, 12))
        ttk.Label(date_row, text="(YYYY-MM or YYYY-MM-DD — blank = all)",
                  style="Hint.TLabel").pack(side="left")
        self._form_row(body, 1, "Date range:", date_row)

        file_row = ttk.Frame(body)
        self.traffic_file_entry = ttk.Entry(file_row, textvariable=self.traffic_csv_path)
        self.traffic_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.btn_traffic_browse = ttk.Button(
            file_row, text="Browse…", command=self._traffic_browse_csv, width=10)
        self.btn_traffic_browse.pack(side="left")
        self._form_row(body, 2, "Data file:", file_row)

        body2 = self._section(parent, "View")
        for value, label in (
            ("country", "By country"),
            ("airport", "By airport"),
            ("route", "By route (origin → destination)"),
            ("period", "By period (time series)"),
        ):
            ttk.Radiobutton(
                body2, text=label, variable=self.traffic_view, value=value,
            ).pack(anchor="w", padx=2, pady=1)

        body3 = self._section(parent, "Output")
        out_row = ttk.Frame(body3)
        ttk.Entry(out_row, textvariable=self.traffic_output_dir).pack(
            side="left", fill="x", expand=True, padx=(0, 8))
        ttk.Button(out_row, text="Browse…", command=self._traffic_browse_out,
                   width=10).pack(side="left")
        self._form_row(body3, 0, "Folder:", out_row)

        ctrl = ttk.Frame(parent)
        ctrl.pack(fill="x", pady=(8, 4), padx=4)
        self.btn_traffic_run = ttk.Button(
            ctrl, text="Run traffic report", command=self._traffic_run,
            style="Primary.TButton")
        self.btn_traffic_run.pack(side="left")
        self.btn_traffic_export = ttk.Button(
            ctrl, text="Export to Excel...", command=self._traffic_export,
            state="disabled")
        self.btn_traffic_export.pack(side="left", padx=(8, 0))
        self.traffic_status_label = ttk.Label(
            ctrl, text="Pick a source and view, then Run.", style="Hint.TLabel")
        self.traffic_status_label.pack(side="left", padx=12)

        body4 = self._section(parent, "Results")
        self.traffic_tree_frame = ttk.Frame(body4)
        self.traffic_tree_frame.pack(fill="both", expand=True)
        self.traffic_tree = ttk.Treeview(
            self.traffic_tree_frame, show="headings", height=14)
        self.traffic_tree.pack(side="left", fill="both", expand=True)
        t_scroll = ttk.Scrollbar(self.traffic_tree_frame, command=self.traffic_tree.yview)
        t_scroll.pack(side="right", fill="y")
        self.traffic_tree.configure(yscrollcommand=t_scroll.set)
        self._traffic_set_columns([("#", 40), ("Info", 600)])
        self.traffic_tree.insert("", "end", values=("—", "Pick a source and click Run."))

        self._traffic_on_source_change()

    def _traffic_current_source(self):
        sid = self._traffic_label_to_id.get(self.traffic_source.get())
        return TRAFFIC_SOURCES.get(sid) if sid else None

    def _traffic_on_source_change(self) -> None:
        src = self._traffic_current_source()
        state = "normal" if getattr(src, "needs_file", False) else "disabled"
        try:
            self.traffic_file_entry.configure(state=state)
            self.btn_traffic_browse.configure(state=state)
        except tk.TclError:
            pass

    def _traffic_browse_csv(self) -> None:
        path = filedialog.askopenfilename(
            title="Pick the data CSV (e.g. BTS T-100 Segment)",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if path:
            self.traffic_csv_path.set(path)

    def _traffic_browse_out(self) -> None:
        d = filedialog.askdirectory(title="Output folder")
        if d:
            self.traffic_output_dir.set(d)

    def _traffic_set_columns(self, cols: "list[tuple[str, int]]") -> None:
        ids = [c[0] for c in cols]
        self.traffic_tree.configure(columns=ids)
        for col_id, width in cols:
            self.traffic_tree.heading(col_id, text=col_id)
            anchor = "e" if any(
                t in col_id.lower() for t in ("value", "%", "rank", "#")
            ) else "w"
            self.traffic_tree.column(col_id, width=width, anchor=anchor, stretch=True)

    def _traffic_run(self) -> None:
        if self._traffic_worker is not None and self._traffic_worker.is_alive():
            return
        src = self._traffic_current_source()
        if src is None:
            messagebox.showerror("Traffic", "Pick a data source first.")
            return
        if getattr(src, "needs_file", False) and not self.traffic_csv_path.get().strip():
            messagebox.showerror(
                "Traffic",
                f"{src.label} needs a data file — click Browse and pick the CSV.",
            )
            return
        cfg = {
            "source_id": src.id,
            "source_label": src.label,
            "date_from": self.traffic_date_from.get().strip(),
            "date_to": self.traffic_date_to.get().strip(),
            "csv_path": self.traffic_csv_path.get().strip(),
            "view": self.traffic_view.get(),
        }
        self.btn_traffic_run.configure(state="disabled")
        self.btn_traffic_export.configure(state="disabled")
        self.traffic_status_label.configure(text="Fetching…")
        self._traffic_worker = threading.Thread(
            target=self._traffic_worker_run, args=(cfg,), daemon=True)
        self._traffic_worker.start()

    def _traffic_worker_run(self, cfg: dict) -> None:
        try:
            src = TRAFFIC_SOURCES[cfg["source_id"]]
            session = traffic_client.build_session()
            filters = {
                "date_from": cfg["date_from"], "date_to": cfg["date_to"],
                "csv_path": cfg["csv_path"],
            }

            def progress(_i, _total, label):
                self._post(MSG_TRAFFIC_BUSY, str(label))

            rows = src.fetch(filters, session=session, progress_cb=progress)
            if not rows:
                self._post(MSG_TRAFFIC_ERROR, "No data returned for that source / filter.")
                return
            self._post(MSG_TRAFFIC_DONE, {
                "rows": rows, "view": cfg["view"],
                "source_label": cfg["source_label"],
                "date_from": cfg["date_from"], "date_to": cfg["date_to"],
            })
        except Exception as exc:  # noqa: BLE001
            log.exception("Traffic worker crashed")
            self._post(MSG_TRAFFIC_ERROR, f"{type(exc).__name__}: {exc}")

    def _traffic_render_results(self, result: dict) -> None:
        self._traffic_last_result = result
        rows = result["rows"]
        view = result["view"]
        self.traffic_tree.delete(*self.traffic_tree.get_children())

        def _share(v, total):
            return f"{(100.0 * v / total):.1f}" if total else "0.0"

        def _fmt(v):
            return f"{v:,.0f}"

        def _totals(agg):
            tot: dict = {}
            for t in agg:
                tot[t.metric] = tot.get(t.metric, 0.0) + t.value
            return tot

        if view == "route":
            agg = traffic_client.aggregate_by_route(rows)
            tot = _totals(agg)
            self._traffic_set_columns([
                ("Rank", 50), ("Origin", 90), ("Destination", 90),
                ("Metric", 110), ("Value", 120), ("Share %", 80)])
            for i, t in enumerate(agg, start=1):
                self.traffic_tree.insert("", "end", values=(
                    i, t.origin, t.destination, t.metric, _fmt(t.value),
                    _share(t.value, tot.get(t.metric, 0))))
        elif view == "airport":
            agg = traffic_client.aggregate_by_airport(rows)
            tot = _totals(agg)
            self._traffic_set_columns([
                ("Rank", 50), ("Airport", 110), ("Metric", 120),
                ("Value", 130), ("Share %", 80)])
            for i, t in enumerate(agg, start=1):
                self.traffic_tree.insert("", "end", values=(
                    i, t.airport, t.metric, _fmt(t.value),
                    _share(t.value, tot.get(t.metric, 0))))
        elif view == "period":
            agg = traffic_client.aggregate_by_period(rows)
            self._traffic_set_columns([
                ("Period", 120), ("Metric", 140), ("Value", 150)])
            for t in agg:
                self.traffic_tree.insert("", "end", values=(
                    t.period, t.metric, _fmt(t.value)))
        else:  # country
            agg = traffic_client.aggregate_by_country(rows)
            tot = _totals(agg)
            self._traffic_set_columns([
                ("Rank", 50), ("Country", 200), ("Metric", 120),
                ("Value", 130), ("Share %", 80)])
            for i, t in enumerate(agg, start=1):
                self.traffic_tree.insert("", "end", values=(
                    i, t.country, t.metric, _fmt(t.value),
                    _share(t.value, tot.get(t.metric, 0))))

        self.btn_traffic_run.configure(state="normal")
        self.btn_traffic_export.configure(state="normal")
        self.traffic_status_label.configure(
            text=f"{len(rows):,} rows · {result['source_label']}")

    def _traffic_export(self) -> None:
        if not self._traffic_last_result:
            messagebox.showinfo("Traffic", "Run a report first.")
            return
        view = self._traffic_last_result["view"]
        folder = Path(self.traffic_output_dir.get().strip() or str(Path.home()))
        default = excel_io.build_traffic_output_path(folder, view)
        path = filedialog.asksaveasfilename(
            title="Export Traffic report", initialdir=str(folder),
            initialfile=default.name, defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return
        try:
            excel_io.write_traffic_report(
                Path(path),
                source_label=self._traffic_last_result["source_label"],
                date_from=self._traffic_last_result["date_from"],
                date_to=self._traffic_last_result["date_to"],
                view=view,
                rows=self._traffic_last_result["rows"],
            )
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Traffic", f"Export failed: {exc}")
            return
        messagebox.showinfo("Traffic", f"Saved:\n{path}")

    # ==================================================================
    # Bulk Mailer tab
    # ==================================================================

    def _build_mailer_tab(self, parent: ttk.Frame) -> None:
        parent = self._make_scrollable(parent)

        intro = self._section(parent, "Bulk Mailer")
        ttk.Label(
            intro, style="Hint.TLabel", justify="left", wraplength=900,
            text=(
                "Send one personalised email per recipient from a mapping "
                "Excel (Email · Name · File · optional CC · BCC). Files are "
                "attached from a folder you pick. Send via Microsoft 365 "
                "sign-in (Graph — works without desktop Outlook), SMTP, or "
                "Outlook desktop. Default mode creates DRAFTS you review first.\n"
                "No mapping yet? Click 'Use latest report-pack mail list' below — "
                "it loads the recipient list the Target Packs generate."
            ),
        ).pack(anchor="w", padx=4, pady=(0, 4))

        # ----- Inputs -----
        io_body = self._section(parent, "Mapping + attachments")
        map_entry = ttk.Entry(io_body, textvariable=self.mail_mapping_path)
        self._form_row(
            io_body, 0, "Mapping Excel:", map_entry,
            suffix=ttk.Button(io_body, text="Browse...", command=self._mail_pick_mapping),
        )
        attach_entry = ttk.Entry(io_body, textvariable=self.mail_attach_dir)
        self._form_row(
            io_body, 1, "Attachments folder:", attach_entry,
            suffix=ttk.Button(io_body, text="Browse...", command=self._mail_pick_attach),
        )
        # one-click mail-list helpers: load the pack-generated list (Email|Name|File|
        # CC|BCC) + its attachments folder, or open the loaded mapping to edit by hand.
        helper = ttk.Frame(io_body)
        helper.grid(row=2, column=1, sticky="w", padx=2, pady=(4, 0))
        ttk.Button(helper, text="Use latest report-pack mail list",
                   command=self._mail_use_pack_list).pack(side="left")
        ttk.Button(helper, text="Edit in Excel",
                   command=self._mail_edit_mapping).pack(side="left", padx=(8, 0))

        # ----- Message -----
        msg_body = self._section(parent, "Message")
        subj_entry = ttk.Entry(msg_body, textvariable=self.mail_subject)
        self._form_row(msg_body, 0, "Subject:", subj_entry)
        ttk.Label(
            msg_body, text="Body (use {name} or any column as a placeholder):",
            style="Hint.TLabel",
        ).grid(row=1, column=0, columnspan=3, sticky="w", padx=2, pady=(6, 2))
        self.mail_body_text = tk.Text(msg_body, height=8, wrap="word")
        self.mail_body_text.grid(row=2, column=0, columnspan=3, sticky="ew", padx=2)
        self.mail_body_text.insert("1.0", (
            "Dear {name},\n\n"
            "Please find attached your report.\n\n"
            "Best regards,\n"
        ))
        msg_body.columnconfigure(0, weight=1)

        # ----- Transport (how the mail is sent) -----
        tx = self._section(parent, "Send via")
        trow = ttk.Frame(tx)
        trow.pack(fill="x", padx=2, pady=(0, 4))
        ttk.Radiobutton(
            trow, text="Microsoft 365 sign-in (Graph)", value="graph",
            variable=self.mail_transport, command=self._mail_sync_transport,
        ).pack(side="left", padx=(0, 12))
        ttk.Radiobutton(
            trow, text="SMTP (Gmail / Workspace / any)",
            value="smtp", variable=self.mail_transport,
            command=self._mail_sync_transport,
        ).pack(side="left", padx=(0, 12))
        ttk.Radiobutton(
            trow, text="Outlook desktop", value="outlook",
            variable=self.mail_transport, command=self._mail_sync_transport,
        ).pack(side="left")

        # Graph sign-in block — shown when transport == graph.
        self.mail_graph_frame = ttk.Frame(tx)
        gr = self.mail_graph_frame
        ttk.Label(
            gr, style="Hint.TLabel", justify="left", wraplength=820,
            text=(
                "Sends as your Microsoft 365 address over HTTPS — no SMTP, "
                "no desktop Outlook. Click Sign in, then enter the code in "
                "your browser (one-time; MFA supported). Note: some tenants "
                "require admin approval for this."
            ),
        ).pack(anchor="w", pady=(0, 4))
        grow = ttk.Frame(gr)
        grow.pack(fill="x")
        self.btn_mail_graph_signin = ttk.Button(
            grow, text="Sign in to Microsoft 365", command=self._mail_graph_signin,
        )
        self.btn_mail_graph_signin.pack(side="left")
        ttk.Label(grow, textvariable=self.mail_graph_status, style="Hint.TLabel").pack(
            side="left", padx=(10, 0),
        )
        ttk.Button(
            grow, text="Sign out", command=self._mail_graph_signout,
        ).pack(side="left", padx=(10, 0))

        # SMTP credential block — shown only when transport == smtp.
        self.mail_smtp_frame = ttk.Frame(tx)
        self.mail_smtp_frame.pack(fill="x", padx=2)
        sf = self.mail_smtp_frame
        ttk.Label(sf, text="Provider:", width=14, anchor="w").grid(row=0, column=0, sticky="w", pady=3)
        preset_cb = ttk.Combobox(
            sf, textvariable=self.mail_smtp_preset, state="readonly",
            values=list(mailer_client.SMTP_PRESETS.keys()), width=30,
        )
        preset_cb.grid(row=0, column=1, sticky="w", pady=3)
        preset_cb.bind("<<ComboboxSelected>>", lambda _e: self._mail_apply_preset())
        ttk.Label(sf, text="Host:", width=14, anchor="w").grid(row=1, column=0, sticky="w", pady=3)
        ttk.Entry(sf, textvariable=self.mail_smtp_host, width=32).grid(row=1, column=1, sticky="w", pady=3)
        ttk.Label(sf, text="Port:").grid(row=1, column=2, sticky="e", padx=(12, 4))
        ttk.Entry(sf, textvariable=self.mail_smtp_port, width=7).grid(row=1, column=3, sticky="w")
        ttk.Label(sf, text="From (email):", width=14, anchor="w").grid(row=2, column=0, sticky="w", pady=3)
        sender_entry = ttk.Entry(sf, textvariable=self.mail_smtp_sender, width=32)
        sender_entry.grid(row=2, column=1, sticky="w", pady=3)
        sender_entry.bind("<FocusOut>", lambda _e: self._mail_load_saved_password())
        ttk.Button(
            sf, text="Auto-detect host", command=self._mail_autodetect_host,
        ).grid(row=2, column=2, columnspan=2, sticky="w", padx=(12, 0))
        ttk.Label(sf, text="Password:", width=14, anchor="w").grid(row=3, column=0, sticky="w", pady=3)
        ttk.Entry(sf, textvariable=self.mail_smtp_password, show="•", width=32).grid(row=3, column=1, sticky="w", pady=3)
        ttk.Checkbutton(
            sf, text="Remember (Credential Manager)", variable=self.mail_smtp_remember,
        ).grid(row=3, column=2, columnspan=2, sticky="w", padx=(12, 0))
        self.mail_smtp_hint = ttk.Label(
            sf, style="Hint.TLabel", wraplength=820, justify="left",
            text=(
                "Gmail / Workspace & Office 365 require an APP PASSWORD "
                "(enable 2-step verification, then create a 16-char app "
                "password). Your normal login password will be rejected."
            ),
        )
        self.mail_smtp_hint.grid(row=4, column=0, columnspan=4, sticky="w", pady=(2, 2))

        # Outlook account picker — shown only when transport == outlook.
        self.mail_outlook_frame = ttk.Frame(tx)
        ttk.Label(
            self.mail_outlook_frame, text="Send from account:", width=16, anchor="w",
        ).pack(side="left")
        self.mail_outlook_combo = ttk.Combobox(
            self.mail_outlook_frame, textvariable=self.mail_outlook_account,
            state="readonly", width=36, values=[],
        )
        self.mail_outlook_combo.pack(side="left", padx=(4, 8))
        self.mail_outlook_combo.bind(
            "<<ComboboxSelected>>",
            lambda _e: self._mail_save_last_outlook_account(
                self.mail_outlook_account.get()),
        )
        ttk.Button(
            self.mail_outlook_frame, text="Refresh accounts",
            command=self._mail_refresh_outlook_accounts,
        ).pack(side="left")

        # ----- Options -----
        opt = ttk.Frame(parent)
        opt.pack(fill="x", padx=6, pady=(8, 0))
        ttk.Label(opt, text="Mode:").pack(side="left")
        ttk.Radiobutton(
            opt, text="Create drafts (review first)", value="draft",
            variable=self.mail_mode,
        ).pack(side="left", padx=(4, 10))
        ttk.Radiobutton(
            opt, text="Send now", value="send", variable=self.mail_mode,
        ).pack(side="left", padx=(0, 16))
        ttk.Checkbutton(
            opt, text="Skip already-sent rows", variable=self.mail_skip_sent,
        ).pack(side="left", padx=(0, 16))
        ttk.Label(opt, text="Delay (s):").pack(side="left")
        ttk.Spinbox(
            opt, from_=0.0, to=10.0, increment=0.5, width=5,
            textvariable=self.mail_delay_s,
        ).pack(side="left", padx=(4, 0))

        # ----- Actions -----
        act = ttk.Frame(parent)
        act.pack(fill="x", padx=6, pady=(8, 0))
        self.btn_mail_preview = ttk.Button(
            act, text="Load + preview", command=self._mail_preview,
        )
        self.btn_mail_preview.pack(side="left")
        self.btn_mail_test = ttk.Button(
            act, text="Send test to myself", command=self._mail_test,
            state="disabled",
        )
        self.btn_mail_test.pack(side="left", padx=(8, 0))
        self.btn_mail_run = ttk.Button(
            act, text="Create drafts", style="Primary.TButton",
            command=self._mail_run, state="disabled",
        )
        self.btn_mail_run.pack(side="left", padx=(8, 0))
        self.btn_mail_stop = ttk.Button(
            act, text="Stop", command=self._mail_stop,
            state="disabled", style="Danger.TButton",
        )
        self.btn_mail_stop.pack(side="left", padx=(8, 0))
        self.mail_status = ttk.Label(act, text="Idle.", style="Hint.TLabel")
        self.mail_status.pack(side="left", padx=(12, 0))
        # Keep the run button's label in sync with the chosen mode.
        self.mail_mode.trace_add("write", lambda *_a: self._mail_sync_run_label())

        # ----- Preview grid -----
        prev = self._section(parent, "Preview")
        grid = ttk.Frame(prev)
        grid.pack(fill="both", expand=True)
        cols = ("row", "email", "name", "files", "cc", "bcc", "status")
        self.mail_tree = ttk.Treeview(
            grid, columns=cols, show="headings", height=10,
        )
        for cid, txt, w in (
            ("row", "#", 36), ("email", "Email", 200), ("name", "Name", 130),
            ("files", "Attachment(s)", 220), ("cc", "CC", 120),
            ("bcc", "BCC", 120), ("status", "Status", 130),
        ):
            self.mail_tree.heading(cid, text=txt)
            self.mail_tree.column(cid, width=w, anchor="w")
        self.mail_tree.tag_configure("bad", background=self._COLOR_ROW_BAD)
        self.mail_tree.tag_configure("ok", background=self._COLOR_ROW_GOOD)
        self.mail_tree.pack(side="left", fill="both", expand=True, padx=(2, 0), pady=2)
        msb = ttk.Scrollbar(grid, command=self.mail_tree.yview)
        msb.pack(side="left", fill="y")
        self.mail_tree.configure(yscrollcommand=msb.set)

        # ----- Progress -----
        prog = self._section(parent, "Progress")
        self.mail_progress = ttk.Progressbar(prog, mode="determinate", maximum=1)
        self.mail_progress.pack(fill="x", padx=2, pady=2)

        # Show the credential block matching the default transport — layout only;
        # initial=True so we DON'T touch Outlook COM / MSAL at startup (that pops a
        # Microsoft sign-in dialog before the user has even opened the mailer).
        self._mail_sync_transport(initial=True)

    def _mail_sync_run_label(self) -> None:
        self.btn_mail_run.configure(
            text="Create drafts" if self.mail_mode.get() == "draft" else "Send now",
        )

    def _mail_sync_transport(self, initial: bool = False) -> None:
        """Show the credential block that matches the chosen transport.

        ``initial=True`` (the single build-time call) ONLY lays out the frames. It must
        not touch Outlook COM or MSAL: doing so during app startup makes Outlook pop a
        Microsoft (login.microsoftonline.com) sign-in dialog before the user has chosen
        to use the mailer at all. The account-refresh / silent-sign-in side effects run
        only when the user actually selects a transport (a deliberate radio click).
        """
        t = self.mail_transport.get()
        self.mail_smtp_frame.pack_forget()
        self.mail_outlook_frame.pack_forget()
        self.mail_graph_frame.pack_forget()
        if t == "smtp":
            self.mail_smtp_frame.pack(fill="x", padx=2)
        elif t == "outlook":
            self.mail_outlook_frame.pack(fill="x", padx=2, pady=(2, 4))
            if not initial and not self.mail_outlook_combo.cget("values"):
                self._mail_refresh_outlook_accounts()   # Dispatches Outlook COM — user action only
        else:  # graph
            self.mail_graph_frame.pack(fill="x", padx=2, pady=(2, 4))
            # silent sign-in from cached token so returning users don't re-auth —
            # but only on a real selection, never during startup construction.
            if not initial and self._graph_session is None:
                self._mail_graph_try_silent()

    def _mail_graph_try_silent(self) -> None:
        try:
            sess = graph_mailer.GraphSession.try_silent()
        except Exception:  # noqa: BLE001
            sess = None
        if sess is not None:
            self._graph_session = sess
            self.mail_graph_status.set(f"● Signed in: {sess.account}")

    def _mail_graph_signin(self) -> None:
        """Interactive device-code sign-in on a worker thread."""
        self.btn_mail_graph_signin.configure(state="disabled")
        self.mail_graph_status.set("Starting sign-in…")

        def prompt_cb(message: str, code: str) -> None:
            # Surface the instructions + bare code on the GUI thread.
            self._post("mail_graph_prompt", (message, code))

        def worker() -> None:
            try:
                sess = graph_mailer.GraphSession.sign_in(prompt_cb=prompt_cb)
                self._post("mail_graph_signed_in", sess)
            except Exception as exc:  # noqa: BLE001
                self._post("mail_graph_signin_failed", str(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _mail_graph_signout(self) -> None:
        graph_mailer.GraphSession.sign_out()
        self._graph_session = None
        self.mail_graph_status.set("Not signed in")

    def _mail_outlook_pref_path(self):
        return config.APP_DIR / "mailer_outlook_account.txt"

    def _mail_load_last_outlook_account(self) -> str:
        try:
            p = self._mail_outlook_pref_path()
            return p.read_text(encoding="utf-8").strip() if p.exists() else ""
        except Exception:  # noqa: BLE001
            return ""

    def _mail_save_last_outlook_account(self, address: str) -> None:
        try:
            p = self._mail_outlook_pref_path()
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(address.strip(), encoding="utf-8")
        except Exception:  # noqa: BLE001
            pass

    def _mail_refresh_outlook_accounts(self) -> None:
        accts = mailer_client.list_outlook_accounts()
        self.mail_outlook_combo.configure(values=accts)
        if accts and not self.mail_outlook_account.get():
            # Prefer the last account the user picked (persisted), else the
            # first configured one. The dropdown lets them change it.
            last = self._mail_load_last_outlook_account()
            self.mail_outlook_account.set(last if last in accts else accts[0])
        if not accts:
            self.mail_status.configure(
                text="No Outlook accounts found — is Outlook installed/signed in?",
            )

    def _mail_autodetect_host(self) -> None:
        """Look up the From-domain's MX and fill the SMTP host/preset."""
        sender = self.mail_smtp_sender.get().strip()
        if "@" not in sender:
            messagebox.showinfo(
                "Auto-detect", "Type the From email address first.",
            )
            return
        self.mail_status.configure(text="Detecting mail host…")
        self.root.update_idletasks()
        info = mailer_client.detect_mail_host(sender)
        if not info:
            self.mail_status.configure(text="Couldn't detect — enter host manually.")
            messagebox.showwarning(
                "Auto-detect",
                "Couldn't resolve the mail host for that domain. "
                "Enter the SMTP host manually.",
            )
            return
        self.mail_smtp_preset.set(info["preset"])
        if info["host"]:
            self.mail_smtp_host.set(info["host"])
        self.mail_smtp_port.set(info["port"])
        self.mail_status.configure(text=f"MX: {info['mx']}")
        if info["note"]:
            messagebox.showinfo("Auto-detect", info["note"])

    def _mail_apply_preset(self) -> None:
        preset = self.mail_smtp_preset.get()
        host_port = mailer_client.SMTP_PRESETS.get(preset)
        if host_port and host_port[0]:
            self.mail_smtp_host.set(host_port[0])
            self.mail_smtp_port.set(host_port[1])

    def _mail_load_saved_password(self) -> None:
        """When the sender address loses focus, pull any saved password."""
        sender = self.mail_smtp_sender.get().strip()
        if sender and not self.mail_smtp_password.get():
            saved = mailer_client.load_smtp_password(sender)
            if saved:
                self.mail_smtp_password.set(saved)

    def _mail_smtp_settings(self) -> "mailer_client.SMTPSettings | None":
        """Build SMTPSettings from the form, or None (+ error) if incomplete."""
        host = self.mail_smtp_host.get().strip()
        sender = self.mail_smtp_sender.get().strip()
        pwd = self.mail_smtp_password.get()
        if not host or not sender:
            messagebox.showerror(
                "Bulk Mailer", "Enter the SMTP host and the From email address.",
            )
            return None
        if not pwd:
            messagebox.showerror(
                "Bulk Mailer",
                "Enter the SMTP password (an app password for Gmail/O365).",
            )
            return None
        if self.mail_smtp_remember.get():
            try:
                mailer_client.save_smtp_password(sender, pwd)
            except Exception:  # noqa: BLE001
                log.warning("could not save SMTP password to keyring")
        try:
            port = int(self.mail_smtp_port.get())
        except (tk.TclError, ValueError):
            port = 587
        return mailer_client.SMTPSettings(
            host=host, port=port, sender=sender, password=pwd, use_starttls=True,
        )

    def _mail_pick_mapping(self) -> None:
        f = filedialog.askopenfilename(
            title="Pick the mapping Excel",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if f:
            self.mail_mapping_path.set(f)

    def _mail_pick_attach(self) -> None:
        d = filedialog.askdirectory(title="Pick the attachments folder")
        if d:
            self.mail_attach_dir.set(d)

    def _pack_manifest_candidates(self):
        """Folders where a report-pack Email_Manifest.xlsx may have been written
        (the Sales-Person Target Packs generate one in Email|Name|File|CC|BCC form —
        exactly this mailer's mapping format)."""
        import os
        from pathlib import Path
        local = os.environ.get("LOCALAPPDATA", str(Path.home()))
        return [
            Path.home() / "Downloads" / "Sales Person Packs" / "Email_Manifest.xlsx",
            Path(local) / "USBA_InstantReports" / "work" / "logs"
            / "Sales Person Packs" / "Email_Manifest.xlsx",
        ]

    def _mail_use_pack_list(self) -> None:
        """One click: load the newest report-pack mail list and point the
        attachments folder at its packs — ready to send the salesperson packs."""
        found = [p for p in self._pack_manifest_candidates() if p.is_file()]
        if not found:
            messagebox.showinfo(
                "Mail list",
                "No report-pack mail list found yet.\n\n"
                "Build the Sales-Person Target Packs first (Reports > Build from "
                "data, or the morning publish). They write Email_Manifest.xlsx next "
                "to the per-person files, and it loads here automatically.")
            return
        newest = max(found, key=lambda p: p.stat().st_mtime)
        self.mail_mapping_path.set(str(newest))
        self.mail_attach_dir.set(str(newest.parent))
        messagebox.showinfo(
            "Mail list",
            f"Loaded {newest.name}\nfrom {newest.parent}\n\n"
            "Attachments folder set to the same place. Click 'Load + preview', then send.")

    def _mail_edit_mapping(self) -> None:
        """Open the currently-selected mapping Excel to edit recipients by hand."""
        import os
        from pathlib import Path
        p = self.mail_mapping_path.get().strip()
        if not p or not Path(p).is_file():
            messagebox.showinfo("Mail list", "Pick or load a mapping Excel first, then Edit.")
            return
        if hasattr(os, "startfile"):
            os.startfile(p)            # open the user's own mapping in Excel

    def _mail_preview(self) -> None:
        """Read the mapping, resolve files, fill the preview grid."""
        mapping = self.mail_mapping_path.get().strip()
        attach = self.mail_attach_dir.get().strip()
        if not mapping or not Path(mapping).is_file():
            messagebox.showerror("Bulk Mailer", "Pick a valid mapping Excel first.")
            return
        if not attach or not Path(attach).is_dir():
            messagebox.showerror("Bulk Mailer", "Pick a valid attachments folder first.")
            return
        try:
            rows, warnings = mailer_io.read_mapping(mapping, attach)
        except Exception as exc:  # noqa: BLE001
            log.exception("mapping read failed")
            messagebox.showerror("Bulk Mailer", f"Couldn't read mapping: {exc}")
            return
        self._mail_rows = rows
        for child in self.mail_tree.get_children():
            self.mail_tree.delete(child)
        valid = 0
        for r in rows:
            status = "OK" if r.is_valid else "; ".join(r.issues)
            files = ", ".join(p.name for p in r.attachments) or "(none)"
            self.mail_tree.insert(
                "", "end",
                values=(r.row_index, r.email, r.name, files, r.cc, r.bcc, status),
                tags=("ok",) if r.is_valid else ("bad",),
            )
            if r.is_valid:
                valid += 1
        msg = f"{valid} valid / {len(rows)} rows"
        if warnings:
            msg += "  ·  " + "; ".join(warnings)
        self.mail_status.configure(text=msg)
        ready = "normal" if valid else "disabled"
        self.btn_mail_run.configure(state=ready)
        self.btn_mail_test.configure(state=ready)
        if warnings:
            messagebox.showwarning("Bulk Mailer", "\n".join(warnings))

    def _mail_test(self) -> None:
        """Send/draft the FIRST valid row to the user's own address."""
        me = (self._signed_in_user or {}).get("email", "")
        if not me:
            from tkinter import simpledialog
            me = simpledialog.askstring(
                "Send test", "Send the test email to which address?",
            )
        if not me:
            return
        first = next((r for r in self._mail_rows if r.is_valid), None)
        if first is None:
            messagebox.showinfo("Bulk Mailer", "No valid row to test with.")
            return
        subject = self.mail_subject.get().strip() or "(no subject)"
        body_tmpl = self.mail_body_text.get("1.0", "end").rstrip("\n")
        body, _missing = mailer_io.render_template(body_tmpl, first.fields)
        send = self.mail_mode.get() == "send"
        transport = self.mail_transport.get()
        email = mailer_client.OutgoingEmail(
            to=me, subject=f"[TEST] {subject}", body=body,
            attachments=first.attachments,
        )

        smtp_settings = None
        draft_dir = Path(self.mail_attach_dir.get() or str(Path.home())) / "_mail_drafts"
        if transport == "smtp":
            smtp_settings = self._mail_smtp_settings()
            if smtp_settings is None:
                return
        if transport == "graph" and self._graph_session is None:
            messagebox.showerror(
                "Bulk Mailer", "Sign in to Microsoft 365 first (Send via → Sign in).",
            )
            return
        graph_sess = self._graph_session
        self.mail_status.configure(text=f"Testing to {me}…")

        def worker() -> None:
            try:
                if transport == "outlook":
                    acct = self.mail_outlook_account.get().strip()
                    with mailer_client.OutlookSession() as ol:
                        outcome = ol.create(email, send=send, from_account=acct)
                elif transport == "graph":
                    outcome = (graph_sess.send(email) if send
                               else graph_sess.draft(email))
                else:
                    with mailer_client.SMTPMailer(smtp_settings) as sm:
                        outcome = (sm.send(email) if send
                                   else sm.draft(email, draft_dir))
                self._post(MSG_MAIL_DONE, {
                    "test": True, "outcome_status": outcome.status,
                    "error": outcome.error, "to": me,
                    "entry": outcome.entry_id,
                })
            except Exception as exc:  # noqa: BLE001
                log.exception("test mail failed")
                self._post(MSG_MAIL_ERROR, f"{type(exc).__name__}: {exc}")

        threading.Thread(target=worker, daemon=True).start()

    def _mail_stop(self) -> None:
        self._mail_stop_flag.set()
        self.mail_status.configure(text="Stopping…")

    def _mail_run(self) -> None:
        if self._mail_worker and self._mail_worker.is_alive():
            messagebox.showinfo("Bulk Mailer", "A run is already in progress.")
            return
        valid_rows = [r for r in self._mail_rows if r.is_valid]
        if not valid_rows:
            messagebox.showerror("Bulk Mailer", "Load a mapping and fix invalid rows first.")
            return
        subject = self.mail_subject.get().strip()
        if not subject:
            messagebox.showerror("Bulk Mailer", "Enter a subject.")
            return
        mode = self.mail_mode.get()
        send = mode == "send"
        transport = self.mail_transport.get()
        body_tmpl = self.mail_body_text.get("1.0", "end").rstrip("\n")
        campaign = Path(self.mail_mapping_path.get()).name
        skip_sent = self.mail_skip_sent.get()
        delay = max(0.0, float(self.mail_delay_s.get()))

        # SMTP needs validated settings up front; drafts land as .eml files.
        smtp_settings = None
        draft_dir = Path(self.mail_attach_dir.get() or str(Path.home())) / "_mail_drafts"
        if transport == "smtp":
            smtp_settings = self._mail_smtp_settings()
            if smtp_settings is None:
                return
        if transport == "graph" and self._graph_session is None:
            messagebox.showerror(
                "Bulk Mailer", "Sign in to Microsoft 365 first (Send via → Sign in).",
            )
            return
        graph_sess = self._graph_session

        if send:
            if transport == "smtp":
                where = f"from {smtp_settings.sender}"
            elif transport == "graph":
                where = f"from {graph_sess.account}"
            else:
                where = "via Outlook"
            tail = f" — these will actually be sent {where}."
        else:
            if transport == "smtp":
                tail = f" as .eml drafts in {draft_dir}."
            elif transport == "graph":
                tail = f" as drafts in the {graph_sess.account} mailbox."
            else:
                tail = " as drafts in Outlook."
        if not messagebox.askyesno(
            "Bulk Mailer — confirm",
            f"About to {'SEND' if send else 'create drafts for'} "
            f"{len(valid_rows)} email(s){tail}\n\nSubject: {subject}\n\nContinue?",
        ):
            return

        self._mail_stop_flag.clear()
        self.btn_mail_run.configure(state="disabled")
        self.btn_mail_test.configure(state="disabled")
        self.btn_mail_preview.configure(state="disabled")
        self.btn_mail_stop.configure(state="normal")
        self.mail_progress.configure(value=0, maximum=len(valid_rows))

        # Map row_index → tree iid for live status updates.
        iid_by_row = {}
        for iid in self.mail_tree.get_children():
            vals = self.mail_tree.item(iid, "values")
            iid_by_row[int(vals[0])] = iid

        def _process(make_one) -> dict:
            """Run the shared per-row loop; `make_one(row)->SendOutcome`."""
            import time as _t
            counts = {"DRAFTED": 0, "SENT": 0, "FAILED": 0, "SKIPPED": 0}
            for i, r in enumerate(valid_rows, start=1):
                if self._mail_stop_flag.is_set():
                    break
                if skip_sent and self._mail_log.already_sent(campaign, r.email, subject):
                    counts["SKIPPED"] += 1
                    self._post(MSG_MAIL_PROGRESS,
                               (i, len(valid_rows), r.email, "SKIPPED", r.row_index))
                    continue
                body, _missing = mailer_io.render_template(body_tmpl, r.fields)
                email = mailer_client.OutgoingEmail(
                    to=r.email, subject=subject, body=body,
                    attachments=r.attachments, cc=r.cc, bcc=r.bcc,
                )
                outcome = make_one(email)
                counts[outcome.status] = counts.get(outcome.status, 0) + 1
                if outcome.status == "SENT":
                    self._mail_log.record(campaign, r.email, subject, "SENT")
                elif outcome.status == "FAILED":
                    self._mail_log.record(campaign, r.email, subject, "FAILED", outcome.error)
                self._post(
                    MSG_MAIL_PROGRESS,
                    (i, len(valid_rows), r.email,
                     outcome.status + (f": {outcome.error}" if outcome.error else ""),
                     r.row_index),
                )
                if delay > 0 and i < len(valid_rows):
                    _t.sleep(delay)
            return counts

        def worker() -> None:
            try:
                if transport == "outlook":
                    acct = self.mail_outlook_account.get().strip()
                    with mailer_client.OutlookSession() as ol:
                        log.info("Bulk Mailer via Outlook from %s",
                                 acct or ol.verify_account() or "?")
                        counts = _process(
                            lambda e: ol.create(e, send=send, from_account=acct)
                        )
                elif transport == "graph":
                    log.info("Bulk Mailer via Graph as %s", graph_sess.account)
                    counts = _process(
                        (lambda e: graph_sess.send(e)) if send
                        else (lambda e: graph_sess.draft(e))
                    )
                else:
                    with mailer_client.SMTPMailer(smtp_settings) as sm:
                        log.info("Bulk Mailer via SMTP %s as %s",
                                 smtp_settings.host, smtp_settings.sender)
                        counts = _process(
                            (lambda e: sm.send(e)) if send
                            else (lambda e: sm.draft(e, draft_dir))
                        )
                self._post(MSG_MAIL_DONE, {
                    "test": False, "counts": counts, "mode": mode,
                    "draft_dir": str(draft_dir) if transport == "smtp" and not send else "",
                })
            except (mailer_client.OutlookUnavailableError,
                    mailer_client.SMTPConfigError,
                    mailer_client.SMTPAuthError) as exc:
                self._post(MSG_MAIL_ERROR, str(exc))
            except Exception as exc:  # noqa: BLE001
                log.exception("bulk mail run failed")
                self._post(MSG_MAIL_ERROR, f"{type(exc).__name__}: {exc}")

        self._mail_iid_by_row = iid_by_row
        self._mail_worker = threading.Thread(target=worker, daemon=True)
        self._mail_worker.start()

    def _build_iata_tab(self, parent: ttk.Frame) -> None:
        # ----- Input -----
        body = self._section(parent, "Input Excel")

        entry = ttk.Entry(body, textvariable=self.input_path)
        self._form_row(
            body, 0, "File:",
            entry,
            suffix=ttk.Button(body, text="Browse...", command=self._pick_input),
        )

        self.sheet_combo = ttk.Combobox(body, textvariable=self.sheet_name, state="readonly")
        self.sheet_combo.bind("<<ComboboxSelected>>", lambda _e: self._reload_columns())
        self._form_row(body, 1, "Sheet:", self.sheet_combo)

        self.col_combo = ttk.Combobox(body, textvariable=self.column_name, state="readonly")
        self._form_row(body, 2, "IATA column:", self.col_combo)

        rr = ttk.Frame(body)
        ttk.Label(rr, text="from").pack(side="left")
        ttk.Entry(rr, textvariable=self.start_row, width=7).pack(side="left", padx=(4, 12))
        ttk.Label(rr, text="to").pack(side="left")
        ttk.Entry(rr, textvariable=self.end_row, width=7).pack(side="left", padx=(4, 8))
        ttk.Label(rr, text="(blank = all)", style="Hint.TLabel").pack(side="left")
        self._form_row(body, 3, "Row range:", rr)

        # ----- Output -----
        body = self._section(parent, "Output")
        entry = ttk.Entry(body, textvariable=self.output_dir)
        self._form_row(
            body, 0, "Folder:",
            entry,
            suffix=ttk.Button(body, text="Browse...", command=self._pick_output),
        )

        # ----- Controls -----
        ctrl = ttk.Frame(parent)
        ctrl.pack(fill="x", pady=(8, 4), padx=4)
        self.btn_start = ttk.Button(
            ctrl, text="Validate IATA codes", command=self._start,
            style="Primary.TButton",
        )
        self.btn_start.pack(side="left", padx=(0, 8))
        self.btn_pause = ttk.Button(ctrl, text="Pause", command=self._pause, state="disabled")
        self.btn_pause.pack(side="left", padx=4)
        self.btn_resume = ttk.Button(ctrl, text="Resume", command=self._resume, state="disabled")
        self.btn_resume.pack(side="left", padx=4)
        self.btn_stop = ttk.Button(
            ctrl, text="Stop", command=self._stop,
            state="disabled", style="Danger.TButton",
        )
        self.btn_stop.pack(side="left", padx=4)

        # ----- Progress -----
        prog = ttk.Frame(parent)
        prog.pack(fill="x", padx=4, pady=(8, 4))
        self.progress_bar = ttk.Progressbar(prog, mode="determinate")
        self.progress_bar.pack(fill="x")
        self.progress_label = ttk.Label(prog, text="Pick an Excel above, then click Validate IATA codes.", style="Hint.TLabel")
        self.progress_label.pack(anchor="w", pady=(2, 0))

        # ----- Log -----
        body = self._section(parent, "Log")
        log_box = ttk.Frame(body)
        log_box.pack(fill="both", expand=True)
        self.log_text = tk.Text(
            log_box, height=12, wrap="none", font=("Consolas", 9),
            relief="flat", borderwidth=1, highlightthickness=1,
            highlightbackground="#cbd5e1",
        )
        self.log_text.pack(side="left", fill="both", expand=True)
        scroll = ttk.Scrollbar(log_box, command=self.log_text.yview)
        scroll.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=scroll.set, state="disabled")
        self._attach_log_placeholder(
            self.log_text,
            "  Live validation events will appear here once you click "
            "Validate IATA codes.",
        )

    # ------------------------------------------------------------------
    # BD Travel Agency tab
    # ------------------------------------------------------------------

    def _build_bd_tab(self, parent: ttk.Frame) -> None:
        # ----- Cached data + Refresh action -----
        body = self._section(parent, "Cached agency list  ·  regtravelagency.gov.bd")
        row = ttk.Frame(body)
        row.pack(fill="x")
        self.bd_status_label = ttk.Label(row, text="…", style="Hint.TLabel")
        self.bd_status_label.pack(side="left", padx=2)
        self.btn_bd_refresh = ttk.Button(
            row, text="↻  Refresh now", command=self._bd_refresh,
        )
        self.btn_bd_refresh.pack(side="right")

        # ----- Mode -----
        body = self._section(parent, "What to run")
        ttk.Radiobutton(
            body,
            text="Export full list to Excel (all cached agencies)",
            variable=self.bd_mode, value="full",
            command=self._toggle_bd_mode,
        ).pack(anchor="w", padx=2, pady=2)
        ttk.Radiobutton(
            body,
            text="Lookup names from Excel (match each row against the cached list)",
            variable=self.bd_mode, value="lookup",
            command=self._toggle_bd_mode,
        ).pack(anchor="w", padx=2, pady=2)

        # ----- Input (lookup mode) -----
        body = self._section(parent, "Input Excel  ·  lookup mode only")
        self.bd_input_frm = body
        self.bd_input_entry = ttk.Entry(body, textvariable=self.bd_input_path)
        self.bd_input_btn = ttk.Button(
            body, text="Browse...", command=self._bd_pick_input,
        )
        self._form_row(
            body, 0, "File:", self.bd_input_entry, suffix=self.bd_input_btn,
        )

        self.bd_sheet_combo = ttk.Combobox(
            body, textvariable=self.bd_sheet_name, state="readonly",
        )
        self.bd_sheet_combo.bind("<<ComboboxSelected>>", lambda _e: self._bd_reload_columns())
        self._form_row(body, 1, "Sheet:", self.bd_sheet_combo)

        self.bd_col_combo = ttk.Combobox(
            body, textvariable=self.bd_column_name, state="readonly",
        )
        self._form_row(body, 2, "Input column:", self.bd_col_combo)

        # Match-against checkboxes — same row, no nested frame.
        match_row = ttk.Frame(body)
        ttk.Checkbutton(
            match_row, text="Agency Name", variable=self.bd_match_name,
        ).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(
            match_row, text="License Number", variable=self.bd_match_license,
        ).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(
            match_row, text="Address", variable=self.bd_match_address,
        ).pack(side="left")
        self._form_row(body, 3, "Match against:", match_row)

        # ----- Output + filter -----
        body = self._section(parent, "Output")
        ttk.Checkbutton(
            body,
            text="Include EXPIRED-PENDING agencies (license expired but still in active list)",
            variable=self.bd_include_expired,
        ).pack(anchor="w", padx=2, pady=(0, 6))
        out_row = ttk.Frame(body)
        out_row.pack(fill="x")
        out_entry = ttk.Entry(out_row, textvariable=self.bd_output_dir)
        ttk.Label(out_row, text="Folder:", width=16, anchor="w").pack(
            side="left", padx=(2, 8)
        )
        out_entry.pack(side="left", fill="x", expand=True, padx=(0, 4))
        ttk.Button(
            out_row, text="Browse...", command=self._bd_pick_output,
        ).pack(side="right", padx=(4, 2))

        # ----- Run -----
        ctrl = ttk.Frame(parent)
        ctrl.pack(fill="x", pady=(8, 4), padx=4)
        self.btn_bd_run = ttk.Button(
            ctrl, text="Run agency lookup",
            command=self._bd_run, style="Primary.TButton",
        )
        self.btn_bd_run.pack(side="left")

        # ----- Progress -----
        prog = ttk.Frame(parent)
        prog.pack(fill="x", padx=4, pady=(8, 4))
        self.bd_progress_bar = ttk.Progressbar(prog, mode="determinate")
        self.bd_progress_bar.pack(fill="x")
        self.bd_progress_label = ttk.Label(prog, text="Refresh once to cache the agency list, then run the lookup.", style="Hint.TLabel")
        self.bd_progress_label.pack(anchor="w", pady=(2, 0))

        # ----- Log -----
        body = self._section(parent, "Log")
        log_box = ttk.Frame(body)
        log_box.pack(fill="both", expand=True)
        self.bd_log_text = tk.Text(
            log_box, height=10, wrap="none", font=("Consolas", 9),
            relief="flat", borderwidth=1, highlightthickness=1,
            highlightbackground="#cbd5e1",
        )
        self.bd_log_text.pack(side="left", fill="both", expand=True)
        bd_scroll = ttk.Scrollbar(log_box, command=self.bd_log_text.yview)
        bd_scroll.pack(side="right", fill="y")
        self.bd_log_text.configure(yscrollcommand=bd_scroll.set, state="disabled")
        self._attach_log_placeholder(
            self.bd_log_text,
            "  Agency lookup events will appear here once a run starts.",
        )

        self._toggle_bd_mode()

    # ------------------------------------------------------------------
    # OEP tab (BD Overseas Movement — oep.gov.bd)
    # ------------------------------------------------------------------

    def _build_oep_tab(self, parent: ttk.Frame) -> None:
        # Wrap the whole tab so a smaller window still reaches every control.
        parent = self._make_scrollable(parent)

        # ----- Description -----
        self._section(
            parent,
            "Where are Bangladeshi workers going?  ·  oep.gov.bd",
            help_text=(
                "Pulls clearance data straight from the Overseas Employment "
                "Platform. Pick a date range and view, then Run."
            ),
        )

        # ----- Filters -----
        body = self._section(parent, "Filters")
        date_row = ttk.Frame(body)
        ttk.Label(date_row, text="From:", width=6, anchor="w").pack(side="left")
        ttk.Entry(date_row, textvariable=self.oep_date_from, width=12).pack(
            side="left", padx=(0, 12)
        )
        ttk.Label(date_row, text="To:", width=4, anchor="w").pack(side="left")
        ttk.Entry(date_row, textvariable=self.oep_date_to, width=12).pack(
            side="left", padx=(0, 12)
        )
        ttk.Label(
            date_row,
            text="(YYYY-MM-DD — leave full range for all historical data)",
            style="Hint.TLabel",
        ).pack(side="left")
        self._form_row(body, 0, "Date range:", date_row)

        # Gender + country filter
        gender_row = ttk.Frame(body)
        for value, label in oep_client.GENDER_OPTIONS:
            ttk.Radiobutton(
                gender_row, text=label, variable=self.oep_gender, value=value,
            ).pack(side="left", padx=(0, 12))
        self._form_row(body, 1, "Gender:", gender_row)

        self.oep_country_combo = ttk.Combobox(
            body, textvariable=self.oep_country_filter, state="readonly",
            values=("(loading…)",),
        )
        self._form_row(body, 2, "Country filter:", self.oep_country_combo)

        # Multi-country selector (used by time-series + pivot only)
        multi_row = ttk.Frame(body)
        self.oep_country_listbox = tk.Listbox(
            multi_row, selectmode="extended", height=5, exportselection=False,
        )
        self.oep_country_listbox.pack(side="left", fill="x", expand=True)
        multi_scroll = ttk.Scrollbar(multi_row, command=self.oep_country_listbox.yview)
        multi_scroll.pack(side="left", fill="y")
        self.oep_country_listbox.configure(yscrollcommand=multi_scroll.set)
        multi_btns = ttk.Frame(multi_row)
        multi_btns.pack(side="left", padx=(8, 0))
        for n in (5, 10, 20):
            ttk.Button(
                multi_btns, text=f"Top {n}",
                command=lambda n=n: self._oep_set_top_n(n),
                width=8,
            ).pack(anchor="w", pady=1)
        ttk.Button(
            multi_btns, text="All",
            command=self._oep_select_all,
            width=8,
        ).pack(anchor="w", pady=1)
        ttk.Button(
            multi_btns, text="Clear",
            command=lambda: self.oep_country_listbox.selection_clear(0, "end"),
            width=8,
        ).pack(anchor="w", pady=1)
        self._form_row(body, 3, "Countries (multi):", multi_row)

        # ----- Presets -----
        body = self._section(parent, "Saved filter presets")
        preset_row = ttk.Frame(body)
        preset_row.pack(fill="x")
        ttk.Label(preset_row, text="Preset:", width=8, anchor="w").pack(side="left")
        self.oep_preset_combo = ttk.Combobox(
            preset_row, textvariable=self.oep_preset_name, state="readonly", width=30,
        )
        self.oep_preset_combo.pack(side="left", padx=(0, 8))
        ttk.Button(preset_row, text="Load", command=self._oep_preset_load).pack(
            side="left", padx=(0, 4)
        )
        ttk.Button(preset_row, text="Save current as...", command=self._oep_preset_save).pack(
            side="left", padx=(0, 4)
        )
        ttk.Button(preset_row, text="Delete", command=self._oep_preset_delete).pack(
            side="left"
        )
        self._oep_refresh_preset_list()

        # ----- View mode -----
        body = self._section(parent, "View")
        for value, label in (
            ("country", "Top destination countries"),
            ("division", "Top source districts (Bangladesh)"),
            ("category", "Top job categories"),
            ("gender", "Gender breakdown per destination"),
            ("timeseries", "Monthly time series (chart)  ·  needs multi-country selection"),
            ("pivot", "Country × Division pivot  ·  needs multi-country selection"),
            ("full", "Full report — every view combined into one Excel  ·  needs multi-country"),
            ("cdt", "Country × Division × Month — single flat sheet  ·  HEAVY, needs multi-country"),
            ("cdtd", "Country × District × Month — district granularity  ·  HEAVY, needs multi-country"),
        ):
            ttk.Radiobutton(
                body, text=label, variable=self.oep_mode, value=value,
            ).pack(anchor="w", padx=2, pady=1)

        # ----- Run / Export -----
        ctrl = ttk.Frame(parent)
        ctrl.pack(fill="x", pady=(8, 4), padx=4)
        self.btn_oep_run = ttk.Button(
            ctrl, text="Run overseas-movement report",
            command=self._oep_run, style="Primary.TButton",
        )
        self.btn_oep_run.pack(side="left")
        self.btn_oep_export = ttk.Button(
            ctrl, text="Export to Excel...", command=self._oep_export, state="disabled",
        )
        self.btn_oep_export.pack(side="left", padx=(8, 0))

        self.oep_status_label = ttk.Label(ctrl, text="Pick a date range and a view, then Run.", style="Hint.TLabel")
        self.oep_status_label.pack(side="left", padx=12)

        # ----- Results panel — tree OR chart, depending on mode -----
        body = self._section(parent, "Results")
        # `oep_results_holder` holds whichever widget is currently visible
        # (tree by default, chart for time-series). Swapping uses pack/forget.
        self.oep_results_holder = ttk.Frame(body)
        self.oep_results_holder.pack(fill="both", expand=True)

        self.oep_tree_frame = ttk.Frame(self.oep_results_holder)
        self.oep_tree = ttk.Treeview(self.oep_tree_frame, show="headings", height=14)
        self.oep_tree.pack(side="left", fill="both", expand=True)
        tree_scroll = ttk.Scrollbar(self.oep_tree_frame, command=self.oep_tree.yview)
        tree_scroll.pack(side="right", fill="y")
        self.oep_tree.configure(yscrollcommand=tree_scroll.set)
        # Double-click on a country row → category drilldown popup.
        self.oep_tree.bind("<Double-1>", self._oep_on_double_click)

        # Default columns; replaced when results arrive.
        self._oep_set_columns([("#", 40), ("Info", 600)])
        self.oep_tree.insert(
            "", "end",
            values=("—", "Click Run to fetch data. Double-click a country to drill into job categories."),
        )

        # Chart frame is created lazily the first time time-series view runs.
        self.oep_chart_frame = ttk.Frame(self.oep_results_holder)

        self._oep_show_tree()

        # Load country list async so the form is usable immediately.
        # The thread fetches; the main loop polls — see _oep_poll_country_load.
        threading.Thread(target=self._oep_load_countries, daemon=True).start()
        self.root.after(200, self._oep_poll_country_load)

    # --- helpers for the multi-select listbox + presets ---

    def _oep_set_top_n(self, n: int) -> None:
        """Select the first `n` entries in the country listbox."""
        if not self.oep_country_listbox.size():
            messagebox.showinfo("OEP", "Country list still loading — try again in a moment.")
            return
        self.oep_country_listbox.selection_clear(0, "end")
        for i in range(min(n, self.oep_country_listbox.size())):
            self.oep_country_listbox.selection_set(i)
        self.oep_country_listbox.see(0)

    def _oep_select_all(self) -> None:
        """Select every country in the listbox.

        Time-series still does one call per month (cheap regardless of
        country count); pivot does one call per country, so 200 countries
        means ~200 round-trips. The progress bar handles the wait.
        """
        size = self.oep_country_listbox.size()
        if not size:
            messagebox.showinfo("OEP", "Country list still loading — try again in a moment.")
            return
        self.oep_country_listbox.selection_set(0, "end")

    def _oep_selected_countries(self) -> list[oep_client.Option]:
        """Return the currently selected items as Option objects."""
        out: list[oep_client.Option] = []
        for idx in self.oep_country_listbox.curselection():
            label = self.oep_country_listbox.get(idx)
            for opt in self._oep_country_options:
                if opt.label == label:
                    out.append(opt)
                    break
        return out

    def _oep_apply_country_selection(self, country_ids: list[str]) -> None:
        """Highlight matching country IDs in the listbox (used by preset load)."""
        self.oep_country_listbox.selection_clear(0, "end")
        for i in range(self.oep_country_listbox.size()):
            label = self.oep_country_listbox.get(i)
            opt = next((o for o in self._oep_country_options if o.label == label), None)
            if opt and opt.value in country_ids:
                self.oep_country_listbox.selection_set(i)

    # --- chart/tree panel swap ---

    def _oep_show_tree(self) -> None:
        try:
            self.oep_chart_frame.pack_forget()
        except Exception:
            pass
        self.oep_tree_frame.pack(fill="both", expand=True)

    def _oep_show_chart(self) -> None:
        try:
            self.oep_tree_frame.pack_forget()
        except Exception:
            pass
        self.oep_chart_frame.pack(fill="both", expand=True)

    # --- preset bar ---

    def _oep_refresh_preset_list(self) -> None:
        try:
            names = self._oep_preset_store.list_names()
        except Exception as exc:  # noqa: BLE001
            log.warning("Preset list failed: %s", exc)
            names = []
        self.oep_preset_combo.configure(values=names)

    def _oep_preset_load(self) -> None:
        name = self.oep_preset_name.get().strip()
        if not name:
            messagebox.showinfo("OEP", "Pick a preset from the dropdown first.")
            return
        preset = self._oep_preset_store.get(name)
        if not preset:
            messagebox.showerror("OEP", f"Preset {name!r} not found.")
            return
        self.oep_date_from.set(preset.date_from)
        self.oep_date_to.set(preset.date_to)
        self.oep_mode.set(preset.mode)
        self.oep_gender.set(preset.gender_id or "")
        if preset.country_ids:
            self._oep_apply_country_selection(preset.country_ids)

    def _oep_preset_save(self) -> None:
        from tkinter import simpledialog
        default = self.oep_preset_name.get() or self.oep_mode.get().title()
        name = simpledialog.askstring(
            "Save preset",
            "Preset name:",
            initialvalue=default,
            parent=self.root,
        )
        if not name:
            return
        name = name.strip()
        if not name:
            return
        selected = self._oep_selected_countries()
        preset = oep_presets.OEPPreset(
            name=name,
            mode=self.oep_mode.get(),
            date_from=self.oep_date_from.get(),
            date_to=self.oep_date_to.get(),
            gender_id=self.oep_gender.get(),
            country_ids=[o.value for o in selected],
            country_labels=[o.label for o in selected],
        )
        try:
            self._oep_preset_store.save(preset)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("OEP", f"Could not save preset:\n{exc}")
            return
        self._oep_refresh_preset_list()
        self.oep_preset_name.set(name)
        messagebox.showinfo("OEP", f"Preset {name!r} saved.")

    def _oep_preset_delete(self) -> None:
        name = self.oep_preset_name.get().strip()
        if not name:
            return
        if not messagebox.askyesno("OEP", f"Delete preset {name!r}?"):
            return
        removed = self._oep_preset_store.delete(name)
        if not removed:
            messagebox.showinfo("OEP", "That preset doesn't exist anymore.")
        self._oep_refresh_preset_list()
        self.oep_preset_name.set("")

    # --- category drilldown popup ---

    def _oep_on_double_click(self, _event) -> None:
        """If we're in 'country' mode, open a category drilldown for the row."""
        if not self._oep_last_result:
            return
        if self._oep_last_result.get("mode") != "country":
            return
        sel = self.oep_tree.selection()
        if not sel:
            return
        values = self.oep_tree.item(sel[0], "values")
        if not values or len(values) < 2:
            return
        country_name = values[1]
        rows = self._oep_last_result.get("raw") or []
        cats = oep_client.categories_for_country(rows, country_name)
        if not cats:
            messagebox.showinfo("OEP", f"No category data for {country_name}.")
            return
        self._oep_open_category_popup(country_name, cats)

    def _oep_open_category_popup(self, country_name: str, cats: list) -> None:
        win = tk.Toplevel(self.root)
        win.title(f"Job categories — {country_name}")
        win.geometry("560x500")
        win.transient(self.root)

        header = ttk.Label(
            win,
            text=f"Top jobs for {country_name}  ·  {sum(c.total_employee for c in cats):,} total workers",
            style="Section.TLabel",
        )
        header.pack(anchor="w", padx=10, pady=(10, 4))

        tree_frame = ttk.Frame(win)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        tree = ttk.Treeview(
            tree_frame, show="headings",
            columns=("Rank", "Job Category", "Total", "Share %"),
        )
        tree.heading("Rank", text="Rank")
        tree.heading("Job Category", text="Job Category")
        tree.heading("Total", text="Total")
        tree.heading("Share %", text="Share %")
        tree.column("Rank", width=50, anchor="e")
        tree.column("Job Category", width=300, anchor="w")
        tree.column("Total", width=100, anchor="e")
        tree.column("Share %", width=80, anchor="e")
        tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(tree_frame, command=tree.yview)
        sb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=sb.set)

        grand = sum(c.total_employee for c in cats) or 1
        for rank, c in enumerate(cats, start=1):
            tree.insert("", "end", values=(
                rank, c.category_name, f"{c.total_employee:,}",
                f"{100 * c.total_employee / grand:.2f}",
            ))

        ttk.Button(win, text="Close", command=win.destroy).pack(pady=(0, 10))

    def _oep_set_columns(self, cols: list[tuple[str, int]]) -> None:
        """Reset the Treeview columns + headings."""
        ids = [c[0] for c in cols]
        self.oep_tree.configure(columns=ids)
        for col_id, width in cols:
            self.oep_tree.heading(col_id, text=col_id)
            anchor = "e" if any(t in col_id.lower() for t in ("total", "%", "count", "male", "female", "other", "rank", "#")) else "w"
            self.oep_tree.column(col_id, width=width, anchor=anchor, stretch=True)

    def _oep_load_countries(self) -> None:
        """Fetch country list off-thread and stash the result.

        We deliberately *don't* call `root.after` from this thread —
        when the sign-in modal is active during app startup, the
        scheduled callback sometimes never fires. Instead the main
        thread polls `_oep_pending_country_payload` via
        `_oep_poll_country_load`, which is rock-solid because it runs
        on the Tk event loop itself.
        """
        log.info("OEP loading country list from oep.gov.bd…")
        error_msg: str | None = None
        try:
            opts = oep_client.list_countries()
            log.info("OEP loaded %d country options", len(opts))
        except Exception as exc:  # noqa: BLE001 — show fallback in UI
            log.warning("OEP countries fetch failed: %s", exc)
            opts = []
            error_msg = f"{type(exc).__name__}: {exc}"
        self._oep_country_options = opts
        labels = ["(All countries)"] + [o.label for o in opts]
        # Set the pending payload — the main-thread poller will pick it up.
        self._oep_pending_country_payload = (labels, error_msg)

    def _oep_poll_country_load(self) -> None:
        """Run on the Tk main loop; applies country options once available."""
        payload = self._oep_pending_country_payload
        if payload is not None:
            self._oep_pending_country_payload = None
            labels, error_msg = payload
            try:
                self._oep_apply_country_options(labels, error_msg)
            except Exception:  # noqa: BLE001 — never let a UI hiccup kill the poll
                log.exception("OEP country-apply failed")
            return
        # Still waiting — keep polling. 200ms is fast enough that the user
        # never notices, slow enough to be free CPU-wise.
        try:
            self.root.after(200, self._oep_poll_country_load)
        except RuntimeError:
            return

    def _oep_apply_country_options(
        self, labels: list[str], error_msg: str | None = None,
    ) -> None:
        log.info("OEP applying %d country options to UI", len(self._oep_country_options))
        self.oep_country_combo.configure(values=labels)
        if self.oep_country_filter.get() in ("", "(loading…)"):
            self.oep_country_filter.set("(All countries)")
        # The listbox holds only the real options — no "(All countries)" entry.
        self.oep_country_listbox.delete(0, "end")
        for o in self._oep_country_options:
            self.oep_country_listbox.insert("end", o.label)
        if error_msg:
            self.oep_status_label.configure(
                text=f"⚠ Could not load country list: {error_msg}",
            )
        elif not self._oep_country_options:
            self.oep_status_label.configure(
                text="⚠ Country list returned zero entries — oep.gov.bd may be down",
            )
        else:
            self.oep_status_label.configure(
                text=f"Ready · {len(self._oep_country_options)} countries loaded",
            )

    def _oep_resolve_country_id(self) -> str:
        label = self.oep_country_filter.get().strip()
        if not label or label.startswith("(All"):
            return ""
        for opt in self._oep_country_options:
            if opt.label == label:
                return opt.value
        return ""

    def _oep_run(self) -> None:
        if self._oep_worker is not None and self._oep_worker.is_alive():
            messagebox.showinfo("OEP", "A request is already running.")
            return
        date_from = self.oep_date_from.get().strip()
        date_to = self.oep_date_to.get().strip()
        try:
            oep_client._validate_date(date_from, "date_from")
            oep_client._validate_date(date_to, "date_to")
        except ValueError as exc:
            messagebox.showerror("OEP", str(exc))
            return
        if date_from > date_to:
            messagebox.showerror("OEP", "'From' date must be on or before 'To' date.")
            return

        mode = self.oep_mode.get()
        gender = self.oep_gender.get() or ""
        country_id = self._oep_resolve_country_id()
        selected_multi = self._oep_selected_countries()
        if mode in ("timeseries", "pivot", "full", "cdt", "cdtd") and not selected_multi:
            messagebox.showerror(
                "OEP",
                f"The {mode} view needs at least one country selected in the "
                "multi-select box. Click 'Top 5' or 'All' to start.",
            )
            return
        if mode in ("cdt", "cdtd", "full"):
            n_months = len(oep_client.iter_year_months(date_from, date_to))
            # Full report caps the heavy sections to 15 countries internally
            # — see `_oep_worker_run`. Standalone CDT honours the full selection.
            FULL_REPORT_HEAVY_CAP = 15
            effective_countries = (
                min(len(selected_multi), FULL_REPORT_HEAVY_CAP)
                if mode == "full" else len(selected_multi)
            )
            est_calls = effective_countries * n_months
            est_minutes = max(1, int(est_calls * 5 / 60))
            if est_calls > 50:
                if mode == "cdt":
                    section_label = "The country×division×month sheet"
                    cap_note = ""
                elif mode == "cdtd":
                    section_label = "The country×district×month sheet"
                    cap_note = ""
                else:
                    section_label = "The Full report's heavy sections (pivot, time-series, country×division×month)"
                    cap_note = (
                        f"\n\nNote: Full report auto-caps to the top {FULL_REPORT_HEAVY_CAP} "
                        f"destinations by volume (you selected {len(selected_multi)})."
                    )
                proceed = messagebox.askyesno(
                    "OEP — heavy fetch",
                    f"{section_label} needs "
                    f"{effective_countries} countries × {n_months} months = "
                    f"{est_calls} HTTP calls (~{est_minutes} min).{cap_note}\n\n"
                    "Continue?",
                )
                if not proceed:
                    return
        cfg = {
            "mode": mode,
            "date_from": date_from,
            "date_to": date_to,
            "gender_id": gender,
            "country_id": country_id,
            "country_options": selected_multi,
        }
        self._oep_last_result = None
        self.btn_oep_run.configure(state="disabled")
        self.btn_oep_export.configure(state="disabled")
        self._oep_show_tree()
        self.oep_status_label.configure(text="Fetching from oep.gov.bd…")
        self.oep_tree.delete(*self.oep_tree.get_children())
        self._oep_set_columns([("#", 40), ("Info", 600)])
        self.oep_tree.insert("", "end", values=("…", "Working — this may take 30-60s for wide date ranges."))

        self._oep_worker = threading.Thread(
            target=self._oep_worker_run, args=(cfg,), daemon=True,
        )
        self._oep_worker.start()

    def _oep_worker_run(self, cfg: dict) -> None:
        mode = cfg["mode"]
        try:
            if mode == "country":
                rows = oep_client.fetch_country_clearance(
                    cfg["date_from"], cfg["date_to"],
                    gender_id=cfg["gender_id"],
                    country_id=cfg["country_id"],
                )
                summary = oep_client.aggregate_by_country(rows)
                self._post(MSG_OEP_DONE, {
                    "mode": "country",
                    "summary": summary,
                    "raw": rows,
                    "date_from": cfg["date_from"],
                    "date_to": cfg["date_to"],
                })
            elif mode == "category":
                rows = oep_client.fetch_country_clearance(
                    cfg["date_from"], cfg["date_to"],
                    gender_id=cfg["gender_id"],
                    country_id=cfg["country_id"],
                )
                summary = oep_client.aggregate_by_category(rows)
                self._post(MSG_OEP_DONE, {
                    "mode": "category",
                    "summary": summary,
                    "raw": rows,
                    "date_from": cfg["date_from"],
                    "date_to": cfg["date_to"],
                })
            elif mode == "division":
                country_ids = [cfg["country_id"]] if cfg["country_id"] else None
                rows = oep_client.fetch_division_clearance(
                    cfg["date_from"], cfg["date_to"],
                    gender_id=cfg["gender_id"],
                    country_ids=country_ids,
                )
                div_summary = oep_client.aggregate_by_division(rows)
                self._post(MSG_OEP_DONE, {
                    "mode": "division",
                    "summary": div_summary,
                    "raw": rows,
                    "date_from": cfg["date_from"],
                    "date_to": cfg["date_to"],
                })
            elif mode == "gender":
                self._post(MSG_OEP_BUSY, "Fetching All…")
                all_rows = oep_client.fetch_country_clearance(
                    cfg["date_from"], cfg["date_to"],
                    country_id=cfg["country_id"],
                )
                self._post(MSG_OEP_BUSY, "Fetching Male…")
                male_rows = oep_client.fetch_country_clearance(
                    cfg["date_from"], cfg["date_to"],
                    gender_id="1", country_id=cfg["country_id"],
                )
                self._post(MSG_OEP_BUSY, "Fetching Female…")
                female_rows = oep_client.fetch_country_clearance(
                    cfg["date_from"], cfg["date_to"],
                    gender_id="2", country_id=cfg["country_id"],
                )
                summary = oep_client.merge_gender_breakdowns(
                    all_rows, male_rows, female_rows,
                )
                self._post(MSG_OEP_DONE, {
                    "mode": "gender",
                    "summary": summary,
                    "raw": all_rows,
                    "date_from": cfg["date_from"],
                    "date_to": cfg["date_to"],
                })
            elif mode == "timeseries":
                opts = cfg["country_options"]

                def progress(idx: int, total: int, label: str) -> None:
                    self._post(MSG_OEP_BUSY, f"{label}  ({idx}/{total})")

                rows = oep_client.fetch_monthly_timeseries(
                    cfg["date_from"], cfg["date_to"],
                    country_ids=[o.value for o in opts],
                    gender_id=cfg["gender_id"],
                    progress_cb=progress,
                )
                months, series = oep_client.pivot_timeseries(rows)
                self._post(MSG_OEP_DONE, {
                    "mode": "timeseries",
                    "months": months,
                    "series": series,
                    "raw": rows,
                    "country_labels": [o.label for o in opts],
                    "date_from": cfg["date_from"],
                    "date_to": cfg["date_to"],
                })
            elif mode == "pivot":
                opts = cfg["country_options"]

                def progress(idx: int, total: int, label: str) -> None:
                    self._post(MSG_OEP_BUSY, f"{label}  ({idx}/{total})")

                cells = oep_client.fetch_country_division_pivot(
                    cfg["date_from"], cfg["date_to"],
                    country_options=opts,
                    gender_id=cfg["gender_id"],
                    progress_cb=progress,
                )
                divisions, countries, table = oep_client.pivot_country_division(cells)
                self._post(MSG_OEP_DONE, {
                    "mode": "pivot",
                    "divisions": divisions,
                    "countries": countries,
                    "table": table,
                    "raw": cells,
                    "date_from": cfg["date_from"],
                    "date_to": cfg["date_to"],
                })
            elif mode == "full":
                user_opts = cfg["country_options"]

                def progress(idx: int, total: int, label: str) -> None:
                    self._post(MSG_OEP_BUSY, f"{label}  ({idx}/{total})")

                # 1) all-rows (no country filter) — powers country/category/gender views
                self._post(MSG_OEP_BUSY, "[1/7] All-country totals…")
                all_rows = oep_client.fetch_country_clearance(
                    cfg["date_from"], cfg["date_to"],
                    gender_id=cfg["gender_id"],
                )
                # 2) male-only and female-only — for gender breakdown
                self._post(MSG_OEP_BUSY, "[2/7] Male totals…")
                male_rows = oep_client.fetch_country_clearance(
                    cfg["date_from"], cfg["date_to"], gender_id="1",
                )
                self._post(MSG_OEP_BUSY, "[3/7] Female totals…")
                female_rows = oep_client.fetch_country_clearance(
                    cfg["date_from"], cfg["date_to"], gender_id="2",
                )
                # 4) division totals (Bangladesh-side)
                self._post(MSG_OEP_BUSY, "[4/7] Division totals…")
                div_rows = oep_client.fetch_division_clearance(
                    cfg["date_from"], cfg["date_to"], gender_id=cfg["gender_id"],
                )
                # Cap the heavy per-country sections (time-series, pivot, CDT)
                # to the top destinations by volume. Iterating 200 countries
                # gets us IP-rate-limited on oep.gov.bd, and a 200-column
                # pivot is unreadable anyway. The user's multi-selection is
                # used only as an upper bound — if they picked Top 5 we
                # honour that, but if they picked All we silently cap to 15.
                FULL_REPORT_HEAVY_CAP = 15
                country_summary = oep_client.aggregate_by_country(all_rows)
                top_country_names = [
                    c.country_name for c in country_summary[:FULL_REPORT_HEAVY_CAP]
                ]
                # Limit to the intersection of user's selection × top-N by
                # volume, preserving the country-options shape for downstream
                # calls.
                user_names = {o.label for o in user_opts}
                heavy_opts = [
                    o for o in user_opts
                    if o.label in top_country_names
                ]
                if not heavy_opts:
                    # User picked countries with zero volume in this window —
                    # fall back to top-N from the data so the report isn't empty.
                    label_to_opt = {o.label: o for o in user_opts}
                    heavy_opts = [
                        label_to_opt[name] for name in top_country_names
                        if name in label_to_opt
                    ][:FULL_REPORT_HEAVY_CAP]
                # Order heavy_opts by volume descending for predictable charts.
                rank = {name: i for i, name in enumerate(top_country_names)}
                heavy_opts = sorted(heavy_opts, key=lambda o: rank.get(o.label, 999))
                log.info(
                    "Full report heavy sections: user picked %d countries, "
                    "capped to %d for pivot/CDT/time-series",
                    len(user_opts), len(heavy_opts),
                )

                # 5) monthly time series for the heavy-capped country set
                self._post(MSG_OEP_BUSY, "[5/7] Monthly time series…")
                ts_rows = oep_client.fetch_monthly_timeseries(
                    cfg["date_from"], cfg["date_to"],
                    country_ids=[o.value for o in heavy_opts],
                    gender_id=cfg["gender_id"],
                    progress_cb=progress,
                )
                months, series = oep_client.pivot_timeseries(ts_rows)
                # 6) country × division pivot for the heavy-capped country set
                self._post(MSG_OEP_BUSY, "[6/7] Country × Division pivot…")
                cells = oep_client.fetch_country_division_pivot(
                    cfg["date_from"], cfg["date_to"],
                    country_options=heavy_opts,
                    gender_id=cfg["gender_id"],
                    progress_cb=progress,
                )
                divisions, pivot_countries, table = oep_client.pivot_country_division(cells)
                # 7) heavy: country × division × month flat sheet
                self._post(MSG_OEP_BUSY, "[7/7] Country × Division × Month…")
                cdt_cells = oep_client.fetch_country_division_timeseries(
                    cfg["date_from"], cfg["date_to"],
                    country_options=heavy_opts,
                    gender_id=cfg["gender_id"],
                    progress_cb=progress,
                )
                cdt_months, cdt_pairs, cdt_table = oep_client.pivot_country_division_timeseries(
                    cdt_cells,
                )

                self._post(MSG_OEP_DONE, {
                    "mode": "full",
                    "country_summary": country_summary,
                    "category_summary": oep_client.aggregate_by_category(all_rows),
                    "division_summary": oep_client.aggregate_by_division(div_rows),
                    "gender_summary": oep_client.merge_gender_breakdowns(
                        all_rows, male_rows, female_rows,
                    ),
                    "raw_country": all_rows,
                    "raw_division": div_rows,
                    "months": months,
                    "series": series,
                    "divisions": divisions,
                    "pivot_countries": pivot_countries,
                    "table": table,
                    # Country × Division × Month (flat sheet) — same shape as
                    # the standalone 'cdt' view but bundled here too.
                    "cdt_months": cdt_months,
                    "cdt_pairs": cdt_pairs,
                    "cdt_table": cdt_table,
                    "country_labels": [o.label for o in heavy_opts],
                    "user_country_labels": [o.label for o in user_opts],
                    "heavy_country_cap": FULL_REPORT_HEAVY_CAP,
                    "date_from": cfg["date_from"],
                    "date_to": cfg["date_to"],
                    "gender_id": cfg["gender_id"],
                })
            elif mode == "cdt":
                opts = cfg["country_options"]

                def progress(idx: int, total: int, label: str) -> None:
                    self._post(MSG_OEP_BUSY, f"{label}  ({idx}/{total})")

                cells = oep_client.fetch_country_division_timeseries(
                    cfg["date_from"], cfg["date_to"],
                    country_options=opts,
                    gender_id=cfg["gender_id"],
                    progress_cb=progress,
                )
                months, pairs, table = oep_client.pivot_country_division_timeseries(cells)
                self._post(MSG_OEP_DONE, {
                    "mode": "cdt",
                    "months": months,
                    "pairs": pairs,
                    "table": table,
                    "raw": cells,
                    "country_labels": [o.label for o in opts],
                    "date_from": cfg["date_from"],
                    "date_to": cfg["date_to"],
                })
            elif mode == "cdtd":
                opts = cfg["country_options"]

                def progress(idx: int, total: int, label: str) -> None:
                    self._post(MSG_OEP_BUSY, f"{label}  ({idx}/{total})")

                cells = oep_client.fetch_country_district_timeseries(
                    cfg["date_from"], cfg["date_to"],
                    country_options=opts,
                    gender_id=cfg["gender_id"],
                    progress_cb=progress,
                )
                months, triples, table = oep_client.pivot_country_district_timeseries(cells)
                self._post(MSG_OEP_DONE, {
                    "mode": "cdtd",
                    "months": months,
                    "triples": triples,
                    "table": table,
                    "raw": cells,
                    "country_labels": [o.label for o in opts],
                    "date_from": cfg["date_from"],
                    "date_to": cfg["date_to"],
                })
            else:
                self._post(MSG_OEP_ERROR, f"Unknown mode: {mode}")
        except Exception as exc:  # noqa: BLE001 — surface in UI
            log.exception("OEP worker failed")
            self._post(MSG_OEP_ERROR, f"{type(exc).__name__}: {exc}")

    def _oep_render_results(self, result: dict) -> None:
        self._oep_last_result = result
        mode = result["mode"]
        # Modes with a flat "summary" list (country/division/category/gender)
        # share a rendering branch below. Timeseries + pivot carry richer
        # payloads instead.
        summary = result.get("summary") or []
        self.oep_tree.delete(*self.oep_tree.get_children())

        if mode == "country":
            cols = [("Rank", 50), ("Destination", 240), ("Total", 110), ("Categories", 100), ("Share %", 90)]
            self._oep_set_columns(cols)
            grand = sum(c.total_employee for c in summary) or 1
            for rank, c in enumerate(summary, start=1):
                self.oep_tree.insert("", "end", values=(
                    rank, c.country_name, f"{c.total_employee:,}",
                    c.category_count, f"{100 * c.total_employee / grand:.2f}",
                ))
            self.oep_status_label.configure(
                text=f"{len(summary):,} destinations  ·  {grand:,} total workers"
            )
        elif mode == "division":
            cols = [("Rank", 50), ("Division", 200), ("Total", 110), ("Districts", 90), ("Share %", 90)]
            self._oep_set_columns(cols)
            grand = sum(d.total_employee for d in summary) or 1
            for rank, d in enumerate(summary, start=1):
                self.oep_tree.insert("", "end", values=(
                    rank, d.division, f"{d.total_employee:,}",
                    d.district_count, f"{100 * d.total_employee / grand:.2f}",
                ))
            self.oep_status_label.configure(
                text=f"{len(summary)} divisions  ·  {len(result['raw'])} districts  ·  {grand:,} total"
            )
        elif mode == "category":
            cols = [("Rank", 50), ("Job Category", 260), ("Total", 110), ("Countries", 90), ("Share %", 90)]
            self._oep_set_columns(cols)
            grand = sum(c.total_employee for c in summary) or 1
            for rank, c in enumerate(summary, start=1):
                self.oep_tree.insert("", "end", values=(
                    rank, c.category_name, f"{c.total_employee:,}",
                    c.country_count, f"{100 * c.total_employee / grand:.2f}",
                ))
            self.oep_status_label.configure(
                text=f"{len(summary):,} job categories  ·  {grand:,} total workers"
            )
        elif mode == "gender":
            cols = [
                ("Rank", 50), ("Destination", 220),
                ("Male", 90), ("Female", 90), ("Other", 70),
                ("Total", 100), ("Female %", 80),
            ]
            self._oep_set_columns(cols)
            for rank, g in enumerate(summary, start=1):
                pct = (100 * g.female / g.total) if g.total else 0.0
                self.oep_tree.insert("", "end", values=(
                    rank, g.country_name,
                    f"{g.male:,}", f"{g.female:,}", f"{g.other:,}",
                    f"{g.total:,}", f"{pct:.1f}",
                ))
            self.oep_status_label.configure(
                text=f"{len(summary):,} destinations  ·  gender breakdown"
            )
        elif mode == "timeseries":
            months = result["months"]
            series = result["series"]
            self._oep_render_timeseries(months, series, result)
        elif mode == "pivot":
            self._oep_render_pivot(result)
        elif mode == "full":
            self._oep_render_full_report(result)
        elif mode == "cdt":
            self._oep_render_cdt(result)
        elif mode == "cdtd":
            self._oep_render_cdtd(result)

        self.btn_oep_run.configure(state="normal")
        has_data = bool(
            summary if mode in ("country", "division", "category", "gender") else
            (result.get("series") if mode == "timeseries" else result.get("table"))
        )
        if has_data:
            self.btn_oep_export.configure(state="normal")

    def _oep_render_timeseries(
        self, months: list[str], series: dict[str, list[int]], result: dict
    ) -> None:
        # Lazy import so non-OEP users don't pay the matplotlib startup cost.
        import matplotlib
        matplotlib.use("TkAgg", force=False)
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        if self._oep_chart_canvas is not None:
            self._oep_chart_canvas.get_tk_widget().destroy()
            self._oep_chart_canvas = None

        fig = Figure(figsize=(8, 4.5), dpi=100)
        ax = fig.add_subplot(111)
        # Preserve user-selected order for legend stacking.
        order = [c for c in result.get("country_labels", []) if c in series]
        order += [c for c in series if c not in order]
        for country in order:
            values = series[country]
            ax.plot(months, values, marker="o", label=country, linewidth=1.6)
        ax.set_title(
            f"Monthly clearance volume  ·  {result['date_from']} → {result['date_to']}"
        )
        ax.set_ylabel("Workers cleared")
        ax.set_xlabel("Month")
        ax.grid(True, alpha=0.3)
        ax.legend(loc="best", fontsize=8)
        # Rotate ticks so YYYY-MM labels don't collide.
        for tick in ax.get_xticklabels():
            tick.set_rotation(45)
            tick.set_ha("right")
        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.oep_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        self._oep_chart_canvas = canvas
        self._oep_chart_figure = fig

        self._oep_show_chart()
        total = sum(sum(v) for v in series.values())
        self.oep_status_label.configure(
            text=(
                f"{len(series)} countries  ·  {len(months)} months  ·  "
                f"{total:,} total workers"
            )
        )

    def _oep_render_cdtd(self, result: dict) -> None:
        """Render the flat country×district×month table.

        Same shape as the division-level view but each row carries a
        Division + District pair. Many more rows; same column count.
        """
        months: list[str] = result["months"]
        triples: list[tuple[str, str, str]] = result["triples"]
        table: dict[tuple[str, str, str, str], int] = result["table"]

        cols = (
            [("Country", 140), ("Division", 110), ("District", 130)]
            + [(ym, 70) for ym in months]
            + [("Total", 90)]
        )
        self._oep_set_columns(cols)

        col_totals = [0] * len(months)
        grand_total = 0
        for country, division, district in triples:
            row_values = [country, division, district]
            row_total = 0
            for ci, ym in enumerate(months):
                v = table.get((country, division, district, ym), 0)
                row_total += v
                col_totals[ci] += v
                row_values.append(f"{v:,}" if v else "—")
            row_values.append(f"{row_total:,}")
            grand_total += row_total
            self.oep_tree.insert("", "end", values=row_values)

        totals_row = (
            ["Total", "", ""]
            + [f"{t:,}" for t in col_totals]
            + [f"{grand_total:,}"]
        )
        self.oep_tree.insert("", "end", values=totals_row, tags=("total",))
        self.oep_tree.tag_configure("total", font=("Segoe UI", 9, "bold"))

        self.oep_status_label.configure(
            text=(
                f"{len(triples)} (country, division, district) rows × "
                f"{len(months)} months  ·  {grand_total:,} workers"
            )
        )

    def _oep_render_cdt(self, result: dict) -> None:
        """Render the flat country×division×month table.

        With many countries and months this is wide — the on-screen view
        is the same shape as the exported Excel, so users can spot-check
        before saving.
        """
        months: list[str] = result["months"]
        pairs: list[tuple[str, str]] = result["pairs"]
        table: dict[tuple[str, str, str], int] = result["table"]

        cols = (
            [("Country", 160), ("Division", 130)]
            + [(ym, 70) for ym in months]
            + [("Total", 90)]
        )
        self._oep_set_columns(cols)

        col_totals = [0] * len(months)
        grand_total = 0
        for country, division in pairs:
            row_values = [country, division]
            row_total = 0
            for ci, ym in enumerate(months):
                v = table.get((country, division, ym), 0)
                row_total += v
                col_totals[ci] += v
                row_values.append(f"{v:,}" if v else "—")
            row_values.append(f"{row_total:,}")
            grand_total += row_total
            self.oep_tree.insert("", "end", values=row_values)

        # Totals row
        totals_row = ["Total", ""] + [f"{t:,}" for t in col_totals] + [f"{grand_total:,}"]
        self.oep_tree.insert("", "end", values=totals_row, tags=("total",))
        self.oep_tree.tag_configure("total", font=("Segoe UI", 9, "bold"))

        self.oep_status_label.configure(
            text=(
                f"{len(pairs)} (country, division) pairs × {len(months)} months  ·  "
                f"{grand_total:,} workers"
            )
        )

    def _oep_render_full_report(self, result: dict) -> None:
        """Show a summary of every section the export will contain.

        The on-screen view is intentionally a compact index — the real
        deliverable is the multi-sheet Excel produced by Export.
        """
        cols = [("Section", 280), ("Rows", 80), ("Total", 140)]
        self._oep_set_columns(cols)

        country_summary = result["country_summary"]
        category_summary = result["category_summary"]
        division_summary = result["division_summary"]
        gender_summary = result["gender_summary"]
        months = result["months"]
        series = result["series"]
        divisions = result["divisions"]
        pivot_countries = result["pivot_countries"]
        table = result["table"]

        grand_total = sum(c.total_employee for c in country_summary)

        cdt_months = result.get("cdt_months", [])
        cdt_pairs = result.get("cdt_pairs", [])
        cdt_table = result.get("cdt_table", {})
        sections = [
            (
                "Cover (date range, headline totals)",
                "—",
                f"{grand_total:,} workers",
            ),
            (
                "By Country (every destination)",
                f"{len(country_summary)}",
                f"{grand_total:,}",
            ),
            (
                "By Division (Bangladesh source)",
                f"{len(division_summary)}",
                f"{sum(d.total_employee for d in division_summary):,}",
            ),
            (
                "By Category (every job)",
                f"{len(category_summary)}",
                f"{sum(c.total_employee for c in category_summary):,}",
            ),
            (
                "By Gender (per destination)",
                f"{len(gender_summary)}",
                f"{sum(g.total for g in gender_summary):,}",
            ),
            (
                "Monthly Time Series (selected countries)",
                f"{len(series)} × {len(months)} mo",
                f"{sum(sum(v) for v in series.values()):,}",
            ),
            (
                "Country × Division Pivot (selected countries)",
                f"{len(divisions)} × {len(pivot_countries)}",
                f"{sum(table.values()):,}",
            ),
            (
                "Country × Division × Month (flat sheet)",
                f"{len(cdt_pairs)} pairs × {len(cdt_months)} mo",
                f"{sum(cdt_table.values()):,}",
            ),
        ]
        for name, rows_count, total in sections:
            self.oep_tree.insert("", "end", values=(name, rows_count, total))

        self.oep_status_label.configure(
            text=(
                f"Full report ready — {grand_total:,} workers across "
                f"{len(country_summary)} destinations. Click Export to save the workbook."
            )
        )

    def _oep_render_pivot(self, result: dict) -> None:
        divisions: list[str] = result["divisions"]
        countries: list[str] = result["countries"]
        table: dict[tuple[str, str], int] = result["table"]

        # Trim long country labels for column headers so the grid fits.
        def _short(name: str, n: int = 14) -> str:
            return name if len(name) <= n else name[: n - 1] + "…"

        cols = [("Division", 140)] + [(_short(c), 110) for c in countries] + [("Total", 110)]
        self._oep_set_columns(cols)

        grand_total = 0
        for div in divisions:
            row_values = [div]
            row_total = 0
            for country in countries:
                v = table.get((div, country), 0)
                row_total += v
                row_values.append(f"{v:,}" if v else "—")
            row_values.append(f"{row_total:,}")
            grand_total += row_total
            self.oep_tree.insert("", "end", values=row_values)

        # Totals row
        totals = ["Total"]
        for country in countries:
            t = sum(table.get((d, country), 0) for d in divisions)
            totals.append(f"{t:,}")
        totals.append(f"{grand_total:,}")
        self.oep_tree.insert("", "end", values=totals, tags=("total",))
        self.oep_tree.tag_configure("total", font=("Segoe UI", 9, "bold"))

        self.oep_status_label.configure(
            text=(
                f"{len(divisions)} divisions × {len(countries)} destinations  ·  "
                f"{grand_total:,} total workers"
            )
        )

    def _oep_export(self) -> None:
        if not self._oep_last_result:
            return
        result = self._oep_last_result
        mode = result["mode"]
        suggested_dir = Path(self.oep_output_dir.get()) if self.oep_output_dir.get() else Path.home()
        target = filedialog.asksaveasfilename(
            title="Save OEP report",
            defaultextension=".xlsx",
            initialdir=str(suggested_dir),
            initialfile=excel_io.build_oep_output_path(suggested_dir, mode).name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not target:
            return
        path = Path(target)
        try:
            if mode == "country":
                excel_io.write_oep_country_report(
                    path,
                    date_from=result["date_from"], date_to=result["date_to"],
                    summary=result["summary"], raw_rows=result["raw"],
                )
            elif mode == "division":
                excel_io.write_oep_division_report(
                    path,
                    date_from=result["date_from"], date_to=result["date_to"],
                    summary=result["summary"], raw_rows=result["raw"],
                )
            elif mode == "category":
                excel_io.write_oep_category_report(
                    path,
                    date_from=result["date_from"], date_to=result["date_to"],
                    summary=result["summary"], raw_rows=result["raw"],
                )
            elif mode == "gender":
                excel_io.write_oep_gender_report(
                    path,
                    date_from=result["date_from"], date_to=result["date_to"],
                    summary=result["summary"],
                )
            elif mode == "timeseries":
                excel_io.write_oep_timeseries_report(
                    path,
                    date_from=result["date_from"], date_to=result["date_to"],
                    months=result["months"], series=result["series"],
                )
            elif mode == "pivot":
                excel_io.write_oep_pivot_report(
                    path,
                    date_from=result["date_from"], date_to=result["date_to"],
                    divisions=result["divisions"], countries=result["countries"],
                    table=result["table"],
                )
            elif mode == "cdt":
                excel_io.write_oep_country_division_timeseries(
                    path,
                    date_from=result["date_from"], date_to=result["date_to"],
                    months=result["months"], pairs=result["pairs"],
                    table=result["table"],
                )
            elif mode == "cdtd":
                excel_io.write_oep_country_district_timeseries(
                    path,
                    date_from=result["date_from"], date_to=result["date_to"],
                    months=result["months"], triples=result["triples"],
                    table=result["table"],
                )
            elif mode == "full":
                excel_io.write_oep_full_report(
                    path,
                    date_from=result["date_from"],
                    date_to=result["date_to"],
                    gender_id=result.get("gender_id", ""),
                    country_summary=result["country_summary"],
                    category_summary=result["category_summary"],
                    division_summary=result["division_summary"],
                    gender_summary=result["gender_summary"],
                    raw_country=result["raw_country"],
                    raw_division=result["raw_division"],
                    months=result["months"],
                    series=result["series"],
                    divisions=result["divisions"],
                    pivot_countries=result["pivot_countries"],
                    table=result["table"],
                    country_labels=result["country_labels"],
                    cdt_months=result.get("cdt_months"),
                    cdt_pairs=result.get("cdt_pairs"),
                    cdt_table=result.get("cdt_table"),
                )
        except Exception as exc:  # noqa: BLE001 — surface to user
            messagebox.showerror("OEP export", f"Could not save:\n{exc}")
            return
        # Remember the folder for next time
        self.oep_output_dir.set(str(path.parent))
        messagebox.showinfo("OEP export", f"Saved:\n{path}")

    # ------------------------------------------------------------------
    # File pickers
    # ------------------------------------------------------------------

    def _pick_input(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Excel with IATA numbers",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if not path:
            return
        self.input_path.set(path)
        self._reload_sheets()

    def _pick_output(self) -> None:
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.output_dir.set(path)

    def _reload_sheets(self) -> None:
        try:
            sheets = excel_io.list_sheet_names(Path(self.input_path.get()))
        except Exception as e:
            messagebox.showerror("Cannot read Excel", str(e))
            return
        self.sheet_combo["values"] = sheets
        if sheets:
            self.sheet_combo.current(0)
            self._reload_columns()

    def _reload_columns(self) -> None:
        try:
            cols = excel_io.list_columns(Path(self.input_path.get()), self.sheet_name.get())
        except Exception as e:
            messagebox.showerror("Cannot read columns", str(e))
            return
        self._sheet_columns = cols
        self.col_combo["values"] = cols
        # Try to auto-pick a column whose name looks like IATA
        guess = next(
            (c for c in cols if "iata" in c.lower() or "code" in c.lower() or "number" in c.lower()),
            cols[0] if cols else "",
        )
        self.column_name.set(guess)

    # ------------------------------------------------------------------
    # Run control
    # ------------------------------------------------------------------

    def _start(self) -> None:
        if self._worker is not None and self._worker.is_alive():
            return
        cfg = self._validate_inputs()
        if cfg is None:
            return

        self._stop_flag.clear()
        self._pause_event.set()
        self._captcha_clear_flag.clear()

        self.btn_start.configure(state="disabled")
        self.btn_pause.configure(state="normal")
        self.btn_stop.configure(state="normal")

        self._worker = threading.Thread(
            target=self._run_worker,
            args=(cfg,),
            daemon=True,
        )
        self._worker.start()

    def _pause(self) -> None:
        self._pause_event.clear()
        self.btn_pause.configure(state="disabled")
        self.btn_resume.configure(state="normal")
        self._log("Paused. Click Resume when ready.")

    def _resume(self) -> None:
        self._pause_event.set()
        self._captcha_clear_flag.set()
        self.btn_pause.configure(state="normal")
        self.btn_resume.configure(state="disabled")
        self._log("Resumed.")

    def _stop(self) -> None:
        if not messagebox.askyesno("Stop?", "Stop the run? Partial results are already saved."):
            return
        self._stop_flag.set()
        self._pause_event.set()  # un-pause so the worker exits its wait
        self._captcha_clear_flag.set()
        if self._validator is not None:
            self._validator.stop()

    def _on_close(self) -> None:
        any_alive = (
            (self._worker is not None and self._worker.is_alive())
            or (self._bd_worker is not None and self._bd_worker.is_alive())
        )
        if any_alive:
            if not messagebox.askyesno("Quit?", "A run is in progress. Quit anyway?"):
                return
            self._stop_flag.set()
            self._pause_event.set()
            self._captcha_clear_flag.set()
            if self._validator is not None:
                self._validator.stop()
            if self._worker is not None:
                self._worker.join(timeout=5)
            # BD worker is HTTP-only and short-lived; just let it die.
        self.root.destroy()

    def _validate_inputs(self) -> dict | None:
        input_path = self.input_path.get().strip()
        if not input_path or not Path(input_path).exists():
            messagebox.showerror("Invalid input", "Pick a valid Excel file.")
            return None
        if not self.sheet_name.get():
            messagebox.showerror("Invalid input", "Pick a sheet.")
            return None
        if not self.column_name.get() or self.column_name.get() not in self._sheet_columns:
            messagebox.showerror("Invalid input", "Pick the IATA column.")
            return None
        try:
            start = int(self.start_row.get())
        except ValueError:
            messagebox.showerror("Invalid input", "Start row must be an integer.")
            return None
        if start < 2:
            messagebox.showerror("Invalid input", "Start row must be >= 2 (row 1 = header).")
            return None
        end_str = self.end_row.get().strip()
        end: int | None = None
        if end_str:
            try:
                end = int(end_str)
            except ValueError:
                messagebox.showerror("Invalid input", "End row must be empty or an integer.")
                return None
            if end < start:
                messagebox.showerror("Invalid input", "End row must be >= start row.")
                return None
        out_dir = Path(self.output_dir.get())
        if not out_dir.exists():
            messagebox.showerror("Invalid output", "Output folder does not exist.")
            return None

        column_index = self._sheet_columns.index(self.column_name.get())
        return {
            "input_path": Path(input_path),
            "sheet": self.sheet_name.get(),
            "column_index": column_index,
            "start_row": start,
            "end_row": end,
            "output_dir": out_dir,
        }

    # ------------------------------------------------------------------
    # Worker thread
    # ------------------------------------------------------------------

    def _run_worker(self, cfg: dict) -> None:
        try:
            self._post(MSG_LOG, "Reading IATA numbers from Excel...")
            rows = excel_io.read_iata_numbers(
                cfg["input_path"],
                cfg["sheet"],
                cfg["column_index"],
                cfg["start_row"],
                cfg["end_row"],
            )
            if not rows:
                self._post(MSG_ERROR, "No IATA numbers found in the selected range.")
                return

            output_path = excel_io.build_output_path(cfg["output_dir"])
            self._post(MSG_LOG, f"Writing results to: {output_path}")

            cache = Cache(config.CACHE_DB)
            self._post(MSG_LOG, f"Cache: {cache.count()} prior lookups available.")

            total = len(rows)
            self._post(MSG_PROGRESS, (0, total, "Starting browser..."))

            # ResultWriter as context manager guarantees the final flush even
            # if the validator raises during shutdown.
            with excel_io.ResultWriter(output_path) as writer, make_validator(
                profile_dir=config.PROFILE_DIR,
                on_log=lambda m: self._post(MSG_LOG, m),
            ) as validator:
                self._validator = validator
                started = time.monotonic()

                for idx, (excel_row, iata) in enumerate(rows, start=1):
                    if self._stop_flag.is_set():
                        self._post(MSG_LOG, "Stop requested — exiting.")
                        break
                    # Honor pause
                    self._pause_event.wait()

                    # Cache hit
                    cached = cache.get(iata)
                    if cached is not None:
                        writer.append(cached)
                        self._post(MSG_RESULT, cached)
                        self._post_progress(idx, total, started, "cached")
                        continue

                    result = self._lookup_with_retry(validator, iata)
                    cache.put(result)
                    writer.append(result)
                    self._post(MSG_RESULT, result)
                    self._post_progress(idx, total, started, result.status)

            self._post(MSG_LOG, f"Done. Output: {output_path}")
            self._post(MSG_DONE, str(output_path))

            # Best-effort usage log. Never breaks the run on failure.
            try:
                auth.log_lookup_event(
                    action="iata_validate",
                    target=str(cfg["input_path"].name),
                    count=total,
                )
            except Exception as exc:  # noqa: BLE001
                log.warning("usage log failed: %s", exc)

        except Exception as e:
            log.exception("worker crashed")
            self._post(MSG_ERROR, f"{type(e).__name__}: {e}")
        finally:
            self._validator = None

    def _lookup_with_retry(self, validator: IATAValidator, iata: str) -> LookupResult:
        """One CAPTCHA retry — alert user, wait for green check, then resume.

        Uses complete_after_captcha (NOT lookup) on the second pass so the
        user's puzzle solve isn't thrown away by a re-navigation.
        """
        try:
            return validator.lookup(iata)
        except CaptchaChallenge as e:
            self._post(MSG_CAPTCHA, str(e))
            self._captcha_clear_flag.clear()
            ok = validator.wait_for_user_captcha()
            if not ok:
                return LookupResult(
                    iata_number=iata,
                    trading_name="",
                    country="",
                    accredited="",
                    status="ERROR",
                    checked_at=now_iso(),
                    notes="captcha not solved within timeout",
                )
            # User solved it — resume from the submit step on the SAME page.
            try:
                return validator.complete_after_captcha(iata)
            except (CaptchaChallenge, ValidatorStopped):
                raise
            except Exception as e2:  # noqa: BLE001 — keep loop alive on unknown errors
                log.warning("resume failed for %s: %s", iata, e2)
                return LookupResult(
                    iata_number=iata,
                    trading_name="",
                    country="",
                    accredited="",
                    status="ERROR",
                    checked_at=now_iso(),
                    notes=f"resume failed: {e2}",
                )
        except ValidatorStopped:
            raise

    # ------------------------------------------------------------------
    # Worker → GUI messaging
    # ------------------------------------------------------------------

    def _post(self, kind: str, payload: object) -> None:
        self._msg_queue.put((kind, payload))

    def _post_progress(
        self, idx: int, total: int, started: float, last_status: str
    ) -> None:
        elapsed = max(time.monotonic() - started, 0.001)
        rate = idx / elapsed
        remaining = (total - idx) / rate if rate > 0 else 0
        msg = (
            f"{idx} / {total} ({100 * idx / total:.1f}%)  "
            f"— last: {last_status}  — ETA: {_fmt_dur(remaining)}"
        )
        self._post(MSG_PROGRESS, (idx, total, msg))

    def _poll_queue(self) -> None:
        try:
            while True:
                kind, payload = self._msg_queue.get_nowait()
                self._handle_msg(kind, payload)
        except queue.Empty:
            pass
        self.root.after(100, self._poll_queue)

    def _handle_msg(self, kind: str, payload: object) -> None:
        # Central usage telemetry — fire for any registered completion message
        # before the feature-specific handling, so a handler that raises can't
        # swallow the event. Best-effort; never affects the UI.
        _spec = _USAGE_EVENTS.get(kind)
        if _spec is not None:
            _action, _count_key = _spec
            _count = 0
            if _count_key and isinstance(payload, dict):
                try:
                    _count = int(payload.get(_count_key) or 0)
                except (TypeError, ValueError):
                    _count = 0
            self._track(action=_action, count=_count)
        _err_action = _USAGE_ERRORS.get(kind)
        if _err_action is not None:
            _note = str(payload)[:120] if payload else None
            self._track(action=_err_action, notes=_note)

        if kind == MSG_LOG:
            self._log(str(payload))
        elif kind == MSG_PROGRESS:
            idx, total, msg = payload  # type: ignore[misc]
            self.progress_bar["maximum"] = total
            self.progress_bar["value"] = idx
            self.progress_label.configure(text=msg)
        elif kind == MSG_RESULT:
            r: LookupResult = payload  # type: ignore[assignment]
            line = (
                f"{r.checked_at}  {r.iata_number:<12}  {r.status:<8}  "
                f"{r.trading_name}  ({r.country})"
            )
            self._log(line)
        elif kind == MSG_CAPTCHA:
            self._on_captcha_alert(str(payload))
        elif kind == MSG_DONE:
            self._log(f"Finished. File: {payload}")
            messagebox.showinfo("Done", f"Finished.\n\nOutput:\n{payload}")
            self._reset_buttons()
        elif kind == MSG_ERROR:
            self._log(f"ERROR: {payload}")
            messagebox.showerror("Error", str(payload))
            self._reset_buttons()
        # ------ BD tab messages ------
        elif kind == MSG_BD_LOG:
            self._bd_log(str(payload))
        elif kind == MSG_BD_PROGRESS:
            idx, total, msg = payload  # type: ignore[misc]
            self.bd_progress_bar["maximum"] = total
            self.bd_progress_bar["value"] = idx
            self.bd_progress_label.configure(text=msg)
        elif kind == MSG_BD_REFRESHED:
            count, last_refresh = payload  # type: ignore[misc]
            self._bd_log(f"Refreshed: {count:,} agencies cached at {last_refresh}.")
            self._refresh_bd_status_label()
            self._bd_reset_buttons()
        elif kind == MSG_BD_DONE:
            path = str(payload)
            self._bd_log(f"Done. File: {path}")
            messagebox.showinfo("BD Agency Lookup — Done", f"Finished.\n\nOutput:\n{path}")
            self._bd_reset_buttons()
        elif kind == MSG_BD_ERROR:
            self._bd_log(f"ERROR: {payload}")
            messagebox.showerror("BD Agency Lookup — Error", str(payload))
            self._bd_reset_buttons()
        # ------ OEP messages ------
        elif kind == MSG_OEP_BUSY:
            self.oep_status_label.configure(text=str(payload))
        elif kind == MSG_OEP_LOG:
            self.oep_status_label.configure(text=str(payload))
        elif kind == MSG_OEP_DONE:
            self._oep_render_results(payload)  # type: ignore[arg-type]
        elif kind == MSG_OEP_ERROR:
            self.oep_tree.delete(*self.oep_tree.get_children())
            self._oep_set_columns([("#", 40), ("Info", 600)])
            self.oep_tree.insert("", "end", values=("✕", str(payload)))
            self.oep_status_label.configure(text="Failed.")
            self.btn_oep_run.configure(state="normal")
            self.btn_oep_export.configure(state="disabled")
            messagebox.showerror("OEP", str(payload))
        # ------ Traffic tab messages ------
        elif kind in (MSG_TRAFFIC_BUSY, MSG_TRAFFIC_LOG):
            self.traffic_status_label.configure(text=str(payload))
        elif kind == MSG_TRAFFIC_DONE:
            self._traffic_render_results(payload)  # type: ignore[arg-type]
        elif kind == MSG_TRAFFIC_ERROR:
            self.btn_traffic_run.configure(state="normal")
            self.btn_traffic_export.configure(state="disabled")
            self.traffic_status_label.configure(text="Failed.")
            messagebox.showerror("Traffic", str(payload))
        # ------ Update messages ------
        elif kind == MSG_UPDATE_FOUND:
            self._on_update_found(payload)  # type: ignore[arg-type]
        elif kind == MSG_UPDATE_NONE:
            self.btn_check_updates.configure(state="normal", text="Check for updates")
            messagebox.showinfo(
                "Up to date",
                f"You're on the latest version (v{__version__}).",
            )
        elif kind == MSG_UPDATE_PROGRESS:
            downloaded, total = payload  # type: ignore[misc]
            if total > 0:
                pct = 100 * downloaded / total
                self.btn_check_updates.configure(
                    text=f"Downloading… {pct:.0f}%"
                )
            else:
                mb = downloaded / 1_000_000
                self.btn_check_updates.configure(text=f"Downloading… {mb:.0f} MB")
        elif kind == MSG_UPDATE_DOWNLOADED:
            self._on_update_downloaded(str(payload))
        elif kind == MSG_UPDATE_ERROR:
            self.btn_check_updates.configure(state="normal", text="Check for updates")
            messagebox.showerror("Update", str(payload))
        # ------ Zenith tab messages ------
        elif kind == MSG_ZENITH_LOGGED_IN:
            sess = payload  # type: ignore[assignment]
            self._zenith_session = sess
            sv = sess.state_values
            self.zenith_login_status.configure(
                text=(
                    f"Signed in · user={sv.get('ID_ADMIN')} · "
                    f"company={sv.get('ID_SOCIETE')} · app={sv.get('ID_APPLICATION')}"
                ),
                foreground=self._COLOR_SUCCESS,
                font=("Segoe UI Semibold", 10),
            )
            self.btn_zenith_login.configure(state="normal", text="Sign in again")
            # Don't keep the password in the widget after success.
            self.zenith_pwd_entry.delete(0, "end")
            self._zenith_log("Signed in to Zenith.")
            # Collapse the form to a one-line strip — the credentials
            # row is dead UI once auth is established.
            self._zenith_collapse_login()
        elif kind == MSG_ZENITH_LOGIN_FAILED:
            self.btn_zenith_login.configure(state="normal", text="Sign in to Zenith")
            self.zenith_login_status.configure(
                text="Not signed in",
                foreground=self._COLOR_MUTED, font=("Segoe UI", 9),
            )
            messagebox.showerror("Zenith login failed", str(payload))
        elif kind == MSG_ZENITH_PROGRESS:
            completed, total_n, ok, nf, err = payload  # type: ignore[misc]
            self.zenith_progress_bar.configure(maximum=total_n, value=completed)
            pct = 100 * completed / total_n if total_n else 0
            self.zenith_progress_label.configure(
                text=(
                    f"{completed:,} / {total_n:,} ({pct:.1f}%)  ·  "
                    f"{ok:,} OK · {nf:,} not found · {err:,} errors"
                ),
            )
        elif kind == MSG_ZENITH_RESULT:
            r = payload  # type: ignore[assignment]
            if r.status == zenith_client.STATUS_OK and r.record is not None:
                self._zenith_log(
                    f"{r.checked_at}  {r.customer_id}  OK   "
                    f"{r.record.first_name} {r.record.last_name}  ·  {r.record.email}",
                )
            elif r.status == zenith_client.STATUS_NOT_FOUND:
                self._zenith_log(
                    f"{r.checked_at}  {r.customer_id}  NOT_FOUND",
                )
            else:
                self._zenith_log(
                    f"{r.checked_at}  {r.customer_id}  ERROR  {r.error}",
                )
        elif kind == MSG_ZENITH_DONE:
            path = str(payload)
            self._zenith_log(f"Done. Wrote {path}")
            self._refresh_zenith_cache_label()
            self._zenith_reset_buttons()
            messagebox.showinfo(
                "Zenith — Run Done", f"Finished.\n\nOutput:\n{path}",
            )
        elif kind == MSG_ZENITH_ERROR:
            self._zenith_log(f"ERROR: {payload}")
            self._zenith_reset_buttons()
            messagebox.showerror("Zenith — Run Error", str(payload))
        # ------ Zenith Flight Loads messages ------
        elif kind == MSG_ZENITH_FL_PROGRESS:
            chunk_label, done, total, rows = payload  # type: ignore[misc]
            self.zenith_fl_progress_bar.configure(maximum=total, value=done)
            self.zenith_fl_progress_label.configure(
                text=f"Chunk {done}/{total}  ·  {chunk_label}  ·  {rows:,} rows so far",
            )
            self._zenith_fl_log(
                f"Chunk {done}/{total} done: {chunk_label}  ({rows:,} rows total)",
            )
        elif kind == MSG_ZENITH_FL_DONE:
            path = str(payload)
            self._zenith_fl_log(f"Done. Wrote {path}")
            self._zenith_fl_reset_buttons()
            # The Ordered-Report appender is now actionable.
            self.btn_zenith_fl_append_ordered.configure(state="normal")
            # Populate the per-leg drill-down grid for passenger detail.
            self._zenith_fl_populate_legs()
            messagebox.showinfo(
                "Zenith Flight Loads — Done",
                f"Finished.\n\nOutput:\n{path}",
            )
        # ----- Passenger manifest drill-down -----
        elif kind == MSG_ZENITH_PAX_PROGRESS:
            done, total, label, pax = payload  # type: ignore[misc]
            self.zenith_fl_progress_bar.configure(maximum=total, value=done)
            self.zenith_fl_pax_status.configure(
                text=f"Leg {done}/{total}  ·  {label}  ·  {pax:,} pax so far",
            )
            self._zenith_fl_log(f"  manifest {done}/{total}: {label} ({pax:,} pax)")
        elif kind == MSG_ZENITH_PAX_DONE:
            info = payload  # type: ignore[assignment]
            self.btn_zenith_fl_pax.configure(state="normal")
            self.btn_zenith_fl_pax_all.configure(state="normal")
            self.zenith_fl_pax_status.configure(
                text=f"Done — {info['pax']:,} passengers from {info['legs']} legs",
            )
            self._zenith_fl_log(f"Passenger manifest written: {info['path']}")
            messagebox.showinfo(
                "Passenger detail — Done",
                f"Saved to: {info['path']}\n\n"
                f"  Legs pulled: {info['legs']}\n"
                f"  Passengers:  {info['pax']:,}",
            )
        elif kind == MSG_ZENITH_PAX_ERROR:
            self.btn_zenith_fl_pax.configure(state="normal")
            self.btn_zenith_fl_pax_all.configure(state="normal")
            self.zenith_fl_pax_status.configure(text=f"⚠ {payload}")
            self._zenith_fl_log(f"ERROR (passenger detail): {payload}")
            messagebox.showerror("Passenger detail — Error", str(payload))
        elif kind == MSG_ZENITH_FL_ERROR:
            self._zenith_fl_log(f"ERROR: {payload}")
            self._zenith_fl_reset_buttons()
            messagebox.showerror("Zenith Flight Loads — Error", str(payload))
        # ----- Flight History Analyzer -----
        elif kind == MSG_ZENITH_FH_PROGRESS:
            i, total, name = payload  # type: ignore[misc]
            self.zenith_fh_progress.configure(maximum=total, value=i)
            self.zenith_fh_status_label.configure(
                text=f"Parsing {i}/{total}: {name}",
            )
        elif kind == MSG_ZENITH_FH_PARSED:
            count = int(payload)  # type: ignore[arg-type]
            self.zenith_fh_status_label.configure(
                text=f"Parsed {count:,} events. Running audit…",
            )
        elif kind == MSG_ZENITH_FH_DONE:
            report = payload
            self._zenith_fh_last_report = report
            self._zenith_fh_render(report)
            self.btn_zenith_fh_run.configure(state="normal")
            self.btn_zenith_fh_export.configure(state="normal")
            self.btn_zenith_fh_pnr.configure(state="normal")
            self.btn_zenith_fh_misuse.configure(state="normal")
            self.zenith_fh_status_label.configure(
                text=(
                    f"Done — {report.event_count:,} events / "
                    f"{report.file_count} files. Click Export to save."
                )
            )
        elif kind == MSG_ZENITH_FH_ERROR:
            self.btn_zenith_fh_run.configure(state="normal")
            self.btn_zenith_fh_misuse.configure(state="normal")
            self.zenith_fh_status_label.configure(text=f"⚠ {payload}")
            messagebox.showerror("Flight History Analyzer — Error", str(payload))
        elif kind == MSG_ZENITH_PNRMISUSE_DONE:
            self.btn_zenith_fh_misuse.configure(state="normal")
            self.btn_zenith_fh_run.configure(state="normal")
            p = payload
            self.zenith_fh_status_label.configure(
                text=f"PNR misuse: {p['flags']} flag(s) → {Path(p['path']).name}")
            if messagebox.askyesno(
                    "PNR Misuse audit",
                    f"Analysed {p['events']:,} events across {p['pnrs']:,} PNRs.\n"
                    f"Flags: {p['flags']}  (critical {p['critical']}, high {p['high']}).\n"
                    f"Top risk: {p['top']}\n\nSaved to:\n{p['path']}\n\n"
                    "Flags are review leads, not findings. Open the workbook now?"):
                try:
                    import os
                    os.startfile(p["path"])  # noqa: S606
                except Exception:  # noqa: BLE001
                    pass
        elif kind == MSG_ZENITH_PNRMISUSE_ERROR:
            self.btn_zenith_fh_misuse.configure(state="normal")
            self.btn_zenith_fh_run.configure(state="normal")
            self.zenith_fh_status_label.configure(text=f"⚠ {payload}")
            messagebox.showerror("PNR Misuse audit — Error", str(payload))
        # ----- PNR Dossier audit (Phase 2: payment / contact) -----
        elif kind == MSG_ZENITH_DOSSIER_PROGRESS:
            i, n, pnr = payload
            self.zenith_bulk_progress.configure(value=i, maximum=max(n, 1))
            self.zenith_bulk_status_label.configure(text=f"Dossier {i}/{n}: {pnr}")
        elif kind == MSG_ZENITH_DOSSIER_DONE:
            self.btn_zenith_bulk_run.configure(state="normal")
            self.btn_zenith_dossier.configure(state="normal")
            self.btn_zenith_bulk_stop.configure(state="disabled")
            p = payload
            note = " (stopped early)" if p.get("aborted") else ""
            self.zenith_bulk_status_label.configure(
                text=f"Dossier audit: {p['flags']} flag(s) → {Path(p['path']).name}{note}")
            self._zenith_bulk_log_line(
                f"Done{note}: {p['events']} events, {p['pnrs']} PNRs · "
                f"scraped {p['scraped']}, cached {p['cached']}, failed {p['failed']} · "
                f"{p['flags']} flags (crit {p['critical']}, high {p['high']}), "
                f"{p['txn']} distinct txn")
            if messagebox.askyesno(
                    "Dossier audit",
                    f"Scraped {p['pnrs']:,} PNRs ({p['scraped']} live, {p['cached']} cached, "
                    f"{p['failed']} failed){note}.\n"
                    f"Flags: {p['flags']} (critical {p['critical']}, high {p['high']}); "
                    f"{p['txn']} distinct transaction ids.\n\nSaved to:\n{p['path']}\n\n"
                    "Flags are review leads, not findings. Open the workbook now?"):
                try:
                    import os
                    os.startfile(p["path"])  # noqa: S606
                except Exception:  # noqa: BLE001
                    pass
        elif kind == MSG_ZENITH_DOSSIER_ERROR:
            self.btn_zenith_bulk_run.configure(state="normal")
            self.btn_zenith_dossier.configure(state="normal")
            self.btn_zenith_bulk_stop.configure(state="disabled")
            self.zenith_bulk_status_label.configure(text=f"⚠ {payload}")
            messagebox.showerror("Dossier audit — Error", str(payload))
        # ----- Flight History downloader (Phase 3) -----
        elif kind == MSG_ZENITH_DL_STATUS:
            self.zenith_fh_status_label.configure(text=str(payload))
        elif kind == MSG_ZENITH_DL_PROGRESS:
            idx, total, label, status_code = payload  # type: ignore[misc]
            self.zenith_fh_progress.configure(maximum=total, value=idx)
            self.zenith_fh_status_label.configure(
                text=f"Downloading {idx}/{total}  ·  {label}  [{status_code}]",
            )
        elif kind == MSG_ZENITH_DL_DONE:
            info = payload  # type: ignore[assignment]
            folder = info["folder"]
            summary = info["summary"]
            total = info["total"]
            self.btn_zenith_fh_download.configure(state="normal")
            ok = summary.get("OK", 0)
            skipped = summary.get("SKIP_EXISTS", 0)
            empty = summary.get("EMPTY", 0)
            errors = summary.get("ERROR", 0)
            self.zenith_fh_status_label.configure(
                text=(
                    f"Download done — {ok} new, {skipped} skipped, "
                    f"{empty} empty, {errors} errors / {total} flights"
                )
            )
            messagebox.showinfo(
                "Download from Zenith — Done",
                f"Saved to: {folder}\n\n"
                f"  OK:      {ok}\n"
                f"  Skipped: {skipped}\n"
                f"  Empty:   {empty}\n"
                f"  Errors:  {errors}\n"
                f"  Total:   {total}",
            )
        elif kind == MSG_ZENITH_DL_ERROR:
            self.btn_zenith_fh_download.configure(state="normal")
            self.zenith_fh_status_label.configure(text=f"⚠ Download: {payload}")
            messagebox.showerror("Download from Zenith — Error", str(payload))
        # ----- PNR enrichment (Phase 2) -----
        elif kind == MSG_ZENITH_PNR_PROGRESS:
            idx, total, code, status_code = payload  # type: ignore[misc]
            self.zenith_fh_progress.configure(maximum=total, value=idx)
            self.zenith_fh_status_label.configure(
                text=f"PNR {idx}/{total}  ·  {code}  [{status_code}]",
            )
        elif kind == MSG_ZENITH_PNR_DONE:
            enriched = payload
            self._zenith_fh_last_report = enriched
            self._zenith_fh_render(enriched)
            self.btn_zenith_fh_run.configure(state="normal")
            self.btn_zenith_fh_pnr.configure(state="normal")
            self.zenith_fh_status_label.configure(
                text=(
                    f"PNR enrichment done — {len(enriched.pnr_routes)} PNRs "
                    f"in the route audit. Click Export to save."
                )
            )
        elif kind == MSG_ZENITH_PNR_ERROR:
            self.btn_zenith_fh_run.configure(state="normal")
            self.btn_zenith_fh_pnr.configure(state="normal")
            self.zenith_fh_status_label.configure(text=f"⚠ PNR enrich: {payload}")
            messagebox.showerror("PNR enrichment — Error", str(payload))
        # ----- PNR Bulk Lookup -----
        elif kind == MSG_ZENITH_BULK_PROGRESS:
            idx, total, code, status_code = payload  # type: ignore[misc]
            self.zenith_bulk_progress.configure(maximum=total, value=idx)
            # Live counters so users see exactly where the money is going.
            counters = self._zenith_bulk_counters
            counters[status_code] = counters.get(status_code, 0) + 1
            ok = counters.get("OK", 0)
            cached = counters.get("CACHED", 0)
            errors = counters.get("NOT_FOUND", 0) + sum(
                v for k, v in counters.items()
                if k not in ("OK", "CACHED", "CANCELLED")
            )
            # Throughput + ETA from a sliding wall clock.
            import time
            now = time.monotonic()
            if self._zenith_bulk_started_at is None:
                self._zenith_bulk_started_at = now
            elapsed = max(now - self._zenith_bulk_started_at, 0.001)
            rate = idx / elapsed
            remaining = (total - idx) / rate if rate > 0 else 0
            eta = _fmt_dur(remaining) if remaining else "—"
            self.zenith_bulk_status_label.configure(
                text=(
                    f"{idx:,}/{total:,}  ·  "
                    f"{ok:,} fetched · {cached:,} cached · {errors:,} errors  ·  "
                    f"{rate:.1f}/s  ·  ETA {eta}"
                ),
                style="Hint.TLabel",
            )
            self._zenith_bulk_log_line(f"  {idx}/{total} {code}: {status_code}")
        elif kind == MSG_ZENITH_BULK_DONE:
            info = payload  # type: ignore[assignment]
            self.btn_zenith_bulk_run.configure(state="normal")
            self.btn_zenith_bulk_stop.configure(state="disabled")
            self.zenith_bulk_status_label.configure(
                text=(
                    f"Done — {info['ok']}/{info['total']} resolved, "
                    f"{info['errors']} errors"
                ),
                style="Success.TLabel" if info["errors"] == 0 else "Warning.TLabel",
            )
            # Reset for the next run.
            self._zenith_bulk_counters = {}
            self._zenith_bulk_started_at = None
            self._zenith_bulk_log_line(f"Wrote: {info['path']}")
            messagebox.showinfo(
                "PNR Bulk Lookup — Done",
                f"Saved to: {info['path']}\n\n"
                f"  Total:    {info['total']}\n"
                f"  Resolved: {info['ok']}\n"
                f"  Errors:   {info['errors']}",
            )
        elif kind == MSG_ZENITH_BULK_ERROR:
            self.btn_zenith_bulk_run.configure(state="normal")
            self.btn_zenith_bulk_stop.configure(state="disabled")
            self.zenith_bulk_status_label.configure(text=f"⚠ {payload}")
            self._zenith_bulk_log_line(f"ERROR: {payload}")
            messagebox.showerror("PNR Bulk Lookup — Error", str(payload))
        # ----- Bulk Mailer -----
        elif kind == MSG_MAIL_PROGRESS:
            idx, total, to, status, row_index = payload  # type: ignore[misc]
            self.mail_progress.configure(maximum=total, value=idx)
            self.mail_status.configure(text=f"{idx}/{total}  ·  {to}  [{status}]")
            iid = getattr(self, "_mail_iid_by_row", {}).get(row_index)
            if iid is not None:
                vals = list(self.mail_tree.item(iid, "values"))
                vals[6] = status
                tag = "ok" if status in ("DRAFTED", "SENT", "SKIPPED") else "bad"
                self.mail_tree.item(iid, values=vals, tags=(tag,))
        elif kind == MSG_MAIL_DONE:
            info = payload  # type: ignore[assignment]
            self.btn_mail_run.configure(state="normal")
            self.btn_mail_test.configure(state="normal")
            self.btn_mail_preview.configure(state="normal")
            self.btn_mail_stop.configure(state="disabled")
            if info.get("test"):
                st = info.get("outcome_status", "?")
                err = info.get("error", "")
                self.mail_status.configure(text=f"Test {st} → {info.get('to','')}")
                if st == "FAILED":
                    messagebox.showerror("Bulk Mailer — Test", err or "Test failed.")
                else:
                    messagebox.showinfo(
                        "Bulk Mailer — Test",
                        f"Test {st.lower()} to {info.get('to','')}.\n\n"
                        + ("Check your Outlook Drafts." if st == "DRAFTED"
                           else "Check your inbox."),
                    )
            else:
                c = info.get("counts", {})
                verb = "drafted" if info.get("mode") == "draft" else "sent"
                summary = (
                    f"{c.get('DRAFTED',0)} drafted · {c.get('SENT',0)} sent · "
                    f"{c.get('SKIPPED',0)} skipped · {c.get('FAILED',0)} failed"
                )
                self.mail_status.configure(text=summary)
                if info.get("mode") == "draft":
                    dd = info.get("draft_dir")
                    if dd:
                        review = (f"\n\n.eml drafts written to:\n{dd}\n"
                                  "Open any to review; double-click to send.")
                    else:
                        review = ("\n\nReview them in your Drafts folder "
                                  "(Outlook desktop or Outlook web), then send.")
                else:
                    review = ""
                messagebox.showinfo(
                    "Bulk Mailer — Done", f"Run complete ({verb}).\n\n{summary}{review}",
                )
        elif kind == MSG_MAIL_ERROR:
            self.btn_mail_run.configure(state="normal")
            self.btn_mail_test.configure(state="normal")
            self.btn_mail_preview.configure(state="normal")
            self.btn_mail_stop.configure(state="disabled")
            self.mail_status.configure(text=f"⚠ {payload}")
            messagebox.showerror("Bulk Mailer — Error", str(payload))
        # ----- Graph (M365) sign-in -----
        elif kind == "mail_graph_prompt":
            # payload = (full_message, user_code). Keep the code visible on
            # screen (status bar) so dismissing the popup doesn't lose it,
            # and RE-ENABLE the sign-in button so a fumbled code can be
            # retried without restarting the app.
            msg, code = payload  # type: ignore[misc]
            self.mail_graph_status.set(f"Code: {code}  →  microsoft.com/devicelogin")
            self.btn_mail_graph_signin.configure(state="normal")
            import webbrowser
            webbrowser.open("https://microsoft.com/devicelogin")
            messagebox.showinfo(
                "Microsoft 365 sign-in",
                f"{msg}\n\nThe code is also shown on the tab "
                f"({code}) in case you need it again.",
            )
        elif kind == "mail_graph_signed_in":
            self._graph_session = payload
            acct = getattr(payload, "account", "")
            self.mail_graph_status.set(f"● Signed in: {acct}")
            self.btn_mail_graph_signin.configure(state="normal")
            messagebox.showinfo("Microsoft 365", f"Signed in as {acct}.")
        elif kind == "mail_graph_signin_failed":
            self.btn_mail_graph_signin.configure(state="normal")
            # A stale/expired device-code worker can fire after a newer
            # sign-in already succeeded — don't clobber that good state.
            if self._graph_session is None:
                self.mail_graph_status.set("Sign-in failed — click Sign in to retry")
                messagebox.showerror("Microsoft 365 sign-in", str(payload))

    def _on_captcha_alert(self, message: str) -> None:
        self._log(f"CAPTCHA: {message}")
        try:
            winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
            for _ in range(2):
                winsound.Beep(880, 250)
                time.sleep(0.05)
        except Exception:
            pass
        # Bring window to front to nudge the user
        try:
            self.root.attributes("-topmost", True)
            self.root.after(100, lambda: self.root.attributes("-topmost", False))
            self.root.bell()
        except Exception:
            pass
        self.progress_label.configure(text="⚠ CAPTCHA — solve it in the browser, then keep working.")

    def _reset_buttons(self) -> None:
        self.btn_start.configure(state="normal")
        self.btn_pause.configure(state="disabled")
        self.btn_resume.configure(state="disabled")
        self.btn_stop.configure(state="disabled")

    def _log(self, msg: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    # ==================================================================
    # BD Travel Agency tab — actions + worker
    # ==================================================================

    def _refresh_bd_status_label(self) -> None:
        """Update the 'Last fetched: ...' label using the cache."""
        last = self._bd_cache.last_refresh()
        count = self._bd_cache.count()
        if not last or count == 0:
            self.bd_status_label.configure(
                text="No cached data yet. Click Refresh to download the agency list."
            )
            return
        self.bd_status_label.configure(
            text=f"Last fetched: {last}  ·  {count:,} cached records"
        )

    def _toggle_bd_mode(self) -> None:
        """Enable/disable input pickers based on current mode."""
        is_lookup = self.bd_mode.get() == "lookup"
        state = "normal" if is_lookup else "disabled"
        for widget in (
            self.bd_input_entry,
            self.bd_input_btn,
            self.bd_sheet_combo,
            self.bd_col_combo,
        ):
            widget.configure(state=state if widget is not self.bd_sheet_combo and widget is not self.bd_col_combo else ("readonly" if is_lookup else "disabled"))

    def _bd_pick_input(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Excel with agency names",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if not path:
            return
        self.bd_input_path.set(path)
        self._bd_reload_sheets()

    def _bd_pick_output(self) -> None:
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.bd_output_dir.set(path)

    def _bd_reload_sheets(self) -> None:
        try:
            sheets = excel_io.list_sheet_names(Path(self.bd_input_path.get()))
        except Exception as e:
            messagebox.showerror("Cannot read Excel", str(e))
            return
        self.bd_sheet_combo["values"] = sheets
        if sheets:
            self.bd_sheet_combo.current(0)
            self._bd_reload_columns()

    def _bd_reload_columns(self) -> None:
        try:
            cols = excel_io.list_columns(
                Path(self.bd_input_path.get()), self.bd_sheet_name.get()
            )
        except Exception as e:
            messagebox.showerror("Cannot read columns", str(e))
            return
        self._bd_sheet_columns = cols
        self.bd_col_combo["values"] = cols
        guess = next(
            (c for c in cols if "name" in c.lower() or "agency" in c.lower() or "license" in c.lower()),
            cols[0] if cols else "",
        )
        self.bd_column_name.set(guess)

    # ------------------------------------------------------------------
    # Refresh worker
    # ------------------------------------------------------------------

    def _bd_refresh(self) -> None:
        if self._bd_worker is not None and self._bd_worker.is_alive():
            messagebox.showinfo("Busy", "Another BD task is already running.")
            return
        self.btn_bd_refresh.configure(state="disabled")
        self.btn_bd_run.configure(state="disabled")
        self._post(MSG_BD_LOG, "Refreshing agency list from regtravelagency.gov.bd ...")
        self._bd_worker = threading.Thread(
            target=self._bd_run_refresh,
            daemon=True,
        )
        self._bd_worker.start()

    def _bd_run_refresh(self) -> None:
        try:
            agencies = fetch_all_agencies()
            self._bd_cache.replace_all(agencies)
            self._post(
                MSG_BD_REFRESHED,
                (len(agencies), self._bd_cache.last_refresh()),
            )
            try:
                auth.log_lookup_event(
                    action="bd_refresh",
                    target="get-list",
                    count=len(agencies),
                )
            except Exception as exc:  # noqa: BLE001
                log.warning("usage log failed: %s", exc)
        except Exception as e:
            log.exception("BD refresh failed")
            self._post(MSG_BD_ERROR, f"Refresh failed: {type(e).__name__}: {e}")

    # ------------------------------------------------------------------
    # Run worker (full export OR lookup)
    # ------------------------------------------------------------------

    def _bd_run(self) -> None:
        if self._bd_worker is not None and self._bd_worker.is_alive():
            messagebox.showinfo("Busy", "Another BD task is already running.")
            return
        if self._bd_cache.count() == 0:
            messagebox.showerror(
                "No cached data",
                "Click Refresh first to download the agency list.",
            )
            return

        out_dir = Path(self.bd_output_dir.get())
        if not out_dir.exists():
            messagebox.showerror("Invalid output", "Output folder does not exist.")
            return

        mode = self.bd_mode.get()
        cfg: dict = {
            "mode": mode,
            "output_dir": out_dir,
            "include_expired": bool(self.bd_include_expired.get()),
        }
        if mode == "lookup":
            input_path = self.bd_input_path.get().strip()
            if not input_path or not Path(input_path).exists():
                messagebox.showerror("Invalid input", "Pick a valid Excel file.")
                return
            if not self.bd_sheet_name.get():
                messagebox.showerror("Invalid input", "Pick a sheet.")
                return
            if (
                not self.bd_column_name.get()
                or self.bd_column_name.get() not in self._bd_sheet_columns
            ):
                messagebox.showerror("Invalid input", "Pick the column.")
                return

            # Field selector — pick the checked ones in priority order.
            fields: list[str] = []
            if self.bd_match_name.get():
                fields.append(FIELD_NAME)
            if self.bd_match_license.get():
                fields.append(FIELD_LICENSE)
            if self.bd_match_address.get():
                fields.append(FIELD_ADDRESS)
            if not fields:
                messagebox.showerror(
                    "Invalid input",
                    "Tick at least one field to match against "
                    "(Agency Name / License Number / Address).",
                )
                return

            cfg["input_path"] = Path(input_path)
            cfg["sheet"] = self.bd_sheet_name.get()
            cfg["column_index"] = self._bd_sheet_columns.index(self.bd_column_name.get())
            cfg["match_fields"] = tuple(fields)

        self.btn_bd_refresh.configure(state="disabled")
        self.btn_bd_run.configure(state="disabled")
        self._bd_worker = threading.Thread(
            target=self._bd_worker_run,
            args=(cfg,),
            daemon=True,
        )
        self._bd_worker.start()

    def _bd_worker_run(self, cfg: dict) -> None:
        try:
            agencies = self._bd_cache.all()
            agencies = filter_status(agencies, cfg["include_expired"])
            self._post(MSG_BD_LOG, f"Loaded {len(agencies):,} cached agencies after filter.")

            if cfg["mode"] == "full":
                output_path = excel_io.build_bd_output_path(cfg["output_dir"], kind="full")
                self._post(MSG_BD_LOG, f"Writing full list to {output_path} ...")
                excel_io.write_bd_full_list(output_path, agencies)
                self._post(MSG_BD_DONE, str(output_path))
                try:
                    auth.log_lookup_event(
                        action="bd_export",
                        target="full-list",
                        count=len(agencies),
                    )
                except Exception as exc:  # noqa: BLE001
                    log.warning("usage log failed: %s", exc)
                return

            # Lookup mode
            inputs = excel_io.read_iata_numbers(  # reuses generic Excel reader
                cfg["input_path"],
                cfg["sheet"],
                cfg["column_index"],
                start_row=2,
                end_row=None,
            )
            # read_iata_numbers strips and normalises — but BD names need
            # less normalisation. Re-read raw values so names like
            # "ZEPHYR TOURS & TRAVELS" survive.
            inputs = self._read_bd_inputs(
                cfg["input_path"], cfg["sheet"], cfg["column_index"]
            )
            if not inputs:
                self._post(MSG_BD_ERROR, "No values found in the selected column.")
                return

            self._post(MSG_BD_LOG, f"Building search index over {len(agencies):,} agencies ...")
            index = AgencyIndex(agencies)

            fields = cfg.get("match_fields") or (FIELD_NAME, FIELD_LICENSE)
            self._post(
                MSG_BD_LOG,
                f"Matching {len(inputs):,} input rows against fields: {', '.join(fields)} ...",
            )
            results = []
            total = len(inputs)
            for i, (_, value) in enumerate(inputs, start=1):
                results.append(index.lookup(value, fields=fields))
                if i % 50 == 0 or i == total:
                    self._post(MSG_BD_PROGRESS, (i, total, f"{i}/{total} matched"))

            output_path = excel_io.build_bd_output_path(cfg["output_dir"], kind="lookup")
            self._post(MSG_BD_LOG, f"Writing lookup results to {output_path} ...")
            excel_io.write_bd_lookup_results(output_path, results)

            # Summary
            counts: dict[str, int] = {}
            for r in results:
                counts[r.match_method] = counts.get(r.match_method, 0) + 1
            summary = ", ".join(
                f"{k}={v}"
                for k, v in sorted(counts.items(), key=lambda kv: -kv[1])
            )
            self._post(MSG_BD_LOG, f"Match summary: {summary}")
            self._post(MSG_BD_DONE, str(output_path))

            try:
                auth.log_lookup_event(
                    action="bd_lookup",
                    target=str(cfg["input_path"].name),
                    count=len(inputs),
                    notes="fields=" + "/".join(fields),
                )
            except Exception as exc:  # noqa: BLE001
                log.warning("usage log failed: %s", exc)

        except Exception as e:
            log.exception("BD worker crashed")
            self._post(MSG_BD_ERROR, f"{type(e).__name__}: {e}")

    @staticmethod
    def _read_bd_inputs(
        path: Path, sheet: str, column_index: int
    ) -> list[tuple[int, str]]:
        """Read a column of free-form text from Excel.

        Unlike `read_iata_numbers`, we don't strip hyphens/spaces — agency
        names need to keep their formatting for matching.
        """
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet]
            rows: list[tuple[int, str]] = []
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                if column_index >= len(row):
                    continue
                value = row[column_index]
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                rows.append((row_idx, text))
            return rows
        finally:
            wb.close()

    # ------------------------------------------------------------------
    # BD message handling
    # ------------------------------------------------------------------

    def _bd_log(self, msg: str) -> None:
        self.bd_log_text.configure(state="normal")
        self.bd_log_text.insert("end", msg + "\n")
        self.bd_log_text.see("end")
        self.bd_log_text.configure(state="disabled")

    def _bd_reset_buttons(self) -> None:
        self.btn_bd_refresh.configure(state="normal")
        self.btn_bd_run.configure(state="normal")

    # ==================================================================
    # Auth + status bar
    # ==================================================================

    def ensure_signed_in(self) -> bool:
        """Show the login dialog if no valid session, return True on success."""
        # 1. If we already have a token, validate it server-side.
        if auth.is_signed_in():
            user = auth.whoami()
            if user:
                self._signed_in_user = user
                self._refresh_user_label()
                self._track(action="login", notes="resumed")
                return True
            # Stale token — drop it and prompt fresh sign-in.
            auth.clear_token()

        # 2. Show modal sign-in dialog.
        ok = self._show_login_dialog()
        if ok:
            self._track(action="login", notes="interactive")
        return ok

    def _show_login_dialog(self) -> bool:
        dlg = tk.Toplevel(self.root)
        dlg.title("Sign in — IATA Code Validator")
        dlg.geometry("480x230")
        dlg.resizable(False, False)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.protocol("WM_DELETE_WINDOW", lambda: dlg.destroy())

        ttk.Label(
            dlg,
            text="Sign in to continue",
            font=("Segoe UI", 14, "bold"),
        ).pack(pady=(20, 6), padx=16)
        ttk.Label(
            dlg,
            text=(
                "The IATA Code Validator is licensed to your team. "
                "Click below and complete the Google sign-in in your browser."
            ),
            wraplength=440,
            justify="left",
        ).pack(padx=16, pady=(0, 16))

        status_var = tk.StringVar(value="")
        status_label = ttk.Label(dlg, textvariable=status_var, foreground="#b91c1c")
        status_label.pack(padx=16)

        result = {"ok": False}

        def do_signin() -> None:
            btn_signin.configure(state="disabled", text="Opening browser…")
            status_var.set("")
            dlg.update_idletasks()
            try:
                oauth_result = auth.run_google_oauth_flow()
                auth.save_token(oauth_result.session_token)
                self._signed_in_user = {
                    "email": oauth_result.email,
                    "full_name": oauth_result.name,
                    "user_id": oauth_result.user_id,
                }
                result["ok"] = True
                dlg.destroy()
            except auth.AuthError as exc:
                status_var.set(str(exc))
                btn_signin.configure(state="normal", text="Sign in with Google")
            except Exception as exc:  # noqa: BLE001
                status_var.set(f"Unexpected error: {exc}")
                btn_signin.configure(state="normal", text="Sign in with Google")

        btn_signin = ttk.Button(dlg, text="Sign in with Google", command=do_signin)
        btn_signin.pack(pady=(12, 8))

        ttk.Button(dlg, text="Quit", command=dlg.destroy).pack()

        self.root.wait_window(dlg)
        if result["ok"]:
            self._refresh_user_label()
        return result["ok"]

    def _refresh_user_label(self) -> None:
        # sv_ttk overrides named-style foregrounds for tk.Label children,
        # so set the colour directly on the widget — that always wins.
        font_bold = ("Segoe UI Semibold", 10)
        if self._signed_in_user:
            email = self._signed_in_user.get("email") or "(signed in)"
            self.status_user_label.configure(
                text=f"●  Signed in: {email}",
                foreground=self._COLOR_SUCCESS, font=font_bold,
            )
            self.btn_sign_in.pack_forget()
            self.btn_sign_out.pack(
                side="right", padx=4, pady=2, before=self.btn_check_updates,
            )
        else:
            self.btn_sign_out.pack_forget()
            if auth.GOOGLE_CLIENT_ID:
                self.status_user_label.configure(
                    text="○  Not signed in",
                    foreground=self._COLOR_WARNING, font=font_bold,
                )
                self.btn_sign_in.pack(
                    side="right", padx=4, pady=2, before=self.btn_check_updates,
                )
            else:
                self.status_user_label.configure(
                    text="○  Unauthenticated (dev build)",
                    foreground=self._COLOR_MUTED,
                    font=("Segoe UI", 9),
                )
                self.btn_sign_in.pack_forget()

    def _on_sign_in(self) -> None:
        """Status-bar entry point for signing in after launch."""
        if self._signed_in_user:
            return
        if self._show_login_dialog():
            # _show_login_dialog already updates the label on success.
            pass

    def _on_sign_out(self) -> None:
        if not messagebox.askyesno(
            "Sign out?",
            "Sign out of the IATA Code Validator? You will need to sign in again on next launch.",
        ):
            return
        try:
            auth.logout()
        except Exception as exc:  # noqa: BLE001
            log.warning("logout error: %s", exc)
        self._signed_in_user = None
        self._refresh_user_label()
        messagebox.showinfo("Signed out", "Signed out. Please restart the app.")

    # ==================================================================
    # Self-updater
    # ==================================================================

    def _toggle_theme(self) -> None:
        """Flip the sv_ttk theme + relabel the toggle.

        sv_ttk handles re-painting every widget; we just have to nudge
        a few of our custom-styled labels back to their semantic colors
        in case the theme reset their foregrounds.
        """
        try:
            import sv_ttk
            self._theme_is_dark = not self._theme_is_dark
            sv_ttk.set_theme("dark" if self._theme_is_dark else "light")
            self.btn_theme.configure(
                text="☀  Light" if self._theme_is_dark else "☾  Dark",
            )
            # Re-apply our semantic styles — sv_ttk's set_theme resets
            # every style table including the ones we configured.
            self._setup_styles()
            # Force a refresh of the auth status chip so its color survives.
            self._refresh_user_label()
        except Exception as exc:  # noqa: BLE001
            log.warning("Theme toggle failed: %s", exc)

    def _on_check_for_updates(self) -> None:
        if self._update_worker is not None and self._update_worker.is_alive():
            return
        self.btn_check_updates.configure(state="disabled", text="Checking…")
        self._update_worker = threading.Thread(
            target=self._update_check_worker, daemon=True
        )
        self._update_worker.start()

    def _update_check_worker(self) -> None:
        info = updater.check_for_update()
        if info is None:
            self._post(MSG_UPDATE_ERROR, "Could not check for updates (network error or no published release).")
            return
        if info.is_newer:
            self._post(MSG_UPDATE_FOUND, info)
        else:
            self._post(MSG_UPDATE_NONE, info)

    def _on_update_found(self, info: "updater.UpdateInfo") -> None:
        self.btn_check_updates.configure(state="normal", text="Check for updates")
        msg = (
            f"A new version is available.\n\n"
            f"Current : v{__version__}\n"
            f"Latest  : v{info.latest_version}\n\n"
            f"Download (~380 MB) and restart now?"
        )
        if not messagebox.askyesno("Update available", msg):
            return
        self.btn_check_updates.configure(state="disabled", text="Downloading…")
        self._update_worker = threading.Thread(
            target=self._update_download_worker,
            args=(info.download_url, info.sha256),
            daemon=True,
        )
        self._update_worker.start()

    def _update_download_worker(self, url: str, expected_sha256: str = "") -> None:
        def progress(downloaded: int, total: int) -> None:
            self._post(MSG_UPDATE_PROGRESS, (downloaded, total))
        try:
            staged = updater.download_update(
                url, on_progress=progress, expected_sha256=expected_sha256,
            )
            self._post(MSG_UPDATE_DOWNLOADED, str(staged))
        except Exception as exc:  # noqa: BLE001
            log.exception("update download failed")
            self._post(MSG_UPDATE_ERROR, f"Download failed: {exc}")

    def _on_update_downloaded(self, staged_path: str) -> None:
        self.btn_check_updates.configure(text="Restart to update", state="normal")
        if not messagebox.askyesno(
            "Download complete",
            "The new version is ready. The app will close and the new version will launch automatically. Continue?",
        ):
            return
        ok = updater.apply_update_and_exit()
        if not ok:
            messagebox.showwarning(
                "Cannot self-update",
                "Self-update is only supported in the bundled .exe. "
                f"Your downloaded file is at:\n{staged_path}",
            )


    # ==================================================================
    # Zenith Customer Lookup tab — UI
    # ==================================================================

    def _build_zenith_tab(self, parent: ttk.Frame) -> None:
        parent = self._make_scrollable(parent)

        self._section(
            parent, "Zenith  ·  usba.ttinteractive.com",
            help_text=(
                "Four sub-tabs: Customer Lookup, Flight Loads, "
                "Flight History Analyzer, PNR Bulk Lookup. "
                "Sign in once below; the session is shared across all four."
            ),
        )

        # ----- Login -----
        # Wrap the login section in its own frame so we can collapse it
        # to a one-line summary once the user is authenticated.
        self.zenith_login_section = ttk.Frame(parent)
        self.zenith_login_section.pack(fill="x", padx=4, pady=(8, 4))
        login_body = self._section(self.zenith_login_section, "Sign in to Zenith")
        login_row = ttk.Frame(login_body)
        login_row.pack(fill="x", padx=2)
        self.zenith_user_entry = ttk.Entry(
            login_row, textvariable=self.zenith_username, width=24,
        )
        self.zenith_pwd_entry = ttk.Entry(
            login_row, show="•", width=24,
        )
        self.zenith_company_entry = ttk.Entry(
            login_row, textvariable=self.zenith_company, width=10,
        )
        ttk.Label(login_row, text="Username:").grid(
            row=0, column=0, padx=(0, 4), pady=4, sticky="w",
        )
        self.zenith_user_entry.grid(row=0, column=1, padx=(0, 12), pady=4)
        ttk.Label(login_row, text="Password:").grid(
            row=0, column=2, padx=(0, 4), pady=4, sticky="w",
        )
        self.zenith_pwd_entry.grid(row=0, column=3, padx=(0, 12), pady=4)
        ttk.Label(login_row, text="Company:").grid(
            row=0, column=4, padx=(0, 4), pady=4, sticky="w",
        )
        self.zenith_company_entry.grid(row=0, column=5, padx=(0, 12), pady=4)
        self.btn_zenith_login = ttk.Button(
            login_row, text="Sign in to Zenith",
            command=self._zenith_login, width=18,
        )
        self.btn_zenith_login.grid(row=0, column=6, padx=(0, 4), pady=4)
        self.zenith_login_status = ttk.Label(
            login_body, text="Not signed in", style="Hint.TLabel",
        )
        self.zenith_login_status.pack(anchor="w", padx=2, pady=(0, 4))

        # Compact one-line strip shown in place of the login form once
        # the user is authenticated. Hidden by default — the
        # `_refresh_zenith_login_collapse` method swaps the two views.
        self.zenith_login_compact = ttk.Frame(parent)
        self.zenith_login_compact_label = ttk.Label(
            self.zenith_login_compact, text="",
            foreground=self._COLOR_SUCCESS,
            font=("Segoe UI Semibold", 10),
        )
        self.zenith_login_compact_label.pack(side="left", padx=(4, 12))
        ttk.Button(
            self.zenith_login_compact, text="Sign in again",
            command=self._zenith_show_login_form,
        ).pack(side="left")

        # ----- Inner notebook so each Zenith feature gets its own
        # canvas without polluting the others. Login above remains
        # shared across all sub-tabs.
        inner_nb = ttk.Notebook(parent)
        self.zenith_inner_notebook = inner_nb
        inner_nb.pack(fill="both", expand=True, padx=4, pady=(8, 0))
        customer_inner = ttk.Frame(inner_nb)
        flight_inner = ttk.Frame(inner_nb)
        history_inner = ttk.Frame(inner_nb)
        pnr_bulk_inner = ttk.Frame(inner_nb)
        reports_inner = ttk.Frame(inner_nb)
        inner_nb.add(customer_inner, text="Customer Lookup")
        inner_nb.add(flight_inner, text="Flight Loads")
        inner_nb.add(history_inner, text="Flight History Analyzer")
        inner_nb.add(pnr_bulk_inner, text="PNR Bulk Lookup")
        inner_nb.add(reports_inner, text="Reports")
        self._build_zenith_history_tab(history_inner)
        self._build_zenith_pnr_bulk_tab(pnr_bulk_inner)
        self._build_zenith_reports_tab(reports_inner)

        # From here, the existing Customer Lookup form goes into
        # `customer_inner` instead of the outer scroll frame.
        parent = customer_inner

        # ----- Input picker -----
        body = self._section(parent, "Input Excel")
        self.zenith_input_entry = ttk.Entry(
            body, textvariable=self.zenith_input_path,
        )
        ttk.Label(body, text="File:", width=14, anchor="w").grid(
            row=0, column=0, sticky="w", padx=(2, 4), pady=4,
        )
        self.zenith_input_entry.grid(row=0, column=1, sticky="ew", padx=(0, 4), pady=4)
        ttk.Button(body, text="Browse…", command=self._zenith_pick_input).grid(
            row=0, column=2, padx=(4, 2), pady=4,
        )
        self.zenith_sheet_combo = ttk.Combobox(
            body, textvariable=self.zenith_sheet_name, state="readonly",
            postcommand=self._zenith_reload_columns,
        )
        ttk.Label(body, text="Sheet:", width=14, anchor="w").grid(
            row=1, column=0, sticky="w", padx=(2, 4), pady=4,
        )
        self.zenith_sheet_combo.grid(row=1, column=1, sticky="ew", padx=(0, 4), pady=4)
        self.zenith_col_combo = ttk.Combobox(
            body, textvariable=self.zenith_column_name, state="readonly",
        )
        ttk.Label(body, text="ID column:", width=14, anchor="w").grid(
            row=2, column=0, sticky="w", padx=(2, 4), pady=4,
        )
        self.zenith_col_combo.grid(row=2, column=1, sticky="ew", padx=(0, 4), pady=4)
        body.columnconfigure(1, weight=1)

        # ----- Throughput controls -----
        speed_body = self._section(parent, "Throughput")
        ttk.Label(
            speed_body, style="Hint.TLabel",
            text=(
                "Concurrency = parallel HTTP connections (1 = polite, 10 = aggressive). "
                "Delay = pause between calls per worker. Start at 3 / 0.8 s; "
                "watch for any errors before increasing."
            ),
            wraplength=900, justify="left",
        ).pack(anchor="w", padx=2, pady=(0, 6))
        knobs = ttk.Frame(speed_body)
        knobs.pack(fill="x", padx=2)
        ttk.Label(knobs, text="Concurrency:").grid(
            row=0, column=0, padx=(0, 4), pady=4, sticky="w",
        )
        self.zenith_conc_scale = ttk.Scale(
            knobs, from_=1, to=10, orient="horizontal",
            variable=self.zenith_concurrency, length=200,
            command=lambda _v: self.zenith_concurrency.set(
                int(self.zenith_concurrency.get())
            ),
        )
        self.zenith_conc_scale.grid(row=0, column=1, padx=(0, 8), pady=4)
        self.zenith_conc_label = ttk.Label(knobs, text="3 workers")
        self.zenith_conc_label.grid(row=0, column=2, padx=(0, 24), pady=4)
        self.zenith_concurrency.trace_add(
            "write",
            lambda *_a: self.zenith_conc_label.configure(
                text=f"{int(self.zenith_concurrency.get())} workers",
            ),
        )

        ttk.Label(knobs, text="Delay (sec):").grid(
            row=0, column=3, padx=(0, 4), pady=4, sticky="w",
        )
        self.zenith_delay_scale = ttk.Scale(
            knobs, from_=0.1, to=2.0, orient="horizontal",
            variable=self.zenith_delay_s, length=200,
        )
        self.zenith_delay_scale.grid(row=0, column=4, padx=(0, 8), pady=4)
        self.zenith_delay_label = ttk.Label(knobs, text="0.8 s")
        self.zenith_delay_label.grid(row=0, column=5, padx=(0, 4), pady=4)
        self.zenith_delay_s.trace_add(
            "write",
            lambda *_a: self.zenith_delay_label.configure(
                text=f"{float(self.zenith_delay_s.get()):.1f} s",
            ),
        )

        # Cache + safety options
        opts = ttk.Frame(speed_body)
        opts.pack(fill="x", padx=2, pady=(8, 0))
        ttk.Checkbutton(
            opts, text="Skip IDs already in local cache",
            variable=self.zenith_skip_cached,
        ).pack(side="left", padx=(0, 16))
        ttk.Checkbutton(
            opts, text="Test mode — first 100 IDs only",
            variable=self.zenith_test_mode,
        ).pack(side="left", padx=(0, 16))
        ttk.Button(
            opts, text="Reset cache…", command=self._zenith_reset_cache,
        ).pack(side="right")
        ttk.Button(
            opts, text="Retry failures", command=self._zenith_clear_errors,
        ).pack(side="right", padx=(0, 4))

        # ----- Output -----
        out_body = self._section(parent, "Output")
        self.zenith_output_entry = ttk.Entry(
            out_body, textvariable=self.zenith_output_dir,
        )
        ttk.Label(out_body, text="Folder:", width=14, anchor="w").grid(
            row=0, column=0, sticky="w", padx=(2, 4), pady=4,
        )
        self.zenith_output_entry.grid(
            row=0, column=1, sticky="ew", padx=(0, 4), pady=4,
        )
        ttk.Button(out_body, text="Browse…", command=self._zenith_pick_output).grid(
            row=0, column=2, padx=(4, 2), pady=4,
        )
        out_body.columnconfigure(1, weight=1)

        # ----- Run controls -----
        ctl = ttk.Frame(parent)
        ctl.pack(fill="x", padx=4, pady=(8, 0))
        self.btn_zenith_run = ttk.Button(
            ctl, text="Run", style="Primary.TButton",
            command=self._zenith_run,
        )
        self.btn_zenith_run.pack(side="left")
        self.btn_zenith_pause = ttk.Button(
            ctl, text="Pause", command=self._zenith_pause, state="disabled",
        )
        self.btn_zenith_pause.pack(side="left", padx=(8, 0))
        self.btn_zenith_resume = ttk.Button(
            ctl, text="Resume", command=self._zenith_resume, state="disabled",
        )
        self.btn_zenith_resume.pack(side="left", padx=(4, 0))
        self.btn_zenith_stop = ttk.Button(
            ctl, text="Stop", command=self._zenith_stop,
            state="disabled", style="Danger.TButton",
        )
        self.btn_zenith_stop.pack(side="left", padx=(4, 0))
        self.btn_zenith_export = ttk.Button(
            ctl, text="Export cache to Excel", command=self._zenith_export_cache,
        )
        self.btn_zenith_export.pack(side="left", padx=(16, 0))

        # ----- Progress -----
        prog_body = self._section(parent, "Progress")
        self.zenith_progress_bar = ttk.Progressbar(
            prog_body, mode="determinate", length=200,
        )
        self.zenith_progress_bar.pack(fill="x", padx=2, pady=(0, 4))
        self.zenith_progress_label = ttk.Label(
            prog_body,
            text="Sign in to Zenith, then pick a Customer-ID Excel and Run.",
            style="Hint.TLabel",
        )
        self.zenith_progress_label.pack(anchor="w", padx=2)

        # ----- Log -----
        log_body = self._section(parent, "Log")
        log_frame = ttk.Frame(log_body)
        log_frame.pack(fill="both", expand=True, padx=2)
        self.zenith_log_text = tk.Text(
            log_frame, height=10, wrap="none", state="disabled",
            font=("Consolas", 9),
        )
        self.zenith_log_text.pack(side="left", fill="both", expand=True)
        log_scroll = ttk.Scrollbar(log_frame, command=self.zenith_log_text.yview)
        log_scroll.pack(side="right", fill="y")
        self.zenith_log_text.configure(yscrollcommand=log_scroll.set)
        self._attach_log_placeholder(
            self.zenith_log_text,
            "  Customer-lookup events will appear here once a run starts.",
        )

        self._refresh_zenith_cache_label()

        # ==============================================================
        # Flight Loads inner tab
        # ==============================================================
        self._build_zenith_flight_subtab(flight_inner)

    def _build_zenith_reports_tab(self, parent: ttk.Frame) -> None:
        """Reports sub-tab — download pre-built analytics workbooks, gated by a
        14-day rotating password. Reports are generated by the separate analytics
        pipeline and published to a shared folder; this tab only authenticates and
        downloads. No data engine is bundled — the shared `reporting` core is
        stdlib-only. See reporting/INTEGRATION.md."""
        try:
            from reporting.reports_tab import ReportsFrame
        except ImportError as exc:
            ttk.Label(
                parent,
                text=(
                    "Reports module not installed in this build.\n\n"
                    "Install the shared library into the app environment:\n"
                    "    pip install usba_reporting-<ver>-py3-none-any.whl\n\n"
                    f"({exc})"
                ),
                justify="left",
                padding=20,
            ).pack(anchor="w")
            return
        ReportsFrame(
            parent,
            reports_dir=config.REPORTS_DIR,
            auth_path=config.REPORTS_AUTH_FILE,
        ).pack(fill="both", expand=True)

    def _build_zenith_flight_subtab(self, parent: ttk.Frame) -> None:
        """Inner tab — date range pull from the View PNLs report."""
        # ----- Description (tucked into a tooltip on the section title) -----
        self._section(
            parent, "Flight Loads  ·  View PNLs",
            help_text=(
                "Pull flight-load data (tickets, seats, load %, status) "
                "for a date range. The server caps each search at 10 "
                "pages; this tab auto-chunks longer ranges into smaller "
                "windows so you can ask for up to ~12 months."
            ),
        )

        # ----- Range + page-size form -----
        form = self._section(parent, "Range")
        # State vars
        from datetime import date, timedelta
        today = date.today()
        a_week_ago = today - timedelta(days=6)
        self.zenith_fl_date_from = tk.StringVar(value=a_week_ago.strftime("%d/%m/%Y"))
        self.zenith_fl_date_to = tk.StringVar(value=today.strftime("%d/%m/%Y"))
        self.zenith_fl_page_size = tk.StringVar(value="100")
        self.zenith_fl_chunk_days = tk.IntVar(value=5)
        self.zenith_fl_delay_s = tk.DoubleVar(value=1.0)

        row1 = ttk.Frame(form)
        row1.pack(fill="x", padx=2)
        ttk.Label(row1, text="From (DD/MM/YYYY):").grid(
            row=0, column=0, padx=(0, 4), pady=4, sticky="w",
        )
        ttk.Entry(row1, textvariable=self.zenith_fl_date_from, width=14).grid(
            row=0, column=1, padx=(0, 12), pady=4,
        )
        ttk.Label(row1, text="To:").grid(
            row=0, column=2, padx=(0, 4), pady=4, sticky="w",
        )
        ttk.Entry(row1, textvariable=self.zenith_fl_date_to, width=14).grid(
            row=0, column=3, padx=(0, 12), pady=4,
        )
        ttk.Label(row1, text="Page size:").grid(
            row=0, column=4, padx=(0, 4), pady=4, sticky="w",
        )
        ttk.Combobox(
            row1, textvariable=self.zenith_fl_page_size,
            values=("20", "50", "100"), width=6, state="readonly",
        ).grid(row=0, column=5, padx=(0, 12), pady=4)
        ttk.Label(row1, text="Chunk (days):").grid(
            row=0, column=6, padx=(0, 4), pady=4, sticky="w",
        )
        ttk.Spinbox(
            row1, from_=1, to=30, textvariable=self.zenith_fl_chunk_days,
            width=5,
        ).grid(row=0, column=7, padx=(0, 12), pady=4)

        # ----- Throughput -----
        speed = self._section(
            parent, "Throughput",
            help_text=(
                "Polite delay between paginated calls. The View PNLs page "
                "is heavier than the Customer page (~360 KB per call); "
                "1.0 s is a safe default."
            ),
        )
        knobs = ttk.Frame(speed)
        knobs.pack(fill="x", padx=2)
        ttk.Label(knobs, text="Delay (sec):").grid(
            row=0, column=0, padx=(0, 4), pady=4, sticky="w",
        )
        self.zenith_fl_delay_scale = ttk.Scale(
            knobs, from_=0.3, to=3.0, orient="horizontal",
            variable=self.zenith_fl_delay_s, length=200,
        )
        self.zenith_fl_delay_scale.grid(row=0, column=1, padx=(0, 8), pady=4)
        self.zenith_fl_delay_label = ttk.Label(knobs, text="1.0 s")
        self.zenith_fl_delay_label.grid(row=0, column=2, padx=(0, 4), pady=4)
        self.zenith_fl_delay_s.trace_add(
            "write",
            lambda *_a: self.zenith_fl_delay_label.configure(
                text=f"{float(self.zenith_fl_delay_s.get()):.1f} s",
            ),
        )

        # ----- Output -----
        out_body = self._section(parent, "Output")
        self.zenith_fl_output_dir = tk.StringVar(
            value=str(Path.home() / "Documents"),
        )
        self.zenith_fl_output_entry = ttk.Entry(
            out_body, textvariable=self.zenith_fl_output_dir,
        )
        ttk.Label(out_body, text="Folder:", width=14, anchor="w").grid(
            row=0, column=0, sticky="w", padx=(2, 4), pady=4,
        )
        self.zenith_fl_output_entry.grid(
            row=0, column=1, sticky="ew", padx=(0, 4), pady=4,
        )
        ttk.Button(
            out_body, text="Browse…", command=self._zenith_fl_pick_output,
        ).grid(row=0, column=2, padx=(4, 2), pady=4)
        out_body.columnconfigure(1, weight=1)

        # ----- Run controls -----
        ctl = ttk.Frame(parent)
        ctl.pack(fill="x", padx=4, pady=(8, 0))
        self.btn_zenith_fl_run = ttk.Button(
            ctl, text="Pull flight loads", style="Primary.TButton",
            command=self._zenith_fl_run,
        )
        self.btn_zenith_fl_run.pack(side="left")
        self.btn_zenith_fl_stop = ttk.Button(
            ctl, text="Stop", command=self._zenith_fl_stop,
            state="disabled", style="Danger.TButton",
        )
        self.btn_zenith_fl_stop.pack(side="left", padx=(8, 0))
        # Append the LAST run's data into a user-picked Ordered Report
        # (cross-tab one-date-per-block format). Enabled only after a
        # successful run because we need the rows in memory.
        self.btn_zenith_fl_append_ordered = ttk.Button(
            ctl, text="Append to Ordered Report…",
            command=self._zenith_fl_append_ordered,
            state="disabled",
        )
        self.btn_zenith_fl_append_ordered.pack(side="left", padx=(8, 0))

        # ----- Per-leg drill-down: passenger manifest -----
        legs_body = self._section(
            parent, "Passenger detail (drill-down)",
            help_text=(
                "After a flight-loads pull, the legs appear below. Select "
                "one or more (Ctrl/Shift-click) and pull the full passenger "
                "manifest for just those legs — name, PRBD, fare basis, "
                "ticket, PNR, seat, passport, DOB, agency. One row per "
                "passenger, written to its own Excel."
            ),
        )
        # Nested filters: Region (Domestic/International) + Direction
        # (Inbound/Outbound). They subset the grid without re-fetching.
        filt = ttk.Frame(parent)
        filt.pack(fill="x", padx=6, pady=(2, 0))
        self.zenith_fl_filter_region = tk.StringVar(value="All regions")
        self.zenith_fl_filter_dir = tk.StringVar(value="All directions")
        ttk.Label(filt, text="Region:").pack(side="left")
        rcb = ttk.Combobox(
            filt, textvariable=self.zenith_fl_filter_region, width=16,
            state="readonly",
            values=["All regions", "Domestic", "International"],
        )
        rcb.pack(side="left", padx=(4, 12))
        ttk.Label(filt, text="Direction:").pack(side="left")
        dcb = ttk.Combobox(
            filt, textvariable=self.zenith_fl_filter_dir, width=16,
            state="readonly",
            values=["All directions", "Outbound", "Inbound"],
        )
        dcb.pack(side="left", padx=(4, 12))
        rcb.bind("<<ComboboxSelected>>", lambda _e: self._zenith_fl_populate_legs())
        dcb.bind("<<ComboboxSelected>>", lambda _e: self._zenith_fl_populate_legs())
        self.zenith_fl_filter_count = ttk.Label(filt, text="", style="Hint.TLabel")
        self.zenith_fl_filter_count.pack(side="left")

        grid_wrap = ttk.Frame(legs_body)
        grid_wrap.pack(fill="both", expand=True)
        cols = ("flight", "date", "route", "region", "direction", "cabin", "load", "issued")
        self.zenith_fl_legs_tree = ttk.Treeview(
            grid_wrap, columns=cols, show="headings", height=8,
            selectmode="extended",
        )
        for cid, txt, w in (
            ("flight", "Flight", 64), ("date", "Date", 88),
            ("route", "Leg", 80), ("region", "Region", 92),
            ("direction", "Direction", 84), ("cabin", "Cabin", 76),
            ("load", "Load", 104), ("issued", "Issued", 74),
        ):
            self.zenith_fl_legs_tree.heading(cid, text=txt)
            self.zenith_fl_legs_tree.column(cid, width=w, anchor="w")
        # Load-factor traffic light (matches the Ordered Report thresholds).
        self.zenith_fl_legs_tree.tag_configure("load_hi", background=self._COLOR_ROW_GOOD)
        self.zenith_fl_legs_tree.tag_configure("load_mid", background=self._COLOR_ROW_WARN)
        self.zenith_fl_legs_tree.tag_configure("load_lo", background=self._COLOR_ROW_BAD)
        self.zenith_fl_legs_tree.pack(side="left", fill="both", expand=True, padx=(2, 0), pady=2)
        legs_scroll = ttk.Scrollbar(
            grid_wrap, command=self.zenith_fl_legs_tree.yview,
        )
        legs_scroll.pack(side="left", fill="y")
        self.zenith_fl_legs_tree.configure(yscrollcommand=legs_scroll.set)

        pax_ctl = ttk.Frame(parent)
        pax_ctl.pack(fill="x", padx=4, pady=(4, 0))
        self.btn_zenith_fl_pax = ttk.Button(
            pax_ctl, text="Pull passenger detail (selected legs)",
            command=self._zenith_fl_pull_pax, state="disabled",
        )
        self.btn_zenith_fl_pax.pack(side="left")
        self.btn_zenith_fl_pax_all = ttk.Button(
            pax_ctl, text="Select all legs",
            command=lambda: self.zenith_fl_legs_tree.selection_set(
                self.zenith_fl_legs_tree.get_children()
            ),
            state="disabled",
        )
        self.btn_zenith_fl_pax_all.pack(side="left", padx=(8, 0))
        self.zenith_fl_pax_status = ttk.Label(
            pax_ctl, text="", style="Hint.TLabel",
        )
        self.zenith_fl_pax_status.pack(side="left", padx=(12, 0))

        # ----- Progress + log -----
        prog_body = self._section(parent, "Progress")
        self.zenith_fl_progress_bar = ttk.Progressbar(
            prog_body, mode="determinate", length=200,
        )
        self.zenith_fl_progress_bar.pack(fill="x", padx=2, pady=(0, 4))
        self.zenith_fl_progress_label = ttk.Label(
            prog_body,
            text="Pick a date range and page size, then Pull flight loads.",
            style="Hint.TLabel",
        )
        self.zenith_fl_progress_label.pack(anchor="w", padx=2)

        log_body = self._section(parent, "Log")
        log_frame = ttk.Frame(log_body)
        log_frame.pack(fill="both", expand=True, padx=2)
        self.zenith_fl_log_text = tk.Text(
            log_frame, height=10, wrap="none", state="disabled",
            font=("Consolas", 9),
        )
        self.zenith_fl_log_text.pack(side="left", fill="both", expand=True)
        fl_scroll = ttk.Scrollbar(
            log_frame, command=self.zenith_fl_log_text.yview,
        )
        fl_scroll.pack(side="right", fill="y")
        self.zenith_fl_log_text.configure(yscrollcommand=fl_scroll.set)
        self._attach_log_placeholder(
            self.zenith_fl_log_text,
            "  Flight-loads pagination events will appear here once a run starts.",
        )

        # ----- Worker state -----
        self._zenith_fl_worker: threading.Thread | None = None
        self._zenith_fl_stop_flag = threading.Event()
        # Last run's rows — re-used by 'Append to Ordered Report' and the
        # passenger drill-down. tree iid (str index) maps back into this.
        self._zenith_fl_last_rows: list = []
        self._zenith_pax_worker: threading.Thread | None = None
        self._zenith_pax_stop_flag = threading.Event()

    # ==================================================================
    # Zenith Flight History Analyzer sub-tab — UI
    # ==================================================================

    def _build_zenith_history_tab(self, parent: ttk.Frame) -> None:
        """Wire up the Flight History Analyzer inner tab.

        Reads downloaded ModificationHistory*.xls files from a folder
        and produces a multi-sheet audit Excel. No network calls — all
        work is local, so this sub-tab is usable even when Zenith is
        unreachable.
        """
        parent = self._make_scrollable(parent)

        self._section(
            parent, "Flight History Analyzer  ·  audits downloaded .xls logs",
            help_text=(
                "Reads the ModificationHistory .xls files exported from "
                "Zenith's Inventory → Flight List → History view. "
                "Builds a 7-sheet audit: class downgrades, downgrade "
                "leaders, G-class issuance, agent activity, revenue "
                "mgmt, suspicious activity, and raw events."
            ),
        )

        # ----- Input folder + output -----
        io_body = self._section(parent, "Folders")
        default_input = Path(r"E:\US Bangla\Flight History Logs")
        self.zenith_fh_input_dir = tk.StringVar(
            value=str(default_input) if default_input.exists() else "",
        )
        self.zenith_fh_output_dir = tk.StringVar(
            value=str(Path.home() / "Documents"),
        )

        input_entry = ttk.Entry(io_body, textvariable=self.zenith_fh_input_dir)
        input_btn = ttk.Button(
            io_body, text="Browse…", command=self._zenith_fh_pick_input,
        )
        self._form_row(io_body, 0, "Logs folder:", input_entry, suffix=input_btn)

        output_entry = ttk.Entry(io_body, textvariable=self.zenith_fh_output_dir)
        output_btn = ttk.Button(
            io_body, text="Browse…", command=self._zenith_fh_pick_output,
        )
        self._form_row(io_body, 1, "Output folder:", output_entry, suffix=output_btn)

        # Optional Flight Loads Excel — when provided, enriches the audit
        # with a "Downgrade Justification" sheet (was the fare cut justified
        # by low load, or was it on an already-full flight?).
        self.zenith_fh_loads_path = tk.StringVar(value="")
        loads_entry = ttk.Entry(io_body, textvariable=self.zenith_fh_loads_path)
        loads_btn = ttk.Button(
            io_body, text="Browse…", command=self._zenith_fh_pick_loads,
        )
        self._form_row(
            io_body, 2, "Flight Loads (optional):", loads_entry, suffix=loads_btn,
        )

        # Tunable verdict thresholds. Only used when a Flight Loads file
        # is supplied. Defaults match zenith_loads_index module constants.
        self.zenith_fh_high_threshold = tk.DoubleVar(
            value=zenith_loads_index.HIGH_LOAD_THRESHOLD,
        )
        self.zenith_fh_low_threshold = tk.DoubleVar(
            value=zenith_loads_index.LOW_LOAD_THRESHOLD,
        )
        thresh_row = ttk.Frame(io_body)
        ttk.Label(thresh_row, text="QUESTIONABLE at ≥").pack(side="left")
        ttk.Spinbox(
            thresh_row, from_=50.0, to=100.0, increment=1.0,
            textvariable=self.zenith_fh_high_threshold, width=6,
            format="%.1f",
        ).pack(side="left", padx=(4, 2))
        ttk.Label(thresh_row, text="%   JUSTIFIED below").pack(side="left", padx=(8, 0))
        ttk.Spinbox(
            thresh_row, from_=20.0, to=90.0, increment=1.0,
            textvariable=self.zenith_fh_low_threshold, width=6,
            format="%.1f",
        ).pack(side="left", padx=(4, 2))
        ttk.Label(thresh_row, text="%").pack(side="left")
        ttk.Label(
            thresh_row,
            text=" (in between = SITUATIONAL)",
            style="Hint.TLabel",
        ).pack(side="left", padx=(8, 0))
        self._form_row(io_body, 3, "Load verdict thresholds:", thresh_row)

        # ----- Controls -----
        ctl = ttk.Frame(parent)
        ctl.pack(fill="x", padx=4, pady=(8, 4))
        self.btn_zenith_fh_run = ttk.Button(
            ctl, text="Run history audit", style="Primary.TButton",
            command=self._zenith_fh_run,
        )
        self.btn_zenith_fh_run.pack(side="left")
        self.btn_zenith_fh_export = ttk.Button(
            ctl, text="Export to Excel…",
            command=self._zenith_fh_export, state="disabled",
        )
        self.btn_zenith_fh_export.pack(side="left", padx=(8, 0))
        self.btn_zenith_fh_download = ttk.Button(
            ctl, text="Download from Zenith…",
            command=self._zenith_fh_download_dialog,
        )
        self.btn_zenith_fh_download.pack(side="left", padx=(8, 0))
        # Enrich the most-recent audit by fetching each PNR from Zenith.
        self.btn_zenith_fh_pnr = ttk.Button(
            ctl, text="Enrich with PNR details",
            command=self._zenith_fh_enrich_pnrs,
            state="disabled",
        )
        self.btn_zenith_fh_pnr.pack(side="left", padx=(8, 0))
        # Re-pivot the SAME logs into a PNR-centric misuse audit (refund-of-flown,
        # self-refund, off-hours, downgrades, reissue churn) — no network, no extra
        # download. Flags are review leads, not determinations.
        self.btn_zenith_fh_misuse = ttk.Button(
            ctl, text="PNR Misuse audit",
            command=self._zenith_pnr_misuse_run,
        )
        self.btn_zenith_fh_misuse.pack(side="left", padx=(8, 0))
        self.zenith_fh_status_label = ttk.Label(
            ctl,
            text="Browse a logs folder, then Run history audit.",
            style="Hint.TLabel",
        )
        self.zenith_fh_status_label.pack(side="left", padx=(12, 0))

        # ----- Progress + results summary -----
        prog_body = self._section(parent, "Progress")
        self.zenith_fh_progress = ttk.Progressbar(
            prog_body, mode="determinate", maximum=1, value=0,
        )
        self.zenith_fh_progress.pack(fill="x", padx=2, pady=2)

        summary_body = self._section(parent, "Audit summary")
        cols = (
            ("metric", "Metric", 280),
            ("value", "Value", 200),
        )
        self.zenith_fh_tree = ttk.Treeview(
            summary_body,
            columns=[c[0] for c in cols],
            show="headings", height=14,
        )
        for cid, label, width in cols:
            self.zenith_fh_tree.heading(cid, text=label)
            self.zenith_fh_tree.column(cid, width=width, anchor="w")
        self.zenith_fh_tree.pack(fill="both", expand=True, padx=2, pady=4)

        # ----- Worker state -----
        self._zenith_fh_worker: threading.Thread | None = None
        self._zenith_fh_last_report = None
        self._zenith_dl_worker: threading.Thread | None = None
        self._zenith_dl_stop_flag = threading.Event()
        self._zenith_pnr_worker: threading.Thread | None = None
        # SQLite cache shared by all PNR-enrichment runs.
        self._zenith_pnr_cache = ZenithPNRCache(config.APP_DIR / "zenith_pnr.sqlite")

    # ==================================================================
    # Zenith PNR Bulk Lookup sub-tab — UI
    # ==================================================================

    def _build_zenith_pnr_bulk_tab(self, parent: ttk.Frame) -> None:
        """Standalone bulk lookup: Excel of PNRs in → Excel with route/customer out."""
        parent = self._make_scrollable(parent)

        self._section(
            parent, "PNR Bulk Lookup  ·  paste a list, get every detail",
            help_text=(
                "Reads a column of PNR codes from an Excel file, looks up "
                "each one against Zenith, and writes the route, customer, "
                "status, fares, and per-segment breakdown back to a new "
                "workbook. Cached locally so re-runs are instant."
            ),
        )

        # ----- Files -----
        io_body = self._section(parent, "Files")
        self.zenith_bulk_input_path = tk.StringVar(value="")
        self.zenith_bulk_sheet_name = tk.StringVar(value="")
        self.zenith_bulk_column_name = tk.StringVar(value="PNR")
        self.zenith_bulk_output_dir = tk.StringVar(
            value=str(Path.home() / "Documents"),
        )

        input_entry = ttk.Entry(io_body, textvariable=self.zenith_bulk_input_path)
        input_btn = ttk.Button(
            io_body, text="Browse…",
            command=self._zenith_bulk_pick_input,
        )
        self._form_row(io_body, 0, "Input Excel:", input_entry, suffix=input_btn)

        # Sheet + column drop into Comboboxes the moment a workbook is picked.
        # User can still type, but the lists keep them from guessing.
        self.zenith_bulk_sheet_combo = ttk.Combobox(
            io_body, textvariable=self.zenith_bulk_sheet_name, state="readonly",
        )
        self.zenith_bulk_sheet_combo.bind(
            "<<ComboboxSelected>>",
            lambda _e: self._zenith_bulk_reload_columns(),
        )
        self._form_row(io_body, 1, "Sheet:", self.zenith_bulk_sheet_combo)

        self.zenith_bulk_column_combo = ttk.Combobox(
            io_body, textvariable=self.zenith_bulk_column_name, state="readonly",
        )
        self.zenith_bulk_column_combo.bind(
            "<<ComboboxSelected>>",
            lambda _e: self._zenith_bulk_refresh_caption(),
        )
        self._form_row(io_body, 2, "PNR column:", self.zenith_bulk_column_combo)

        # Live caption: "1,841 PNRs detected  ·  427 already cached  ·  1,414 to fetch"
        # Updates the moment the user picks a sheet + column, so they
        # know what they're about to ask for before clicking Run.
        self.zenith_bulk_caption = ttk.Label(
            io_body, text="", style="Hint.TLabel",
        )
        self.zenith_bulk_caption.grid(
            row=3, column=1, sticky="w", padx=(0, 4), pady=(0, 4),
        )

        out_entry = ttk.Entry(io_body, textvariable=self.zenith_bulk_output_dir)
        out_btn = ttk.Button(
            io_body, text="Browse…",
            command=self._zenith_bulk_pick_output,
        )
        self._form_row(io_body, 3, "Output folder:", out_entry, suffix=out_btn)

        # ----- Controls -----
        ctl = ttk.Frame(parent)
        ctl.pack(fill="x", padx=4, pady=(8, 4))
        self.btn_zenith_bulk_run = ttk.Button(
            ctl, text="Look up PNRs", style="Primary.TButton",
            command=self._zenith_bulk_run,
        )
        self.btn_zenith_bulk_run.pack(side="left")
        self.btn_zenith_bulk_stop = ttk.Button(
            ctl, text="Stop", command=self._zenith_bulk_stop,
            state="disabled", style="Danger.TButton",
        )
        self.btn_zenith_bulk_stop.pack(side="left", padx=(8, 0))
        self.btn_zenith_dossier = ttk.Button(
            ctl, text="Dossier audit (payment/contact)",
            command=self._zenith_dossier_run,
        )
        self.btn_zenith_dossier.pack(side="left", padx=(8, 0))
        self.zenith_bulk_status_label = ttk.Label(
            ctl,
            text="Pick an Excel above, then Look up PNRs.",
            style="Hint.TLabel",
        )
        self.zenith_bulk_status_label.pack(side="left", padx=(12, 0))

        # ----- Progress -----
        prog_body = self._section(parent, "Progress")
        self.zenith_bulk_progress = ttk.Progressbar(
            prog_body, mode="determinate", maximum=1, value=0,
        )
        self.zenith_bulk_progress.pack(fill="x", padx=2, pady=2)

        log_body = self._section(parent, "Log")
        self.zenith_bulk_log = tk.Text(
            log_body, height=12, wrap="none", state="disabled",
        )
        self.zenith_bulk_log.pack(side="left", fill="both", expand=True, padx=2)
        sb = ttk.Scrollbar(log_body, command=self.zenith_bulk_log.yview)
        sb.pack(side="left", fill="y")
        self.zenith_bulk_log.configure(yscrollcommand=sb.set)
        self._attach_log_placeholder(
            self.zenith_bulk_log,
            "  PNR-lookup events will appear here once a run starts.",
        )

        # ----- Worker state -----
        self._zenith_bulk_worker: threading.Thread | None = None
        self._zenith_bulk_stop_flag = threading.Event()
        # Live progress counters for the status line (E enhancement).
        self._zenith_bulk_counters: dict[str, int] = {}
        self._zenith_bulk_started_at: float | None = None

    def _zenith_bulk_pick_input(self) -> None:
        f = filedialog.askopenfilename(
            title="Pick the Excel with your PNR list",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")],
        )
        if not f:
            return
        self.zenith_bulk_input_path.set(f)
        self._zenith_bulk_reload_sheets()

    def _zenith_bulk_reload_sheets(self) -> None:
        """Fill the sheet dropdown from the picked workbook."""
        path = self.zenith_bulk_input_path.get().strip()
        if not path or not Path(path).is_file():
            self.zenith_bulk_sheet_combo["values"] = ()
            self.zenith_bulk_column_combo["values"] = ()
            return
        try:
            sheets = excel_io.list_sheet_names(Path(path))
        except Exception as exc:  # noqa: BLE001 — surface to UI
            messagebox.showerror(
                "PNR Bulk Lookup",
                f"Couldn't read sheet names:\n{type(exc).__name__}: {exc}",
            )
            return
        self.zenith_bulk_sheet_combo["values"] = sheets
        if sheets:
            self.zenith_bulk_sheet_combo.current(0)
        self._zenith_bulk_reload_columns()

    def _zenith_bulk_reload_columns(self) -> None:
        """Fill the PNR-column dropdown from the chosen sheet's header row.

        Auto-selects a column whose header starts with "PNR" (case-insensitive)
        when the user hasn't already picked one — so 90% of the time the
        defaults are right and no manual fiddling is needed.
        """
        path = self.zenith_bulk_input_path.get().strip()
        sheet = self.zenith_bulk_sheet_name.get().strip()
        if not path or not sheet:
            self.zenith_bulk_column_combo["values"] = ()
            return
        try:
            cols = excel_io.list_columns(Path(path), sheet)
        except Exception as exc:  # noqa: BLE001 — surface to UI
            messagebox.showerror(
                "PNR Bulk Lookup",
                f"Couldn't read columns from sheet {sheet!r}:\n"
                f"{type(exc).__name__}: {exc}",
            )
            return
        self.zenith_bulk_column_combo["values"] = cols
        current = self.zenith_bulk_column_name.get().strip()
        if current not in cols:
            # Prefer a header that looks like a PNR column.
            pnr_like = next(
                (c for c in cols if c.strip().lower().startswith("pnr")), None,
            )
            if pnr_like:
                self.zenith_bulk_column_name.set(pnr_like)
            elif cols:
                self.zenith_bulk_column_combo.current(0)
        # Trigger the live caption now that sheet + column are set.
        self._zenith_bulk_refresh_caption()

    def _zenith_bulk_refresh_caption(self) -> None:
        """Show 'N PNRs detected · M cached · K to fetch' below the picker.

        Reads the Excel column the user just chose (fast — just the
        single column, no full Dossier fetch) and intersects with the
        local SQLite cache to surface what a Run would actually cost.
        """
        path = self.zenith_bulk_input_path.get().strip()
        sheet = self.zenith_bulk_sheet_name.get().strip()
        column = self.zenith_bulk_column_name.get().strip()
        if not (path and sheet and column and Path(path).is_file()):
            self.zenith_bulk_caption.configure(text="")
            return
        try:
            codes = excel_io.read_pnr_codes_from_excel(
                Path(path), sheet_name=sheet, column_name=column,
            )
        except Exception as exc:  # noqa: BLE001 — surface compactly
            self.zenith_bulk_caption.configure(
                text=f"⚠ Couldn't read column: {exc}",
                style="Error.TLabel",
            )
            return
        total = len(codes)
        if total == 0:
            self.zenith_bulk_caption.configure(
                text="No PNR codes found in that column.",
                style="Warning.TLabel",
            )
            return
        cached = sum(1 for c in codes if self._zenith_pnr_cache.get(c) is not None)
        to_fetch = total - cached
        # Rough time estimate at 1s per uncached PNR.
        est = ""
        if to_fetch > 0:
            mins = max(1, to_fetch // 60)
            est = f"  ·  ~{mins} min"
        self.zenith_bulk_caption.configure(
            text=(
                f"{total:,} PNRs detected  ·  "
                f"{cached:,} already cached  ·  "
                f"{to_fetch:,} to fetch{est}"
            ),
            style="Hint.TLabel",
        )

    def _zenith_bulk_pick_output(self) -> None:
        d = filedialog.askdirectory(
            title="Pick the output folder",
            initialdir=self.zenith_bulk_output_dir.get() or str(Path.home()),
        )
        if d:
            self.zenith_bulk_output_dir.set(d)

    def _zenith_bulk_log_line(self, line: str) -> None:
        self.zenith_bulk_log.configure(state="normal")
        self.zenith_bulk_log.insert("end", line + "\n")
        self.zenith_bulk_log.see("end")
        self.zenith_bulk_log.configure(state="disabled")

    def _zenith_bulk_stop(self) -> None:
        self._zenith_bulk_stop_flag.set()
        self.zenith_bulk_status_label.configure(text="Stopping…")

    def _zenith_dossier_run(self) -> None:
        """Phase-2 dossier audit: scrape each PNR's CHANGES history for payment-txn reuse
        and contact churn/funnel. Reuses this tab's input Excel + session + output folder."""
        if not getattr(self, "_zenith_session", None):
            messagebox.showerror("Dossier audit", "Sign in to Zenith first (top of this tab).")
            return
        if self._zenith_bulk_worker and self._zenith_bulk_worker.is_alive():
            messagebox.showinfo("Dossier audit", "A job is already running on this tab.")
            return
        input_path = self.zenith_bulk_input_path.get().strip()
        if not input_path or not Path(input_path).is_file():
            messagebox.showerror(
                "Dossier audit", f"Pick a valid Excel of PNRs first.\nCurrent: {input_path!r}")
            return
        out_folder = Path(self.zenith_bulk_output_dir.get().strip() or str(Path.home()))
        sheet = self.zenith_bulk_sheet_name.get().strip() or None
        column = self.zenith_bulk_column_name.get().strip() or None
        try:
            codes = excel_io.read_pnr_codes_from_excel(
                Path(input_path), sheet_name=sheet, column_name=column)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror(
                "Dossier audit", f"Couldn't read the input Excel:\n{type(exc).__name__}: {exc}")
            return
        if not codes:
            messagebox.showerror(
                "Dossier audit", f"No PNRs found in column {column!r} of {Path(input_path).name}.")
            return
        if not messagebox.askyesno(
            "Dossier audit",
            f"Scrape dossier payment/contact history for {len(codes)} PNRs from Zenith?\n\n"
            f"~2 requests per uncached PNR (budget {config.PNR_HISTORY_MAX_REQUESTS}); "
            "cached locally so re-runs are cheap.\n\nFlags are review leads, not findings."):
            return

        self.zenith_bulk_log.configure(state="normal")
        self.zenith_bulk_log.delete("1.0", "end")
        self.zenith_bulk_log.configure(state="disabled")
        self._zenith_bulk_log_line(f"Dossier audit: scraping {len(codes)} PNRs…")
        self._zenith_bulk_stop_flag.clear()
        self.btn_zenith_bulk_run.configure(state="disabled")
        self.btn_zenith_dossier.configure(state="disabled")
        self.btn_zenith_bulk_stop.configure(state="normal")
        self.zenith_bulk_progress.configure(value=0, maximum=len(codes))
        self.zenith_bulk_status_label.configure(text=f"Scraping {len(codes)} dossiers…")

        sess = self._zenith_session
        stop = self._zenith_bulk_stop_flag

        def worker() -> None:
            try:
                cache = ZenithPNRHistoryCache(config.ZENITH_PNR_HISTORY_CACHE_DB)
                events, stats = zenith_pnr_history_downloader.scrape_dossier_events(
                    sess, codes, cache=cache, stop_flag=stop.is_set,
                    progress_cb=lambda i, n, pnr: self._post(
                        MSG_ZENITH_DOSSIER_PROGRESS, (i, n, pnr)))
                report = zenith_pnr_history_analyzer.run_dossier_audit(events)
                out_path = excel_io.build_zenith_dossier_output_path(out_folder)
                excel_io.write_zenith_dossier_audit(out_path, report)
                sev: dict[str, int] = {}
                for f in report.flags:
                    sev[f.severity] = sev.get(f.severity, 0) + 1
                self._post(MSG_ZENITH_DOSSIER_DONE, {
                    "path": str(out_path), "events": report.event_count,
                    "pnrs": report.pnr_count, "flags": len(report.flags),
                    "critical": sev.get("critical", 0), "high": sev.get("high", 0),
                    "txn": report.distinct_txn, "scraped": stats.scraped,
                    "cached": stats.from_cache, "failed": stats.failed,
                    "aborted": stats.aborted,
                })
            except Exception as exc:  # noqa: BLE001 — surface to UI
                log.exception("dossier audit failed")
                self._post(MSG_ZENITH_DOSSIER_ERROR, f"{type(exc).__name__}: {exc}")

        self._zenith_bulk_worker = threading.Thread(target=worker, daemon=True)
        self._zenith_bulk_worker.start()

    def _zenith_bulk_run(self) -> None:
        if not getattr(self, "_zenith_session", None):
            messagebox.showerror(
                "PNR Bulk Lookup",
                "Sign in to Zenith first (top of this tab).",
            )
            return
        if self._zenith_bulk_worker and self._zenith_bulk_worker.is_alive():
            messagebox.showinfo(
                "PNR Bulk Lookup", "A lookup is already running.",
            )
            return
        input_path = self.zenith_bulk_input_path.get().strip()
        if not input_path or not Path(input_path).is_file():
            messagebox.showerror(
                "PNR Bulk Lookup",
                f"Pick a valid Excel file first.\nCurrent: {input_path!r}",
            )
            return
        out_folder = Path(self.zenith_bulk_output_dir.get().strip() or str(Path.home()))
        sheet = self.zenith_bulk_sheet_name.get().strip() or None
        column = self.zenith_bulk_column_name.get().strip() or None

        try:
            codes = excel_io.read_pnr_codes_from_excel(
                Path(input_path),
                sheet_name=sheet,
                column_name=column,
            )
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror(
                "PNR Bulk Lookup",
                f"Couldn't read the input Excel:\n{type(exc).__name__}: {exc}",
            )
            return
        if not codes:
            messagebox.showerror(
                "PNR Bulk Lookup",
                f"No PNRs found in column {column!r} of {Path(input_path).name}.",
            )
            return

        cache_hits = sum(
            1 for c in codes if self._zenith_pnr_cache.get(c) is not None
        )
        misses = len(codes) - cache_hits
        est_min = max(1, misses // 60)
        if misses > 200:
            proceed = messagebox.askyesno(
                "PNR Bulk Lookup",
                f"{len(codes)} PNRs in input.\n"
                f"  · {cache_hits} already cached\n"
                f"  · {misses} need fetching from Zenith\n\n"
                f"Estimated time: ~{est_min} min.\n\nContinue?",
            )
            if not proceed:
                return

        self.zenith_bulk_log.configure(state="normal")
        self.zenith_bulk_log.delete("1.0", "end")
        self.zenith_bulk_log.configure(state="disabled")
        self._zenith_bulk_log_line(
            f"Reading {len(codes)} PNRs ({cache_hits} cached, {misses} to fetch)…",
        )
        self._zenith_bulk_stop_flag.clear()
        # Reset live counters + start-time at the top of every fresh run
        # so the throughput / ETA display starts from zero each time.
        self._zenith_bulk_counters = {}
        import time as _t
        self._zenith_bulk_started_at = _t.monotonic()
        self.btn_zenith_bulk_run.configure(state="disabled")
        self.btn_zenith_bulk_stop.configure(state="normal")
        self.zenith_bulk_progress.configure(value=0, maximum=len(codes))
        self.zenith_bulk_status_label.configure(
            text=f"Looking up {len(codes)} PNRs…",
        )

        cache = self._zenith_pnr_cache
        sess = self._zenith_session
        stop = self._zenith_bulk_stop_flag

        # Bounded concurrency + in-worker 504 retry beats the old serial loop:
        # transient 504s are now retried (recovered) instead of lost, and the
        # slow Dossier waits overlap across workers. 3 workers is the same
        # proven range as the customer-lookup bulk path (fetch_many).
        _BULK_CONCURRENCY = 3
        _BULK_DELAY_S = 0.8

        def worker() -> None:
            try:
                got: dict[str, object] = {}
                errors: dict[str, str] = {}

                def on_result(code, details, status) -> None:
                    # Fires as each PNR lands (serially, in this thread) — so we
                    # checkpoint every success to the cache immediately, making
                    # a Stop / crash / session-expiry resume-safe on re-run.
                    if status == "OK":
                        if details is not None:
                            try:
                                cache.put(details)
                            except Exception:  # noqa: BLE001
                                log.exception("cache.put failed for %s", code)
                            got[code] = details
                    elif status == "CACHED":
                        got[code] = details
                    elif status == "NOT_FOUND":
                        got[code] = None
                    elif isinstance(status, str) and status.startswith("ERROR"):
                        got[code] = None
                        errors[code] = status.split("ERROR:", 1)[-1].strip() or status

                def progress(done, total, code, status) -> None:
                    self._post(
                        MSG_ZENITH_BULK_PROGRESS, (done, total, code, status),
                    )

                zenith_pnr_client.lookup_many(
                    sess, codes,
                    concurrency=_BULK_CONCURRENCY,
                    delay_s=_BULK_DELAY_S,
                    skip_cached=cache.get,
                    on_result=on_result,
                    progress_cb=progress,
                    stop_event=stop,
                )

                # Rebuild output rows in INPUT order (completion order is
                # arbitrary under concurrency). PNRs not processed (Stop) are
                # simply absent from `got` and dropped from the sheet.
                seen: set = set()
                results: list = []
                for raw in codes:
                    c = raw.strip().upper()
                    if c and c in got and c not in seen:
                        seen.add(c)
                        results.append((c, got[c]))

                out_path = excel_io.build_zenith_pnr_bulk_output_path(out_folder)
                excel_io.write_zenith_pnr_bulk(out_path, results, errors=errors)
                summary = {
                    "path": str(out_path),
                    "total": len(results),
                    "ok": sum(1 for _, d in results if d is not None),
                    "errors": len(errors),
                }
                self._post(MSG_ZENITH_BULK_DONE, summary)
            except zenith_client.SessionExpiredError as exc:
                self._post(
                    MSG_ZENITH_BULK_ERROR,
                    "Session expired — sign in again and re-run. PNRs already "
                    f"fetched are cached and will be skipped. ({exc})",
                )
            except Exception as exc:  # noqa: BLE001
                log.exception("PNR bulk lookup failed")
                self._post(
                    MSG_ZENITH_BULK_ERROR, f"{type(exc).__name__}: {exc}",
                )

        self._zenith_bulk_worker = threading.Thread(target=worker, daemon=True)
        self._zenith_bulk_worker.start()

    def _zenith_fh_pick_input(self) -> None:
        d = filedialog.askdirectory(
            title="Pick the Flight History Logs folder",
            initialdir=self.zenith_fh_input_dir.get() or str(Path.home()),
        )
        if d:
            self.zenith_fh_input_dir.set(d)

    def _zenith_fh_pick_loads(self) -> None:
        f = filedialog.askopenfilename(
            title="Pick a Flight Loads Excel exported by this app",
            initialdir=str(Path.home() / "Documents"),
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if f:
            self.zenith_fh_loads_path.set(f)

    def _zenith_fh_pick_output(self) -> None:
        d = filedialog.askdirectory(
            title="Pick the output folder",
            initialdir=self.zenith_fh_output_dir.get() or str(Path.home()),
        )
        if d:
            self.zenith_fh_output_dir.set(d)

    def _zenith_fh_run(self) -> None:
        folder = self.zenith_fh_input_dir.get().strip()
        if not folder or not Path(folder).is_dir():
            messagebox.showerror(
                "Flight History Analyzer",
                f"Pick a valid logs folder first.\n\nCurrent: {folder!r}",
            )
            return
        if self._zenith_fh_worker and self._zenith_fh_worker.is_alive():
            messagebox.showinfo(
                "Flight History Analyzer", "An audit is already running.",
            )
            return
        # Clear previous results
        for child in self.zenith_fh_tree.get_children():
            self.zenith_fh_tree.delete(child)
        self._zenith_fh_last_report = None
        self.btn_zenith_fh_export.configure(state="disabled")
        self.btn_zenith_fh_run.configure(state="disabled")
        self.btn_zenith_fh_misuse.configure(state="disabled")
        self.zenith_fh_status_label.configure(text="Parsing files…")
        self.zenith_fh_progress.configure(value=0, maximum=1)

        loads_path = self.zenith_fh_loads_path.get().strip()
        try:
            high_thresh = float(self.zenith_fh_high_threshold.get())
            low_thresh = float(self.zenith_fh_low_threshold.get())
        except (tk.TclError, ValueError):
            messagebox.showerror(
                "Flight History Analyzer",
                "Verdict thresholds must be numbers. Resetting to defaults.",
            )
            self.zenith_fh_high_threshold.set(
                zenith_loads_index.HIGH_LOAD_THRESHOLD,
            )
            self.zenith_fh_low_threshold.set(
                zenith_loads_index.LOW_LOAD_THRESHOLD,
            )
            self.btn_zenith_fh_run.configure(state="normal")
            return
        if high_thresh <= low_thresh:
            messagebox.showerror(
                "Flight History Analyzer",
                f"QUESTIONABLE threshold ({high_thresh}) must be greater than "
                f"JUSTIFIED threshold ({low_thresh}).",
            )
            self.btn_zenith_fh_run.configure(state="normal")
            return

        def worker() -> None:
            try:
                events: list = []

                def progress(i: int, total: int, name: str) -> None:
                    self._post(MSG_ZENITH_FH_PROGRESS, (i, total, name))

                events.extend(zenith_history_parser.collect_history(
                    folder, progress_cb=progress,
                ))
                self._post(MSG_ZENITH_FH_PARSED, len(events))

                load_lookup = None
                if loads_path:
                    self._post(MSG_ZENITH_FH_PROGRESS,
                               (1, 1, f"Loading {Path(loads_path).name}…"))
                    entries = zenith_loads_index.read_flight_loads_excel(loads_path)
                    load_lookup = zenith_loads_index.LoadLookup.from_entries(entries)

                report = zenith_history_analyzer.run_history_audit(
                    events, include_raw=True, load_lookup=load_lookup,
                    high_threshold=high_thresh, low_threshold=low_thresh,
                )
                self._post(MSG_ZENITH_FH_DONE, report)
            except Exception as exc:  # noqa: BLE001 — surface to UI
                log.exception("Flight History audit failed")
                self._post(MSG_ZENITH_FH_ERROR, f"{type(exc).__name__}: {exc}")

        self._zenith_fh_worker = threading.Thread(target=worker, daemon=True)
        self._zenith_fh_worker.start()

    def _zenith_pnr_misuse_run(self) -> None:
        """Re-pivot the same ModificationHistory corpus into a PNR misuse audit.

        Reuses the Flight History logs + output folders. No network — parses the
        already-downloaded .xls logs, runs the structural misuse detectors, and writes
        a ranked risk worklist + flags workbook. Flags are review leads, not findings.
        """
        folder = self.zenith_fh_input_dir.get().strip()
        if not folder or not Path(folder).is_dir():
            messagebox.showerror(
                "PNR Misuse audit",
                f"Pick a valid logs folder first.\n\nCurrent: {folder!r}")
            return
        out_dir = self.zenith_fh_output_dir.get().strip() or str(Path.home() / "Documents")
        if self._zenith_fh_worker and self._zenith_fh_worker.is_alive():
            messagebox.showinfo("PNR Misuse audit", "An audit is already running.")
            return
        self.btn_zenith_fh_misuse.configure(state="disabled")
        self.btn_zenith_fh_run.configure(state="disabled")
        self.zenith_fh_status_label.configure(text="Running PNR misuse audit…")

        def worker() -> None:
            try:
                events = zenith_history_parser.collect_history(folder)
                if not events:
                    self._post(MSG_ZENITH_PNRMISUSE_ERROR, (
                        f"No ModificationHistory*.xls flight logs found in:\n{folder}\n\n"
                        "The PNR Misuse audit reads the raw flight-history .xls files that the "
                        "Flight History Analyzer downloads from Zenith — NOT report spreadsheets "
                        "(e.g. a Reissues_by_Counter .xlsx has no agent/coupon-status/time detail "
                        "to audit). Download the flight history first, then point this folder at "
                        "those ModificationHistory*.xls files."))
                    return
                report = zenith_pnr_history_analyzer.run_pnr_misuse_audit(events)
                out_path = excel_io.build_zenith_pnr_misuse_output_path(Path(out_dir))
                excel_io.write_zenith_pnr_misuse_audit(out_path, report)
                sev: dict[str, int] = {}
                for f in report.flags:
                    sev[f.severity] = sev.get(f.severity, 0) + 1
                top = report.risk_worklist[0] if report.risk_worklist else None
                self._post(MSG_ZENITH_PNRMISUSE_DONE, {
                    "path": str(out_path),
                    "events": report.event_count,
                    "pnrs": report.pnr_count,
                    "flags": len(report.flags),
                    "critical": sev.get("critical", 0),
                    "high": sev.get("high", 0),
                    "top": (f"{top.grain} {top.entity} (score {top.score})" if top else "—"),
                })
            except Exception as exc:  # noqa: BLE001 — surface to UI
                log.exception("PNR misuse audit failed")
                self._post(MSG_ZENITH_PNRMISUSE_ERROR, f"{type(exc).__name__}: {exc}")

        self._zenith_fh_worker = threading.Thread(target=worker, daemon=True)
        self._zenith_fh_worker.start()

    def _zenith_fh_render(self, report) -> None:
        """Drop the audit headlines into the on-screen summary tree."""
        tree = self.zenith_fh_tree
        for child in tree.get_children():
            tree.delete(child)
        start, end = report.date_range
        rng = (
            f"{start.strftime('%Y-%m-%d')} → {end.strftime('%Y-%m-%d')}"
            if start and end else "—"
        )
        rows = [
            ("Files parsed", f"{report.file_count}"),
            ("Total events", f"{report.event_count:,}"),
            ("Date range", rng),
            ("PNRs with class history", f"{len(report.class_trajectories):,}"),
            (
                "  · with downgrades",
                f"{sum(1 for t in report.class_trajectories if t.total_downgrade_severity > 0):,}",
            ),
            ("Downgrade leaders (agents)", f"{len(report.downgrade_leaders):,}"),
            ("G-class events", f"{len(report.g_class_events):,}"),
            ("Distinct agents", f"{len(report.agent_activity):,}"),
            ("Revenue Mgmt changes", f"{len(report.revenue_mgmt_changes):,}"),
            ("Suspicious flags (high)",
             f"{sum(1 for f in report.suspicious_flags if f.severity == 'high'):,}"),
            ("Suspicious flags (medium)",
             f"{sum(1 for f in report.suspicious_flags if f.severity == 'medium'):,}"),
        ]
        for metric, value in rows:
            tree.insert("", "end", values=(metric, value))
        if report.downgrade_leaders:
            tree.insert("", "end", values=("", ""))
            tree.insert("", "end", values=("Top downgrade leaders", ""))
            for d in report.downgrade_leaders[:5]:
                label = f"  · {d.agent_display_name or d.agent_user_id}"
                tree.insert("", "end", values=(
                    label[:80],
                    f"{d.total_severity} sev / {d.downgrade_event_count} events",
                ))
        if report.downgrade_justifications:
            from collections import Counter
            counts = Counter(j.verdict for j in report.downgrade_justifications)
            # Tag the verdicts with a semantic colour so the user can
            # tell QUESTIONABLE (red) from JUSTIFIED (green) at a glance.
            tree.tag_configure("verdict_bad", background=self._COLOR_ROW_BAD)
            tree.tag_configure("verdict_warn", background=self._COLOR_ROW_WARN)
            tree.tag_configure("verdict_good", background=self._COLOR_ROW_GOOD)
            tree.insert("", "end", values=("", ""))
            tree.insert("", "end", values=("Downgrade verdicts (load-aware)", ""))
            verdict_tags = {
                "QUESTIONABLE": ("verdict_bad",),
                "SITUATIONAL": ("verdict_warn",),
                "JUSTIFIED": ("verdict_good",),
                "UNKNOWN": (),
            }
            for v in ("QUESTIONABLE", "SITUATIONAL", "JUSTIFIED", "UNKNOWN"):
                if counts.get(v):
                    tree.insert(
                        "", "end",
                        values=(f"  · {v}", f"{counts[v]:,}"),
                        tags=verdict_tags[v],
                    )
        if report.pnr_routes:
            tree.insert("", "end", values=("", ""))
            tree.insert("", "end", values=("PNR Routes (enriched)", ""))
            disrupted = sum(
                1 for r in report.pnr_routes
                if r.refunded_count + r.voided_count > 0
            )
            tree.insert(
                "", "end",
                values=("  · Total PNRs", f"{len(report.pnr_routes):,}"),
            )
            tree.insert(
                "", "end",
                values=(
                    "  · With refunds/voids",
                    f"{disrupted:,}",
                ),
            )

    def _zenith_fh_export(self) -> None:
        if self._zenith_fh_last_report is None:
            messagebox.showerror(
                "Flight History Analyzer", "Run an audit first.",
            )
            return
        folder = Path(self.zenith_fh_output_dir.get().strip() or str(Path.home()))
        try:
            out_path = excel_io.build_zenith_history_output_path(folder)
            excel_io.write_zenith_history_audit(out_path, self._zenith_fh_last_report)
        except Exception as exc:  # noqa: BLE001 — surface
            log.exception("Flight History export failed")
            messagebox.showerror(
                "Flight History Analyzer",
                f"Export failed: {type(exc).__name__}: {exc}",
            )
            return
        messagebox.showinfo(
            "Flight History Analyzer — Done",
            f"Wrote audit workbook to:\n\n{out_path}",
        )

    # ------------------------------------------------------------------
    # Phase 3 — download fresh history files from Zenith
    # ------------------------------------------------------------------

    def _zenith_fh_download_dialog(self) -> None:
        """Modal prompt for date range, then kick off the downloader."""
        if not getattr(self, "_zenith_session", None):
            messagebox.showerror(
                "Download from Zenith",
                "Sign in to Zenith first (top of this tab).",
            )
            return
        if self._zenith_dl_worker and self._zenith_dl_worker.is_alive():
            messagebox.showinfo(
                "Download from Zenith", "A download is already running.",
            )
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Download flight history from Zenith")
        dlg.transient(self.root)
        dlg.resizable(False, False)
        # Pre-fill with the last week.
        from datetime import date, timedelta
        today = date.today()
        last_week = today - timedelta(days=7)
        from_var = tk.StringVar(value=last_week.strftime("%d/%m/%Y"))
        to_var = tk.StringVar(value=today.strftime("%d/%m/%Y"))
        folder_var = tk.StringVar(value=self.zenith_fh_input_dir.get())
        skip_var = tk.BooleanVar(value=True)

        frm = ttk.Frame(dlg, padding=12)
        frm.grid(row=0, column=0)
        ttk.Label(frm, text="From (DD/MM/YYYY):").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=from_var, width=14).grid(row=0, column=1, padx=(8, 0))
        ttk.Label(frm, text="To (DD/MM/YYYY):").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(frm, textvariable=to_var, width=14).grid(row=1, column=1, padx=(8, 0), pady=(6, 0))
        ttk.Label(frm, text="Save to folder:").grid(row=2, column=0, sticky="w", pady=(6, 0))
        folder_entry = ttk.Entry(frm, textvariable=folder_var, width=50)
        folder_entry.grid(row=2, column=1, padx=(8, 0), pady=(6, 0))
        ttk.Button(
            frm, text="…",
            command=lambda: folder_var.set(
                filedialog.askdirectory(initialdir=folder_var.get()) or folder_var.get(),
            ),
        ).grid(row=2, column=2, padx=(4, 0), pady=(6, 0))
        ttk.Checkbutton(
            frm, text="Skip files that already exist (recommended)",
            variable=skip_var,
        ).grid(row=3, column=0, columnspan=3, sticky="w", pady=(8, 0))

        btns = ttk.Frame(frm)
        btns.grid(row=4, column=0, columnspan=3, pady=(12, 0), sticky="e")

        def on_start() -> None:
            try:
                date_from = from_var.get().strip()
                date_to = to_var.get().strip()
                folder = Path(folder_var.get().strip())
                folder.mkdir(parents=True, exist_ok=True)
            except Exception as exc:  # noqa: BLE001 — surface to user
                messagebox.showerror("Download from Zenith", f"Bad inputs: {exc}")
                return
            dlg.destroy()
            self._zenith_fh_start_download(
                date_from, date_to, folder, skip_var.get(),
            )

        ttk.Button(btns, text="Cancel", command=dlg.destroy).pack(side="right")
        ttk.Button(
            btns, text="Start download", command=on_start, style="Primary.TButton",
        ).pack(side="right", padx=(0, 8))
        dlg.grab_set()
        self.root.wait_window(dlg)

    def _zenith_fh_start_download(
        self, date_from: str, date_to: str, folder: Path, skip_existing: bool,
    ) -> None:
        """Worker driver — lists flights, then downloads each into `folder`."""
        self._zenith_dl_stop_flag.clear()
        self.zenith_fh_status_label.configure(
            text=f"Downloading from Zenith ({date_from} → {date_to})…",
        )
        self.zenith_fh_progress.configure(value=0, maximum=1)
        self.btn_zenith_fh_download.configure(state="disabled")

        def worker() -> None:
            try:
                self._post(MSG_ZENITH_DL_STATUS, "Listing flights from Zenith…")
                flights = zenith_history_downloader.list_flights(
                    self._zenith_session, date_from, date_to,
                )
                self._post(
                    MSG_ZENITH_DL_STATUS,
                    f"Found {len(flights)} flights — starting downloads",
                )

                def per_progress(result, idx: int, total: int) -> None:
                    label = (
                        f"{result.flight.flight_number} "
                        f"{result.flight.origin}-{result.flight.destination}"
                    )
                    self._post(
                        MSG_ZENITH_DL_PROGRESS,
                        (idx, total, label, result.status),
                    )

                results = zenith_history_downloader.download_history_batch(
                    self._zenith_session, flights, folder,
                    skip_if_exists=skip_existing,
                    progress_cb=per_progress,
                    stop_event=self._zenith_dl_stop_flag,
                )
                summary: dict[str, int] = {}
                for r in results:
                    summary[r.status] = summary.get(r.status, 0) + 1
                self._post(MSG_ZENITH_DL_DONE, {
                    "folder": str(folder),
                    "summary": summary,
                    "total": len(results),
                })
            except Exception as exc:  # noqa: BLE001 — surface
                log.exception("Zenith history download failed")
                self._post(
                    MSG_ZENITH_DL_ERROR, f"{type(exc).__name__}: {exc}",
                )

        self._zenith_dl_worker = threading.Thread(target=worker, daemon=True)
        self._zenith_dl_worker.start()

    # ------------------------------------------------------------------
    # Phase 2/A — PNR enrichment (customer name + full route)
    # ------------------------------------------------------------------

    def _zenith_fh_enrich_pnrs(self) -> None:
        """Walk every PNR in the last audit + fetch its Dossier from Zenith."""
        if self._zenith_fh_last_report is None:
            messagebox.showinfo(
                "PNR enrichment",
                "Run the audit first — then I'll know which PNRs to fetch.",
            )
            return
        if not getattr(self, "_zenith_session", None):
            messagebox.showerror(
                "PNR enrichment",
                "Sign in to Zenith first (top of this tab).",
            )
            return
        if self._zenith_pnr_worker and self._zenith_pnr_worker.is_alive():
            messagebox.showinfo(
                "PNR enrichment", "Enrichment is already running.",
            )
            return

        # Distinct PNRs across every audit row that carries one.
        report = self._zenith_fh_last_report
        unique_pnrs: list[str] = []
        seen: set[str] = set()
        for t in report.class_trajectories:
            if t.pnr and t.pnr not in seen:
                seen.add(t.pnr)
                unique_pnrs.append(t.pnr)
        for g in report.g_class_events:
            if g.pnr and g.pnr not in seen:
                seen.add(g.pnr)
                unique_pnrs.append(g.pnr)
        # Also pull raw events' PNRs — many bookings only appear here.
        for e in report.raw_events:
            if e.pnr and e.pnr not in seen:
                seen.add(e.pnr)
                unique_pnrs.append(e.pnr)
        if not unique_pnrs:
            messagebox.showinfo(
                "PNR enrichment",
                "No PNRs found in this audit — nothing to enrich.",
            )
            return

        # Estimate runtime so the user can decide.
        cache_hits = sum(
            1 for p in unique_pnrs if self._zenith_pnr_cache.get(p) is not None
        )
        misses = len(unique_pnrs) - cache_hits
        est_secs = misses * 1   # ~0.5s fetch + 0.5s delay per miss
        est_min = max(1, est_secs // 60)
        if misses > 50:
            proceed = messagebox.askyesno(
                "PNR enrichment",
                f"This audit has {len(unique_pnrs)} unique PNRs.\n"
                f"  · {cache_hits} already cached locally\n"
                f"  · {misses} need fetching from Zenith\n\n"
                f"Estimated time: ~{est_min} min.\n\n"
                f"Continue?",
            )
            if not proceed:
                return

        self.btn_zenith_fh_pnr.configure(state="disabled")
        self.btn_zenith_fh_run.configure(state="disabled")
        self.zenith_fh_status_label.configure(
            text=f"Enriching {len(unique_pnrs)} PNRs (using local cache for {cache_hits})…",
        )
        self.zenith_fh_progress.configure(value=0, maximum=len(unique_pnrs))

        cache = self._zenith_pnr_cache
        sess = self._zenith_session

        def worker() -> None:
            try:
                def progress(idx: int, total: int, code: str, status: str) -> None:
                    self._post(MSG_ZENITH_PNR_PROGRESS, (idx, total, code, status))

                def skip_cached(code: str):
                    return cache.get(code)

                fresh = zenith_pnr_client.lookup_many(
                    sess, unique_pnrs,
                    delay_s=0.5,
                    skip_cached=skip_cached,
                    progress_cb=progress,
                )
                # Persist any new entries to the cache.
                new_entries = [
                    d for code, d in fresh.items()
                    if cache.get(code) is None
                ]
                if new_entries:
                    cache.put_many(new_entries)

                # Apply enrichment using the union of fresh + cached.
                merged: dict = {}
                for code in unique_pnrs:
                    cached = cache.get(code)
                    if cached is not None:
                        merged[code] = cached
                enriched = zenith_history_analyzer.apply_pnr_enrichment(
                    report, merged,
                )
                self._post(MSG_ZENITH_PNR_DONE, enriched)
            except Exception as exc:  # noqa: BLE001 — surface
                log.exception("PNR enrichment failed")
                self._post(
                    MSG_ZENITH_PNR_ERROR, f"{type(exc).__name__}: {exc}",
                )

        self._zenith_pnr_worker = threading.Thread(target=worker, daemon=True)
        self._zenith_pnr_worker.start()

    # ==================================================================
    # Zenith Customer Lookup tab — actions
    # ==================================================================

    def _refresh_zenith_cache_label(self) -> None:
        counts = self._zenith_cache.counts_by_status()
        ok = counts.get(zenith_client.STATUS_OK, 0)
        nf = counts.get(zenith_client.STATUS_NOT_FOUND, 0)
        err = counts.get(zenith_client.STATUS_ERROR, 0)
        total = ok + nf + err
        if total == 0:
            text = "Cache is empty."
        else:
            text = f"Cache: {ok:,} OK · {nf:,} not found · {err:,} errors  ({total:,} total)"
        self.zenith_progress_label.configure(text=text)

    def _zenith_pick_input(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Excel with Customer IDs",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if not path:
            return
        self.zenith_input_path.set(path)
        try:
            sheets = excel_io.list_sheet_names(Path(path))
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Zenith", f"Could not read sheets: {exc}")
            return
        self.zenith_sheet_combo.configure(values=sheets)
        if sheets:
            self.zenith_sheet_name.set(sheets[0])
            self._zenith_reload_columns()

    def _zenith_pick_output(self) -> None:
        folder = filedialog.askdirectory(title="Choose output folder")
        if folder:
            self.zenith_output_dir.set(folder)

    def _zenith_reload_columns(self) -> None:
        path = self.zenith_input_path.get().strip()
        sheet = self.zenith_sheet_name.get().strip()
        if not path or not sheet:
            return
        try:
            cols = excel_io.list_columns(Path(path), sheet)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Zenith", f"Could not read columns: {exc}")
            return
        self._zenith_sheet_columns = cols
        self.zenith_col_combo.configure(values=cols)
        if cols and not self.zenith_column_name.get():
            # Common defaults the user might have
            for guess in ("Customer ID", "CustomerID", "ID", "id"):
                if guess in cols:
                    self.zenith_column_name.set(guess)
                    break
            else:
                self.zenith_column_name.set(cols[0])

    def _zenith_collapse_login(self) -> None:
        """Hide the full sign-in form; show the compact 'signed in' strip."""
        if hasattr(self, "zenith_login_section"):
            self.zenith_login_section.pack_forget()
        if hasattr(self, "zenith_login_compact"):
            sv = self._zenith_session.state_values if self._zenith_session else {}
            email = (self._zenith_session.state_values.get("ID_ADMIN")
                     if self._zenith_session else "")
            self.zenith_login_compact_label.configure(
                text=(
                    f"●  Signed in to Zenith  ·  user={sv.get('ID_ADMIN')}  ·  "
                    f"company={sv.get('ID_SOCIETE')}"
                ),
            )
            self.zenith_login_compact.pack(
                fill="x", padx=8, pady=(8, 4),
                before=getattr(self, "zenith_inner_notebook", None),
            )

    def _zenith_show_login_form(self) -> None:
        """Re-expand the full form (for 'Sign in again')."""
        if hasattr(self, "zenith_login_compact"):
            self.zenith_login_compact.pack_forget()
        if hasattr(self, "zenith_login_section"):
            self.zenith_login_section.pack(
                fill="x", padx=4, pady=(8, 4),
                before=getattr(self, "zenith_inner_notebook", None),
            )

    def _zenith_login(self) -> None:
        user = self.zenith_username.get().strip()
        pwd = self.zenith_pwd_entry.get()
        company = self.zenith_company.get().strip() or "usba"
        if not user or not pwd:
            messagebox.showerror(
                "Zenith", "Enter both username and password.",
            )
            return
        self.btn_zenith_login.configure(state="disabled", text="Signing in…")
        self.zenith_login_status.configure(text="Connecting…")

        def worker() -> None:
            try:
                sess = zenith_client.ZenithSession.from_credentials(
                    user, pwd, company_code=company,
                )
                self._post(MSG_ZENITH_LOGGED_IN, sess)
            except zenith_client.LoginError as exc:
                self._post(MSG_ZENITH_LOGIN_FAILED, str(exc))
            except Exception as exc:  # noqa: BLE001
                log.exception("Zenith login crashed")
                self._post(MSG_ZENITH_LOGIN_FAILED, f"{type(exc).__name__}: {exc}")

        threading.Thread(target=worker, daemon=True).start()

    def _zenith_run(self) -> None:
        if self._zenith_session is None:
            messagebox.showerror(
                "Zenith", "Sign in to Zenith first.",
            )
            return
        in_path = self.zenith_input_path.get().strip()
        sheet = self.zenith_sheet_name.get().strip()
        col = self.zenith_column_name.get().strip()
        if not in_path or not sheet or not col:
            messagebox.showerror(
                "Zenith", "Pick an input file, sheet, and ID column first.",
            )
            return
        out_dir = Path(self.zenith_output_dir.get().strip() or str(Path.home()))

        # Read IDs
        try:
            ids = excel_io.read_zenith_ids(Path(in_path), sheet, col)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Zenith", f"Could not read IDs: {exc}")
            return
        if not ids:
            messagebox.showerror("Zenith", "No IDs found in that column.")
            return
        if self.zenith_test_mode.get():
            ids = ids[:100]

        # Optionally skip cached
        if self.zenith_skip_cached.get():
            cached = self._zenith_cache.cached_ids(only_ok=True)
            skipped = len([i for i in ids if i in cached])
            ids = [i for i in ids if i not in cached]
            if skipped:
                self._zenith_log(f"Skipping {skipped:,} IDs already in cache (OK).")
            if not ids:
                messagebox.showinfo(
                    "Zenith",
                    "All requested IDs are already cached. Nothing to fetch.\n\n"
                    "Use 'Export cache to Excel' to write them out.",
                )
                return

        # Big-run confirmation
        if len(ids) > 5000 and not messagebox.askyesno(
            "Zenith — large run",
            f"About to fetch {len(ids):,} customer records.\n\n"
            f"At concurrency={int(self.zenith_concurrency.get())} and "
            f"delay={float(self.zenith_delay_s.get()):.1f}s, this could take "
            f"a few hours. Continue?",
        ):
            return

        # Worker setup
        self._zenith_stop_flag.clear()
        self._zenith_pause_flag.clear()
        out_path = excel_io.build_zenith_output_path(out_dir)
        cfg = {
            "ids": ids,
            "concurrency": int(self.zenith_concurrency.get()),
            "delay_s": float(self.zenith_delay_s.get()),
            "out_path": out_path,
        }
        self._zenith_worker = threading.Thread(
            target=self._zenith_worker_run, args=(cfg,), daemon=True,
        )
        self._zenith_log(
            f"Run starting — {len(ids):,} IDs · "
            f"concurrency={cfg['concurrency']} · delay={cfg['delay_s']:.1f}s",
        )
        self.zenith_progress_bar.configure(value=0, maximum=len(ids))
        self.btn_zenith_run.configure(state="disabled")
        self.btn_zenith_pause.configure(state="normal")
        self.btn_zenith_stop.configure(state="normal")
        self._zenith_worker.start()

    def _zenith_worker_run(self, cfg: dict) -> None:
        ids = cfg["ids"]
        total = len(ids)
        ok = nf = err = 0

        def progress_cb(result, completed: int, total_n: int) -> None:
            nonlocal ok, nf, err
            self._zenith_cache.save_result(result)
            if result.status == zenith_client.STATUS_OK:
                ok += 1
            elif result.status == zenith_client.STATUS_NOT_FOUND:
                nf += 1
            else:
                err += 1
            self._post(MSG_ZENITH_PROGRESS, (completed, total_n, ok, nf, err))
            self._post(MSG_ZENITH_RESULT, result)

        try:
            zenith_client.fetch_many(
                self._zenith_session, ids,
                concurrency=cfg["concurrency"],
                delay_s=cfg["delay_s"],
                progress_cb=progress_cb,
                stop_event=self._zenith_stop_flag,
                pause_event=self._zenith_pause_flag,
            )
        except Exception as exc:  # noqa: BLE001
            log.exception("Zenith run crashed")
            self._post(MSG_ZENITH_ERROR, f"{type(exc).__name__}: {exc}")
            return

        # Write everything cached to Excel
        try:
            excel_io.write_zenith_results(
                cfg["out_path"], self._zenith_cache.iter_all(),
            )
        except Exception as exc:  # noqa: BLE001
            log.exception("Zenith Excel write failed")
            self._post(MSG_ZENITH_ERROR, f"Excel write failed: {exc}")
            return
        self._post(MSG_ZENITH_DONE, str(cfg["out_path"]))

    def _zenith_pause(self) -> None:
        self._zenith_pause_flag.set()
        self.btn_zenith_pause.configure(state="disabled")
        self.btn_zenith_resume.configure(state="normal")
        self._zenith_log("Paused.")

    def _zenith_resume(self) -> None:
        self._zenith_pause_flag.clear()
        self.btn_zenith_resume.configure(state="disabled")
        self.btn_zenith_pause.configure(state="normal")
        self._zenith_log("Resumed.")

    def _zenith_stop(self) -> None:
        self._zenith_stop_flag.set()
        self._zenith_pause_flag.clear()
        self._zenith_log("Stopping after current batch finishes…")
        self.btn_zenith_stop.configure(state="disabled")

    def _zenith_reset_cache(self) -> None:
        counts = self._zenith_cache.counts_by_status()
        total = sum(counts.values())
        if not total:
            messagebox.showinfo("Zenith", "Cache already empty.")
            return
        if not messagebox.askyesno(
            "Reset cache?",
            f"Delete all {total:,} cached customer records?\n\n"
            "Useful for starting fresh. The Excel files you've already "
            "exported are NOT affected.",
        ):
            return
        self._zenith_cache.reset()
        self._refresh_zenith_cache_label()
        self._zenith_log("Cache reset.")

    def _zenith_clear_errors(self) -> None:
        dropped = self._zenith_cache.clear_errors()
        self._refresh_zenith_cache_label()
        self._zenith_log(
            f"Cleared {dropped:,} error rows — they'll be retried on the next run.",
        )

    def _zenith_export_cache(self) -> None:
        out_dir = Path(self.zenith_output_dir.get().strip() or str(Path.home()))
        path = excel_io.build_zenith_output_path(out_dir)
        try:
            excel_io.write_zenith_results(path, self._zenith_cache.iter_all())
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Zenith", f"Excel write failed: {exc}")
            return
        self._zenith_log(f"Exported cache to {path}")
        messagebox.showinfo("Zenith — Export Done", f"Wrote:\n\n{path}")

    def _zenith_log(self, text: str) -> None:
        self.zenith_log_text.configure(state="normal")
        self.zenith_log_text.insert("end", text + "\n")
        self.zenith_log_text.see("end")
        self.zenith_log_text.configure(state="disabled")

    def _zenith_reset_buttons(self) -> None:
        self.btn_zenith_run.configure(state="normal")
        self.btn_zenith_pause.configure(state="disabled")
        self.btn_zenith_resume.configure(state="disabled")
        self.btn_zenith_stop.configure(state="disabled")

    # ==================================================================
    # Zenith Flight Loads sub-tab — actions
    # ==================================================================

    def _zenith_fl_log(self, text: str) -> None:
        self.zenith_fl_log_text.configure(state="normal")
        self.zenith_fl_log_text.insert("end", text + "\n")
        self.zenith_fl_log_text.see("end")
        self.zenith_fl_log_text.configure(state="disabled")

    def _zenith_fl_pick_output(self) -> None:
        folder = filedialog.askdirectory(title="Choose flight-loads output folder")
        if folder:
            self.zenith_fl_output_dir.set(folder)

    def _zenith_fl_run(self) -> None:
        if self._zenith_session is None:
            messagebox.showerror(
                "Zenith", "Sign in to Zenith first (top of this tab).",
            )
            return
        date_from = self.zenith_fl_date_from.get().strip()
        date_to = self.zenith_fl_date_to.get().strip()
        if not date_from or not date_to:
            messagebox.showerror(
                "Zenith", "Enter both From and To dates (DD/MM/YYYY).",
            )
            return
        try:
            from datetime import datetime
            datetime.strptime(date_from, "%d/%m/%Y")
            datetime.strptime(date_to, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror(
                "Zenith", "Dates must be in DD/MM/YYYY format.",
            )
            return

        try:
            page_size = int(self.zenith_fl_page_size.get())
        except ValueError:
            page_size = 100
        chunk_days = max(1, int(self.zenith_fl_chunk_days.get()))
        delay_s = float(self.zenith_fl_delay_s.get())
        out_dir = Path(
            self.zenith_fl_output_dir.get().strip() or str(Path.home()),
        )
        out_path = excel_io.build_zenith_flight_output_path(out_dir)

        self._zenith_fl_stop_flag.clear()
        self.btn_zenith_fl_run.configure(state="disabled")
        self.btn_zenith_fl_stop.configure(state="normal")
        self.zenith_fl_progress_bar.configure(value=0, maximum=1)
        self._zenith_fl_log(
            f"Flight loads run starting: {date_from} → {date_to} · "
            f"page_size={page_size} · chunk={chunk_days}d · delay={delay_s:.1f}s",
        )

        cfg = {
            "date_from": date_from, "date_to": date_to,
            "page_size": page_size, "chunk_days": chunk_days,
            "delay_s": delay_s, "out_path": out_path,
        }
        self._zenith_fl_worker = threading.Thread(
            target=self._zenith_fl_worker_run, args=(cfg,), daemon=True,
        )
        self._zenith_fl_worker.start()

    def _zenith_fl_worker_run(self, cfg: dict) -> None:
        def progress_cb(chunk_label, done, total, rows_so_far):
            self._post(
                MSG_ZENITH_FL_PROGRESS,
                (chunk_label, done, total, rows_so_far),
            )

        try:
            rows = zenith_client.fetch_flight_loads(
                self._zenith_session,
                cfg["date_from"], cfg["date_to"],
                page_size=cfg["page_size"],
                chunk_days=cfg["chunk_days"],
                inter_call_delay_s=cfg["delay_s"],
                progress_cb=progress_cb,
                stop_event=self._zenith_fl_stop_flag,
            )
        except zenith_client.SessionExpiredError as exc:
            self._post(MSG_ZENITH_FL_ERROR, f"Session expired — sign in again. ({exc})")
            return
        except Exception as exc:  # noqa: BLE001
            log.exception("Zenith flight-loads run crashed")
            self._post(MSG_ZENITH_FL_ERROR, f"{type(exc).__name__}: {exc}")
            return

        if not rows:
            self._post(
                MSG_ZENITH_FL_ERROR,
                "No flights found for that date range (or the run was stopped).",
            )
            return

        try:
            excel_io.write_zenith_flight_loads(cfg["out_path"], rows)
        except Exception as exc:  # noqa: BLE001
            log.exception("Flight-loads Excel write failed")
            self._post(MSG_ZENITH_FL_ERROR, f"Excel write failed: {exc}")
            return
        # Keep the rows in memory so the user can later append them to
        # an Ordered Report without re-running the fetch.
        self._zenith_fl_last_rows = rows
        self._post(MSG_ZENITH_FL_DONE, str(cfg["out_path"]))

    def _zenith_fl_append_ordered(self) -> None:
        """Append the last run's rows to a user-picked Ordered Report file."""
        if not self._zenith_fl_last_rows:
            messagebox.showinfo(
                "Append to Ordered Report",
                "Run a Flight Loads pull first — there's no data to append yet.",
            )
            return
        # Pick an existing file or a new path; create_if_missing handles either.
        f = filedialog.asksaveasfilename(
            title="Pick the Ordered Report file (existing or new)",
            initialdir=str(Path.home() / "Documents"),
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            confirmoverwrite=False,
        )
        if not f:
            return
        try:
            summary = excel_io.append_flight_loads_to_ordered_report(
                Path(f), self._zenith_fl_last_rows, create_if_missing=True,
            )
        except Exception as exc:  # noqa: BLE001
            log.exception("Ordered Report append failed")
            messagebox.showerror(
                "Append to Ordered Report — Error",
                f"{type(exc).__name__}: {exc}",
            )
            return
        self._zenith_fl_log(
            f"Ordered Report updated: +{summary['dates_added']} new dates, "
            f"{summary['dates_refreshed']} dates refreshed, "
            f"{summary['cells_updated']} cells written.",
        )
        messagebox.showinfo(
            "Append to Ordered Report — Done",
            (
                f"Saved to: {f}\n\n"
                f"  New dates added:      {summary['dates_added']}\n"
                f"  Existing dates refreshed: {summary['dates_refreshed']}\n"
                f"  Total dates in file:  {summary['total_dates']}\n"
                f"  Flights added:        {summary['flights_added']}\n"
                f"  Flights updated:      {summary['flights_updated']}\n"
                f"  Cells written:        {summary['cells_updated']}"
            ),
        )

    def _zenith_fl_stop(self) -> None:
        self._zenith_fl_stop_flag.set()
        self._zenith_fl_log("Stopping after current page finishes…")
        self.btn_zenith_fl_stop.configure(state="disabled")

    def _zenith_fl_reset_buttons(self) -> None:
        self.btn_zenith_fl_run.configure(state="normal")
        self.btn_zenith_fl_stop.configure(state="disabled")

    # ------------------------------------------------------------------
    # Passenger manifest drill-down
    # ------------------------------------------------------------------

    @staticmethod
    def _load_pct_bucket(seats_available: str) -> str:
        """Map a '13/410 97%' string to a traffic-light tag, '' if unknown."""
        import re as _re
        m = _re.search(r"(-?\d+(?:\.\d+)?)\s*%", seats_available or "")
        if not m:
            return ""
        pct = float(m.group(1))
        if pct >= 90:
            return "load_hi"
        if pct >= 70:
            return "load_mid"
        return "load_lo"

    def _zenith_fl_populate_legs(self) -> None:
        """Fill the per-leg grid from the last run, honouring the nested
        Region + Direction filters and load-factor colour tags.

        tree iid = str(absolute index into self._zenith_fl_last_rows), so
        selection maps straight back to the FlightLoadRow (with its
        id_vol/id_leg/id_aero keys) even when rows are filtered out.
        """
        tree = self.zenith_fl_legs_tree
        for child in tree.get_children():
            tree.delete(child)

        region_f = self.zenith_fl_filter_region.get()
        dir_f = self.zenith_fl_filter_dir.get()
        has_keys = False
        shown = 0
        for idx, r in enumerate(self._zenith_fl_last_rows):
            region = zenith_client.classify_leg_region(r.leg_origin, r.leg_destination)
            direction = zenith_client.classify_leg_direction(r.leg_origin, r.leg_destination)
            if region_f != "All regions" and region != region_f:
                continue
            if dir_f != "All directions" and direction != dir_f:
                continue
            tag = self._load_pct_bucket(r.seats_available)
            tree.insert(
                "", "end", iid=str(idx),
                values=(
                    r.flight_number, r.flight_date, r.leg_route,
                    region, direction, r.leg_cabin,
                    r.seats_available, r.tickets_issued,
                ),
                tags=(tag,) if tag else (),
            )
            shown += 1
            if r.leg_id_leg and r.leg_id_aero:
                has_keys = True

        total = len(self._zenith_fl_last_rows)
        self.zenith_fl_filter_count.configure(
            text=(f"  {shown} of {total} legs" if total else ""),
        )
        state = "normal" if (shown and has_keys) else "disabled"
        self.btn_zenith_fl_pax.configure(state=state)
        self.btn_zenith_fl_pax_all.configure(state=state)

    def _zenith_fl_pull_pax(self) -> None:
        if not getattr(self, "_zenith_session", None):
            messagebox.showerror(
                "Passenger detail", "Sign in to Zenith first (top of this tab).",
            )
            return
        if self._zenith_pax_worker and self._zenith_pax_worker.is_alive():
            messagebox.showinfo("Passenger detail", "A pull is already running.")
            return
        sel = self.zenith_fl_legs_tree.selection()
        if not sel:
            messagebox.showinfo(
                "Passenger detail",
                "Select one or more legs in the grid first "
                "(Ctrl/Shift-click), or use 'Select all legs'.",
            )
            return
        # Map selection → FlightLoadRow legs that carry drill-down keys.
        legs = []
        for iid in sel:
            try:
                r = self._zenith_fl_last_rows[int(iid)]
            except (ValueError, IndexError):
                continue
            if r.leg_id_vol and r.leg_id_leg and r.leg_id_aero:
                legs.append(r)
        if not legs:
            messagebox.showerror(
                "Passenger detail",
                "None of the selected legs have a passenger-list link "
                "(missing id_leg/id_aero).",
            )
            return

        est_min = max(1, len(legs) * 3 // 60)
        if len(legs) > 20 and not messagebox.askyesno(
            "Passenger detail",
            f"Pull the full passenger manifest for {len(legs)} legs?\n\n"
            f"Each leg is a separate fetch (~150–400 passengers). "
            f"Estimated time ~{est_min} min.\n\nContinue?",
        ):
            return

        out_dir = Path(self.zenith_fl_output_dir.get().strip() or str(Path.home()))
        self._zenith_pax_stop_flag.clear()
        self.btn_zenith_fl_pax.configure(state="disabled")
        self.btn_zenith_fl_pax_all.configure(state="disabled")
        self.zenith_fl_progress_bar.configure(value=0, maximum=len(legs))
        self.zenith_fl_pax_status.configure(text=f"Pulling {len(legs)} legs…")
        self._zenith_fl_log(f"Passenger detail: pulling {len(legs)} leg(s)…")

        sess = self._zenith_session

        def worker() -> None:
            try:
                all_pax: list = []
                for i, r in enumerate(legs, start=1):
                    if self._zenith_pax_stop_flag.is_set():
                        break
                    label = f"{r.flight_number} {r.leg_route} {r.flight_date}"
                    recs = zenith_client.fetch_passenger_manifest(
                        sess, r.leg_id_vol, r.leg_id_leg, r.leg_id_aero,
                    )
                    all_pax.extend(recs)
                    self._post(
                        MSG_ZENITH_PAX_PROGRESS,
                        (i, len(legs), label, len(all_pax)),
                    )
                    import time
                    time.sleep(1.0)  # polite gap between legs
                out_path = excel_io.build_zenith_pax_output_path(out_dir)
                excel_io.write_passenger_manifest(out_path, all_pax)
                self._post(MSG_ZENITH_PAX_DONE, {
                    "path": str(out_path), "legs": len(legs), "pax": len(all_pax),
                })
            except zenith_client.SessionExpiredError as exc:
                self._post(MSG_ZENITH_PAX_ERROR, f"Session expired — sign in again. ({exc})")
            except Exception as exc:  # noqa: BLE001
                log.exception("Passenger manifest pull failed")
                self._post(MSG_ZENITH_PAX_ERROR, f"{type(exc).__name__}: {exc}")

        self._zenith_pax_worker = threading.Thread(target=worker, daemon=True)
        self._zenith_pax_worker.start()


def _fmt_dur(seconds: float) -> str:
    seconds = int(seconds)
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}h {m:02d}m"
    if m:
        return f"{m}m {s:02d}s"
    return f"{s}s"


def run() -> None:
    # Ensure the log directory exists *before* basicConfig opens the file.
    config.LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    from logging.handlers import RotatingFileHandler
    handler = RotatingFileHandler(
        str(config.LOG_FILE),
        maxBytes=5_000_000,
        backupCount=2,
        encoding="utf-8",
    )
    logging.basicConfig(
        level=logging.INFO,
        handlers=[handler],
        format="%(asctime)s %(levelname)s %(name)s — %(message)s",
    )
    root = tk.Tk()
    try:
        # Better default font on Windows
        from tkinter import font as tkfont
        default = tkfont.nametofont("TkDefaultFont")
        default.configure(family="Segoe UI", size=10)
    except Exception:
        pass
    app = App(root)

    # Sign-in gate. The Cloud Run backend now has GOOGLE_CLIENT_SECRET
    # configured, so the OAuth code exchange works end-to-end.
    # If IATA_GOOGLE_CLIENT_ID isn't baked (dev build / fork), we still
    # let the app run unauthenticated so a developer build keeps working.
    _REQUIRE_SIGN_IN = True
    if _REQUIRE_SIGN_IN and auth.GOOGLE_CLIENT_ID:
        if not app.ensure_signed_in():
            root.destroy()
            return
    elif not auth.GOOGLE_CLIENT_ID:
        log.info("running unauthenticated — IATA_GOOGLE_CLIENT_ID not baked")

    root.mainloop()
