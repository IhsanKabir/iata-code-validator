"""Excel writer for the flight load-factor inspection.

Produces a workbook with a Flights comparison sheet (every flight, key metrics
+ plain-language flags), a flag legend, and a per-flight detail sheet for each
flagged flight (funnel, demand curve, fare-class mix, hold summary). Kept in its
own module so excel_io.py stays focused.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .flight_load_diagnostics import FLAG_REASONS, FlightDiagnostics

_NAVY = PatternFill("solid", fgColor="1F3864")
_BLUE = PatternFill("solid", fgColor="2E5496")
_RED = PatternFill("solid", fgColor="F4B6B6")
_AMBER = PatternFill("solid", fgColor="FFE699")
_GREEN = PatternFill("solid", fgColor="C6E0B4")
_WHITEB = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_TITLE = Font(color="FFFFFF", bold=True, size=13)
_thin = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_WRAP = Alignment(wrap_text=True, vertical="top")

# detail sheets are heavy; cap so a 100-flight run stays manageable
MAX_DETAIL_SHEETS = 25


def _hdr(ws, row, cols, fill=_BLUE):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        cell.font = _WHITEB
        cell.fill = fill
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _put(ws, row, vals, fill=None, bold=False):
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.border = _BORDER
        cell.alignment = _WRAP
        if fill:
            cell.fill = fill
        if bold:
            cell.font = _BOLD


def _widths(ws, w):
    for i, x in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = x


def _title(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = _TITLE
    c.fill = _NAVY
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


def _lf_fill(lf: float, cap: int):
    if not cap:
        return None
    if lf < 55:
        return _RED
    if lf < 85:
        return _AMBER
    return _GREEN


def _safe_sheet_name(name: str, used: set[str]) -> str:
    clean = re.sub(r"[\\/?*\[\]:]", "-", name)[:28]
    base, n = clean, 2
    while clean in used:
        clean = f"{base[:25]}-{n}"
        n += 1
    used.add(clean)
    return clean


def _write_comparison(ws, diags: list[FlightDiagnostics]) -> None:
    _widths(ws, [12, 9, 12, 7, 7, 7, 8, 7, 8, 8, 7, 9, 10, 40])
    _title(ws, "Flight Load Inspection  ·  scan flags to see WHY each flight behaved as it did", 14)
    _hdr(ws, 2, ["Route", "Flight", "Date", "Dep", "Flown", "~Cap", "LF~%",
                 "PNRs", "No-show", "Cancel", "Held", "Held med h",
                 "Final-wk %", "Flags (why)"])
    ordered = sorted(diags, key=lambda d: (d.route, -d.flown))
    for i, d in enumerate(ordered):
        r = 3 + i
        _put(ws, r, [
            d.route, d.flight_number, d.flight_date, d.departure_time,
            d.flown, d.capacity_est or "?", d.load_factor_est or "?",
            d.total_pnrs, d.no_shows, d.cancellations, d.held_coupons,
            d.held_median_hold_h, d.pct_booked_final_week, ", ".join(d.flags),
        ])
        ws.cell(r, 7).fill = _lf_fill(d.load_factor_est, d.capacity_est) or PatternFill()
        if d.flags:
            ws.cell(r, 14).fill = _AMBER
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:N{2 + len(ordered)}"


def _write_legend(ws) -> None:
    _widths(ws, [22, 80])
    _title(ws, "Flag legend — what each flag means", 2)
    _hdr(ws, 2, ["Flag", "Meaning"])
    for i, (flag, reason) in enumerate(FLAG_REASONS.items()):
        _put(ws, 3 + i, [flag, reason], bold=False)
        ws.cell(3 + i, 1).font = _BOLD
    note = ("Capacity & LF~ are ESTIMATES (occupied seat-map max row x 6). Flown / no-show / "
            "cancel / held counts are exact. Hold durations use a coupon-ID creation clock.")
    ws.merge_cells(start_row=4 + len(FLAG_REASONS), start_column=1,
                   end_row=4 + len(FLAG_REASONS), end_column=2)
    ws.cell(4 + len(FLAG_REASONS), 1, note).font = Font(italic=True)


def _write_detail(ws, d: FlightDiagnostics) -> None:
    _widths(ws, [22, 14, 16, 14])
    _title(ws, f"{d.route}  {d.flight_number}  {d.flight_date} {d.departure_time}", 4)
    r = 2
    if d.flags:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, "FLAGS: " + ", ".join(d.flags)).font = Font(bold=True, color="C00000")
        r += 2
    # funnel
    _hdr(ws, r, ["Funnel", "Count", "", ""])
    r += 1
    empty = (d.capacity_est - d.flown) if d.capacity_est else "?"
    for label, val, fill in [
        ("PNRs touched", d.total_pnrs, None),
        ("Flown (pax)", d.flown, _GREEN),
        ("  checked-in / boarded", f"{d.checked_in} / {d.boarded}", None),
        ("No-shows", d.no_shows, _RED if d.no_shows else None),
        ("Segment-cancelled PNRs", d.cancellations, None),
        ("Held then cancelled (coupons)", d.held_coupons, None),
        (f"Capacity (est) / LF (est)", f"{d.capacity_est or '?'} / {d.load_factor_est or '?'}%", None),
        ("Empty seats (est)", empty, _RED),
    ]:
        _put(ws, r, [label, val, "", ""], fill=fill)
        ws.cell(r, 1).font = _BOLD
        r += 1
    r += 1
    # hold summary
    _hdr(ws, r, ["Held-coupon hold duration", "Value", "", ""])
    r += 1
    for label, val in [
        ("median (hours)", d.held_median_hold_h),
        ("max (hours)", d.held_max_hold_h),
        ("% under 24h", d.held_pct_under_24h),
        ("clock error p90 (hours)", d.clock_error_h),
    ]:
        _put(ws, r, [label, val, "", ""])
        r += 1
    r += 1
    # demand curve
    _hdr(ws, r, ["Weeks before dep", "New PNRs", "Cumulative", ""])
    r += 1
    for w, new, cum in d.demand_curve:
        _put(ws, r, [f"{w} wk", new, cum, ""])
        r += 1
    r += 1
    # fare-class mix
    _hdr(ws, r, ["Fare class (flown)", "Coupons", "", ""])
    r += 1
    for cls, n in d.fare_class_mix:
        _put(ws, r, [cls, n, "", ""])
        r += 1
    r += 1
    # channels
    _hdr(ws, r, ["Booking channel", "PNRs", "", ""])
    r += 1
    for ch, n in d.channel_mix:
        _put(ws, r, [ch, n, "", ""])
        r += 1


def write_flight_inspection(path: str | Path, diags: list[FlightDiagnostics]) -> int:
    """Write the inspection workbook. Returns the number of flights written.

    Comparison sheet (all flights) + legend + per-flight detail sheets for
    flagged flights (most-flagged first, capped at MAX_DETAIL_SHEETS)."""
    wb = Workbook()
    _write_comparison(wb.active, diags)
    wb.active.title = "Flights"
    _write_legend(wb.create_sheet("Flag legend"))

    flagged = sorted((d for d in diags if d.flags),
                     key=lambda d: (-len(d.flags), d.load_factor_est))
    used: set[str] = set()
    for d in flagged[:MAX_DETAIL_SHEETS]:
        name = _safe_sheet_name(f"{d.route} {d.flight_date[:5]}", used)
        _write_detail(wb.create_sheet(name), d)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return len(diags)
