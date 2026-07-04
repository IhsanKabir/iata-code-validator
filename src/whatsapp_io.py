"""WhatsApp blast data layer — phone normalization, dedup, row build, image
validation, speed presets.

Pure and browser-free so it is fully unit-testable. The GUI reads a data sheet
(e.g. an FFP export), picks the phone column, and this module turns it into
de-duplicated `WhatsAppRow`s with a rendered message + optional shared image.
Blank/invalid numbers are bucketed, never silently dropped, never sent.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from .mailer_io import render_template

# ---- phone normalization ---------------------------------------------------

_NON_DIGIT_RE = re.compile(r"\D+")
# A WhatsApp-usable international number is country-code + a real national
# number. E.164 caps at 15; we floor at 10 (shortest realistic mobile + cc,
# e.g. Singapore +65 XXXXXXXX = 10) so fragments like a 5-digit stub — even
# after a country code is prepended — are rejected rather than dialled.
_MIN_INTL_DIGITS = 10
_MAX_INTL_DIGITS = 15


def normalize_phone(raw: object, *, default_cc: str = "880") -> str | None:
    """Return a bare international number (digits only, no '+') or None. GLOBAL.

    Any country works. The rules, in order:
      * `+…` or `00…`  -> already international; used as-is (message anyone).
      * leading `0`     -> local trunk number; the `0` becomes `default_cc`.
      * bare, 11+ digits-> assumed to already include a country code; used as-is.
      * bare, <=10 digits-> a short national number; `default_cc` is prepended.

    `default_cc` is only a fallback for clearly-local numbers; it never rewrites
    a number that already carries its own country code. So a sheet mixing
    +91…, +1…, 880… and local 01… all resolve correctly. Set `default_cc` to
    the country your *local* numbers belong to.
    """
    s = ("" if raw is None else str(raw)).strip()
    explicit_intl = s.startswith("+") or s.startswith("00")
    digits = _NON_DIGIT_RE.sub("", s)
    if not digits:
        return None
    if digits.startswith("00"):                 # 00<cc>… international prefix
        digits = digits[2:]
        explicit_intl = True
    cc = default_cc.lstrip("0")
    if explicit_intl:
        pass                                    # trust the given country code
    elif digits.startswith("0"):
        digits = cc + digits[1:]                # local trunk 0 -> country code
    elif len(digits) <= 10:
        digits = cc + digits                    # short national number
    # else: 11+ bare digits already include a country code -> keep as-is
    if not (_MIN_INTL_DIGITS <= len(digits) <= _MAX_INTL_DIGITS):
        return None
    return digits


# ---- rows ------------------------------------------------------------------

@dataclass(frozen=True)
class WhatsAppRecipient:
    """One unique, reachable recipient parsed from the sheet."""

    row_index: int          # 1-based data row of the FIRST occurrence
    phone: str              # bare international, e.g. 8801812377362
    fields: dict[str, str]  # every column value, for {placeholder} templating


@dataclass(frozen=True)
class WhatsAppReadResult:
    source: str
    phone_column: str
    rows: tuple[WhatsAppRecipient, ...]
    unreachable_rows: tuple[int, ...]      # 1-based rows with blank/invalid #
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class WhatsAppRow:
    """A composed, ready-to-send message."""

    phone: str
    text: str
    image_path: str | None = None
    fields: dict[str, str] = field(default_factory=dict)


def _norm(v: object) -> str:
    return "" if v is None else str(v).strip()


def read_whatsapp_rows(
    path: str | Path,
    *,
    phone_column: str,
    sheet_name: str | None = None,
    default_cc: str = "880",
) -> WhatsAppReadResult:
    """Parse the sheet into unique reachable recipients.

    Raises ValueError if the phone column is missing (fail fast at the
    boundary). Duplicate numbers keep the FIRST row's fields. Blank/invalid
    numbers go to `unreachable_rows` + a warning."""
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
        it = ws.iter_rows(values_only=True)
        try:
            headers = [_norm(c) for c in next(it)]
        except StopIteration:
            return WhatsAppReadResult(str(path), phone_column, (), (),
                                      ("Sheet is empty.",))
        wanted = phone_column.strip().lower()
        try:
            phone_i = [h.lower() for h in headers].index(wanted)
        except ValueError:
            raise ValueError(
                f"Phone column {phone_column!r} not found. "
                f"Headers: {', '.join(h for h in headers if h)}") from None

        seen: set[str] = set()
        rows: list[WhatsAppRecipient] = []
        unreachable: list[int] = []
        data_row = 0
        for raw in it:
            if not any(_norm(v) for v in raw):
                continue
            data_row += 1
            cell = raw[phone_i] if phone_i < len(raw) else None
            num = normalize_phone(cell, default_cc=default_cc)
            if num is None:
                unreachable.append(data_row)
                continue
            if num in seen:
                continue                        # dedup by number, first wins
            seen.add(num)
            fields: dict[str, str] = {}
            for i, label in enumerate(headers):
                if not label:
                    continue
                val = _norm(raw[i]) if i < len(raw) else ""
                fields[label] = val
                fields[label.lower()] = val
            fields.setdefault("phone", num)
            rows.append(WhatsAppRecipient(row_index=data_row, phone=num, fields=fields))
    finally:
        wb.close()

    warnings: list[str] = []
    if unreachable:
        warnings.append(
            f"{len(unreachable)} row(s) had a blank/invalid mobile number — "
            "excluded and never messaged.")
    if not rows:
        warnings.append("No reachable numbers found in that column.")
    return WhatsAppReadResult(
        source=str(path), phone_column=phone_column, rows=tuple(rows),
        unreachable_rows=tuple(unreachable), warnings=tuple(warnings))


def build_whatsapp_rows(
    result: WhatsAppReadResult,
    message_template: str,
    *,
    image_path: str | None,
) -> list[WhatsAppRow]:
    """Render the message per recipient (shared image applies to every one)."""
    out: list[WhatsAppRow] = []
    for r in result.rows:
        text, _missing = render_template(message_template, r.fields)
        out.append(WhatsAppRow(phone=r.phone, text=text,
                               image_path=image_path or None, fields=r.fields))
    return out


# ---- image validation ------------------------------------------------------

# WhatsApp accepts these inline as photos; caps media at ~16 MB.
_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
_IMAGE_MAGIC = (b"\xff\xd8\xff", b"\x89PNG\r\n\x1a\n", b"GIF8", b"RIFF")
_MAX_IMAGE_BYTES = 16 * 1024 * 1024


def validate_image(path: str | None) -> tuple[bool, str]:
    """(ok, error). None/'' path is a text-only run and is OK."""
    if not path:
        return True, ""
    p = Path(path)
    if not p.is_file():
        return False, f"Image not found: {path}"
    if p.suffix.lower() not in _IMAGE_EXTS:
        return False, (f"'{p.name}' is not an image "
                       "(use .jpg/.png/.webp/.gif).")
    size = p.stat().st_size
    if size > _MAX_IMAGE_BYTES:
        return False, (f"'{p.name}' is {size / 1024 / 1024:.1f} MB — "
                       "over WhatsApp's 16 MB photo limit.")
    try:
        head = p.read_bytes()[:8]
    except OSError as exc:
        return False, f"Can't read image: {exc}"
    if not any(head.startswith(m) for m in _IMAGE_MAGIC):
        return False, f"'{p.name}' doesn't look like a real image file."
    return True, ""


# ---- speed / risk presets --------------------------------------------------

@dataclass(frozen=True)
class SpeedPreset:
    """A pacing profile. Delays are randomized in [min,max] between messages;
    daily_cap bounds one session. Higher speed = higher ban risk."""

    label: str
    min_delay_s: float
    max_delay_s: float
    daily_cap: int
    risk_level: str          # "low" | "moderate" | "high"
    warning: str


SPEED_PRESETS: dict[str, SpeedPreset] = {
    "Safe": SpeedPreset(
        "Safe", 15.0, 30.0, 50, "low",
        "Slowest and safest. Best for cold lists (members who never messaged "
        "you first)."),
    "Balanced": SpeedPreset(
        "Balanced", 8.0, 15.0, 150, "moderate",
        "Moderate pace and volume. Watch for delivery failures and stop if you "
        "see them climb."),
    "Fast": SpeedPreset(
        "Fast", 3.0, 6.0, 400, "high",
        "HIGH RISK. Fast bulk sending to cold numbers is the most common cause "
        "of a WhatsApp number ban. Use only for small, warm lists."),
}

DEFAULT_PRESET = "Safe"
