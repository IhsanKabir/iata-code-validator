"""Read the Bulk Mailer mapping Excel, resolve attachments, render bodies.

Mapping sheet (one row per recipient), header names matched case-
insensitively, only `Email` + `File` are required:

  Email | Name | File | CC | BCC

`File` is resolved against a user-picked attachments folder. A row can
carry several files separated by `;` or `|` for multi-attachment sends.

The body is a plain-text template with `{column}` placeholders filled
from that row (e.g. `{name}`). Missing placeholders are left as-is and
reported, never crash the run.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

log = logging.getLogger(__name__)


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_FILE_SPLIT_RE = re.compile(r"[;|]")


@dataclass(frozen=True)
class MailRow:
    """One resolved recipient ready to compose.

    `issues` is non-empty when something is wrong (bad email, missing
    file). The GUI shows these in the preview and blocks/flags the row
    so a broken row never silently sends a blank attachment.
    """

    row_index: int                       # 1-based data row (excludes header)
    email: str
    name: str
    attachments: tuple[Path, ...]
    cc: str
    bcc: str
    fields: dict[str, str]               # every column value, for templating
    issues: tuple[str, ...] = ()

    @property
    def is_valid(self) -> bool:
        return not self.issues


def _norm(s: object) -> str:
    return "" if s is None else str(s).strip()


def _column_index_map(header: list) -> dict[str, int]:
    """Map lowercased column name → index."""
    out: dict[str, int] = {}
    for i, cell in enumerate(header):
        key = _norm(cell).lower()
        if key:
            out[key] = i
    return out


def read_mapping(
    path: str | Path,
    attachments_dir: str | Path,
    *,
    sheet_name: str | None = None,
) -> tuple[list[MailRow], list[str]]:
    """Parse the mapping workbook into validated MailRows.

    Returns (rows, header_warnings). Every row is returned — invalid ones
    carry `issues` so the GUI can show them red rather than dropping them
    silently.
    """
    path = Path(path)
    attach_dir = Path(attachments_dir)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        wb.close()
        return [], ["Mapping sheet is empty."]

    cols = _column_index_map(header)
    warnings: list[str] = []
    if "email" not in cols:
        warnings.append("No 'Email' column found — it is required.")
    if "file" not in cols:
        warnings.append("No 'File' column found — it is required.")

    email_i = cols.get("email")
    file_i = cols.get("file")
    name_i = cols.get("name")
    cc_i = cols.get("cc")
    bcc_i = cols.get("bcc")
    header_labels = [_norm(h) for h in header]

    def cell(row: tuple, idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        return _norm(row[idx])

    out: list[MailRow] = []
    data_row = 0
    for raw in rows_iter:
        # Skip fully-blank rows (trailing empties in the sheet).
        if not any(_norm(v) for v in raw):
            continue
        data_row += 1
        email = cell(raw, email_i)
        name = cell(raw, name_i)
        cc = cell(raw, cc_i)
        bcc = cell(raw, bcc_i)
        file_spec = cell(raw, file_i)

        issues: list[str] = []
        if not email:
            issues.append("missing email")
        elif not _EMAIL_RE.match(email):
            issues.append(f"invalid email '{email}'")

        attachments: list[Path] = []
        if not file_spec:
            issues.append("missing file")
        else:
            for piece in _FILE_SPLIT_RE.split(file_spec):
                fname = piece.strip()
                if not fname:
                    continue
                p = Path(fname)
                if not p.is_absolute():
                    p = attach_dir / fname
                if not p.is_file():
                    issues.append(f"file not found: {fname}")
                else:
                    attachments.append(p)

        # Build the per-row field dict for templating (every column).
        fields: dict[str, str] = {}
        for label, idx in zip(header_labels, range(len(header_labels))):
            if not label:
                continue
            fields[label] = cell(raw, idx)
            fields[label.lower()] = cell(raw, idx)
        # Convenience alias used by most templates.
        fields.setdefault("name", name)

        out.append(MailRow(
            row_index=data_row,
            email=email,
            name=name,
            attachments=tuple(attachments),
            cc=cc,
            bcc=bcc,
            fields=fields,
            issues=tuple(issues),
        ))
    wb.close()
    return out, warnings


_PLACEHOLDER_RE = re.compile(r"\{([^{}]+)\}")


def render_template(template: str, fields: dict[str, str]) -> tuple[str, list[str]]:
    """Fill `{placeholder}` tokens from `fields` (case-insensitive).

    Unknown placeholders are left verbatim and reported, so a typo in the
    template never blanks out or crashes a message.
    """
    missing: list[str] = []

    def repl(m: re.Match) -> str:
        key = m.group(1).strip()
        if key in fields:
            return fields[key]
        if key.lower() in fields:
            return fields[key.lower()]
        missing.append(key)
        return m.group(0)

    return _PLACEHOLDER_RE.sub(repl, template), missing
