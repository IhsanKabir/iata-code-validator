"""Shared WhatsApp-blast UI, mixed into both apps' mailer surface.

Self-contained: its own input pickers, speed presets, QR-login flow, preview
grid, progress, worker, and message dispatch — so the combined console and the
standalone mailer get the identical feature with no per-app duplication. The
host class must provide: `self.root` (Tk), `self._section(parent, title)`,
`self._post(kind, payload)` (enqueue a (kind,payload) for the Tk poll loop),
and route messages whose kind starts with "wa_" to `self._wa_handle_msg`.

Every step carries the ban-risk disclaimer (first-select modal, always-visible
note, escalating preset warnings, and a run-confirmation restating the risk).
"""

from __future__ import annotations

import logging
import threading
import time
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook

from . import config, whatsapp_client as wac
from .whatsapp_io import (
    DEFAULT_PRESET,
    SPEED_PRESETS,
    build_whatsapp_rows,
    read_whatsapp_rows,
    validate_image,
)

log = logging.getLogger(__name__)

WA_MSG_LOGIN = "wa_login"        # payload: status str (from whatsapp_client)
WA_MSG_PROGRESS = "wa_progress"  # (i, total, phone, status)
WA_MSG_DONE = "wa_done"          # payload: counts dict
WA_MSG_ERROR = "wa_error"        # payload: str
WA_MSG_LOG = "wa_log"            # payload: str

_DISCLAIMER = (
    "⚠ WhatsApp does NOT allow automation. Sending bulk messages this way can "
    "get your WhatsApp number BANNED — especially fast/large blasts to people "
    "who never messaged you first. You send from YOUR number, at your own risk. "
    "Go slow, keep batches small, and only message your own opted-in contacts."
)


class WhatsAppMixin:
    # -- state / profile ----------------------------------------------------

    def _wa_init_state(self) -> None:
        self.wa_input = tk.StringVar()
        self.wa_sheet = tk.StringVar()
        self.wa_column = tk.StringVar()
        self.wa_country = tk.StringVar(value="880")   # default cc for LOCAL #s
        self.wa_image = tk.StringVar()
        self.wa_preset = tk.StringVar(value=DEFAULT_PRESET)
        self.wa_min_delay = tk.DoubleVar(value=SPEED_PRESETS[DEFAULT_PRESET].min_delay_s)
        self.wa_max_delay = tk.DoubleVar(value=SPEED_PRESETS[DEFAULT_PRESET].max_delay_s)
        self.wa_cap = tk.IntVar(value=SPEED_PRESETS[DEFAULT_PRESET].daily_cap)
        self._wa_rows: list = []
        self._wa_session = None
        self._wa_worker: threading.Thread | None = None
        self._wa_stop = threading.Event()
        self._wa_disclaimer_ack = False

    def _wa_profile_dir(self) -> Path:
        return config.APP_DIR / "whatsapp_profile"

    # -- section build ------------------------------------------------------

    def _build_whatsapp_section(self, parent: tk.Widget) -> None:
        if not hasattr(self, "wa_input"):
            self._wa_init_state()
        body = self._section(parent, "WhatsApp Blast  ·  free · sends from YOUR number")
        warn = ttk.Label(body, text=_DISCLAIMER, style="Hint.TLabel",
                         justify="left", wraplength=900, foreground=self._COLOR_DANGER)
        warn.grid(row=0, column=0, columnspan=4, sticky="w", padx=2, pady=(0, 6))

        self._wa_row(body, 1, "Data Excel:", self.wa_input, self._wa_pick_input)
        # sheet + phone column + country code
        pick = ttk.Frame(body)
        self.wa_sheet_cb = ttk.Combobox(pick, textvariable=self.wa_sheet,
                                        state="readonly", width=22)
        self.wa_sheet_cb.pack(side="left")
        self.wa_sheet_cb.bind("<<ComboboxSelected>>", lambda _e: self._wa_fill_columns())
        ttk.Label(pick, text="  Phone column:").pack(side="left")
        self.wa_column_cb = ttk.Combobox(pick, textvariable=self.wa_column,
                                         state="readonly", width=22)
        self.wa_column_cb.pack(side="left", padx=(4, 0))
        ttk.Label(pick, text="  Country code (local #s):").pack(side="left")
        ttk.Entry(pick, textvariable=self.wa_country, width=6).pack(side="left", padx=(4, 0))
        self._wa_grid_row(body, 2, "Sheet:", pick)

        # optional image
        img = ttk.Frame(body)
        ttk.Entry(img, textvariable=self.wa_image, width=48).pack(side="left")
        ttk.Button(img, text="Browse…", command=self._wa_pick_image).pack(side="left", padx=(4, 0))
        ttk.Button(img, text="Clear", command=lambda: self.wa_image.set("")).pack(
            side="left", padx=(4, 0))
        self._wa_grid_row(body, 3, "Image (optional):", img)

        # message
        ttk.Label(body, text="Message (placeholders: {FIRSTNAME}, {phone}, any column):",
                  style="Hint.TLabel").grid(row=4, column=0, columnspan=4, sticky="w",
                                            padx=2, pady=(6, 2))
        self.wa_message = tk.Text(body, height=5, wrap="word")
        self.wa_message.grid(row=5, column=0, columnspan=4, sticky="ew", padx=2)
        self.wa_message.insert("1.0", "Dear {FIRSTNAME},\n\n")
        body.columnconfigure(0, weight=1)

        # speed preset
        sp = ttk.Frame(body)
        for name in ("Safe", "Balanced", "Fast", "Custom"):
            ttk.Radiobutton(sp, text=name, value=name, variable=self.wa_preset,
                            command=self._wa_on_preset).pack(side="left", padx=(0, 10))
        ttk.Label(sp, text=" delay").pack(side="left")
        ttk.Spinbox(sp, from_=1, to=120, increment=1, width=4,
                    textvariable=self.wa_min_delay).pack(side="left", padx=(4, 0))
        ttk.Label(sp, text="–").pack(side="left")
        ttk.Spinbox(sp, from_=1, to=180, increment=1, width=4,
                    textvariable=self.wa_max_delay).pack(side="left")
        ttk.Label(sp, text="s   cap/day").pack(side="left")
        ttk.Spinbox(sp, from_=1, to=2000, increment=10, width=6,
                    textvariable=self.wa_cap).pack(side="left", padx=(4, 0))
        self._wa_grid_row(body, 6, "Pace:", sp)
        self.wa_preset_warn = ttk.Label(body, text="", style="Hint.TLabel",
                                        wraplength=900, justify="left")
        self.wa_preset_warn.grid(row=7, column=0, columnspan=4, sticky="w", padx=2)
        self._wa_on_preset()

        # login + actions
        act = ttk.Frame(body)
        act.grid(row=8, column=0, columnspan=4, sticky="w", padx=2, pady=(8, 2))
        ttk.Button(act, text="Open WhatsApp & sign in (scan QR)",
                   command=self._wa_login).pack(side="left")
        self.wa_login_label = ttk.Label(act, text="Not signed in", style="Hint.TLabel")
        self.wa_login_label.pack(side="left", padx=(8, 0))
        ttk.Button(act, text="Preview", command=self._wa_preview).pack(side="left", padx=(12, 0))
        self.btn_wa_send = ttk.Button(act, text="Send via WhatsApp", style="Primary.TButton",
                                      command=self._wa_run, state="disabled")
        self.btn_wa_send.pack(side="left", padx=(8, 0))
        self.btn_wa_stop = ttk.Button(act, text="Stop", style="Danger.TButton",
                                      command=lambda: self._wa_stop.set(), state="disabled")
        self.btn_wa_stop.pack(side="left", padx=(8, 0))

        # preview grid + progress
        cols = ("row", "phone", "message", "status")
        self.wa_tree = ttk.Treeview(body, columns=cols, show="headings", height=8)
        for cid, txt, w in (("row", "#", 36), ("phone", "Phone", 150),
                            ("message", "Message", 460), ("status", "Status", 130)):
            self.wa_tree.heading(cid, text=txt)
            self.wa_tree.column(cid, width=w, anchor="w")
        self.wa_tree.tag_configure("bad", background=self._COLOR_ROW_BAD)
        self.wa_tree.tag_configure("ok", background=self._COLOR_ROW_GOOD)
        self.wa_tree.grid(row=9, column=0, columnspan=4, sticky="ew", padx=2, pady=(6, 2))
        if hasattr(self, "_register_result_tree"):
            self._register_result_tree(self.wa_tree)
        self.wa_progress = ttk.Progressbar(body, mode="determinate", maximum=1)
        self.wa_progress.grid(row=10, column=0, columnspan=4, sticky="ew", padx=2, pady=(2, 2))
        self.wa_status = ttk.Label(body, text="Idle.", style="Hint.TLabel")
        self.wa_status.grid(row=11, column=0, columnspan=4, sticky="w", padx=2)

    def _wa_row(self, parent, r, label, var, browse) -> None:
        ttk.Label(parent, text=label, width=18, anchor="w").grid(
            row=r, column=0, sticky="w", padx=(2, 8), pady=2)
        ttk.Entry(parent, textvariable=var).grid(row=r, column=1, columnspan=2,
                                                 sticky="ew", padx=(0, 4), pady=2)
        ttk.Button(parent, text="Browse…", command=browse).grid(
            row=r, column=3, padx=(4, 2), pady=2)

    def _wa_grid_row(self, parent, r, label, widget) -> None:
        ttk.Label(parent, text=label, width=18, anchor="w").grid(
            row=r, column=0, sticky="w", padx=(2, 8), pady=2)
        widget.grid(row=r, column=1, columnspan=3, sticky="w", padx=(0, 4), pady=2)

    # -- pickers ------------------------------------------------------------

    def _wa_pick_input(self) -> None:
        path = filedialog.askopenfilename(
            title="Pick the data Excel (rows carry a phone column)",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")])
        if not path:
            return
        self.wa_input.set(path)
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = list(wb.sheetnames)
            wb.close()
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("WhatsApp", f"Couldn't read workbook: {exc}")
            return
        self.wa_sheet_cb.configure(values=sheets)
        if sheets:
            self.wa_sheet.set(sheets[0])
        self._wa_fill_columns()

    def _wa_fill_columns(self) -> None:
        path = self.wa_input.get().strip()
        if not path or not Path(path).is_file():
            return
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[self.wa_sheet.get()] if self.wa_sheet.get() in wb.sheetnames else wb.active
            headers = [str(c).strip() for c in next(ws.iter_rows(max_row=1, values_only=True))
                       if c is not None and str(c).strip()]
            wb.close()
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("WhatsApp", f"Couldn't read headers: {exc}")
            return
        self.wa_column_cb.configure(values=headers)
        guess = next((h for h in headers if "mobile" in h.lower()), "") or \
            next((h for h in headers if "phone" in h.lower()), "")
        self.wa_column.set(guess or (headers[0] if headers else ""))

    def _wa_pick_image(self) -> None:
        path = filedialog.askopenfilename(
            title="Pick one image to send to everyone",
            filetypes=[("Images", "*.jpg *.jpeg *.png *.webp *.gif"), ("All files", "*.*")])
        if not path:
            return
        ok, msg = validate_image(path)
        if not ok:
            messagebox.showerror("WhatsApp", msg)
            return
        self.wa_image.set(path)

    def _wa_on_preset(self) -> None:
        name = self.wa_preset.get()
        p = SPEED_PRESETS.get(name)
        if p:                                       # Custom leaves the spinboxes alone
            self.wa_min_delay.set(p.min_delay_s)
            self.wa_max_delay.set(p.max_delay_s)
            self.wa_cap.set(p.daily_cap)
            colors = {"low": self._COLOR_SUCCESS, "moderate": self._COLOR_WARNING,
                      "high": self._COLOR_DANGER}
            self.wa_preset_warn.configure(text=p.warning,
                                          foreground=colors.get(p.risk_level, self._COLOR_MUTED))
        else:
            self.wa_preset_warn.configure(
                text="Custom pace — you own the risk. Slower + smaller = safer.",
                foreground=self._COLOR_WARNING)

    # -- disclaimer ---------------------------------------------------------

    def _wa_require_ack(self) -> bool:
        if self._wa_disclaimer_ack:
            return True
        ok = messagebox.askokcancel(
            "WhatsApp — read this first",
            _DISCLAIMER + "\n\nClick OK only if you understand your number can be "
            "banned and you accept that risk.")
        self._wa_disclaimer_ack = bool(ok)
        return self._wa_disclaimer_ack

    # -- login --------------------------------------------------------------

    def _wa_login(self) -> None:
        if not self._wa_require_ack():
            return
        if self._wa_worker and self._wa_worker.is_alive():
            messagebox.showinfo("WhatsApp", "A WhatsApp task is already running.")
            return
        self.wa_login_label.configure(text="Opening WhatsApp Web…")

        def worker() -> None:
            try:
                if self._wa_session is None:
                    self._wa_session = wac.WhatsAppSession(
                        self._wa_profile_dir(),
                        on_log=lambda m: self._post(WA_MSG_LOG, m))
                    self._wa_session.start()
                status = self._wa_session.login_status()
                if status != wac.LOGGED_IN:
                    self._post(WA_MSG_LOGIN, "needs_qr")
                    self._wa_session.wait_until_logged_in(timeout_s=180)
                self._post(WA_MSG_LOGIN, self._wa_session.login_status())
            except wac.WhatsAppBrowserError as exc:
                self._post(WA_MSG_ERROR, str(exc))
            except Exception as exc:  # noqa: BLE001
                log.exception("whatsapp login failed")
                self._post(WA_MSG_ERROR, f"{type(exc).__name__}: {exc}")

        self._wa_worker = threading.Thread(target=worker, daemon=True)
        self._wa_worker.start()

    # -- preview ------------------------------------------------------------

    def _wa_preview(self) -> None:
        path = self.wa_input.get().strip()
        col = self.wa_column.get().strip()
        if not path or not Path(path).is_file():
            messagebox.showerror("WhatsApp", "Pick the data Excel first.")
            return
        if not col:
            messagebox.showerror("WhatsApp", "Pick the phone column.")
            return
        img = self.wa_image.get().strip() or None
        ok, msg = validate_image(img)
        if not ok:
            messagebox.showerror("WhatsApp", msg)
            return
        try:
            res = read_whatsapp_rows(path, phone_column=col,
                                     sheet_name=self.wa_sheet.get().strip() or None,
                                     default_cc=self.wa_country.get().strip() or "880")
        except ValueError as exc:
            messagebox.showerror("WhatsApp", str(exc))
            return
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("WhatsApp", f"{type(exc).__name__}: {exc}")
            return
        template = self.wa_message.get("1.0", "end").rstrip("\n")
        rows = build_whatsapp_rows(res, template, image_path=img)
        self._wa_rows = rows
        for c in self.wa_tree.get_children():
            self.wa_tree.delete(c)
        for i, r in enumerate(rows, 1):
            preview = (r.text[:80] + "…") if len(r.text) > 80 else r.text
            self.wa_tree.insert("", "end", values=(i, "+" + r.phone, preview, "READY"),
                                tags=("ok",))
        if hasattr(self, "_stripe_tree"):
            self._stripe_tree(self.wa_tree)
        warn = ("  ·  " + "; ".join(res.warnings)) if res.warnings else ""
        self.wa_status.configure(text=f"{len(rows)} reachable recipient(s){warn}")
        self.btn_wa_send.configure(state="normal" if rows else "disabled")

    # -- run ----------------------------------------------------------------

    def _wa_run(self) -> None:
        if not self._wa_rows:
            messagebox.showerror("WhatsApp", "Preview first.")
            return
        if not self._wa_require_ack():
            return
        if self._wa_session is None or self._wa_session.login_status() != wac.LOGGED_IN:
            messagebox.showerror("WhatsApp",
                                 "Sign in first: click 'Open WhatsApp & sign in', scan the QR.")
            return
        if self._wa_worker and self._wa_worker.is_alive():
            messagebox.showinfo("WhatsApp", "A WhatsApp task is already running.")
            return
        try:
            lo, hi = float(self.wa_min_delay.get()), float(self.wa_max_delay.get())
            cap = int(self.wa_cap.get())
        except (tk.TclError, ValueError):
            messagebox.showerror("WhatsApp", "Delay and cap must be numbers.")
            return
        if hi < lo:
            lo, hi = hi, lo
        n = min(len(self._wa_rows), cap)
        img_note = " with an image" if self.wa_image.get().strip() else ""
        if not messagebox.askyesno(
                "WhatsApp — confirm send",
                f"About to message {n} recipient(s){img_note} from YOUR WhatsApp number, "
                f"{lo:.0f}–{hi:.0f}s apart.\n\n{_DISCLAIMER}\n\nContinue?"):
            return

        rows = list(self._wa_rows[:cap])
        session = self._wa_session
        self._wa_stop.clear()
        self.btn_wa_send.configure(state="disabled")
        self.btn_wa_stop.configure(state="normal")
        self.wa_progress.configure(value=0, maximum=len(rows))
        iid_by_i = {i: iid for i, iid in enumerate(self.wa_tree.get_children(), 1)}

        def worker() -> None:
            import random
            counts = {wac.SENT: 0, wac.NOT_ON_WHATSAPP: 0, wac.FAILED: 0, wac.SKIPPED: 0}
            log_key = getattr(self, "_mail_log", None)
            campaign = "WA:" + Path(self.wa_input.get()).name
            subj = "whatsapp"
            try:
                for i, r in enumerate(rows, 1):
                    if self._wa_stop.is_set():
                        break
                    if log_key and log_key.already_sent(campaign, r.phone, subj):
                        counts[wac.SKIPPED] += 1
                        self._post(WA_MSG_PROGRESS, (i, len(rows), r.phone, "SKIPPED"))
                        continue
                    res = session.send(r.phone, r.text, r.image_path)
                    counts[res.status] = counts.get(res.status, 0) + 1
                    if res.status == wac.SENT and log_key:
                        log_key.record(campaign, r.phone, subj, "SENT")
                    if res.status == wac.LOGGED_OUT:
                        self._post(WA_MSG_PROGRESS, (i, len(rows), r.phone, "LOGGED_OUT"))
                        self._post(WA_MSG_ERROR,
                                   "WhatsApp logged out mid-run — re-scan the QR and resume "
                                   "(already-sent are skipped).")
                        break
                    self._post(WA_MSG_PROGRESS,
                               (i, len(rows), r.phone,
                                res.status + (f": {res.error}" if res.error else "")))
                    if i < len(rows) and not self._wa_stop.is_set():
                        time.sleep(random.uniform(lo, hi))
                self._post(WA_MSG_DONE, counts)
            except Exception as exc:  # noqa: BLE001
                log.exception("whatsapp run failed")
                self._post(WA_MSG_ERROR, f"{type(exc).__name__}: {exc}")

        self._wa_iid_by_i = iid_by_i
        self._wa_worker = threading.Thread(target=worker, daemon=True)
        self._wa_worker.start()

    # -- message dispatch ---------------------------------------------------

    def _wa_handle_msg(self, kind: str, payload) -> bool:
        if kind == WA_MSG_LOGIN:
            texts = {wac.LOGGED_IN: "Signed in ✓", "needs_qr": "Scan the QR in the browser…",
                     wac.LOADING: "Loading…"}
            self.wa_login_label.configure(text=texts.get(payload, str(payload)))
            return True
        if kind == WA_MSG_LOG:
            self.wa_status.configure(text=str(payload))
            return True
        if kind == WA_MSG_PROGRESS:
            i, total, phone, status = payload
            self.wa_progress.configure(value=i, maximum=total)
            self.wa_status.configure(text=f"{i}/{total}  +{phone}: {status}")
            iid = getattr(self, "_wa_iid_by_i", {}).get(i)
            if iid:
                vals = list(self.wa_tree.item(iid, "values"))
                vals[3] = status
                bad = not status.startswith(wac.SENT)
                self.wa_tree.item(iid, values=vals, tags=("bad",) if bad else ("ok",))
            return True
        if kind == WA_MSG_DONE:
            self.btn_wa_send.configure(state="normal")
            self.btn_wa_stop.configure(state="disabled")
            c = payload
            self.wa_status.configure(
                text=(f"Done — sent {c.get(wac.SENT, 0)}, "
                      f"no-WhatsApp {c.get(wac.NOT_ON_WHATSAPP, 0)}, "
                      f"failed {c.get(wac.FAILED, 0)}, skipped {c.get(wac.SKIPPED, 0)}."))
            messagebox.showinfo("WhatsApp", self.wa_status.cget("text"))
            return True
        if kind == WA_MSG_ERROR:
            self.btn_wa_send.configure(state="normal")
            self.btn_wa_stop.configure(state="disabled")
            self.wa_login_label.configure(text="Not signed in")
            messagebox.showerror("WhatsApp", str(payload))
            return True
        return False
