"""Read input IATA numbers from Excel and write timestamped result file."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .config import (
    BD_OUTPUT_COLUMNS_FULL,
    BD_OUTPUT_COLUMNS_LOOKUP,
    OEP_OUTPUT_COLUMNS_CATEGORY_SUMMARY,
    OEP_OUTPUT_COLUMNS_COUNTRY_RAW,
    OEP_OUTPUT_COLUMNS_COUNTRY_SUMMARY,
    OEP_OUTPUT_COLUMNS_DIVISION_RAW,
    OEP_OUTPUT_COLUMNS_DIVISION_SUMMARY,
    OEP_OUTPUT_COLUMNS_GENDER_SUMMARY,
    OUTPUT_COLUMNS,
    ZENITH_FLIGHT_OUTPUT_COLUMNS,
    ZENITH_OUTPUT_COLUMNS,
)
from .parser import LookupResult


def list_sheet_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def list_columns(path: Path, sheet: str) -> list[str]:
    """Return header row values for the chosen sheet (row 1)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        cols: list[str] = []
        for idx, val in enumerate(first_row):
            label = str(val).strip() if val is not None else ""
            if not label:
                label = f"Column {get_column_letter(idx + 1)}"
            cols.append(label)
        return cols
    finally:
        wb.close()


def read_iata_numbers(
    path: Path,
    sheet: str,
    column_index: int,
    start_row: int = 2,
    end_row: int | None = None,
) -> list[tuple[int, str]]:
    """Read IATA numbers from `column_index` (0-based) on `sheet`.

    Returns list of (excel_row_number, iata_number) tuples, skipping blanks.
    Numbers are normalized to digit-only strings.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        rows: list[tuple[int, str]] = []
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True),
            start=start_row,
        ):
            if column_index >= len(row):
                continue
            value = row[column_index]
            if value is None:
                continue
            normalized = _normalize(value)
            if not normalized:
                continue
            rows.append((row_idx, normalized))
        return rows
    finally:
        wb.close()


def _normalize(value: object) -> str:
    """Strip whitespace; for floats from Excel that look like ints, drop the decimal."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value).strip()
    # Some users paste codes with spaces or hyphens
    s = s.replace(" ", "").replace("-", "")
    return s


def build_output_path(folder: Path, stem: str = "iata_results") -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return folder / f"{stem}_{timestamp}.xlsx"


class ResultWriter:
    """Streaming writer — flushes to disk every 10 rows.

    Always use as a context manager so the final flush runs even if the
    surrounding code raises:

        with ResultWriter(path) as writer:
            writer.append(result)
    """

    def __init__(self, path: Path) -> None:
        self.path = path
        self._wb = Workbook()
        ws = self._wb.active
        ws.title = "Results"
        ws.append(OUTPUT_COLUMNS)
        self._row_count = 0
        self._save()

    def __enter__(self) -> "ResultWriter":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def append(self, result: LookupResult) -> None:
        ws = self._wb.active
        ws.append([
            result.iata_number,
            result.trading_name,
            result.country,
            result.accredited,
            result.status,
            result.checked_at,
            result.notes,
        ])
        self._row_count += 1
        # Save every 10 rows to balance durability vs speed
        if self._row_count % 10 == 0:
            self._save()

    def append_many(self, results: Iterable[LookupResult]) -> None:
        for r in results:
            self.append(r)

    def close(self) -> None:
        self._save()

    def _save(self) -> None:
        self._wb.save(self.path)


# ---------------------------------------------------------------------------
# BD Travel Agency exports
# ---------------------------------------------------------------------------


def write_bd_full_list(path: Path, agencies: list) -> None:
    """Dump the full BD agency list to a fresh Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BD Agencies"
    ws.append(BD_OUTPUT_COLUMNS_FULL)
    for a in agencies:
        ws.append([
            a.agency_name,
            a.license_no,
            a.email,
            a.mobile,
            a.website,
            a.address,
            a.license_expired_date,
            a.status,
        ])
    wb.save(path)


def write_bd_lookup_results(path: Path, results: list) -> None:
    """Write BD lookup results (one row per input)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lookup Results"
    ws.append(BD_OUTPUT_COLUMNS_LOOKUP)
    for r in results:
        a = r.agency
        ws.append([
            r.searched_input,
            r.match_method,
            r.matched_field,
            r.match_score,
            a.agency_name if a else "",
            a.license_no if a else "",
            a.email if a else "",
            a.mobile if a else "",
            a.website if a else "",
            a.address if a else "",
            a.license_expired_date if a else "",
            a.status if a else "",
            r.other_matches,
        ])
    wb.save(path)


def build_bd_output_path(folder: Path, kind: str = "lookup") -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = "bd_agency_full_list" if kind == "full" else "bd_agency_lookup"
    return folder / f"{stem}_{timestamp}.xlsx"


# ---------------------------------------------------------------------------
# OEP overseas-movement exports
# ---------------------------------------------------------------------------


def _safe_share(n: int, total: int) -> float:
    return round(100.0 * n / total, 2) if total > 0 else 0.0


def write_oep_country_report(
    path: Path,
    *,
    date_from: str,
    date_to: str,
    summary: list,
    raw_rows: list,
) -> None:
    """Write the "Top destinations" report (summary + raw)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "By Country"
    ws.append([f"Range: {date_from} → {date_to}"])
    ws.append([])
    ws.append(OEP_OUTPUT_COLUMNS_COUNTRY_SUMMARY)
    grand = sum(c.total_employee for c in summary) or 1
    for rank, c in enumerate(summary, start=1):
        ws.append([
            rank,
            c.country_name,
            c.total_employee,
            c.category_count,
            _safe_share(c.total_employee, grand),
        ])

    raw_ws = wb.create_sheet("Raw")
    raw_ws.append(OEP_OUTPUT_COLUMNS_COUNTRY_RAW)
    for r in raw_rows:
        raw_ws.append([
            r.country_id,
            r.country_name,
            r.category_name,
            r.total_employee,
        ])
    wb.save(path)


def write_oep_division_report(
    path: Path,
    *,
    date_from: str,
    date_to: str,
    summary: list,
    raw_rows: list,
) -> None:
    """Write the "Top source districts" report (summary + raw)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "By Division"
    ws.append([f"Range: {date_from} → {date_to}"])
    ws.append([])
    ws.append(OEP_OUTPUT_COLUMNS_DIVISION_SUMMARY)
    grand = sum(d.total_employee for d in summary) or 1
    for rank, d in enumerate(summary, start=1):
        ws.append([
            rank,
            d.division,
            d.total_employee,
            d.district_count,
            _safe_share(d.total_employee, grand),
        ])

    raw_ws = wb.create_sheet("Raw")
    raw_ws.append(OEP_OUTPUT_COLUMNS_DIVISION_RAW)
    for r in raw_rows:
        raw_ws.append([r.division, r.district, r.total_employee])
    wb.save(path)


def write_oep_category_report(
    path: Path,
    *,
    date_from: str,
    date_to: str,
    summary: list,
    raw_rows: list,
) -> None:
    """Write the "Top job categories" report (summary + raw)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "By Category"
    ws.append([f"Range: {date_from} → {date_to}"])
    ws.append([])
    ws.append(OEP_OUTPUT_COLUMNS_CATEGORY_SUMMARY)
    grand = sum(c.total_employee for c in summary) or 1
    for rank, c in enumerate(summary, start=1):
        ws.append([
            rank,
            c.category_name,
            c.total_employee,
            c.country_count,
            _safe_share(c.total_employee, grand),
        ])

    raw_ws = wb.create_sheet("Raw")
    raw_ws.append(OEP_OUTPUT_COLUMNS_COUNTRY_RAW)
    for r in raw_rows:
        raw_ws.append([
            r.country_id,
            r.country_name,
            r.category_name,
            r.total_employee,
        ])
    wb.save(path)


def write_oep_gender_report(
    path: Path,
    *,
    date_from: str,
    date_to: str,
    summary: list,
) -> None:
    """Write the "Gender breakdown by destination" report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "By Gender"
    ws.append([f"Range: {date_from} → {date_to}"])
    ws.append([])
    ws.append(OEP_OUTPUT_COLUMNS_GENDER_SUMMARY)
    for rank, g in enumerate(summary, start=1):
        ws.append([
            rank,
            g.country_name,
            g.male,
            g.female,
            g.other,
            g.total,
            _safe_share(g.female, g.total),
        ])
    wb.save(path)


def write_oep_timeseries_report(
    path: Path,
    *,
    date_from: str,
    date_to: str,
    months: list,
    series: dict,
) -> None:
    """Write the monthly time-series report.

    Sheet "Time Series" — wide format with months as columns; useful for
    pasting into pivot charts. Sheet "Raw" — long format (month, country,
    total) for downstream analysis.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Series"
    ws.append([f"Range: {date_from} → {date_to}"])
    ws.append([])
    header = ["Country"] + list(months) + ["Total"]
    ws.append(header)
    # Sort countries by their grand total, descending.
    sorted_countries = sorted(
        series.items(), key=lambda kv: sum(kv[1]), reverse=True,
    )
    for name, values in sorted_countries:
        ws.append([name, *values, sum(values)])

    raw_ws = wb.create_sheet("Raw")
    raw_ws.append(["Year-Month", "Country", "Total Employees"])
    for name, values in sorted_countries:
        for ym, total in zip(months, values):
            raw_ws.append([ym, name, total])
    wb.save(path)


def write_oep_pivot_report(
    path: Path,
    *,
    date_from: str,
    date_to: str,
    divisions: list,
    countries: list,
    table: dict,
) -> None:
    """Write the country × division pivot report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot"
    ws.append([f"Range: {date_from} → {date_to}"])
    ws.append([])
    ws.append(["Division"] + list(countries) + ["Total"])
    grand_total = 0
    for div in divisions:
        row = [div]
        row_total = 0
        for country in countries:
            v = table.get((div, country), 0)
            row.append(v)
            row_total += v
        row.append(row_total)
        grand_total += row_total
        ws.append(row)
    totals = ["Total"]
    for country in countries:
        totals.append(sum(table.get((d, country), 0) for d in divisions))
    totals.append(grand_total)
    ws.append(totals)

    raw_ws = wb.create_sheet("Raw")
    raw_ws.append(["Country", "Division", "Total Employees"])
    for div in divisions:
        for country in countries:
            v = table.get((div, country), 0)
            if v:
                raw_ws.append([country, div, v])
    wb.save(path)


def write_oep_full_report(
    path: Path,
    *,
    date_from: str,
    date_to: str,
    gender_id: str,
    country_summary: list,
    category_summary: list,
    division_summary: list,
    gender_summary: list,
    raw_country: list,
    raw_division: list,
    months: list,
    series: dict,
    divisions: list,
    pivot_countries: list,
    table: dict,
    country_labels: list,
    cdt_months: list | None = None,
    cdt_pairs: list | None = None,
    cdt_table: dict | None = None,
) -> None:
    """Mega-report — one workbook with Cover + 6 data sheets + 2 raw sheets.

    Order of sheets (left to right) is what someone scanning the workbook
    will read first. Cover comes first so the file opens to the headline
    numbers; raw sheets are last for power users.
    """
    wb = Workbook()

    # ----- Cover -----
    cover = wb.active
    cover.title = "Cover"
    grand_total = sum(c.total_employee for c in country_summary)
    gender_label = {"1": "Male only", "2": "Female only", "3": "Other only"}.get(
        gender_id, "All genders"
    )

    cover.append(["BD Overseas Workforce Movement — Full Report"])
    cover.append([])
    cover.append(["Date range", f"{date_from}  →  {date_to}"])
    cover.append(["Gender filter", gender_label])
    cover.append(["Countries in time-series / pivot",
                  ", ".join(country_labels) or "—"])
    cover.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    cover.append([])
    cover.append(["Headline numbers"])
    cover.append(["Total workers cleared", grand_total])
    cover.append(["Unique destinations", len(country_summary)])
    cover.append(["Unique job categories", len(category_summary)])
    cover.append(["BD divisions covered", len(division_summary)])
    cover.append([])
    cover.append(["Top 5 destinations"])
    cover.append(["Rank", "Country", "Workers", "Share %"])
    for rank, c in enumerate(country_summary[:5], start=1):
        cover.append([
            rank, c.country_name, c.total_employee,
            _safe_share(c.total_employee, grand_total),
        ])

    # ----- By Country -----
    ws = wb.create_sheet("By Country")
    ws.append(OEP_OUTPUT_COLUMNS_COUNTRY_SUMMARY)
    for rank, c in enumerate(country_summary, start=1):
        ws.append([
            rank, c.country_name, c.total_employee, c.category_count,
            _safe_share(c.total_employee, grand_total),
        ])

    # ----- By Division -----
    ws = wb.create_sheet("By Division")
    div_grand = sum(d.total_employee for d in division_summary) or 1
    ws.append(OEP_OUTPUT_COLUMNS_DIVISION_SUMMARY)
    for rank, d in enumerate(division_summary, start=1):
        ws.append([
            rank, d.division, d.total_employee, d.district_count,
            _safe_share(d.total_employee, div_grand),
        ])

    # ----- By Category -----
    ws = wb.create_sheet("By Category")
    cat_grand = sum(c.total_employee for c in category_summary) or 1
    ws.append(OEP_OUTPUT_COLUMNS_CATEGORY_SUMMARY)
    for rank, c in enumerate(category_summary, start=1):
        ws.append([
            rank, c.category_name, c.total_employee, c.country_count,
            _safe_share(c.total_employee, cat_grand),
        ])

    # ----- By Gender -----
    ws = wb.create_sheet("By Gender")
    ws.append(OEP_OUTPUT_COLUMNS_GENDER_SUMMARY)
    for rank, g in enumerate(gender_summary, start=1):
        ws.append([
            rank, g.country_name, g.male, g.female, g.other, g.total,
            _safe_share(g.female, g.total),
        ])

    # ----- Time Series -----
    ws = wb.create_sheet("Time Series")
    ws.append(["Country", *months, "Total"])
    sorted_ts = sorted(series.items(), key=lambda kv: sum(kv[1]), reverse=True)
    for name, values in sorted_ts:
        ws.append([name, *values, sum(values)])

    # ----- Pivot -----
    ws = wb.create_sheet("Country x Division")
    ws.append(["Division", *pivot_countries, "Total"])
    pivot_grand = 0
    for div in divisions:
        row = [div]
        row_total = 0
        for country in pivot_countries:
            v = table.get((div, country), 0)
            row.append(v)
            row_total += v
        row.append(row_total)
        pivot_grand += row_total
        ws.append(row)
    totals = ["Total"]
    for country in pivot_countries:
        totals.append(sum(table.get((d, country), 0) for d in divisions))
    totals.append(pivot_grand)
    ws.append(totals)

    # ----- Country × Division × Month (flat) -----
    if cdt_pairs and cdt_months and cdt_table is not None:
        ws = wb.create_sheet("Country×Division×Month")
        ws.append(["Country", "Division", *cdt_months, "Total"])
        col_totals = [0] * len(cdt_months)
        cdt_grand = 0
        for country, division in cdt_pairs:
            row = [country, division]
            row_total = 0
            for ci, ym in enumerate(cdt_months):
                v = cdt_table.get((country, division, ym), 0)
                row.append(v)
                row_total += v
                col_totals[ci] += v
            row.append(row_total)
            cdt_grand += row_total
            ws.append(row)
        ws.append(["Total", "", *col_totals, cdt_grand])

    # ----- Raw sheets -----
    raw_c = wb.create_sheet("Raw Country×Category")
    raw_c.append(OEP_OUTPUT_COLUMNS_COUNTRY_RAW)
    for r in raw_country:
        raw_c.append([r.country_id, r.country_name, r.category_name, r.total_employee])

    raw_d = wb.create_sheet("Raw Division")
    raw_d.append(OEP_OUTPUT_COLUMNS_DIVISION_RAW)
    for r in raw_division:
        raw_d.append([r.division, r.district, r.total_employee])

    wb.save(path)


def write_oep_country_district_timeseries(
    path: Path,
    *,
    date_from: str,
    date_to: str,
    months: list,
    triples: list,
    table: dict,
) -> None:
    """One flat sheet at district granularity.

    Schema:
      A: Country | B: Division | C: District | D..N: month columns | last: Total
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Country×District×Month"
    ws.append([f"Range: {date_from} → {date_to}"])
    ws.append([])
    ws.append(["Country", "Division", "District", *months, "Total"])
    col_totals = [0] * len(months)
    grand_total = 0
    for country, division, district in triples:
        row_values = [country, division, district]
        row_total = 0
        for ci, ym in enumerate(months):
            v = table.get((country, division, district, ym), 0)
            row_values.append(v)
            row_total += v
            col_totals[ci] += v
        row_values.append(row_total)
        grand_total += row_total
        ws.append(row_values)
    ws.append(["Total", "", "", *col_totals, grand_total])
    wb.save(path)


def write_oep_country_division_timeseries(
    path: Path,
    *,
    date_from: str,
    date_to: str,
    months: list,
    pairs: list,
    table: dict,
) -> None:
    """One flat sheet: each row is (Country, Division) × monthly columns.

    Schema:
      A: Country | B: Division | C..N: month columns | last: Total

    Tidy/long data — pasteable straight into a PivotTable for any cut
    (heatmaps, % change, sum-over-time, etc.).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Country×Division×Month"
    ws.append([f"Range: {date_from} → {date_to}"])
    ws.append([])
    ws.append(["Country", "Division", *months, "Total"])
    grand_total = 0
    # Column totals row collected as we go
    col_totals = [0] * len(months)
    for country, division in pairs:
        row_values = [country, division]
        row_total = 0
        for ci, ym in enumerate(months):
            v = table.get((country, division, ym), 0)
            row_values.append(v)
            row_total += v
            col_totals[ci] += v
        row_values.append(row_total)
        grand_total += row_total
        ws.append(row_values)
    # Totals row
    totals_row = ["Total", "", *col_totals, grand_total]
    ws.append(totals_row)
    wb.save(path)


def build_oep_output_path(folder: Path, kind: str) -> Path:
    """`kind` ∈ {country, division, category, gender, timeseries, pivot}."""
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return folder / f"oep_{kind}_{timestamp}.xlsx"


# ---------------------------------------------------------------------------
# Traffic tab IO (one writer for every source — rows are unified TrafficRows)
# ---------------------------------------------------------------------------


def write_traffic_report(
    path: Path,
    *,
    source_label: str,
    date_from: str,
    date_to: str,
    view: str,
    rows: list,
) -> None:
    """Write a Traffic report: a per-`view` Summary sheet + a full Raw sheet.

    `rows` are unified TrafficRow objects from any source. The Summary sheet
    aggregates by country / airport / route / period; Share % is computed
    within each metric so passengers and seats never blend.
    """
    from .config import TRAFFIC_OUTPUT_COLUMNS_RAW
    from .traffic_client import (
        aggregate_by_airport,
        aggregate_by_country,
        aggregate_by_period,
        aggregate_by_route,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([f"Source: {source_label}"])
    ws.append([f"Range: {date_from or 'all'} → {date_to or 'all'}    View: {view}"])
    ws.append([])

    def _metric_totals(agg) -> dict:
        tot: dict[str, float] = {}
        for t in agg:
            tot[t.metric] = tot.get(t.metric, 0.0) + t.value
        return tot

    if view == "route":
        agg = aggregate_by_route(rows)
        tot = _metric_totals(agg)
        ws.append(["Rank", "Origin", "Destination", "Metric", "Unit", "Value", "Share % (metric)"])
        for i, t in enumerate(agg, start=1):
            ws.append([i, t.origin, t.destination, t.metric, t.unit, t.value,
                       _safe_share(t.value, tot.get(t.metric, 0))])
    elif view == "airport":
        agg = aggregate_by_airport(rows)
        tot = _metric_totals(agg)
        ws.append(["Rank", "Airport", "Metric", "Unit", "Value", "Share % (metric)"])
        for i, t in enumerate(agg, start=1):
            ws.append([i, t.airport, t.metric, t.unit, t.value,
                       _safe_share(t.value, tot.get(t.metric, 0))])
    elif view == "period":
        agg = aggregate_by_period(rows)
        ws.append(["Period", "Metric", "Unit", "Value"])
        for t in agg:
            ws.append([t.period, t.metric, t.unit, t.value])
    else:  # "country" (default)
        agg = aggregate_by_country(rows)
        tot = _metric_totals(agg)
        ws.append(["Rank", "Country", "Metric", "Unit", "Value", "Share % (metric)"])
        for i, t in enumerate(agg, start=1):
            ws.append([i, t.country, t.metric, t.unit, t.value,
                       _safe_share(t.value, tot.get(t.metric, 0))])

    raw_ws = wb.create_sheet("Raw")
    raw_ws.append(TRAFFIC_OUTPUT_COLUMNS_RAW)
    for r in rows:
        raw_ws.append([
            r.source_label, r.country, r.airport, r.origin, r.destination, r.carrier,
            r.period, r.period_granularity, r.direction, r.flight_type,
            r.metric, r.unit, r.value, r.nationality, r.gender, r.raw_label,
        ])
    wb.save(path)


def build_traffic_output_path(folder: Path, kind: str) -> Path:
    """`kind` ∈ {country, airport, route, period}."""
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return folder / f"traffic_{kind}_{timestamp}.xlsx"


# ---------------------------------------------------------------------------
# Zenith Customer Lookup IO
# ---------------------------------------------------------------------------


def read_zenith_ids(
    path: Path,
    sheet: str,
    column_header: str,
    *,
    start_row: int = 2,
    end_row: int | None = None,
) -> list[str]:
    """Read customer IDs from a column in an Excel file.

    Same shape as `read_iata_numbers` but doesn't reject non-numeric
    or short-length values — Zenith customer IDs are 8 digits but we
    accept anything non-blank.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        header_row = next(
            ws.iter_rows(min_row=1, max_row=1, values_only=True), ()
        )
        try:
            col_idx = list(header_row).index(column_header) + 1
        except ValueError:
            raise ValueError(
                f"Column {column_header!r} not found in sheet {sheet!r}"
            ) from None

        ids: list[str] = []
        for row in ws.iter_rows(
            min_row=start_row, max_row=end_row,
            min_col=col_idx, max_col=col_idx,
            values_only=True,
        ):
            cell = row[0]
            text = _normalize(cell)
            if text:
                ids.append(text)
        return ids
    finally:
        wb.close()


def write_zenith_results(path: Path, results: Iterable) -> None:
    """Write a flat per-customer Excel from cached LookupResult rows.

    `results` is an iterable of zenith_client.LookupResult — typically
    yielded by ZenithCache.iter_all(). Fields are written in
    ZENITH_OUTPUT_COLUMNS order.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Zenith Customers"
    ws.append(ZENITH_OUTPUT_COLUMNS)
    for r in results:
        rec = r.record
        ws.append([
            r.customer_id,
            r.status,
            rec.title if rec else "",
            rec.first_name if rec else "",
            rec.middle_name if rec else "",
            rec.last_name if rec else "",
            rec.date_of_birth if rec else "",
            rec.email if rec else "",
            rec.home_phone if rec else "",
            rec.home_phone_international if rec else "",
            rec.mobile_phone if rec else "",
            rec.mobile_phone_international if rec else "",
            rec.office_phone if rec else "",
            rec.nationality if rec else "",
            rec.language if rec else "",
            rec.spoken_language if rec else "",
            rec.address if rec else "",
            rec.city if rec else "",
            rec.postal_code if rec else "",
            rec.country if rec else "",
            rec.registration_date if rec else "",
            r.error,
            r.checked_at,
        ])
    # Auto-size the Customer ID column for readability.
    ws.column_dimensions[get_column_letter(1)].width = 14
    wb.save(path)


def build_zenith_output_path(folder: Path) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return folder / f"zenith_customers_{timestamp}.xlsx"


def write_zenith_flight_loads(path: Path, rows: Iterable) -> None:
    """Write flight-load rows to a single flat Excel sheet.

    `rows` is an iterable of zenith_client.FlightLoadRow.
    Schema is one row per (flight, leg, cabin).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Flight Loads"
    ws.append(ZENITH_FLIGHT_OUTPUT_COLUMNS)
    for r in rows:
        ws.append([
            r.flight_number,
            r.day_of_week,
            r.flight_date,
            r.departure_time,
            r.aircraft,
            r.registration,
            r.total_tickets_issued,
            r.leg_route,
            r.leg_origin,
            r.leg_destination,
            r.leg_local_time_range,
            r.leg_cabin,
            r.tickets_issued,
            r.tickets_wl,
            r.seats_confirmed,
            r.seats_options,
            r.seats_wl,
            r.seats_available,
            r.inventory_status,
            r.comments,
        ])
    # Reasonable widths for the most-used columns.
    widths = {1: 12, 3: 12, 8: 12, 11: 22, 18: 16, 19: 22}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    wb.save(path)


# ---------------------------------------------------------------------------
# Ordered (cross-tab) Flight Load report
#
# Reverse-engineered from the user's template. One growing workbook holds
# every operating date as a 16-column block side-by-side. Each Flight
# Loads run appends a new block on the right; existing dates + formatting
# stay untouched.
#
# Schema per date (16 cols, in order):
#   Capacity | STD | Aircraft | Starting fare | Flown Seats | Infant
#   | Sold | No-Show | Gate No-show | Immigration Offload
#   | Immigration not faced | Unsold | Value of unsold ticket
#   | Utilization | Load Factor | Note
#
# The auto-fill scope (from FlightLoadRow data) is intentionally narrow:
# Capacity, STD, Aircraft, Flown Seats, Load Factor. Everything else
# (Starting fare, no-shows, infant, etc.) stays blank for manual entry
# — that's how the template is built and what users will expect.
# ---------------------------------------------------------------------------

ORDERED_REPORT_SHEET = "Ordered Data"
ORDERED_REPORT_HEADER_FILL_RGB = "FF1F4E78"   # dark blue, white bold text
ORDERED_REPORT_DATE_BLOCK_WIDTH = 16
ORDERED_REPORT_PER_DATE_HEADERS = (
    "Capacity",
    "STD",
    "Aircraft",
    "Starting fare\nas on {date_short}",
    "Flown Seats",
    "Infant",
    "Sold",
    "No-Show",
    "Gate No-show",
    "Immigration Offload",
    "Immigration not faced",
    "Unsold",
    "Value of unsold ticket",
    "Utilization",
    "Load Factor",
    "Note",
)
# 1-based column offsets within a date block (matches the schema above).
ORDERED_REPORT_COL = {
    "capacity": 1,
    "std": 2,
    "aircraft": 3,
    "starting_fare": 4,
    "flown_seats": 5,
    "infant": 6,
    "sold": 7,
    "no_show": 8,
    "gate_no_show": 9,
    "imm_offload": 10,
    "imm_not_faced": 11,
    "unsold": 12,
    "value_unsold": 13,
    "utilization": 14,
    "load_factor": 15,
    "note": 16,
}


def _seats_available_to_numbers(text: str) -> tuple[int | None, float | None]:
    """Parse '13/410 97%' → (capacity=410, load_pct=0.97).

    Tolerates over-bookings ('-5/152 103%' → (152, 1.03)) and missing %
    suffix. Returns (None, None) on anything we can't read.
    """
    if not text:
        return None, None
    import re as _re
    m = _re.match(
        r"\s*(-?\d+)\s*/\s*(\d+)\s+(-?\d+(?:\.\d+)?)\s*%\s*$", text,
    )
    if not m:
        return None, None
    try:
        return int(m.group(2)), float(m.group(3)) / 100.0
    except ValueError:
        return None, None


def _bracketed_int(text: str) -> int | None:
    """Parse '[152]' → 152. Returns None on anything unparseable."""
    if not text:
        return None
    import re as _re
    m = _re.search(r"-?\d+", text)
    if not m:
        return None
    try:
        return int(m.group(0))
    except ValueError:
        return None


def _aircraft_label(aircraft: str, registration: str) -> str:
    """Format the aircraft cell the way the existing template does it.

    The template uses 'ATR-72-600', 'Boeing 737-800' etc — just the type
    string, no registration. Some legacy rows also include the dash
    style. We keep the aircraft type as-is.
    """
    a = (aircraft or "").strip()
    return a


def _format_report_date_header(date_dmy: str) -> str:
    """'24/05/2026' → '24/05/2026 (Sunday)'. Falls back to bare date if unparseable."""
    try:
        d = datetime.strptime(date_dmy, "%d/%m/%Y")
    except ValueError:
        return date_dmy
    return f"{date_dmy} ({d.strftime('%A')})"


def _short_date(date_dmy: str) -> str:
    """'24/05/2026' → '24/05/26' for header label inside 'Starting fare as on …'."""
    try:
        d = datetime.strptime(date_dmy, "%d/%m/%Y")
    except ValueError:
        return date_dmy
    return d.strftime("%d/%m/%y")


def _read_existing_dates(ws) -> dict[str, int]:
    """Map 'DD/MM/YYYY' → starting column of its date block, from row 1.

    The user's template encodes dates as 'DD/MM/YYYY (Weekday)' in
    merged row-1 cells; we strip the weekday and use the date as key.
    """
    import re as _re
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if not v:
            continue
        m = _re.match(r"\s*(\d{1,2}/\d{1,2}/\d{4})", str(v))
        if m:
            out[m.group(1).zfill(10)] = c   # zfill keeps 2-digit dates stable
            # Also accept the literal value as found
            out[m.group(1)] = c
    return out


def _read_existing_flight_rows(ws) -> dict[tuple[str, str], int]:
    """Map (flight_number, leg_route) → row index from the existing sheet."""
    out: dict[tuple[str, str], int] = {}
    for r in range(3, ws.max_row + 1):
        flight = ws.cell(row=r, column=1).value
        leg = ws.cell(row=r, column=2).value
        if not flight:
            continue
        key = (str(flight).strip(), str(leg or "").strip())
        out.setdefault(key, r)
    return out


def _apply_ordered_header_styles(ws, start_col: int, end_col: int) -> None:
    """Style the date-header cell (row 1) and column-label row (row 2)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill(
        fill_type="solid",
        fgColor=ORDERED_REPORT_HEADER_FILL_RGB,
    )
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Row 1 — merged date title
    head = ws.cell(row=1, column=start_col)
    head.fill = fill
    head.font = white_bold
    head.alignment = center
    if end_col > start_col:
        ws.merge_cells(
            start_row=1, end_row=1, start_column=start_col, end_column=end_col,
        )
    # Row 2 — column labels in the same dark band
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = fill
        cell.font = white_bold
        cell.alignment = center


def _apply_load_factor_cf(ws, col_letter: str, last_row: int) -> None:
    """Recreate the template's traffic-light CF on a Load Factor column.

    Rules (matching the user's template):
      load >= 0.9       — dark green
      0.8 <= load < 0.9 — light green (#92D050)
      0.6 <= load < 0.8 — yellow (#FFFF00)
      load < 0.6        — red (#FF0000)
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    rng = f"{col_letter}3:{col_letter}{max(last_row, 3)}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(
            operator="greaterThanOrEqual", formula=["0.9"],
            fill=PatternFill(fill_type="solid", fgColor="FF00B050"),
        ),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(
            operator="between", formula=["0.8", "0.899999"],
            fill=PatternFill(fill_type="solid", fgColor="FF92D050"),
        ),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(
            operator="between", formula=["0.6", "0.799999"],
            fill=PatternFill(fill_type="solid", fgColor="FFFFFF00"),
        ),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(
            operator="lessThan", formula=["0.6"],
            fill=PatternFill(fill_type="solid", fgColor="FFFF0000"),
        ),
    )


def append_flight_loads_to_ordered_report(
    report_path: Path,
    flight_load_rows: Iterable,
    *,
    create_if_missing: bool = True,
) -> dict[str, int]:
    """Append a Flight Loads run to a cross-tab Ordered Report workbook.

    For each unique flight_date in `flight_load_rows` we append a new
    16-col block on the right of the existing "Ordered Data" sheet,
    fill in the values we can derive (Capacity, STD, Aircraft, Flown
    Seats, Load Factor), and re-apply the traffic-light CF on the new
    Load Factor column. Dates already present in the file are skipped
    so re-runs are idempotent.

    Returns a small summary dict so the GUI can show what changed.
    """
    rows = list(flight_load_rows)
    if not rows:
        return {"dates_added": 0, "flights_added": 0, "flights_updated": 0}

    report_path = Path(report_path)
    if report_path.exists():
        wb = load_workbook(report_path)
        ws = (
            wb[ORDERED_REPORT_SHEET]
            if ORDERED_REPORT_SHEET in wb.sheetnames
            else wb.active
        )
    elif create_if_missing:
        wb = Workbook()
        ws = wb.active
        ws.title = ORDERED_REPORT_SHEET
        # Freeze the three identity columns + the date header rows.
        ws.cell(row=2, column=1, value="Flight No")
        ws.cell(row=2, column=2, value="Leg/Sector")
        ws.cell(row=2, column=3, value="Departure Time")
        for c in (1, 2, 3):
            cell = ws.cell(row=2, column=c)
            from openpyxl.styles import Font, PatternFill, Alignment
            cell.fill = PatternFill(
                fill_type="solid", fgColor=ORDERED_REPORT_HEADER_FILL_RGB,
            )
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "D3"
    else:
        raise FileNotFoundError(report_path)

    existing_dates = _read_existing_dates(ws)
    flight_row_map = _read_existing_flight_rows(ws)

    # Group rows by date so each date is a single block.
    by_date: dict[str, list] = {}
    for r in rows:
        by_date.setdefault(r.flight_date, []).append(r)

    dates_added = 0
    flights_added = 0
    flights_updated = 0
    next_col = ws.max_column + 1
    if next_col < 4:
        next_col = 4

    dates_refreshed = 0
    cells_updated = 0

    for date in sorted(
        by_date.keys(),
        key=lambda d: datetime.strptime(d, "%d/%m/%Y"),
    ):
        is_new = date not in existing_dates
        if is_new:
            block_start = next_col
            block_end = block_start + ORDERED_REPORT_DATE_BLOCK_WIDTH - 1
            # Row 1 — date title
            ws.cell(
                row=1, column=block_start,
                value=_format_report_date_header(date),
            )
            # Row 2 — per-date column labels
            short = _short_date(date)
            for i, label_tmpl in enumerate(ORDERED_REPORT_PER_DATE_HEADERS):
                label = label_tmpl.format(date_short=short)
                ws.cell(row=2, column=block_start + i, value=label)
            _apply_ordered_header_styles(ws, block_start, block_end)
        else:
            # Refresh existing date — flights' counts may have moved
            # as bookings came in. We only touch the 5 auto-fill cells
            # (Capacity, STD, Aircraft, Flown Seats, Load Factor) so
            # the user's manually-entered Starting Fare / No-Show /
            # Notes stay untouched.
            block_start = existing_dates[date]

        # Per-flight cells
        for fr in by_date[date]:
            key = (
                str(fr.flight_number).strip(),
                str(fr.leg_route or "").strip(),
            )
            row_idx = flight_row_map.get(key)
            if row_idx is None:
                row_idx = ws.max_row + 1 if ws.max_row >= 3 else 3
                ws.cell(row=row_idx, column=1, value=key[0])
                ws.cell(row=row_idx, column=2, value=key[1])
                ws.cell(row=row_idx, column=3, value=fr.departure_time or "")
                flight_row_map[key] = row_idx
                flights_added += 1
            else:
                flights_updated += 1

            capacity, load_pct = _seats_available_to_numbers(fr.seats_available)
            flown = _bracketed_int(fr.seats_confirmed)
            aircraft_label = _aircraft_label(fr.aircraft, fr.registration)

            if capacity is not None:
                ws.cell(
                    row=row_idx,
                    column=block_start + ORDERED_REPORT_COL["capacity"] - 1,
                    value=capacity,
                )
                cells_updated += 1
            if fr.departure_time:
                ws.cell(
                    row=row_idx,
                    column=block_start + ORDERED_REPORT_COL["std"] - 1,
                    value=fr.departure_time,
                )
                cells_updated += 1
            if aircraft_label:
                ws.cell(
                    row=row_idx,
                    column=block_start + ORDERED_REPORT_COL["aircraft"] - 1,
                    value=aircraft_label,
                )
                cells_updated += 1
            if flown is not None:
                ws.cell(
                    row=row_idx,
                    column=block_start + ORDERED_REPORT_COL["flown_seats"] - 1,
                    value=flown,
                )
                cells_updated += 1
            if load_pct is not None:
                cell = ws.cell(
                    row=row_idx,
                    column=block_start + ORDERED_REPORT_COL["load_factor"] - 1,
                    value=round(load_pct, 4),
                )
                cell.number_format = "0.00%"
                cells_updated += 1

        # CF only needs to be (re-)applied for new blocks; the user's
        # template already carries CF rules on existing Load Factor
        # columns so re-running won't pile up duplicates.
        if is_new:
            lf_col_letter = get_column_letter(
                block_start + ORDERED_REPORT_COL["load_factor"] - 1,
            )
            _apply_load_factor_cf(ws, lf_col_letter, ws.max_row)
            existing_dates[date] = block_start
            block_end = block_start + ORDERED_REPORT_DATE_BLOCK_WIDTH - 1
            next_col = block_end + 1
            dates_added += 1
        else:
            dates_refreshed += 1

    wb.save(report_path)
    return {
        "dates_added": dates_added,
        "dates_refreshed": dates_refreshed,
        "flights_added": flights_added,
        "flights_updated": flights_updated,
        "cells_updated": cells_updated,
        "total_dates": len(existing_dates),
    }


def build_zenith_flight_output_path(folder: Path) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return folder / f"zenith_flight_loads_{timestamp}.xlsx"


# ---------------------------------------------------------------------------
# Passenger manifest (per-leg drill-down) export
# ---------------------------------------------------------------------------

ZENITH_PAX_OUTPUT_COLUMNS = [
    "Flight", "Flight Date", "Flight Time", "Route", "Leg", "Direction",
    "Title", "Passenger Name", "Pax Type", "Gender",
    "Date of Birth", "Passport No.", "Weight (kg)",
    "Cabin", "PRBD", "Fare Basis", "Web Class",
    "Ticket Number", "Seat", "PNR", "GDS PNR", "Issuing Agency",
    "id_vol", "id_leg",
]


def write_passenger_manifest(path: Path, records: Iterable) -> None:
    """Write passenger-manifest rows to one flat sheet (one row per pax).

    `records` is an iterable of zenith_client.PassengerRecord. Contains
    PII (passport, DOB) by design — the file is confidential.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Passengers"
    ws.append(ZENITH_PAX_OUTPUT_COLUMNS)
    for r in records:
        ws.append([
            r.flight_number, r.flight_date, r.flight_time, r.route_desc,
            r.leg, r.direction,
            r.title, r.full_name, r.pax_type, r.gender,
            r.date_of_birth, r.passport_no, r.weight_kg,
            r.cabin_code, r.prbd, r.fare_basis, r.web_class,
            r.ticket_number, r.seat, r.pnr, r.gds_pnr, r.issuing_agency,
            r.id_vol, r.id_leg,
        ])
    widths = {1: 9, 2: 11, 4: 22, 8: 26, 11: 12, 12: 14, 16: 10, 20: 10, 22: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    wb.save(path)


def build_zenith_pax_output_path(folder: Path) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return folder / f"zenith_passengers_{timestamp}.xlsx"


# ---------------------------------------------------------------------------
# Zenith Flight History audit
# ---------------------------------------------------------------------------


def write_zenith_history_audit(path: Path, report) -> None:
    """Multi-sheet workbook for the Flight History Analyzer.

    Sheets (left → right, so the file opens to the most-skimmed view):
      Cover | Class Downgrades | Downgrade Leaders | G-Class Issuance
            | Agent Activity   | Revenue Mgmt      | Suspicious Activity
            | Raw Events

    The Raw Events sheet is last (largest) so the user lands on the
    summaries first.
    """
    wb = Workbook()

    # ----- Cover -----
    cover = wb.active
    cover.title = "Cover"
    cover.append(["Zenith Flight History Audit"])
    cover.append([])
    start, end = report.date_range
    cover.append(["Files parsed", report.file_count])
    cover.append(["Total events", report.event_count])
    cover.append([
        "Date range",
        f"{start.strftime('%Y-%m-%d') if start else '—'}"
        f"  →  {end.strftime('%Y-%m-%d') if end else '—'}",
    ])
    cover.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    cover.append([])
    cover.append(["Top agents (by event count)"])
    cover.append(["Rank", "Agent user ID", "Events"])
    for rank, (uid, count) in enumerate(report.top_agents, start=1):
        cover.append([rank, uid, count])
    cover.append([])
    cover.append(["Top RBD classes touched"])
    cover.append(["Rank", "Class", "Events"])
    for rank, (rbd, count) in enumerate(report.top_rbds, start=1):
        cover.append([rank, rbd, count])

    # ----- Class Downgrades -----
    ws = wb.create_sheet("Class Downgrades")
    ws.append([
        "Severity", "Steps", "PNR", "Customer", "Passenger",
        "Flight", "Flight Date",
        "Start Class", "End Class", "Trajectory",
        "Last Changed By", "Last Changed At",
    ])
    for t in report.class_trajectories:
        # Skip rows with no downgrade activity — keep the sheet focused.
        if t.total_downgrade_severity == 0:
            continue
        ws.append([
            t.total_downgrade_severity,
            t.downgrade_steps,
            t.pnr,
            t.customer_name,
            t.passenger,
            t.flight_number,
            t.flight_date,
            t.starting_class,
            t.ending_class,
            " → ".join(t.classes_seen),
            t.last_changed_by,
            t.last_changed_at.strftime("%Y-%m-%d %H:%M") if t.last_changed_at else "",
        ])

    # ----- Downgrade Leaders -----
    ws = wb.create_sheet("Downgrade Leaders")
    ws.append([
        "Rank", "Agent User ID", "Display Name", "Department",
        "Downgrade Events", "Total Severity", "Distinct PNRs",
    ])
    for rank, d in enumerate(report.downgrade_leaders, start=1):
        ws.append([
            rank, d.agent_user_id, d.agent_display_name, d.agent_department,
            d.downgrade_event_count, d.total_severity, d.distinct_pnrs,
        ])

    # ----- G-Class Issuance -----
    ws = wb.create_sheet("G-Class Issuance")
    ws.append([
        "Timestamp", "Agent User ID", "Display Name", "Department",
        "PNR", "Customer", "Passenger",
        "Flight", "Flight Date", "Event Type", "Ticket Number",
    ])
    for g in report.g_class_events:
        ws.append([
            g.timestamp.strftime("%Y-%m-%d %H:%M") if g.timestamp else "",
            g.agent_user_id, g.agent_display_name, g.agent_department,
            g.pnr, g.customer_name, g.passenger,
            g.flight_number, g.flight_date, g.event_type, g.ticket_number,
        ])

    # ----- Agent Activity -----
    ws = wb.create_sheet("Agent Activity")
    # Pre-compute the union of event types so columns are stable.
    all_types: list[str] = sorted({
        t for row in report.agent_activity for t in row.by_type
    })
    ws.append([
        "Rank", "Agent User ID", "Display Name", "Department",
        "Total Events", *all_types,
    ])
    for rank, row in enumerate(report.agent_activity, start=1):
        type_cells = [row.by_type.get(t, 0) for t in all_types]
        ws.append([
            rank, row.agent_user_id, row.agent_display_name,
            row.agent_department, row.total_events, *type_cells,
        ])

    # ----- Revenue Mgmt -----
    ws = wb.create_sheet("Revenue Mgmt")
    ws.append([
        "Timestamp", "Agent User ID", "Display Name",
        "Flight", "Flight Date", "Route",
        "Booking Class", "Seats Before", "Seats After", "Delta",
    ])
    for r in report.revenue_mgmt_changes:
        ws.append([
            r.timestamp.strftime("%Y-%m-%d %H:%M") if r.timestamp else "",
            r.agent_user_id, r.agent_display_name,
            r.flight_number, r.flight_date, r.route,
            r.booking_class, r.seats_before, r.seats_after, r.delta,
        ])

    # ----- Suspicious Activity -----
    ws = wb.create_sheet("Suspicious Activity")
    ws.append([
        "Severity", "Timestamp", "Agent User ID",
        "PNR", "Passenger", "Flight", "Event Type", "Reason",
    ])
    for f in report.suspicious_flags:
        ws.append([
            f.severity,
            f.timestamp.strftime("%Y-%m-%d %H:%M") if f.timestamp else "",
            f.agent_user_id,
            f.pnr, f.passenger, f.flight_number, f.event_type, f.reason,
        ])

    # ----- Downgrade Justification -----
    # Only created when a Flight Loads Excel was passed in. Lets the
    # auditor answer: "was the fare reduction logical given the load?"
    if report.downgrade_justifications:
        ws = wb.create_sheet("Downgrade Justification")
        ws.append([
            "Verdict", "Load %", "Capacity",
            "Timestamp", "Agent User ID", "Display Name",
            "PNR", "Passenger",
            "Flight", "Flight Date", "Route",
            "Old Class", "New Class", "Severity",
            "Inventory Status",
        ])
        for j in report.downgrade_justifications:
            ws.append([
                j.verdict,
                j.load_pct if j.load_pct is not None else "",
                j.seats_capacity if j.seats_capacity is not None else "",
                j.timestamp.strftime("%Y-%m-%d %H:%M") if j.timestamp else "",
                j.agent_user_id, j.agent_display_name,
                j.pnr, j.passenger,
                j.flight_number, j.flight_date, j.route,
                j.old_class, j.new_class, j.severity,
                j.inventory_status,
            ])

    # ----- PNR Routes (only when PNR enrichment was run) -----
    if report.pnr_routes:
        ws = wb.create_sheet("PNR Routes")
        ws.append([
            "PNR", "Customer", "Traveler Surname", "Phone",
            "PNR Status", "Pax", "Payment Method",
            "Booked Route", "Flown Route",
            "Segments", "Flown", "Refunded", "Voided", "Other",
            "Total Amount", "Currency",
            "Segments Detail",
        ])
        for r in report.pnr_routes:
            ws.append([
                r.pnr_code, r.customer_name, r.traveler_surname, r.phone,
                r.pnr_status, r.pax_count, r.payment_method,
                r.booked_route, r.flown_route,
                r.segment_count, r.flown_count, r.refunded_count,
                r.voided_count, r.other_status_count,
                r.total_amount, r.currency,
                r.segments_summary,
            ])

    # ----- Raw Events (always last; can be enormous) -----
    ws = wb.create_sheet("Raw Events")
    ws.append([
        "Source File", "Row", "Timestamp", "Agent User ID", "Department",
        "Event Type", "PNR", "Passenger", "Flight", "Flight Date",
        "RBD Class", "Old Status", "New Status",
        "Capacity Class", "Seats Before", "Seats After",
        "Ticket Number", "Description",
    ])
    for e in report.raw_events:
        ws.append([
            e.source_file, e.row_index,
            e.timestamp.strftime("%Y-%m-%d %H:%M") if e.timestamp else "",
            e.agent.user_id, e.agent.department,
            e.event_type, e.pnr, e.passenger,
            e.flight.flight_number, e.flight.flight_date,
            e.rbd_class, e.old_status, e.new_status,
            e.capacity_class, e.capacity_before, e.capacity_after,
            e.ticket_number, e.raw_description,
        ])

    wb.save(path)


def build_zenith_history_output_path(folder: Path) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return folder / f"zenith_history_audit_{timestamp}.xlsx"


# ---------------------------------------------------------------------------
# Zenith Bulk PNR Lookup
# ---------------------------------------------------------------------------


# Header layout for the standalone PNR-lookup Excel output. The
# segments column is a compact human-scannable summary; the per-segment
# detail rows are in a second sheet for power users.
ZENITH_PNR_BULK_COLUMNS = [
    "PNR", "Status", "Customer", "Traveler Surname", "Phone",
    "Payment Method", "Pax",
    "Booked Route", "Flown Route",
    "Segments", "Flown", "Refunded", "Voided",
    "Total Amount", "Currency",
    "Segments Detail", "Lookup Status", "Error",
]


def read_pnr_codes_from_excel(
    path: Path,
    *,
    sheet_name: str | None = None,
    column_name: str | None = None,
) -> list[str]:
    """Read a column of PNR codes from a user-provided Excel.

    If `column_name` is given we look it up by header; otherwise we take
    the first non-empty column. Empty rows and the header itself are
    skipped. PNR codes are upper-cased.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        wb.close()
        return []

    col_idx: int | None = None
    if column_name:
        for i, h in enumerate(header):
            if h is not None and str(h).strip().lower() == column_name.strip().lower():
                col_idx = i
                break
    if col_idx is None:
        col_idx = 0

    out: list[str] = []
    for row in rows:
        if col_idx >= len(row):
            continue
        val = row[col_idx]
        if val is None:
            continue
        code = str(val).strip().upper()
        if code:
            out.append(code)
    wb.close()
    return out


def write_zenith_pnr_bulk(
    path: Path,
    results: list,
    *,
    errors: dict[str, str] | None = None,
) -> None:
    """Write the bulk PNR lookup output workbook.

    `results` is a list of (pnr_code, PNRDetails or None) tuples in the
    user's input order. `errors` maps PNR → human-readable failure when
    the lookup didn't return details.

    Sheets:
      "PNR Lookup"      — one row per PNR (summary view)
      "Segments"        — one row per segment (PNR repeated; pivot-ready)
    """
    errors = errors or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "PNR Lookup"
    ws.append(ZENITH_PNR_BULK_COLUMNS)

    seg_ws = wb.create_sheet("Segments")
    seg_ws.append([
        "PNR", "Direction", "Leg", "RBD Class", "Fare Basis",
        "Coupon Status", "Departure", "Arrival", "Aircraft",
        "Price Ex-Tax", "Price All-Tax", "Ticket Number", "Passenger",
    ])

    for code, details in results:
        if details is None:
            ws.append([
                code, "", "", "", "", "", "",
                "", "",
                "", "", "", "",
                "", "",
                "",
                "NOT_FOUND" if code not in errors else "ERROR",
                errors.get(code, "PNR not resolved"),
            ])
            continue
        not_flown_keys = {"voided", "refunded", "cancelled", "canceled", "no show"}
        flown = sum(
            1 for s in details.segments
            if s.coupon_status and s.coupon_status.lower() not in not_flown_keys
        )
        refunded = sum(1 for s in details.segments if s.coupon_status.lower() == "refunded")
        voided = sum(1 for s in details.segments if s.coupon_status.lower() == "voided")
        seg_summary = " ; ".join(
            f"{s.leg_route or '?'}/{s.rbd_class or '?'}/"
            f"{s.coupon_status or '?'}/{s.price_ttc or '?'}"
            for s in details.segments
        )
        ws.append([
            details.pnr_code,
            details.pnr_status,
            details.customer_name,
            details.traveler_surname,
            details.phone,
            details.payment_method,
            details.pax_count,
            details.booked_route,
            details.flown_route,
            len(details.segments),
            flown,
            refunded,
            voided,
            details.total_amount,
            details.currency,
            seg_summary,
            "OK",
            "",
        ])
        for s in details.segments:
            seg_ws.append([
                details.pnr_code,
                s.leg_direction,
                s.leg_route,
                s.rbd_class,
                s.fare_basis,
                s.coupon_status,
                s.departure_text,
                s.arrival_text,
                s.aircraft,
                s.price_ht,
                s.price_ttc,
                s.ticket_number,
                s.passenger,
            ])

    # Column widths tuned for the headline columns.
    widths = {1: 10, 3: 22, 4: 18, 5: 16, 8: 18, 9: 18, 14: 14, 16: 40}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    wb.save(path)


def build_zenith_pnr_bulk_output_path(folder: Path) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return folder / f"zenith_pnr_lookup_{timestamp}.xlsx"
