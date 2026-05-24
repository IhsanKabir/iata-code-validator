"""Read a Flight Loads Excel and build a fast lookup by (flight, date, leg).

The Flight Loads sub-tab exports a workbook with one row per
(flight, leg, cabin). This module reads that workbook back in and
returns a dict the History Analyzer can query while enriching audit
rows with load-factor context.

Used by Phase 4 of the Flight History Analyzer:
  "Was the fare reduction justified by low load, or was it discounted
  on an already-full flight?"
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Verdict thresholds — tunable in one place if user feedback shifts.
# ---------------------------------------------------------------------------

# Above this load %, a class downgrade is suspect (flight near full).
HIGH_LOAD_THRESHOLD = 90.0

# Below this load %, a class downgrade is reasonable (flight under-booked).
LOW_LOAD_THRESHOLD = 70.0

VERDICT_QUESTIONABLE = "QUESTIONABLE"  # high load → discount looks unjustified
VERDICT_SITUATIONAL = "SITUATIONAL"    # mid range — managers' call
VERDICT_JUSTIFIED = "JUSTIFIED"        # low load → discount made sense
VERDICT_UNKNOWN = "UNKNOWN"            # no load data for this flight


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class LoadEntry:
    """Load-factor snapshot for one (flight, date, leg, cabin) cell."""

    flight_number: str
    flight_date: str          # 'DD/MM/YYYY'
    origin: str
    destination: str
    cabin: str
    seats_available: int      # negative when over-booked
    seats_capacity: int
    load_pct: float           # 0.0 - ~105.0 (rarely above 100 due to overbook)
    inventory_status: str
    raw_seats_available: str  # e.g. '13/410 97%' — kept for traceability


# ---------------------------------------------------------------------------
# Cell parsing
# ---------------------------------------------------------------------------


# Matches Zenith's per-leg "Seats Available" string. Examples:
#   '13/410 97%'
#   '-5/152 103%'
#   '0/152 100%'
_SEATS_AVAIL_RE = re.compile(
    r"^\s*(?P<avail>-?\d+)\s*/\s*(?P<cap>\d+)\s+(?P<pct>-?\d+(?:\.\d+)?)\s*%\s*$"
)


def parse_seats_available(text: str) -> tuple[int, int, float] | None:
    """Parse '13/410 97%' → (13, 410, 97.0). Returns None on failure."""
    if not text:
        return None
    m = _SEATS_AVAIL_RE.match(text)
    if not m:
        return None
    try:
        return int(m.group("avail")), int(m.group("cap")), float(m.group("pct"))
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------


# Column names we depend on. These mirror ZENITH_FLIGHT_OUTPUT_COLUMNS
# from config.py — kept in sync by hand because changing the writer's
# layout already means breaking downstream consumers.
_REQUIRED_COLUMNS = (
    "Flight Number", "Flight Date", "Origin", "Destination",
    "Cabin", "Seats Available", "Inventory Status",
)


def _column_index_map(header_row: list) -> dict[str, int]:
    """Map column name → 0-based index, ignoring case + extra whitespace."""
    return {
        (str(cell).strip() if cell is not None else ""): i
        for i, cell in enumerate(header_row)
    }


def read_flight_loads_excel(path: str | Path) -> list[LoadEntry]:
    """Parse a Flight Loads workbook into LoadEntry rows.

    Tolerates missing optional columns but logs a warning. Returns
    rows in source order; the indexing step deduplicates.
    """
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Flight Loads" not in wb.sheetnames:
        raise ValueError(
            f"{path.name} doesn't contain a 'Flight Loads' sheet "
            f"(found: {wb.sheetnames}). Was it exported from this app?"
        )
    ws = wb["Flight Loads"]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        log.warning("Flight Loads sheet has no rows in %s", path.name)
        return []

    cols = _column_index_map(header)
    missing = [c for c in _REQUIRED_COLUMNS if c not in cols]
    if missing:
        raise ValueError(
            f"{path.name} is missing required columns: {missing}. "
            f"Header was: {list(cols.keys())}",
        )

    def cell(row: tuple, name: str) -> str:
        i = cols.get(name)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else str(v).strip()

    out: list[LoadEntry] = []
    for row in rows_iter:
        flight_number = cell(row, "Flight Number")
        flight_date = cell(row, "Flight Date")
        if not flight_number or not flight_date:
            continue
        parsed = parse_seats_available(cell(row, "Seats Available"))
        if parsed is None:
            # Skip rows where we can't read the load factor — they
            # provide no value to the audit and only confuse the index.
            continue
        avail, cap, pct = parsed
        out.append(LoadEntry(
            flight_number=flight_number,
            flight_date=flight_date,
            origin=cell(row, "Origin"),
            destination=cell(row, "Destination"),
            cabin=cell(row, "Cabin") or "Economy",
            seats_available=avail,
            seats_capacity=cap,
            load_pct=pct,
            inventory_status=cell(row, "Inventory Status"),
            raw_seats_available=cell(row, "Seats Available"),
        ))
    wb.close()
    log.info("Loaded %d Flight Loads rows from %s", len(out), path.name)
    return out


# ---------------------------------------------------------------------------
# Lookup helpers
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class LoadLookup:
    """In-memory index over LoadEntry rows.

    Keys collapse on (flight_number, flight_date, origin, destination,
    cabin). When the history audit doesn't know the cabin it falls back
    to 'Economy' (the overwhelming majority of US-Bangla loads).
    """

    by_leg: dict[tuple[str, str, str, str, str], LoadEntry]
    by_flight_date: dict[tuple[str, str], list[LoadEntry]]

    @classmethod
    def from_entries(cls, entries: list[LoadEntry]) -> "LoadLookup":
        by_leg: dict[tuple[str, str, str, str, str], LoadEntry] = {}
        by_flight_date: dict[tuple[str, str], list[LoadEntry]] = {}
        for e in entries:
            key = (
                e.flight_number, e.flight_date,
                e.origin, e.destination, e.cabin,
            )
            # Last write wins — duplicates in the workbook are rare.
            by_leg[key] = e
            by_flight_date.setdefault(
                (e.flight_number, e.flight_date), [],
            ).append(e)
        return cls(by_leg=by_leg, by_flight_date=by_flight_date)

    def find(
        self,
        flight_number: str,
        flight_date: str,
        origin: str = "",
        destination: str = "",
        cabin: str = "Economy",
    ) -> LoadEntry | None:
        """Look up a leg's load. Falls back to flight+date if leg unknown."""
        if not flight_number or not flight_date:
            return None
        # Exact match on leg + cabin
        if origin and destination:
            exact = self.by_leg.get(
                (flight_number, flight_date, origin, destination, cabin),
            )
            if exact is not None:
                return exact
            # Some history rows have route reversed (return leg) — try.
            reversed_ = self.by_leg.get(
                (flight_number, flight_date, destination, origin, cabin),
            )
            if reversed_ is not None:
                return reversed_
        # Fall back: any leg on this (flight, date), prefer Economy.
        candidates = self.by_flight_date.get((flight_number, flight_date), [])
        if not candidates:
            return None
        economy = [c for c in candidates if c.cabin.lower() == "economy"]
        return (economy or candidates)[0]


# ---------------------------------------------------------------------------
# Verdict
# ---------------------------------------------------------------------------


def load_verdict(
    load_pct: float | None,
    *,
    high_threshold: float = HIGH_LOAD_THRESHOLD,
    low_threshold: float = LOW_LOAD_THRESHOLD,
) -> str:
    """Bucket a load% into a justification verdict for audit reports.

    Thresholds default to the module-level constants but callers
    (notably the GUI) can override them at audit time without
    editing this file.
    """
    if load_pct is None:
        return VERDICT_UNKNOWN
    if load_pct >= high_threshold:
        return VERDICT_QUESTIONABLE
    if load_pct < low_threshold:
        return VERDICT_JUSTIFIED
    return VERDICT_SITUATIONAL
