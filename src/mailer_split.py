"""Split one main sheet into per-recipient workbooks for the Bulk Mailer.

The user keeps ONE Excel where every data row carries an email address in some
column. `split_by_email` groups the rows by each address found there
(case-insensitive; a cell may list several addresses separated by ; , or |,
in which case the row goes to EACH), writes one workbook per address into a
chosen folder, and parks blank/invalid-address rows in `_UNMATCHED_ROWS.xlsx`
so nothing is ever silently dropped. `build_mail_rows` then wraps each split
file as a ready-to-send `mailer_io.MailRow` — the exact shape the existing
Bulk Mailer preview/run pipeline consumes — stamping a GUI-level CC/BCC on
every message and exposing {email}/{name}/{rows}/{file} template fields.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .mailer_io import MailRow

log = logging.getLogger(__name__)

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_ADDR_SPLIT_RE = re.compile(r"[;,|]")
_UNSAFE_FILENAME_RE = re.compile(r'[\\/:*?"<>|\s]+')
UNMATCHED_FILENAME = "_UNMATCHED_ROWS.xlsx"
_MAX_STEM_LEN = 100


@dataclass(frozen=True)
class SplitGroup:
    """One recipient's slice: their address, how many rows, and the file."""

    email: str
    row_count: int
    path: str


@dataclass(frozen=True)
class SplitResult:
    """Outcome of one split run. `unmatched_rows` are 1-based data-row numbers
    whose email cell was blank/invalid — written to `unmatched_path`."""

    source: str
    email_column: str
    groups: tuple[SplitGroup, ...]
    unmatched_rows: tuple[int, ...]
    unmatched_path: str | None
    warnings: tuple[str, ...]


def read_headers(path: str | Path, sheet_name: str | None = None) -> list[str]:
    """First-row header labels of the sheet (for the email-column picker)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
        for row in ws.iter_rows(max_row=1, values_only=True):
            return ["" if c is None else str(c).strip() for c in row]
        return []
    finally:
        wb.close()


def _addresses_in(cell: str) -> list[str]:
    """Valid, lowercased addresses in one cell (may hold several)."""
    return [a.strip().lower() for a in _ADDR_SPLIT_RE.split(cell)
            if a.strip() and _EMAIL_RE.match(a.strip())]


def _safe_stem(email: str, used: set[str]) -> str:
    """Filesystem-safe unique file stem for an address ('@' is legal on Windows)."""
    stem = _UNSAFE_FILENAME_RE.sub("_", email)[:_MAX_STEM_LEN] or "recipient"
    base, n = stem, 2
    while stem.lower() in used:
        stem = f"{base}_{n}"
        n += 1
    used.add(stem.lower())
    return stem


def _write_rows(path: Path, headers: list, rows: list[tuple]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rows"
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def split_by_email(
    path: str | Path,
    out_dir: str | Path,
    *,
    email_column: str,
    sheet_name: str | None = None,
) -> SplitResult:
    """Split the sheet into one workbook per email address in `email_column`.

    Raises ValueError when the column is missing (fail fast at the boundary);
    data problems (blank/invalid addresses, empty sheet) come back as warnings
    + the unmatched file instead of exceptions, so one bad row never kills a run.
    """
    path = Path(path)
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = ["" if c is None else str(c).strip() for c in next(rows_iter)]
        except StopIteration:
            return SplitResult(str(path), email_column, (), (), None,
                               ("Sheet is empty — nothing to split.",))
        wanted = email_column.strip().lower()
        try:
            email_i = [h.lower() for h in headers].index(wanted)
        except ValueError:
            raise ValueError(
                f"Email column {email_column!r} not found. "
                f"Sheet headers: {', '.join(h for h in headers if h)}") from None

        by_email: dict[str, list[tuple]] = {}
        unmatched: list[tuple[int, tuple]] = []
        data_row = 0
        for raw in rows_iter:
            if not any(str(v).strip() for v in raw if v is not None):
                continue                                    # trailing blank rows
            data_row += 1
            cell = "" if (email_i >= len(raw) or raw[email_i] is None) \
                else str(raw[email_i]).strip()
            addresses = _addresses_in(cell)
            if not addresses:
                unmatched.append((data_row, raw))
                continue
            for addr in addresses:                          # shared row -> each address
                by_email.setdefault(addr, []).append(raw)
    finally:
        wb.close()

    warnings: list[str] = []
    if not by_email and not unmatched:
        warnings.append("Sheet is empty — nothing to split.")

    used: set[str] = set()
    groups: list[SplitGroup] = []
    for addr in sorted(by_email):
        gpath = out / f"{_safe_stem(addr, used)}.xlsx"
        _write_rows(gpath, headers, by_email[addr])
        groups.append(SplitGroup(email=addr, row_count=len(by_email[addr]),
                                 path=str(gpath)))

    unmatched_path: str | None = None
    if unmatched:
        unmatched_path = str(out / UNMATCHED_FILENAME)
        _write_rows(Path(unmatched_path), headers, [r for _, r in unmatched])
        warnings.append(
            f"{len(unmatched)} row(s) had a blank/invalid address in "
            f"{email_column!r} — parked in {UNMATCHED_FILENAME}, NOT sent.")

    log.info("split %s: %d recipients, %d unmatched rows",
             path.name, len(groups), len(unmatched))
    return SplitResult(
        source=str(path), email_column=email_column, groups=tuple(groups),
        unmatched_rows=tuple(i for i, _ in unmatched),
        unmatched_path=unmatched_path, warnings=tuple(warnings))


def build_mail_rows(result: SplitResult, *, cc: str = "", bcc: str = "") -> list[MailRow]:
    """Convert split groups into MailRows the existing mailer pipeline consumes.

    `cc`/`bcc` (the GUI's global fields) are stamped on EVERY message. Template
    fields per row: {email}, {name} (address local part), {rows}, {file}.
    """
    rows: list[MailRow] = []
    for i, g in enumerate(result.groups, start=1):
        p = Path(g.path)
        issues: tuple[str, ...] = () if p.is_file() else (f"file not found: {p.name}",)
        local = g.email.split("@", 1)[0]
        fields = {
            "email": g.email, "name": local, "Name": local,
            "rows": str(g.row_count), "file": p.name,
        }
        rows.append(MailRow(
            row_index=i, email=g.email, name=local, attachments=(p,),
            cc=cc.strip(), bcc=bcc.strip(), fields=fields, issues=issues))
    return rows
