"""Tests for the passenger-manifest parser + Excel export.

All fixtures are inline with fully fake passenger data (names, passport,
DOB) so no real PII lives in the repo. The HTML mirrors the
liste_passager_vol.asp markup: each passenger is one <tr> with
<td class="Col_*"> cells whose value sits in an xs-right div after a
repeated xs-left column label.
"""

from __future__ import annotations

from openpyxl import load_workbook

from src.excel_io import write_passenger_manifest
from src.zenith_client import parse_passenger_list_html


def _cell(col: str, label: str, value: str) -> str:
    return (
        f'<td class="{col}">'
        f'<div class="visible-xs-inline xs-left">{label}</div>'
        f'<div class="visible-xs-inline xs-right">{value}</div>'
        f"</td>"
    )


def _pax_row(
    *, title="Mr.", name="TEST PAXONE", ptype="AD", gender="M",
    dob="25/01/1988", passport="XX0000001", weight="75",
    cabin="Y", prbd="E", fare="EDXBO", web="Economy Lite",
    ticket="7792000000001", seat="", pnr="AAA111", gds="BBB222",
    leg="DXB DAC", agency="TA: 00000000 Example Travel LLC", direction="Outbound",
    bare_nbsp=False,
) -> str:
    sep = "&nbsp" if bare_nbsp else "&nbsp;"  # exercise the malformed-entity path
    name_html = (
        f'<div class="visible-xs-inline xs-left">Surname, First name, Weight</div>'
        f'<div class="visible-xs-inline xs-right">'
        f'<i class="material-icons">&#xE8A6;</i>'
        f"<span><strong> {title} {name} </strong></span>"
        f"<span> {ptype} - {gender} </span>"
        f"<br>Born on:{sep}{dob}<br>Passport{sep}No.:{passport}<br>"
        f'<input type="hidden" name="x" value="{weight}">{weight}{sep}Kg'
        f"</div>"
    )
    pnr_val = f"{pnr} {gds}".strip()
    return (
        '<tr valign="Top">'
        f'<td class="Col_ChronoVol"><div class="visible-xs-inline xs-left">Nr.</div>'
        f'<div class="visible-xs-inline xs-right"> </div></td>'
        f'<td class="Col_NomPrenomAge">{name_html}</td>'
        + _cell("Col_CabineClasseCode", "Cabin code", cabin)
        + _cell("Col_PRBD", "PRBD", prbd)
        + _cell("Col_Classe", "Cl.", fare)
        + _cell("Col_Webclasse", "Web Cl.", web)
        + _cell("Col_Billet", "Ticket", f"TKNE: {ticket}")
        + _cell("Col_SEAT", "", seat)
        + _cell("Col_PNR", "PNR", pnr_val)
        + _cell("Col_Leg", "Leg", leg)
        + _cell("Col_AgenceEmettrice", "Issuing agency", agency)
        + _cell("Col_AR", "OUT/IN", direction)
        + "</tr>"
    )


def _page(rows: list[str]) -> str:
    return (
        "<html><body>"
        "<h2>Flight : BS342 Dubai Intl - Dhaka The 01/06/2026 at 00:25</h2>"
        "<table><tbody>" + "".join(rows) + "</tbody></table></body></html>"
    )


def test_parse_single_passenger_all_fields():
    recs = parse_passenger_list_html(_page([_pax_row()]), id_vol="331712", id_leg="245")
    assert len(recs) == 1
    r = recs[0]
    assert r.flight_number == "BS342"
    assert r.route_desc == "Dubai Intl - Dhaka"
    assert r.flight_date == "01/06/2026"
    assert r.flight_time == "00:25"
    assert r.id_vol == "331712" and r.id_leg == "245"
    assert r.title == "Mr."
    assert r.full_name == "TEST PAXONE"
    assert r.pax_type == "AD" and r.gender == "M"
    assert r.date_of_birth == "25/01/1988"
    assert r.passport_no == "XX0000001"
    assert r.weight_kg == "75"
    assert r.cabin_code == "Y"
    assert r.prbd == "E"
    assert r.fare_basis == "EDXBO"
    assert r.web_class == "Economy Lite"
    assert r.ticket_number == "7792000000001"
    assert r.pnr == "AAA111" and r.gds_pnr == "BBB222"
    assert r.leg == "DXB DAC"
    assert r.issuing_agency == "TA: 00000000 Example Travel LLC"
    assert r.direction == "Outbound"


def test_parse_handles_malformed_nbsp_in_dob():
    """Real markup sometimes emits `Born on:&nbsp25/01/1988` (no semicolon)."""
    recs = parse_passenger_list_html(_page([_pax_row(bare_nbsp=True)]))
    assert recs[0].date_of_birth == "25/01/1988"
    assert recs[0].passport_no == "XX0000001"


def test_parse_single_pnr_no_gds():
    recs = parse_passenger_list_html(_page([_pax_row(pnr="CCC333", gds="")]))
    assert recs[0].pnr == "CCC333"
    assert recs[0].gds_pnr == ""


def test_parse_child_and_infant_types():
    rows = [
        _pax_row(name="TEST CHILD", ptype="CH", gender="F", dob="27/06/2022"),
        _pax_row(name="TEST INFANT", ptype="INF", gender="M", dob="01/01/2026"),
    ]
    recs = parse_passenger_list_html(_page(rows))
    assert [r.pax_type for r in recs] == ["CH", "INF"]
    assert recs[0].full_name == "TEST CHILD"


def test_parse_multiple_passengers():
    rows = [_pax_row(name=f"PAX NUMBER{i}", ticket=f"77920000000{i:02d}",
                     pnr=f"PNR{i:03d}", prbd="K", fare="KDXBO")
            for i in range(12)]
    recs = parse_passenger_list_html(_page(rows))
    assert len(recs) == 12
    assert all(r.prbd == "K" for r in recs)
    assert recs[5].full_name == "PAX NUMBER5"


def test_parse_empty_page():
    assert parse_passenger_list_html("<html><body>nothing</body></html>") == []


def test_write_passenger_manifest_round_trip(tmp_path):
    recs = parse_passenger_list_html(_page([
        _pax_row(),
        _pax_row(name="OTHER PAX", pnr="DDD444", gds="", prbd="X", fare="XDXB6M"),
    ]), id_vol="331712", id_leg="245")
    p = tmp_path / "pax.xlsx"
    write_passenger_manifest(p, recs)
    wb = load_workbook(p)
    ws = wb["Passengers"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "Flight"          # header
    assert rows[0][7] == "Passenger Name"
    assert rows[1][0] == "BS342"
    assert rows[1][7] == "TEST PAXONE"
    assert rows[1][14] == "E"              # PRBD column
    assert rows[2][7] == "OTHER PAX"
    assert rows[2][15] == "XDXB6M"         # Fare Basis column
    assert len(rows) == 1 + 2
