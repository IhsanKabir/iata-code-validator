"""The Flight Loads pull can emit several report formats from ONE set of rows.

The flat workbook is always written; the picker adds a cross-tab, the seats
analysis, or the ops team's daily-snapshot shape. Because every format reads the
same parsed rows, their totals must agree — that is what these tests pin, along
with the parsing of Zenith's composite cells ("13/410 97%", "[152]").
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from src.excel_io import (
    FLIGHT_LOAD_FORMATS,
    build_flight_load_format_path,
    flight_load_leg_totals,
    is_domestic_leg_route,
    write_flight_loads_daily_snapshot,
    write_flight_loads_daily_snapshots,
)


class _Row:
    """Stands in for zenith_client.FlightLoadRow (only the fields used here)."""

    def __init__(self, date, flight, leg, seats_available, seats_confirmed,
                 cabin="Economy", dep="07:00", aircraft="Boeing 737-800"):
        self.flight_date = date
        self.flight_number = flight
        self.leg_route = leg
        self.seats_available = seats_available
        self.seats_confirmed = seats_confirmed
        self.leg_cabin = cabin
        self.departure_time = dep
        self.aircraft = aircraft


def _rows():
    return [
        # 410-seat leg, 397 sold  (13 available)
        _Row("12/08/2026", "BS343", "DAC-CGP", "13/410 97%", "[397]"),
        _Row("12/08/2026", "BS343", "CGP-DXB", "10/410 98%", "[400]"),
        _Row("12/08/2026", "BS101", "DAC-CGP", "2/72 97%", "[70]"),
        _Row("13/08/2026", "BS101", "DAC-CGP", "12/72 83%", "[60]"),
        # a second cabin on the same leg — must be SUMMED, not counted twice
        _Row("13/08/2026", "BS101", "DAC-CGP", "0/8 100%", "[8]", cabin="Business"),
    ]


def test_leg_totals_parse_capacity_and_sold():
    t = flight_load_leg_totals(_rows())
    assert t[("12/08/2026", "BS343", "DAC-CGP")] == {
        "capacity": 410, "sold": 397, "std": "07:00", "aircraft": "Boeing 737-800"}
    # cabins summed into one leg total
    assert t[("13/08/2026", "BS101", "DAC-CGP")]["capacity"] == 80
    assert t[("13/08/2026", "BS101", "DAC-CGP")]["sold"] == 68


def test_leg_totals_skip_unreadable_rows_rather_than_zero_fill():
    rows = _rows() + [_Row("14/08/2026", "BS999", "DAC-ZYL", "n/a", "")]
    t = flight_load_leg_totals(rows)
    assert ("14/08/2026", "BS999", "DAC-ZYL") not in t


def test_flight_numbers_normalised():
    t = flight_load_leg_totals([_Row("12/08/2026", "BS 343", "dac-cgp",
                                     "13/410 97%", "[397]")])
    assert ("12/08/2026", "BS343", "DAC-CGP") in t


def test_daily_snapshot_round_trips_through_the_ops_format_reader(tmp_path):
    """The snapshot we WRITE must be readable by the reader built for the files we
    RECEIVE — that is the proof the format actually matches."""
    reader = pytest.importorskip("reporting.daily_load_reader")
    p = tmp_path / "Daily International Flight Load out.xlsx"
    write_flight_loads_daily_snapshot(p, _rows())

    wb = load_workbook(p)
    assert "12-AUG-2026" in wb.sheetnames and "13-AUG-2026" in wb.sheetnames
    wb.close()

    back = {(r.flight_date.strftime("%d/%m/%Y"), r.flight_number, r.leg_route): r.sold
            for r in reader.read_workbook(p)}
    assert back[("12/08/2026", "BS101", "DAC-CGP")] == 70
    assert back[("13/08/2026", "BS101", "DAC-CGP")] == 68     # both cabins
    assert back[("12/08/2026", "BS343", "CGP-DXB")] == 400


def test_daily_snapshot_refuses_an_empty_pull(tmp_path):
    with pytest.raises(ValueError):
        write_flight_loads_daily_snapshot(tmp_path / "x.xlsx", [])


def test_format_paths_are_distinct_and_labelled():
    assert FLIGHT_LOAD_FORMATS[0] == "Flat rows only"
    made = {build_flight_load_format_path(__import__("pathlib").Path("."), f).name
            for f in FLIGHT_LOAD_FORMATS[1:]}
    assert len(made) == 3                       # one filename per format, no clashes
    assert any(n.startswith("load_factor_seats_") for n in made)
    assert any(n.startswith("daily_flight_load_") for n in made)
    assert any(n.startswith("flight_loads_ordered_") for n in made)


def test_seats_report_totals_match_the_pulled_rows(tmp_path):
    """The seats analysis must not drift from the flat data it came from."""
    pytest.importorskip("reporting.builders.load_factor_seats")
    from src.excel_io import write_flight_loads_seats_report

    p = tmp_path / "seats.xlsx"
    write_flight_loads_seats_report(p, _rows())
    ws = load_workbook(p, data_only=True)["Seats Summary"]

    totals = flight_load_leg_totals(_rows())
    want_cap = sum(v["capacity"] for v in totals.values())
    want_sold = sum(v["sold"] for v in totals.values())

    found = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(r, c).value or "").strip() == "All flights":
                found = (r, c)
    assert found, "headline 'All flights' row missing"
    r, c = found
    assert ws.cell(r, c + 7).value == want_cap       # Total Seats Offered
    assert ws.cell(r, c + 8).value == want_sold      # Total Seats Filled
    assert ws.cell(r, c + 9).value == want_cap - want_sold   # Total Empty Seats


# --- the ops team's files come as a Domestic + an International workbook ---------

def test_domestic_leg_classification():
    assert is_domestic_leg_route("DAC-CGP") and is_domestic_leg_route("CXB-DAC")
    assert not is_domestic_leg_route("CGP-DXB") and not is_domestic_leg_route("DAC-SIN")
    assert not is_domestic_leg_route("DAC")            # malformed is never domestic


def test_snapshot_writes_a_sector_split_pair(tmp_path):
    """One mixed workbook would be read back as entirely international, because the
    sector of these files rides on the FILE NAME — so the pair must be split."""
    reader = pytest.importorskip("reporting.daily_load_reader")
    made = write_flight_loads_daily_snapshots(tmp_path, _rows())
    names = sorted(p.name for p in made)
    assert len(made) == 2
    assert any(n.startswith("Daily Domestic Flight Load") for n in names)
    assert any(n.startswith("Daily International Flight Load") for n in names)

    by_sector = {}
    for p in made:
        for r in reader.read_workbook(p):
            by_sector.setdefault(r.sector, set()).add(r.leg_route)
    assert by_sector["DOM"] == {"DAC-CGP"}          # BS101 + BS343's DAC-CGP sector
    assert by_sector["INTL"] == {"CGP-DXB"}


def test_snapshot_pair_skips_a_sector_with_no_legs(tmp_path):
    only_dom = [r for r in _rows() if is_domestic_leg_route(r.leg_route)]
    made = write_flight_loads_daily_snapshots(tmp_path, only_dom)
    assert len(made) == 1 and made[0].name.startswith("Daily Domestic")


def test_snapshot_pair_refuses_an_empty_pull(tmp_path):
    with pytest.raises(ValueError):
        write_flight_loads_daily_snapshots(tmp_path, [])
