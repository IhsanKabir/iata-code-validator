"""NATTA (Nepal) member-directory extractor — parsers, bulk fetch, Excel out.

Fixtures mirror natta.org.np's real markup (pre.member-detail-link index rows;
dt/dd detail block + owner h4/span widget) with synthetic businesses. The pure
parsers were additionally validated against the live site (644 index rows, all
member fields) before these fixtures were frozen.
"""

from __future__ import annotations

import threading

import pytest
from openpyxl import load_workbook

from src.config import NATTA_OUTPUT_COLUMNS
from src.excel_io import build_natta_output_path, write_natta_members
from src.natta_client import (
    NattaMember,
    fetch_all_members,
    parse_member_html,
    parse_member_index,
)

_INDEX = """
<span class="anchor" id="s"></span><div class="sub-menu"><h3> S </h3>
<pre class="member-detail-link"><a href="https://natta.org.np/member/sherpa-sky-tours/">Sherpa Sky Tours &#038; Travels</a>  [Member ID: NATTA-111/11]</pre>
<pre class="member-detail-link"><a href="https://natta.org.np/member/summit-holidays/">Summit Holidays P. Ltd</a>  [Member ID: NATTA-222/22]</pre>
<pre class="member-detail-link"><a href="https://natta.org.np/member/gone-away-travels/">Gone Away Travels</a>  </pre>
</div>
"""

_MEMBER = """
<div class="member-image-widgets">
  <img src="http://natta.org.np/wp-content/uploads/x/owner.jpg" alt="no-img">
  <h4>Ms. Pema Sherpa\n<span>(&nbsp;Managing Director&nbsp;)</span></h4>
</div>
<div class="single-member">
<dl><dt>Natta Member ID :</dt><dd>NATTA-111/11<br/></dd></dl>
<dl><dt>Company Name:</dt><dd>Sherpa Sky Tours &#038; Travels</dd></dl>
<dl><dt>Telephone:</dt><dd>4411223, 4411224<br/></dd></dl>
<dl><dt>Office Address:</dt><dd>Thamel, Kathmandu<br/></dd></dl>
<dl><dt>Email Address:</dt><dd>info@sherpasky.example<br/></dd></dl>
<dl><dt>Website:</dt><dd><a href="http://sherpasky.example">www.sherpasky.example<br/></a></dd></dl>
<dl><dt>FAX:</dt><dd>01-4411225</dd></dl>
</div>
"""


# --- index parser -------------------------------------------------------------

def test_index_parses_rows_with_and_without_member_id():
    rows = parse_member_index(_INDEX)
    assert rows == [
        ("NATTA-111/11", "Sherpa Sky Tours & Travels",
         "https://natta.org.np/member/sherpa-sky-tours/"),
        ("NATTA-222/22", "Summit Holidays P. Ltd",
         "https://natta.org.np/member/summit-holidays/"),
        ("", "Gone Away Travels",
         "https://natta.org.np/member/gone-away-travels/"),
    ]


def test_index_dedups_repeated_urls():
    assert len(parse_member_index(_INDEX + _INDEX)) == 3


def test_index_empty_page_returns_empty():
    assert parse_member_index("<html>maintenance</html>") == []


# --- member parser ------------------------------------------------------------

def test_member_parses_all_core_fields_and_extras():
    f = parse_member_html(_MEMBER)
    assert f["member_id_page"] == "NATTA-111/11"
    assert f["company_name"] == "Sherpa Sky Tours & Travels"   # entity decoded
    assert f["telephone"] == "4411223, 4411224"
    assert f["office_address"] == "Thamel, Kathmandu"
    assert f["email"] == "info@sherpasky.example"
    assert f["website"] == "www.sherpasky.example"
    assert f["owner_name"] == "Ms. Pema Sherpa"
    assert f["designation"] == "Managing Director"
    assert f["photo_url"].endswith("owner.jpg")
    assert f["extras"] == {"FAX": "01-4411225"}                # unknown label kept


def test_member_page_without_owner_block():
    f = parse_member_html("<dl><dt>Telephone:</dt><dd>123</dd></dl>")
    assert f["telephone"] == "123" and "owner_name" not in f


# --- bulk fetch (fake session) ------------------------------------------------

class _FakeHttp:
    def __init__(self, routes: dict[str, object]) -> None:
        self.routes = routes
        self.calls: list[str] = []

    def get(self, url, headers=None, timeout=None):
        self.calls.append(url)
        item = next((v for k, v in self.routes.items() if k in url), (404, ""))
        status, text = item if isinstance(item, tuple) else (200, item)

        class _R:
            pass
        r = _R()
        r.status_code = status
        r.text = text
        return r


def test_fetch_all_members_index_then_details_in_order():
    sess = _FakeHttp({
        "/members/": _INDEX,
        "sherpa-sky-tours": _MEMBER,
        "summit-holidays": (404, ""),
        "gone-away-travels": (200, "<html>page without the detail block</html>"),
    })
    got: list[tuple] = []
    out = fetch_all_members(session=sess, delay_s=0.0, concurrency=2,
                            progress_cb=lambda d, t, m: got.append((d, t)))
    assert [m.status for m in out] == ["OK", "NOT_FOUND", "NO_DETAIL_BLOCK"]
    ok = out[0]
    assert ok.member_id == "NATTA-111/11" and ok.name == "Sherpa Sky Tours & Travels"
    assert ok.owner_name == "Ms. Pema Sherpa" and ok.email == "info@sherpasky.example"
    assert ok.extras == {"FAX": "01-4411225"}
    assert got and got[-1][0] == 3                     # progress reached the total


def test_fetch_all_members_raises_when_index_unavailable():
    with pytest.raises(RuntimeError):
        fetch_all_members(session=_FakeHttp({"/members/": (500, "boom")}), delay_s=0.0)


def test_fetch_all_members_stop_event_returns_partial():
    stop = threading.Event()
    stop.set()                                          # stop before anything runs
    out = fetch_all_members(session=_FakeHttp({"/members/": _INDEX}),
                            delay_s=0.0, stop_event=stop)
    assert out == []                                    # nothing fetched, no crash


# --- Excel out ----------------------------------------------------------------

def test_write_natta_members_columns_and_rows(tmp_path):
    members = [
        NattaMember(member_id="NATTA-111/11", name="Sherpa Sky Tours & Travels",
                    url="https://natta.org.np/member/sherpa-sky-tours/",
                    owner_name="Ms. Pema Sherpa", designation="Managing Director",
                    telephone="4411223", office_address="Thamel, Kathmandu",
                    email="info@sherpasky.example", website="www.sherpasky.example",
                    company_name="Sherpa Sky Tours & Travels",
                    member_id_page="NATTA-111/11", status="OK",
                    extras={"FAX": "01-4411225"}),
        NattaMember(member_id="NATTA-222/22", name="Summit Holidays P. Ltd",
                    url="https://natta.org.np/member/summit-holidays/",
                    status="NOT_FOUND"),
    ]
    p = tmp_path / "natta.xlsx"
    write_natta_members(p, members)
    ws = load_workbook(p).active
    assert [c.value for c in ws[1]] == NATTA_OUTPUT_COLUMNS
    row1 = {k: v for k, v in zip(NATTA_OUTPUT_COLUMNS, (c.value for c in ws[2]))}
    assert row1["Owner Name"] == "Ms. Pema Sherpa"
    assert row1["Other Fields"] == "FAX: 01-4411225"
    assert row1["Status"] == "OK"
    row2 = {k: v for k, v in zip(NATTA_OUTPUT_COLUMNS, (c.value for c in ws[3]))}
    assert row2["Status"] == "NOT_FOUND" and (row2["Owner Name"] in ("", None))


def test_build_natta_output_path_pattern(tmp_path):
    p = build_natta_output_path(tmp_path)
    assert p.name.startswith("natta_members_") and p.suffix == ".xlsx"
