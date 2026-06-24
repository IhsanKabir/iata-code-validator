"""Tests for the Zenith PNR client + cache + analyzer enrichment.

All fixtures are inline with fully fake passenger/PNR data so no real
booking information lives in the repo. The HTML structure mirrors the
ASP.NET WebForms field-naming used by Dossier.aspx.
"""

from __future__ import annotations

import textwrap
from datetime import datetime

import pytest

from src.zenith_history_analyzer import (
    apply_pnr_enrichment,
    build_pnr_routes,
    run_history_audit,
)
from src.zenith_history_parser import Agent, FlightRef, HistoryEvent
from src.zenith_pnr_cache import ZenithPNRCache
from src.zenith_pnr_client import (
    PNRDetails,
    PNRNotFoundError,
    PNRSegment,
    parse_dossier_html,
)


# ---------------------------------------------------------------------------
# Synthetic Dossier HTML builder
# ---------------------------------------------------------------------------


def _lbl(field: str, value: str) -> str:
    """Render an ASP.NET WebForms label span the parser anchors on."""
    return (
        f'<span id="instanceCtrlContent_'
        f'UsrDossierSynthese_rptSegments_ctl00_rptVols_ctl00_lbl{field}">'
        f"{value}</span>"
    )


def _seg_lbl(seg_idx: int, field: str, value: str) -> str:
    """Per-passenger / per-segment label (rptPassagers under rptVols)."""
    return (
        f'<span id="instanceCtrlContent_'
        f'UsrDossierSynthese_rptSegments_ctl00_rptVols_ctl00_'
        f'rptPassagers_ctl0{seg_idx}_lbl{field}">{value}</span>'
    )


def _etat(seg_idx: int, status: str) -> str:
    return (
        f'<a id="instanceCtrlContent_'
        f'UsrDossierSynthese_rptSegments_ctl00_rptVols_ctl00_'
        f'rptPassagers_ctl0{seg_idx}_hlEtat">{status}</a>'
    )


def _pax(seg_idx: int, name: str) -> str:
    return (
        f'<a id="instanceCtrlContent_'
        f'UsrDossierSynthese_rptSegments_ctl00_rptVols_ctl00_'
        f'rptPassagers_ctl0{seg_idx}_linkPassager">{name}</a>'
    )


def _build_dossier_html(
    pnr: str = "TEST01",
    dossier_id: str = "99999999",
    customer: str = "Test Customer Ltd",
    surname: str = "TESTSURNAME",
    phone: str = "+8801000000000",
    out_route: str = "DAC - DXB",
    return_route: str = "DXB - DAC",
    out_status: str = "Issued",
    return_status: str = "Issued",
    out_fare: str = "20,000 BDT",
    return_fare: str = "22,000 BDT",
    pax_count: str = "1 pax",
) -> str:
    parts = [
        "<html><body>",
        f"PNR : {dossier_id} | {pnr} - {customer}",
        _lbl("PNRCode", pnr),
        _lbl("CustomerName", customer),
        _lbl("NomProprio", surname),
        _lbl("TelMobile", phone),
        _lbl("NbPaxForPNR", pax_count),
        _lbl("EtatDossier", "Issued"),
        _lbl("Paiement", "On account"),
        _lbl("DeviseDossier", "BDT"),
        _lbl("PrixTotalTTC", "42,000 BDT"),
        _lbl("TaxesTotal", "2,000 BDT"),
        _lbl("LegAller", out_route),
        _lbl("LegRetour", return_route),
        _lbl("DatesAller", "Departure Tue 01/01/26"),
        _lbl("DatesRetour", "Departure Sun 07/01/26"),
        _seg_lbl(1, "DepartVol", "From DAC 22:30"),
        _seg_lbl(1, "ArriveeVol", "- To DXB 04:00+1"),
        _seg_lbl(1, "AircraftNumber", "Boeing 737-800 (S2-AJE)"),
        _seg_lbl(1, "Classe", "GTESTR6M-AD"),
        _seg_lbl(1, "PrixHT", "18,000 BDT"),
        _seg_lbl(1, "PrixTTC", out_fare),
        _seg_lbl(1, "TicketNumber", "7792000000001 6/1"),
        _pax(1, surname),
        _etat(1, out_status),
        _seg_lbl(2, "DepartVol", "From DXB 09:00"),
        _seg_lbl(2, "ArriveeVol", "- To DAC 19:00"),
        _seg_lbl(2, "AircraftNumber", "Boeing 737-800 (S2-AJF)"),
        _seg_lbl(2, "Classe", "MTESTRT-AD"),
        _seg_lbl(2, "PrixHT", "20,000 BDT"),
        _seg_lbl(2, "PrixTTC", return_fare),
        _seg_lbl(2, "TicketNumber", "7792000000001 6/2"),
        _pax(2, surname),
        _etat(2, return_status),
        "</body></html>",
    ]
    return "".join(parts)


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


class TestParseDossier:
    def test_extracts_basic_pnr_metadata(self) -> None:
        d = parse_dossier_html(_build_dossier_html())
        assert d.pnr_code == "TEST01"
        assert d.customer_name == "Test Customer Ltd"
        assert d.traveler_surname == "TESTSURNAME"
        assert d.phone == "+8801000000000"
        assert d.pax_count == 1
        assert d.pnr_status == "Issued"
        assert d.currency == "BDT"

    def test_extracts_segments_with_status(self) -> None:
        d = parse_dossier_html(_build_dossier_html(
            out_status="Flown", return_status="Refunded",
        ))
        assert len(d.segments) == 2
        out, ret = d.segments
        assert out.leg_route == "DAC-DXB"
        assert out.leg_direction == "OUT"
        assert out.coupon_status == "Flown"
        assert out.rbd_class == "G"
        assert ret.leg_route == "DXB-DAC"
        assert ret.leg_direction == "RETURN"
        assert ret.coupon_status == "Refunded"
        assert ret.rbd_class == "M"

    def test_booked_route_collapses_repeats(self) -> None:
        d = parse_dossier_html(_build_dossier_html())
        assert d.booked_route == "DAC-DXB-DAC"

    def test_flown_route_drops_refunded_segments(self) -> None:
        d = parse_dossier_html(_build_dossier_html(
            out_status="Flown", return_status="Refunded",
        ))
        # Outbound flown, return refunded → flown route ends at DXB.
        assert d.flown_route == "DAC-DXB"

    def test_flown_route_empty_when_everything_voided(self) -> None:
        d = parse_dossier_html(_build_dossier_html(
            out_status="Voided", return_status="Voided",
        ))
        assert d.flown_route == ""
        # But booked route stays intact.
        assert d.booked_route == "DAC-DXB-DAC"

    def test_raises_when_no_pnr_field_present(self) -> None:
        with pytest.raises(PNRNotFoundError):
            parse_dossier_html("<html><body>nothing</body></html>")


# ---------------------------------------------------------------------------
# Cache
# ---------------------------------------------------------------------------


class TestCache:
    def _details(self, pnr: str = "TEST99") -> PNRDetails:
        return parse_dossier_html(_build_dossier_html(pnr=pnr))

    def test_round_trip(self, tmp_path) -> None:
        cache = ZenithPNRCache(tmp_path / "p.sqlite")
        d = self._details()
        cache.put(d)
        got = cache.get("TEST99")
        assert got is not None
        assert got.pnr_code == "TEST99"
        assert len(got.segments) == 2
        assert got.segments[0].coupon_status == "Issued"

    def test_case_insensitive_lookup(self, tmp_path) -> None:
        cache = ZenithPNRCache(tmp_path / "p.sqlite")
        cache.put(self._details("UpPeR1"))
        # Stored uppercase, retrievable case-insensitive.
        assert cache.get("upper1") is not None

    def test_put_many_then_count(self, tmp_path) -> None:
        cache = ZenithPNRCache(tmp_path / "p.sqlite")
        cache.put_many([self._details(f"PNR{i:03d}") for i in range(5)])
        assert cache.count() == 5


# ---------------------------------------------------------------------------
# Analyzer enrichment
# ---------------------------------------------------------------------------


def _evt(pnr: str, ts: str, rbd: str) -> HistoryEvent:
    return HistoryEvent(
        source_file="t.xls", row_index=0,
        raw_date=ts, raw_created_by="A (a)",
        raw_description="", event_type="Ticket Modification",
        pnr=pnr, customer="", raw_flight="BS999 DAC DXB 01/01/2026 00:00",
        passenger="PAX",
        timestamp=datetime.strptime(ts, "%d/%m/%Y %H:%M"),
        agent=Agent(raw="A (a)", display_name="A", user_id="a", department=""),
        flight=FlightRef(raw="", flight_number="BS999",
                        origin="DAC", destination="DXB",
                        flight_date="01/01/2026", departure_time="00:00"),
        rbd_class=rbd,
    )


def test_build_pnr_routes_orders_disrupted_first() -> None:
    """PNRs with refund/void activity should sort above clean PNRs."""
    clean = parse_dossier_html(_build_dossier_html(pnr="CLEAN1"))
    disrupted = parse_dossier_html(_build_dossier_html(
        pnr="DISRP1", out_status="Flown", return_status="Refunded",
    ))
    rows = build_pnr_routes({"CLEAN1": clean, "DISRP1": disrupted})
    assert rows[0].pnr_code == "DISRP1"
    assert rows[0].refunded_count == 1
    assert rows[1].pnr_code == "CLEAN1"


def test_apply_pnr_enrichment_fills_customer_name() -> None:
    events = [
        _evt("PNR001", "01/01/2026 10:00", "Y"),
        _evt("PNR001", "01/01/2026 14:00", "G"),
    ]
    report = run_history_audit(events)
    # Pre-enrichment: customer_name is empty on the trajectory.
    assert report.class_trajectories[0].customer_name == ""

    details = parse_dossier_html(_build_dossier_html(
        pnr="PNR001", customer="Acme Travels",
    ))
    enriched = apply_pnr_enrichment(report, {"PNR001": details})
    assert enriched.class_trajectories[0].customer_name == "Acme Travels"
    assert len(enriched.pnr_routes) == 1
    assert enriched.pnr_routes[0].customer_name == "Acme Travels"


def test_read_pnr_codes_from_excel_uses_named_column(tmp_path) -> None:
    from openpyxl import Workbook
    from src.excel_io import read_pnr_codes_from_excel

    p = tmp_path / "input.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Note", "PNR", "Other"])
    ws.append(["x", "abc123", "y"])
    ws.append(["x", None, "y"])
    ws.append(["x", "  xyz999  ", "y"])
    wb.save(p)
    codes = read_pnr_codes_from_excel(p, column_name="PNR")
    assert codes == ["ABC123", "XYZ999"]


def test_read_pnr_codes_from_excel_defaults_to_first_column(tmp_path) -> None:
    from openpyxl import Workbook
    from src.excel_io import read_pnr_codes_from_excel

    p = tmp_path / "input.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Anything"])
    for code in ["aaa111", "bbb222", "ccc333"]:
        ws.append([code])
    wb.save(p)
    assert read_pnr_codes_from_excel(p) == ["AAA111", "BBB222", "CCC333"]


def test_write_zenith_pnr_bulk_creates_both_sheets(tmp_path) -> None:
    from openpyxl import load_workbook
    from src.excel_io import write_zenith_pnr_bulk

    details = parse_dossier_html(_build_dossier_html(
        pnr="ABC123", out_status="Flown", return_status="Refunded",
    ))
    p = tmp_path / "out.xlsx"
    # MISSING has no entry in `errors` → writer flags it as NOT_FOUND
    # (the worker treats PNRNotFoundError as distinct from real errors).
    write_zenith_pnr_bulk(
        p,
        [("ABC123", details), ("MISSING", None)],
    )
    wb = load_workbook(p)
    assert wb.sheetnames == ["PNR Lookup", "Segments"]
    # First data row is the resolved PNR.
    lookup = wb["PNR Lookup"]
    rows = list(lookup.iter_rows(values_only=True))
    assert rows[0][0] == "PNR"  # header
    assert rows[1][0] == "ABC123"
    assert rows[1][2] == "Test Customer Ltd"
    assert rows[1][8] == "DAC-DXB"      # flown route — return refunded
    assert rows[1][-2] == "OK"          # lookup status
    # The unresolved PNR gets a NOT_FOUND status.
    assert rows[2][0] == "MISSING"
    assert rows[2][-2] == "NOT_FOUND"

    segments = wb["Segments"]
    seg_rows = list(segments.iter_rows(values_only=True))
    assert seg_rows[0][0] == "PNR"  # header
    # 2 segments for ABC123, 0 for MISSING (no details)
    assert len(seg_rows) == 1 + 2


def test_apply_pnr_enrichment_keeps_old_report_immutable() -> None:
    """Enrichment returns a NEW report — the original is not mutated."""
    events = [
        _evt("PNR002", "01/01/2026 10:00", "Y"),
        _evt("PNR002", "01/01/2026 14:00", "G"),
    ]
    report = run_history_audit(events)
    details = parse_dossier_html(_build_dossier_html(pnr="PNR002"))
    enriched = apply_pnr_enrichment(report, {"PNR002": details})
    assert enriched is not report
    # Original trajectory still has empty customer_name.
    assert report.class_trajectories[0].customer_name == ""


# ---------------------------------------------------------------------------
# Network resilience — lookup_pnr retry ladder + concurrent lookup_many
# ---------------------------------------------------------------------------

import src.zenith_pnr_client as zpc  # noqa: E402
from src.zenith_pnr_client import QUICK_SEARCH_URL  # noqa: E402


class _Resp:
    """Minimal stand-in for a requests.Response."""

    def __init__(self, status_code: int = 200, text: str = "", url: str | None = None):
        self.status_code = status_code
        self.text = text
        self.url = url or QUICK_SEARCH_URL


class _Inner:
    """Scripted requests.Session: each .get() pops the next queued item;
    an Exception is raised, a _Resp is returned. Exhausted → repeat last."""

    def __init__(self, items):
        self._items = list(items)
        self._last = self._items[-1] if self._items else _Resp()
        self.headers: dict = {}
        self.calls = 0

    def get(self, url, params=None, timeout=None):
        self.calls += 1
        item = self._items.pop(0) if self._items else self._last
        if isinstance(item, BaseException):
            raise item
        return item


class _Sess:
    """Stand-in for ZenithSession (only `.session` is used by lookup_pnr)."""

    def __init__(self, inner: _Inner):
        self.session = inner


class TestLookupPnrRetry:
    def _ok(self, pnr: str = "RETRY1") -> str:
        return _build_dossier_html(pnr=pnr)

    def test_retries_transient_504_then_succeeds(self, monkeypatch) -> None:
        monkeypatch.setattr(zpc, "_backoff_with_jitter", lambda *a, **k: 0.0)
        inner = _Inner([_Resp(504), _Resp(200, self._ok("RETRY1"))])
        d = zpc.lookup_pnr(_Sess(inner), "RETRY1")
        assert d.pnr_code == "RETRY1"
        assert inner.calls == 2  # one retry recovered it

    def test_persistent_5xx_raises_after_attempts(self, monkeypatch) -> None:
        monkeypatch.setattr(zpc, "_backoff_with_jitter", lambda *a, **k: 0.0)
        inner = _Inner([_Resp(504)])
        with pytest.raises(zpc.ZenithError):
            zpc.lookup_pnr(_Sess(inner), "X", max_attempts=3)
        assert inner.calls == 3

    def test_401_is_session_expired_and_not_retried(self) -> None:
        inner = _Inner([_Resp(401)])
        with pytest.raises(zpc.SessionExpiredError):
            zpc.lookup_pnr(_Sess(inner), "X")
        assert inner.calls == 1

    def test_otds_redirect_is_session_expired(self) -> None:
        inner = _Inner([_Resp(200, "dash", url="https://x/otds/index.asp")])
        with pytest.raises(zpc.SessionExpiredError):
            zpc.lookup_pnr(_Sess(inner), "X")
        assert inner.calls == 1

    def test_429_raises_rate_limited_after_attempts(self, monkeypatch) -> None:
        monkeypatch.setattr(zpc, "_backoff_with_jitter", lambda *a, **k: 0.0)
        inner = _Inner([_Resp(429)])
        with pytest.raises(zpc.RateLimitedError):
            zpc.lookup_pnr(_Sess(inner), "X", max_attempts=2)
        assert inner.calls == 2

    def test_network_error_retried_then_succeeds(self, monkeypatch) -> None:
        import requests
        monkeypatch.setattr(zpc, "_backoff_with_jitter", lambda *a, **k: 0.0)
        inner = _Inner([requests.RequestException("boom"),
                        _Resp(200, self._ok("NET001"))])
        d = zpc.lookup_pnr(_Sess(inner), "NET001")
        assert d.pnr_code == "NET001"
        assert inner.calls == 2

    def test_unknown_pnr_not_retried(self) -> None:
        inner = _Inner([_Resp(200, "<html>dashboard only</html>")])
        with pytest.raises(zpc.PNRNotFoundError):
            zpc.lookup_pnr(_Sess(inner), "X")
        assert inner.calls == 1  # NOT_FOUND is terminal, not a transient


class TestLookupMany:
    def test_collects_ok_notfound_error_via_on_result(self, monkeypatch) -> None:
        def fake(_session, code, **_kw):
            if code == "NOPE01":
                raise zpc.PNRNotFoundError("x")
            if code == "ERR001":
                raise zpc.ZenithError("500 boom")
            return parse_dossier_html(_build_dossier_html(pnr=code))

        monkeypatch.setattr(zpc, "lookup_pnr", fake)
        seen: dict = {}
        out = zpc.lookup_many(
            object(), ["AAA111", "NOPE01", "ERR001", "BBB222"],
            concurrency=2, delay_s=0.0, retry_cooldown_s=0,
            on_result=lambda c, d, s: seen.__setitem__(c, s),
        )
        assert set(out) == {"AAA111", "BBB222"}
        assert seen["AAA111"] == "OK"
        assert seen["NOPE01"] == "NOT_FOUND"
        assert seen["ERR001"].startswith("ERROR")

    def test_skip_cached_short_circuits_network(self, monkeypatch) -> None:
        fetched: list = []

        def fake(_session, code, **_kw):
            fetched.append(code)
            return parse_dossier_html(_build_dossier_html(pnr=code))

        monkeypatch.setattr(zpc, "lookup_pnr", fake)
        cached = parse_dossier_html(_build_dossier_html(pnr="HIT001"))
        out = zpc.lookup_many(
            object(), ["HIT001", "MISS01"], concurrency=1, delay_s=0.0,
            skip_cached=lambda c: cached if c == "HIT001" else None,
        )
        assert "HIT001" not in fetched  # served from cache, never fetched
        assert "MISS01" in fetched
        assert out["HIT001"].pnr_code == "HIT001"

    def test_hard_down_attempts_all_no_abort_then_sweeps(self, monkeypatch) -> None:
        # No more hard abort: a fully-down Zenith attempts every PNR, then the retry
        # sweeps re-attempt the failures (still nothing if it never recovers).
        monkeypatch.setattr(zpc, "_backoff_with_jitter", lambda *a, **k: 0.0)
        calls = {"n": 0}

        def boom(_session, _code, **_kw):
            calls["n"] += 1
            raise zpc.ZenithError("504")

        monkeypatch.setattr(zpc, "lookup_pnr", boom)
        codes = [f"P{i:04d}" for i in range(20)]
        out = zpc.lookup_many(object(), codes, concurrency=3, delay_s=0.0,
                              retry_passes=1, retry_cooldown_s=0)
        assert out == {}
        assert calls["n"] == 40                            # 20 main pass + 20 in one sweep

    def test_retry_sweep_recovers_transient_failures(self, monkeypatch) -> None:
        # The 504-storm fix: each PNR 504s on its first attempt, succeeds on the retry.
        monkeypatch.setattr(zpc, "_backoff_with_jitter", lambda *a, **k: 0.0)
        seen: dict = {}

        def flaky(_session, code, **_kw):
            seen[code] = seen.get(code, 0) + 1
            if seen[code] == 1:                            # fail once, then succeed
                raise zpc.ZenithError("504")
            return parse_dossier_html(_build_dossier_html(pnr=code))

        monkeypatch.setattr(zpc, "lookup_pnr", flaky)
        codes = [f"P{i:04d}" for i in range(10)]
        notices: list = []
        out = zpc.lookup_many(object(), codes, concurrency=2, delay_s=0.0,
                              retry_passes=2, retry_cooldown_s=0,
                              on_notice=notices.append)
        assert len(out) == 10                              # all recovered on the sweep
        assert any("retry sweep" in n.lower() for n in notices)

    def test_sweep_does_not_retry_not_found(self, monkeypatch) -> None:
        monkeypatch.setattr(zpc, "_backoff_with_jitter", lambda *a, **k: 0.0)
        calls = {"n": 0}

        def fake(_session, code, **_kw):
            calls["n"] += 1
            raise zpc.PNRNotFoundError("nope")

        monkeypatch.setattr(zpc, "lookup_pnr", fake)
        out = zpc.lookup_many(object(), ["X1"], concurrency=1, delay_s=0.0,
                              retry_passes=2, retry_cooldown_s=0)
        assert out == {} and calls["n"] == 1               # not-found never re-swept

    def test_retry_passes_zero_disables_sweeps(self, monkeypatch) -> None:
        monkeypatch.setattr(zpc, "_backoff_with_jitter", lambda *a, **k: 0.0)
        calls = {"n": 0}

        def fail_once(_session, code, **_kw):
            calls["n"] += 1
            raise zpc.ZenithError("504")

        monkeypatch.setattr(zpc, "lookup_pnr", fail_once)
        out = zpc.lookup_many(object(), ["A", "B"], concurrency=1, delay_s=0.0,
                              retry_passes=0)
        assert out == {} and calls["n"] == 2               # main pass only, no sweep

    def test_mostly_healthy_run_all_attempted(self, monkeypatch) -> None:
        monkeypatch.setattr(zpc, "_backoff_with_jitter", lambda *a, **k: 0.0)
        calls = {"n": 0}

        def mostly_ok(_session, _code, **_kw):
            calls["n"] += 1
            if calls["n"] % 20 == 0:                       # ~5% transient failures
                raise zpc.ZenithError("504")
            return parse_dossier_html(_build_dossier_html(pnr=f"OK{calls['n']:05d}"))

        monkeypatch.setattr(zpc, "lookup_pnr", mostly_ok)
        codes = [f"P{i:05d}" for i in range(80)]
        out = zpc.lookup_many(object(), codes, concurrency=1, delay_s=0.0,
                              retry_passes=0)                # main-pass-only for a clean count
        assert calls["n"] == 80                            # every PNR attempted, no abort
        assert len(out) == 76                              # 4 of 80 failed (the 20ths)

    def test_session_loss_raises_after_checkpointing(self, monkeypatch) -> None:
        def fake(_session, code, **_kw):
            if code == "DEAD01":
                raise zpc.SessionExpiredError("gone")
            return parse_dossier_html(_build_dossier_html(pnr=code))

        monkeypatch.setattr(zpc, "lookup_pnr", fake)
        seen: list = []
        with pytest.raises(zpc.SessionExpiredError):
            zpc.lookup_many(
                object(), ["AAA111", "DEAD01", "BBB222", "CCC333"],
                concurrency=1, delay_s=0.0,
                on_result=lambda c, d, s: seen.append((c, s)),
            )
        # Whatever resolved before the session died was reported (so the
        # caller's cache.put already checkpointed it → resume-safe).
        assert ("AAA111", "OK") in seen

    def test_empty_input_returns_empty(self) -> None:
        assert zpc.lookup_many(object(), [], concurrency=3) == {}
