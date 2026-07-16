"""Customer lookup by NAME (CustomerSearch.ashx) — parser, resolution, plumbing.

The name search GETs CustomerSearch.ashx?CustomerName=<name>, whose 302 chain
lands on a Customer.aspx results page (masterUserControl=UsrCustomerSearch)
listing each match's Customer Code / Name / Phone / Type / E-mail. Fixtures
mirror that markup exactly (mobile label divs, duplicated href quirk on the
edit anchor) with synthetic data — no real customers.

Also covers the two latent issues fixed alongside: the cache dropping agency
fields on round-trip, and the Excel ID reader mangling names (space-stripping).
"""

from __future__ import annotations

from pathlib import Path

import pytest

from src import zenith_client
from src.zenith_client import (
    CustomerAmbiguousError,
    CustomerNotFoundError,
    CustomerRecord,
    CustomerSearchMatch,
    LookupResult,
    SessionExpiredError,
    ZenithError,
    ZenithSession,
    parse_customer_search_html,
)


def _result_row(cid: str, name: str, phone: str, ctype: str, email: str) -> str:
    # Mirrors the real row: value inside xs-right div, label in xs-left div,
    # and the edit anchor's doubled href attribute (as captured in the HAR).
    return f"""
    <tr>
      <td class="center th-alert"><div class="visible-xs-inline xs-left">Alert</div></td>
      <td class="center th-code">
        <div class="visible-xs-inline xs-left">Customer Code</div>
        <div class="visible-xs-inline xs-right">
          <a class="btn btn-edit" type="Button"
             href="/TTIDotNet/Transport/TransportNetBO2/Sales2/CustomerViews/TravelAgency.ashx?IdCustomer={cid}"
             href="#">{cid}</a>
        </div>
      </td>
      <td class="center th-name">
        <div class="visible-xs-inline xs-left">Name</div>
        <div class="visible-xs-inline xs-right"><span title="{name}">{name}</span></div>
      </td>
      <td class="center th-phone">
        <div class="visible-xs-inline xs-left">Phone</div>
        <div class="visible-xs-inline xs-right">{phone}</div>
      </td>
      <td class="center th-type">
        <div class="visible-xs-inline xs-left">Type</div>
        <div class="visible-xs-inline xs-right">{ctype}</div>
      </td>
      <td class="center th-email">
        <div class="visible-xs-inline xs-left">E-mail</div>
        <div class="visible-xs-inline xs-right">{email}</div>
      </td>
    </tr>"""


def _results_page(*rows: str) -> str:
    return (
        '<html><body><div id="mUsrMain_UsrCustomerSearch">'
        '<input name="mUsrMain$UsrCustomerSearch$txtName" value="">'
        f'<table><thead><tr><th>Customer Code</th></tr></thead>'
        f'<tbody>{"".join(rows)}</tbody></table></div></body></html>'
    )


_ROW_A = _result_row("11111111", "SKY HIGH TRAVELS &amp; TOURS", "+8801700000001",
                     "Travel Agency", "skyhigh@example.com")
_ROW_B = _result_row("22222222", "SKY HIGH TRAVELS", "+8801700000002",
                     "Travel Agency", "skyhightravels@example.com")
_TWO_MATCHES = _results_page(_ROW_A, _ROW_B)
_ONE_MATCH = _results_page(_ROW_A)
_NO_MATCHES = _results_page()


# --- pure parser ----------------------------------------------------------------

def test_parse_search_rows_all_fields():
    out = parse_customer_search_html(_TWO_MATCHES)
    assert [m.customer_id for m in out] == ["11111111", "22222222"]
    a = out[0]
    assert a.name == "SKY HIGH TRAVELS & TOURS"       # entity decoded, label stripped
    assert a.phone == "+8801700000001"
    assert a.customer_type == "Travel Agency"
    assert a.email == "skyhigh@example.com"


def test_parse_search_no_matches_returns_empty():
    assert parse_customer_search_html(_NO_MATCHES) == []


def test_parse_search_rejects_non_results_page():
    # A degraded/error page must raise (retryable) — NOT read as "no matches".
    with pytest.raises(ZenithError):
        parse_customer_search_html("<html>ERROR PAGE</html>")


def test_parse_search_dedups_repeated_anchor_ids():
    # The edit anchor carries the id twice (doubled href) — one match per row.
    assert len(parse_customer_search_html(_ONE_MATCH)) == 1


# --- session-level search + resolution --------------------------------------------

_CUSTOMER_PAGE = """
<input name="m$txtCompanyName" value="SKY HIGH TRAVELS &amp; TOURS">
<input name="m$txtIATANumber" value="1234567">
<input name="m$txtEmail" value="skyhigh@example.com">
"""


class _FakeHttp:
    """requests.Session stand-in routing by URL substring."""

    def __init__(self, routes: dict[str, object]) -> None:
        self.routes = routes            # substring -> text | (status, text) | [seq]
        self.requests: list[tuple[str, dict]] = []

    def get(self, url, params=None, allow_redirects=True, timeout=None, headers=None):
        self.requests.append((url, dict(params or {})))
        item = next((v for k, v in self.routes.items() if k in url), "")
        if isinstance(item, list):
            item = item.pop(0) if item else ""
        status, text = item if isinstance(item, tuple) else (200, item)

        class _R:
            pass
        r = _R()
        r.status_code = status
        r.text = text
        r.url = url
        return r


def _session(routes: dict) -> ZenithSession:
    return ZenithSession(session=_FakeHttp(routes), state_values={}, company_code="X")


def test_search_customers_returns_matches():
    sess = _session({"CustomerSearch.ashx": _TWO_MATCHES})
    out = sess.search_customers("SKY HIGH")
    assert len(out) == 2 and all(isinstance(m, CustomerSearchMatch) for m in out)
    url, params = sess.session.requests[0]
    assert params == {"CustomerName": "SKY HIGH"}


def test_search_customers_session_expired_on_otds_redirect():
    sess = _session({"CustomerSearch.ashx": _ONE_MATCH})
    sess.session.routes = {"CustomerSearch.ashx": _ONE_MATCH}

    class _OtdsHttp(_FakeHttp):
        def get(self, url, **kw):
            r = super().get(url, **kw)
            r.url = "https://z/otds/index.asp"       # bounced to login
            return r
    sess = ZenithSession(session=_OtdsHttp({"CustomerSearch.ashx": ""}),
                         state_values={}, company_code="X")
    with pytest.raises(SessionExpiredError):
        sess.search_customers("SKY HIGH")


def test_search_customers_retries_degraded_page_then_succeeds():
    sess = _session({"CustomerSearch.ashx": ["<html>degraded</html>", _ONE_MATCH]})
    out = sess.search_customers("SKY HIGH")
    assert len(out) == 1 and len(sess.session.requests) == 2


def test_query_digits_goes_straight_to_id_lookup():
    sess = _session({"FinalCustomer.ashx": _CUSTOMER_PAGE})
    rec = sess.fetch_customer_by_query("  1111-1111 ".replace("-", "").strip())
    assert rec.customer_id == "11111111"
    assert all("CustomerSearch" not in u for u, _ in sess.session.requests)


def test_query_name_single_match_resolves_and_fetches():
    sess = _session({
        "CustomerSearch.ashx": _ONE_MATCH,
        "FinalCustomer.ashx": _CUSTOMER_PAGE,
    })
    rec = sess.fetch_customer_by_query("SKY HIGH TRAVELS & TOURS")
    assert rec.customer_id == "11111111"              # the REAL numeric id
    assert rec.company_name == "SKY HIGH TRAVELS & TOURS"


def test_query_name_multi_match_unique_exact_wins():
    sess = _session({
        "CustomerSearch.ashx": _TWO_MATCHES,
        "FinalCustomer.ashx": _CUSTOMER_PAGE,
    })
    rec = sess.fetch_customer_by_query("sky high travels")   # exact (case-insensitive) = row B
    fetched = [p.get("IdCustomer") for u, p in sess.session.requests if "FinalCustomer" in u]
    assert fetched == ["22222222"]


def test_query_name_ambiguous_lists_candidates():
    sess = _session({"CustomerSearch.ashx": _TWO_MATCHES})
    with pytest.raises(CustomerAmbiguousError) as ei:
        sess.fetch_customer_by_query("SKY")               # matches both, no exact
    msg = str(ei.value)
    assert "AMBIGUOUS" in msg and "11111111" in msg and "22222222" in msg


def test_query_name_no_match_raises_not_found():
    sess = _session({"CustomerSearch.ashx": _NO_MATCHES})
    with pytest.raises(CustomerNotFoundError):
        sess.fetch_customer_by_query("NOBODY AT ALL")


# --- fetch_many integration --------------------------------------------------------

class _FakePlanSession:
    def __init__(self, outcomes: dict) -> None:
        self.outcomes = outcomes
        self.calls: list[str] = []

    def fetch_customer_by_query(self, q, **_kw):
        self.calls.append(q)
        out = self.outcomes[q]
        if isinstance(out, Exception):
            raise out
        return out


def test_fetch_many_name_ambiguous_is_error_and_never_retried():
    sess = _FakePlanSession({
        "SKY": CustomerAmbiguousError("AMBIGUOUS: 2 customers match 'SKY' — …"),
        "33333333": CustomerRecord(customer_id="33333333"),
    })
    results = zenith_client.fetch_many(
        sess, ["SKY", "33333333"], concurrency=1, delay_s=0.0,
        retry_passes=2, retry_cooldown_s=0.0)
    by_id = {r.customer_id: r for r in results}
    assert by_id["SKY"].status == zenith_client.STATUS_ERROR
    assert "AMBIGUOUS" in by_id["SKY"].error
    assert by_id["33333333"].status == zenith_client.STATUS_OK
    assert sess.calls.count("SKY") == 1               # retry sweep skipped it


# --- cache round-trip (regression: agency fields + resolved id) --------------------

def test_cache_roundtrips_agency_fields_and_resolved_id(tmp_path):
    from src.zenith_cache import ZenithCache
    cache = ZenithCache(tmp_path / "z.sqlite")
    rec = CustomerRecord(
        customer_id="11111111", customer_type="Travel Agency",
        company_name="SKY HIGH TRAVELS & TOURS", administrative_name="SKY HIGH",
        iata_number="1234567", email="skyhigh@example.com")
    cache.save_result(LookupResult(
        customer_id="SKY HIGH TRAVELS & TOURS",       # name-keyed row
        status=zenith_client.STATUS_OK, record=rec, checked_at="2026-07-16 11:00:00"))
    (out,) = list(cache.iter_all())
    assert out.customer_id == "SKY HIGH TRAVELS & TOURS"
    assert out.record.customer_id == "11111111"       # resolved id survives
    assert out.record.company_name == "SKY HIGH TRAVELS & TOURS"
    assert out.record.iata_number == "1234567"        # was silently dropped before
    assert out.record.customer_type == "Travel Agency"


def test_cache_migrates_old_schema_in_place(tmp_path):
    import sqlite3
    from src.zenith_cache import ZenithCache
    db = tmp_path / "old.sqlite"
    conn = sqlite3.connect(db)                        # a pre-migration cache file
    conn.executescript("""
        CREATE TABLE zenith_customers (
            customer_id TEXT PRIMARY KEY, status TEXT NOT NULL DEFAULT '',
            title TEXT NOT NULL DEFAULT '', first_name TEXT NOT NULL DEFAULT '',
            middle_name TEXT NOT NULL DEFAULT '', last_name TEXT NOT NULL DEFAULT '',
            date_of_birth TEXT NOT NULL DEFAULT '', email TEXT NOT NULL DEFAULT '',
            home_phone TEXT NOT NULL DEFAULT '', home_phone_international TEXT NOT NULL DEFAULT '',
            mobile_phone TEXT NOT NULL DEFAULT '', mobile_phone_international TEXT NOT NULL DEFAULT '',
            office_phone TEXT NOT NULL DEFAULT '', nationality TEXT NOT NULL DEFAULT '',
            language TEXT NOT NULL DEFAULT '', spoken_language TEXT NOT NULL DEFAULT '',
            address TEXT NOT NULL DEFAULT '', city TEXT NOT NULL DEFAULT '',
            postal_code TEXT NOT NULL DEFAULT '', country TEXT NOT NULL DEFAULT '',
            registration_date TEXT NOT NULL DEFAULT '',
            error TEXT NOT NULL DEFAULT '', checked_at TEXT NOT NULL DEFAULT '');
        INSERT INTO zenith_customers (customer_id, status, email)
        VALUES ('99999999', 'OK', 'kept@example.com');
    """)
    conn.commit(); conn.close()
    cache = ZenithCache(db)                            # triggers migration
    (row,) = list(cache.iter_all())
    assert row.record.email == "kept@example.com"      # old data preserved
    assert row.record.iata_number == ""                # new column defaulted
    cache.save_result(LookupResult(                    # new columns writable
        customer_id="99999999", status=zenith_client.STATUS_OK,
        record=CustomerRecord(customer_id="99999999", iata_number="7654321")))
    (row,) = list(cache.iter_all())
    assert row.record.iata_number == "7654321"


# --- Excel plumbing -----------------------------------------------------------------

def test_read_zenith_ids_preserves_names_and_cleans_ids(tmp_path):
    from openpyxl import Workbook
    from src.excel_io import read_zenith_ids
    wb = Workbook(); ws = wb.active
    ws.append(["Customer"])
    ws.append([12492298.0])                    # Excel float id
    ws.append(["1249-2298"])                   # pasted with hyphen
    ws.append(["  AJMERI TOUR  AND TRAVELS "])  # a NAME — spaces must survive
    p = tmp_path / "in.xlsx"; wb.save(p)
    ids = read_zenith_ids(p, "Sheet", "Customer")
    assert ids == ["12492298", "12492298", "AJMERI TOUR AND TRAVELS"]


def test_write_zenith_results_resolved_id_and_search_input(tmp_path):
    from openpyxl import load_workbook
    from src.excel_io import write_zenith_results
    rec = CustomerRecord(customer_id="11111111", customer_type="Travel Agency",
                         company_name="SKY HIGH TRAVELS & TOURS")
    rows = [
        LookupResult(customer_id="SKY HIGH TRAVELS & TOURS",
                     status=zenith_client.STATUS_OK, record=rec),
        LookupResult(customer_id="22222222", status=zenith_client.STATUS_OK,
                     record=CustomerRecord(customer_id="22222222")),
    ]
    p = tmp_path / "out.xlsx"
    write_zenith_results(p, rows)
    ws = load_workbook(p).active
    header = [c.value for c in ws[1]]
    assert header[0] == "Customer ID" and header[-1] == "Search Input"
    name_row = [c.value for c in ws[2]]
    assert name_row[0] == "11111111"                       # REAL id up front
    assert name_row[-1] == "SKY HIGH TRAVELS & TOURS"      # query kept for traceability
    id_row = [c.value for c in ws[3]]
    assert id_row[0] == "22222222" and (id_row[-1] in ("", None))


def test_parse_search_accepts_person_row_handler():
    # Person rows may use FinalCustomer.ashx in the edit anchor — must match too.
    person = _result_row("44444444", "RAHIM UDDIN", "+8801700000004",
                         "Final Customer", "rahim@example.com").replace(
        "TravelAgency.ashx", "FinalCustomer.ashx")
    out = parse_customer_search_html(_results_page(person))
    assert [m.customer_id for m in out] == ["44444444"]
    assert out[0].customer_type == "Final Customer"
