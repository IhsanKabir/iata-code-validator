"""Tests for OEP Excel writers — round-trip the workbook in-memory."""

from pathlib import Path

from openpyxl import load_workbook

from src.excel_io import (
    write_oep_full_report,
    write_oep_pivot_report,
    write_oep_timeseries_report,
)
from src.oep_client import (
    CategoryTotal,
    CountryClearance,
    CountryTotal,
    DivisionClearance,
    DivisionTotal,
    GenderBreakdown,
)


def test_write_oep_timeseries_report_creates_both_sheets(tmp_path):
    path = tmp_path / "ts.xlsx"
    months = ["2025-01", "2025-02", "2025-03"]
    series = {
        "Saudi Arabia": [100, 150, 200],
        "Qatar": [10, 20, 30],
    }
    write_oep_timeseries_report(
        path,
        date_from="2025-01-01", date_to="2025-03-31",
        months=months, series=series,
    )
    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"Time Series", "Raw"}

    ts = wb["Time Series"]
    # Row 1: date range header; row 2: blank; row 3: column headers
    assert "2025-01-01" in str(ts["A1"].value)
    assert ts["A3"].value == "Country"
    assert ts["B3"].value == "2025-01"
    # Saudi Arabia has the larger total — should sort to row 4
    assert ts["A4"].value == "Saudi Arabia"
    assert ts["B4"].value == 100
    # Columns: A=Country, B/C/D=months, E=Total
    last_col = 2 + len(months)
    assert ts.cell(row=4, column=last_col).value == 450

    raw = wb["Raw"]
    assert raw["A1"].value == "Year-Month"
    assert raw.max_row == 1 + 2 * 3  # header + 2 countries × 3 months


def test_write_oep_pivot_report_includes_totals_row(tmp_path):
    path = tmp_path / "pivot.xlsx"
    divisions = ["Chattagram", "Dhaka", "Sylhet"]
    countries = ["Saudi Arabia", "Qatar"]
    table = {
        ("Chattagram", "Saudi Arabia"): 1000,
        ("Chattagram", "Qatar"): 200,
        ("Dhaka", "Saudi Arabia"): 800,
        ("Dhaka", "Qatar"): 300,
        ("Sylhet", "Saudi Arabia"): 50,
    }
    write_oep_pivot_report(
        path,
        date_from="2025-01-01", date_to="2025-06-30",
        divisions=divisions, countries=countries, table=table,
    )
    wb = load_workbook(path)
    ws = wb["Pivot"]
    # Header row 3
    assert ws["A3"].value == "Division"
    assert ws["B3"].value == "Saudi Arabia"
    # Final row should be the Total row
    total_row = ws.max_row
    assert ws.cell(row=total_row, column=1).value == "Total"
    # Total Saudi = 1000+800+50 = 1850
    assert ws.cell(row=total_row, column=2).value == 1850
    # Total Qatar = 200+300+0 = 500
    assert ws.cell(row=total_row, column=3).value == 500
    # Grand total in last column = 2350
    assert ws.cell(row=total_row, column=4).value == 2350

    raw = wb["Raw"]
    assert raw["A1"].value == "Country"
    # Zero cells aren't written
    assert raw.max_row == 1 + 5  # 5 non-zero entries


def test_write_oep_full_report_has_every_sheet(tmp_path):
    path = tmp_path / "full.xlsx"
    raw_country = [
        CountryClearance(154, "Saudi Arabia", "Driver", 100),
        CountryClearance(154, "Saudi Arabia", "Cook", 50),
        CountryClearance(144, "Qatar", "Driver", 80),
    ]
    raw_division = [DivisionClearance("Dhaka", "Dhaka", 200)]
    write_oep_full_report(
        path,
        date_from="2025-01-01", date_to="2025-12-31", gender_id="",
        country_summary=[
            CountryTotal("Saudi Arabia", 150, 2),
            CountryTotal("Qatar", 80, 1),
        ],
        category_summary=[CategoryTotal("Driver", 180, 2), CategoryTotal("Cook", 50, 1)],
        division_summary=[DivisionTotal("Dhaka", 200, 1)],
        gender_summary=[GenderBreakdown("Saudi Arabia", 130, 15, 5, 150)],
        raw_country=raw_country,
        raw_division=raw_division,
        months=["2025-01", "2025-02"],
        series={"Saudi Arabia": [60, 90], "Qatar": [40, 40]},
        divisions=["Dhaka"],
        pivot_countries=["Saudi Arabia", "Qatar"],
        table={("Dhaka", "Saudi Arabia"): 150, ("Dhaka", "Qatar"): 80},
        country_labels=["Saudi Arabia", "Qatar"],
    )
    wb = load_workbook(path)
    assert wb.sheetnames == [
        "Cover", "By Country", "By Division", "By Category",
        "By Gender", "Time Series", "Country x Division",
        "Raw Country×Category", "Raw Division",
    ]
    # Cover holds the headline numbers
    cover = wb["Cover"]
    headline_total_row = next(
        r for r in cover.iter_rows(values_only=True)
        if r and r[0] == "Total workers cleared"
    )
    assert headline_total_row[1] == 230  # 150 + 80
