"""Which sheet does 'Append to Ordered Report' write into?

Field bug: appending to a workbook with no "Ordered Data" sheet silently fell back
to `wb.active` and wrote nine 16-column date blocks onto an unrelated sheet (a
hand-built summary that happened to be active). The run looked successful, the
real per-year sheets were untouched, and the target sheet was left 144 columns
wider. These tests pin the replacement: find a genuine cross-tab load sheet, or
raise — never guess.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from src.excel_io import (
    ORDERED_REPORT_SHEET,
    _find_ordered_sheet,
    _looks_like_ordered_sheet,
)


def _ordered_sheet(wb, title: str, dates: list[str]):
    """A cross-tab load sheet: dates across row 1, per-date labels in row 2."""
    ws = wb.create_sheet(title)
    ws.cell(2, 1, "Flight No")
    ws.cell(2, 2, "Leg/Sector")
    col = 3
    for d in dates:
        ws.cell(1, col, f"{d} (Tuesday)")
        for i, lbl in enumerate(("Capacity", "STD", "Flown Seats", "Load Factor")):
            ws.cell(2, col + i, lbl)
        col += 4
    ws.cell(3, 1, "BS101")
    ws.cell(3, 2, "DAC-CGP")
    return ws


def _seats_grid(wb, title: str):
    """A generated grid — row 2 ALSO starts 'Flight No'/'Leg/Sector', but the dates
    live in row 2 and there are no per-date 'Capacity' labels."""
    ws = wb.create_sheet(title)
    ws.cell(1, 1, "2026 — SEATS FLOWN per flight per day")
    for j, h in enumerate(["Flight No", "Leg/Sector", "Avg Seats"], 1):
        ws.cell(2, j, h)
    ws.cell(2, 6, "28/07/2026\nTuesday")
    ws.cell(3, 1, "BS101")
    ws.cell(3, 2, "DAC-CGP")
    return ws


def test_recognises_a_real_ordered_sheet():
    wb = Workbook()
    ws = _ordered_sheet(wb, "2026", ["28/07/2026", "29/07/2026"])
    assert _looks_like_ordered_sheet(ws)


def test_rejects_a_generated_seats_grid():
    # the exact shape that made the old fallback dangerous
    wb = Workbook()
    assert not _looks_like_ordered_sheet(_seats_grid(wb, "2026 Seats"))


def test_rejects_an_unrelated_summary_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Load Factor (Seats)"
    ws.cell(1, 1, "LOAD FACTOR IN SEATS")
    ws.cell(2, 1, "Counted in SEATS, not a percentage.")
    ws.cell(5, 1, "Sector")
    assert not _looks_like_ordered_sheet(ws)


def test_canonical_sheet_wins_when_present():
    wb = Workbook()
    _ordered_sheet(wb, "2026", ["28/07/2026"])
    _ordered_sheet(wb, ORDERED_REPORT_SHEET, ["01/01/2026"])
    assert _find_ordered_sheet(wb, {"28/07/2026"}).title == ORDERED_REPORT_SHEET


def test_routes_to_the_sheet_holding_that_year():
    wb = Workbook()
    wb.remove(wb.active)
    _ordered_sheet(wb, "2025", ["03/05/2025", "04/05/2025"])
    _ordered_sheet(wb, "2026", ["28/07/2026"])
    assert _find_ordered_sheet(wb, {"29/07/2026", "05/08/2026"}).title == "2026"
    assert _find_ordered_sheet(wb, {"06/05/2025"}).title == "2025"


def test_ignores_generated_sheets_when_choosing():
    wb = Workbook()
    wb.remove(wb.active)
    _seats_grid(wb, "2026 Seats")          # would have been picked if active
    _ordered_sheet(wb, "2026", ["28/07/2026"])
    assert _find_ordered_sheet(wb, {"29/07/2026"}).title == "2026"


def test_raises_instead_of_writing_to_the_active_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Load Factor (Seats)"       # the sheet that actually got corrupted
    ws.cell(1, 1, "LOAD FACTOR IN SEATS")
    with pytest.raises(ValueError) as ei:
        _find_ordered_sheet(wb, {"28/07/2026"})
    assert ORDERED_REPORT_SHEET in str(ei.value)
