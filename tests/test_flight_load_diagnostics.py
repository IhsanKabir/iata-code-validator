"""Tests for the per-flight load-factor diagnostics engine.

Synthetic HistoryEvent records (no files, no network) are fed through the real
engine so counts, the coupon-ID hold clock, flags, and the comparison pass are
all exercised end-to-end.
"""

from __future__ import annotations

from datetime import datetime, timedelta

from src.flight_load_diagnostics import (
    FLAG_REASONS,
    compare_flights,
    diagnose_flight,
)
from src.flight_inspection_excel import write_flight_inspection
from src.zenith_history_parser import Agent, HistoryEvent, parse_flight

FLIGHT = "BS 217 DAC BKK 27/06/2026 10:00"
DEP = datetime(2026, 6, 27, 10, 0)


def ev(desc, ts=None, pnr="", flight=FLIGHT, agent_id="AGT1"):
    return HistoryEvent(
        source_file="t.xls", row_index=0, raw_date="", raw_created_by=agent_id,
        raw_description=desc, event_type="", pnr=pnr, customer="",
        raw_flight=flight, passenger="", timestamp=ts,
        agent=Agent(raw=agent_id, display_name=agent_id, user_id=agent_id, department=""),
        flight=parse_flight(flight),
    )


def flown(cid, pnr, book_time, fare="I"):
    """Full lifecycle of one flown passenger: booking -> check-in -> board -> flown."""
    return [
        ev(f"Coupon information: [{cid}] 779241100000{cid}C1 BS 217 /{fare} DAC BKK "
           f"27/06/2026 Coupon status :Open", ts=book_time, pnr=pnr),
        ev(f"Check-in of the coupon {cid} : Previous status = I , New status = CK QDCS",
           ts=DEP - timedelta(hours=3), pnr=pnr),
        ev(f"Boarding of the coupon {cid} : Previous status = CK , New status = BD QDCS",
           ts=DEP - timedelta(hours=1), pnr=pnr),
        ev(f"Coupon status change:{cid} Old status:Boarded New status:Flown",
           ts=DEP + timedelta(hours=2), pnr=pnr),
    ]


def test_funnel_counts():
    events = []
    events += flown(1001, "P1", DEP - timedelta(days=20))
    events += flown(1002, "P2", DEP - timedelta(days=10))
    events += flown(1003, "P3", DEP - timedelta(days=5))
    events += [ev("No Show of the coupon 1009 : Previous status = I , New status = NS",
                  ts=DEP - timedelta(hours=4), pnr="P9")]
    events += [ev("Segment Cancelled: BS217 27Jun DACBKK", ts=DEP - timedelta(days=3), pnr="P8")]
    events += [ev("Coupon information: [2001] BS 217 /O DAC BKK 27/06/2026 "
                  "Coupon status :Option->Cancelled", ts=DEP - timedelta(days=8), pnr="P7")]
    events += [ev("Seat(s) 30A-30F blocked. QDCS", ts=DEP - timedelta(hours=2))]

    d = diagnose_flight(events, source_name="t")
    assert d is not None
    assert d.flown == 3 and d.checked_in == 3 and d.boarded == 3
    assert d.no_shows == 1
    assert d.cancellations == 1
    assert d.held_coupons == 1
    assert d.route == "DAC-BKK" and d.flight_number == "BS217"
    assert d.capacity_est == 180          # row 30 x 6
    assert d.load_factor_est == round(100 * 3 / 180, 1)


def test_no_flown_returns_none():
    # only a held option, nobody flew -> no load story
    events = [ev("Coupon information: [2001] BS 217 /O DAC BKK 27/06/2026 "
                 "Coupon status :Option->Cancelled", ts=DEP - timedelta(days=2), pnr="P7")]
    assert diagnose_flight(events, source_name="t") is None


def test_hold_duration_via_clock():
    # two flown anchors define the creation clock; a held coupon midway in ID
    # space interpolates to a creation time -> known hold.
    events = []
    events += flown(1000, "A", DEP - timedelta(days=10))   # id 1000 @ -10d
    events += flown(3000, "B", DEP - timedelta(days=2))    # id 3000 @ -2d
    # held id 2000 (midpoint) cancelled at -4d -> creation interp ~ -6d -> hold ~2d
    events += [ev("Coupon information: [2000] BS 217 /O DAC BKK 27/06/2026 "
                  "Coupon status :Option->Cancelled", ts=DEP - timedelta(days=4), pnr="H")]
    d = diagnose_flight(events, source_name="t")
    assert d.held_coupons == 1
    assert abs(d.held_median_hold_h - 48.0) < 8.0      # ~2 days, within clock noise


def test_late_curve_flag():
    events = []
    # everyone books in the final week -> LATE_CURVE
    for i in range(4):
        events += flown(1000 + i, f"P{i}", DEP - timedelta(days=2))
    events += [ev("Seat(s) 10A-10F blocked. QDCS", ts=DEP)]
    d = diagnose_flight(events, source_name="t")
    assert d.pct_booked_final_week == 100.0
    assert "LATE_CURVE" in d.flags


def test_low_lf_flag_and_reasons_exist():
    events = []
    for i in range(2):
        events += flown(1000 + i, f"P{i}", DEP - timedelta(days=40))
    events += [ev("Seat(s) 30A-30F blocked. QDCS", ts=DEP)]   # cap 180, 2 flown -> ~1%
    d = diagnose_flight(events, source_name="t")
    assert "VERY_LOW_LF" in d.flags
    for fl in d.flags:                       # every emitted flag is documented
        assert fl in FLAG_REASONS


def test_compare_flights_below_route_norm():
    def make(nflown, date):
        events = []
        for i in range(nflown):
            events += flown(1000 + i, f"P{i}", DEP - timedelta(days=30),)
        events += [ev("Seat(s) 32A-32F blocked. QDCS", ts=DEP,
                      flight=f"BS 217 DAC BKK {date} 10:00")]
        # retag flight cell on every event so route/date are consistent
        return [e for e in (ev(x.raw_description, x.timestamp, x.pnr,
                               f"BS 217 DAC BKK {date} 10:00") for x in events)]
    full = diagnose_flight(make(30, "01/06/2026"), source_name="full")
    full2 = diagnose_flight(make(28, "08/06/2026"), source_name="full2")
    weak = diagnose_flight(make(5, "27/06/2026"), source_name="weak")
    out = compare_flights([full, full2, weak])
    weak_out = next(d for d in out if d.source_name == "weak")
    assert "BELOW_ROUTE_NORM" in weak_out.flags


def test_write_inspection_excel_smoke(tmp_path):
    events = []
    for i in range(3):
        events += flown(1000 + i, f"P{i}", DEP - timedelta(days=2))
    events += [ev("Seat(s) 30A-30F blocked. QDCS", ts=DEP)]
    d = diagnose_flight(events, source_name="27 JUN DAC-BKK")
    out = tmp_path / "inspection.xlsx"
    n = write_flight_inspection(out, [d])
    assert n == 1 and out.exists()
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "Flights" in wb.sheetnames and "Flag legend" in wb.sheetnames
    # flagged flight gets a detail sheet
    assert len(wb.sheetnames) >= 3
