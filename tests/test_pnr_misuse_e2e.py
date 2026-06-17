"""End-to-end test of the PNR misuse pipeline:

    real-format ModificationHistory HTML  ->  parse_history_file (the REAL parser)
        ->  run_pnr_misuse_audit (the analyzer)  ->  write_zenith_pnr_misuse_audit (Excel)

No live Zenith needed: we synthesize a ModificationHistory file in the exact HTML the
parser expects (coupon-status transitions), plant known misuse patterns, and assert the
flags + the produced workbook. This exercises every layer except the network download.
"""
from __future__ import annotations

import textwrap

from openpyxl import load_workbook

from src.excel_io import build_zenith_pnr_misuse_output_path, write_zenith_pnr_misuse_audit
from src.zenith_history_parser import parse_history_file
from src.zenith_pnr_history_analyzer import run_pnr_misuse_audit


def _row(date, by, desc, pnr, flight="BS 100 DAC DXB 01/01/2026 19:00",
         passenger="PAX", etype="Ticket Modification") -> str:
    return (f"<tr><td>{date}</td><td>{by}</td><td>{desc}</td><td>{etype}</td>"
            f"<td>{pnr}</td><td></td><td>{flight}</td><td>{passenger}</td></tr>")


def _coupon(ticket, rbd, status) -> str:
    return (f"Coupon information: [x] {ticket}C1 BS 100 /{rbd} DAC DXB 01/01/2026 "
            f"| Coupon status :{status}")


def _fixture(*rows: str) -> str:
    return textwrap.dedent(f"""
    <html><body><table>
      <thead><tr><th>Date</th><th>Created by</th><th>Description</th><th>Type</th>
        <th>PNR</th><th>Customer</th><th>Flight</th><th>Passenger</th></tr></thead>
      <tbody>{''.join(rows)}</tbody>
    </table></body></html>
    """).strip()


def _make_corpus(tmp_path):
    html = _fixture(
        # 1. Refund of a FLOWN coupon (critical) — agent flyagent
        _row("01/01/2026 21:00", "Fly Agent (flyagent)",
             _coupon("7792000000001", "Y", "Issued->Flown"), "FLOWN1"),
        _row("02/01/2026 09:00", "Fly Agent (flyagent)",
             _coupon("7792000000001", "Y", "Flown->Refunded"), "FLOWN1"),
        # 2. Self-refund / segregation-of-duties — sodagent issues AND refunds T2
        _row("01/01/2026 10:00", "Sod Agent (sodagent)",
             _coupon("7792000000002", "T", "Option->Issued"), "SELF01"),
        _row("01/01/2026 13:00", "Sod Agent (sodagent)",
             _coupon("7792000000002", "T", "Issued->Refunded"), "SELF01"),
        # 3. Off-hours refund at 02:15 — offagent
        _row("02/01/2026 02:15", "Off Agent (offagent)",
             _coupon("7792000000003", "G", "Issued->Refunded"), "OFFH01"),
        # 4. Steep class downgrade Y->G — dgagent
        _row("01/01/2026 09:00", "Dg Agent (dgagent)",
             _coupon("7792000000004", "Y", "Option->Issued"), "DOWN01"),
        _row("01/01/2026 15:00", "Dg Agent (dgagent)",
             _coupon("7792000000004", "G", "Issued->Issued"), "DOWN01"),
        # 5. CLEAN PNR — single business-hours issue, must NOT flag
        _row("01/01/2026 11:00", "Clean Agent (cleanagent)",
             _coupon("7792000000005", "T", "Option->Issued"), "CLEAN1"),
    )
    p = tmp_path / "ModificationHistory e2e.xls"
    p.write_bytes(html.encode("utf-8"))
    return parse_history_file(p)


def test_pipeline_flags_planted_patterns(tmp_path) -> None:
    events = _make_corpus(tmp_path)
    assert len(events) == 8                         # the real parser read every row

    report = run_pnr_misuse_audit(events)
    detectors = {f.detector for f in report.flags}
    assert "refund_of_flown" in detectors
    assert "self_refund_sod" in detectors
    assert "off_hours_value" in detectors
    assert "downgrade" in detectors

    # The clean PNR is never named in a flag.
    assert all(f.pnr != "CLEAN1" for f in report.flags)

    # The critical refund-of-flown PNR tops (or near-tops) the PNR risk worklist.
    pnr_rows = [r for r in report.risk_worklist if r.grain == "pnr"]
    assert pnr_rows and pnr_rows[0].entity == "FLOWN1"


def test_pipeline_writes_valid_workbook(tmp_path) -> None:
    events = _make_corpus(tmp_path)
    report = run_pnr_misuse_audit(events)
    out = build_zenith_pnr_misuse_output_path(tmp_path / "out")
    write_zenith_pnr_misuse_audit(out, report)

    assert out.is_file()
    wb = load_workbook(out)
    assert wb.sheetnames == ["Cover", "Risk Worklist", "Flags", "Agent Activity"]
    flags_ws = wb["Flags"]
    assert flags_ws.max_row > 1                     # header + at least one flag
    # Agent Activity carries the issuers/refunders we planted.
    agents = {row[0].value for row in wb["Agent Activity"].iter_rows(min_row=2)}
    assert {"flyagent", "sodagent", "offagent", "dgagent", "cleanagent"} <= agents
