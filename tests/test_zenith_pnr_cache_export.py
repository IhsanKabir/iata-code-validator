"""ZenithPNRCache.iter_all + excel_io.write_zenith_pnr_bulk_from_details — the
'cache as source of truth' export that lets a storm-killed run still yield a sheet.
"""
from __future__ import annotations

from openpyxl import load_workbook

from src import excel_io
from src.zenith_pnr_cache import ZenithPNRCache
from src.zenith_pnr_client import PNRDetails


def _pnr(code: str) -> PNRDetails:
    return PNRDetails(
        pnr_code=code, dossier_id="D" + code, customer_name="Agency " + code,
        traveler_surname="X", phone="", payment_method="", pax_count=1,
        pnr_status="Issued", currency="BDT", total_amount="100", total_taxes="10",
        segments=(), fetched_at=None)


def test_cache_iter_all_roundtrip_ordered(tmp_path) -> None:
    c = ZenithPNRCache(tmp_path / "c.sqlite")
    c.put(_pnr("BBB"))
    c.put(_pnr("AAA"))
    alld = c.iter_all()
    assert [d.pnr_code for d in alld] == ["AAA", "BBB"]      # ORDER BY pnr_code
    assert alld[0].customer_name == "Agency AAA"
    assert alld[0].dossier_id == "DAAA"


def test_export_from_details_writes_workbook(tmp_path) -> None:
    out = tmp_path / "out.xlsx"
    n = excel_io.write_zenith_pnr_bulk_from_details(out, [_pnr("AAA"), _pnr("BBB")])
    assert n == 2 and out.is_file()
    wb = load_workbook(out)
    assert "PNR Lookup" in wb.sheetnames
    codes = {row[0].value for row in wb["PNR Lookup"].iter_rows(min_row=2) if row[0].value}
    assert {"AAA", "BBB"} <= codes


def test_export_skips_blank_pnr_codes(tmp_path) -> None:
    out = tmp_path / "out2.xlsx"
    blank = _pnr("")
    n = excel_io.write_zenith_pnr_bulk_from_details(out, [_pnr("AAA"), blank])
    assert n == 1                                            # blank-code detail dropped
