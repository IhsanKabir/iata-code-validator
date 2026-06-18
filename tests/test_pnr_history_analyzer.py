"""Tests for src/zenith_pnr_history_analyzer.py — the PNR misuse audit re-pivoted from
the flight ModificationHistory corpus. All offline against synthetic HistoryEvents.
"""
from __future__ import annotations

from datetime import datetime

from src.zenith_history_parser import Agent, FlightRef, HistoryEvent
from src.zenith_pnr_history_analyzer import (
    classify_action,
    classify_actor,
    run_pnr_misuse_audit,
)


def _agent(uid="agt1", dept="DAC Sales", name="Agent One") -> Agent:
    return Agent(raw=f"{name} ({uid}/{dept})", display_name=name, user_id=uid, department=dept)


def _ag(uid: str, dept: str) -> Agent:
    raw = f"X ({uid}/{dept})" if dept else f"X ({uid})"
    return Agent(raw=raw, display_name="X", user_id=uid, department=dept)


_FLIGHT = FlightRef(raw="BS341 DAC DXB", flight_number="BS341", origin="DAC",
                    destination="DXB", flight_date="15/06/2026", departure_time="20:45")


def _ev(*, pnr="ABCDEF", ticket="7792000000001", new_status="", old_status="", rbd="",
        hour=14, day=15, agent=None, etype="Ticket Modification", desc="evt") -> HistoryEvent:
    ts = datetime(2026, 6, day, hour, 0)
    return HistoryEvent(
        source_file="f.xls", row_index=1, raw_date=ts.strftime("%d/%m/%Y %H:%M"),
        raw_created_by=(agent or _agent()).raw, raw_description=desc, event_type=etype,
        pnr=pnr, customer="C", raw_flight="BS341 DAC DXB", passenger="P",
        timestamp=ts, agent=agent or _agent(), flight=_FLIGHT,
        rbd_class=rbd, old_status=old_status, new_status=new_status, ticket_number=ticket,
    )


def _detectors(report) -> set[str]:
    return {f.detector for f in report.flags}


class TestClassify:
    def test_actions(self) -> None:
        assert classify_action(_ev(new_status="Refunded")) == "refund"
        assert classify_action(_ev(new_status="Voided")) == "void"
        assert classify_action(_ev(new_status="Issued")) == "issue"
        assert classify_action(_ev(new_status="Flown")) == "flown"
        assert classify_action(_ev(new_status="")) == "modify"


class TestDetectors:
    def test_refund_of_flown_is_critical(self) -> None:
        evs = [
            _ev(new_status="Flown", hour=21),                 # coupon flown
            _ev(new_status="Refunded", hour=21, day=16),      # later refunded
        ]
        rep = run_pnr_misuse_audit(evs)
        crit = [f for f in rep.flags if f.detector == "refund_of_flown"]
        assert crit and crit[0].severity == "critical"

    def test_self_refund_segregation_of_duties(self) -> None:
        a = _agent(uid="same1")
        evs = [
            _ev(new_status="Issued", agent=a),
            _ev(new_status="Refunded", agent=a, day=16),
        ]
        rep = run_pnr_misuse_audit(evs)
        sod = [f for f in rep.flags if f.detector == "self_refund_sod"]
        assert sod and sod[0].agent_user_id == "same1"

    def test_off_hours_refund(self) -> None:
        # Hours are GMT (corpus is excel-exported); DAC = +6. 20:00 GMT == 02:00 DAC = off-hours.
        rep = run_pnr_misuse_audit([_ev(new_status="Refunded", hour=20)])
        assert "off_hours_value" in _detectors(rep)

    def test_business_hours_refund_not_off_hours(self) -> None:
        # 06:00 GMT == 12:00 DAC = midday business hours.
        rep = run_pnr_misuse_audit([_ev(new_status="Refunded", hour=6)])
        assert "off_hours_value" not in _detectors(rep)

    def test_off_hours_uses_dhaka_local_not_gmt(self) -> None:
        # Regression: 04:00 GMT == 10:00 DAC must NOT be off-hours (the v1.16.16 bug).
        assert "off_hours_value" not in _detectors(
            run_pnr_misuse_audit([_ev(new_status="Voided", hour=4)]))

    def test_downgrade_flagged_with_severity(self) -> None:
        evs = [
            _ev(rbd="Y", new_status="Issued"),
            _ev(rbd="G", hour=15),                            # Y(0) -> G(8) = steep
        ]
        rep = run_pnr_misuse_audit(evs)
        dg = [f for f in rep.flags if f.detector == "downgrade"]
        assert dg and dg[0].severity == "high"               # >=6 tiers

    def test_repeated_class_change_churn(self) -> None:
        evs = [
            _ev(rbd="Y", new_status="Issued"),
            _ev(rbd="H", hour=15), _ev(rbd="R", hour=16), _ev(rbd="N", hour=17),
        ]
        rep = run_pnr_misuse_audit(evs)
        assert "repeated_class_change" in _detectors(rep)

    def test_refund_void_burst_by_agent(self) -> None:
        a = _agent(uid="burst1")
        evs = [_ev(pnr=f"P{i:04d}", ticket=f"77920000000{i:02d}",
                   new_status="Refunded", agent=a, hour=14) for i in range(8)]
        rep = run_pnr_misuse_audit(evs)
        assert "refund_void_burst" in _detectors(rep)

    def test_system_and_api_logins_excluded(self) -> None:
        sysagent = Agent(raw="System (system)", display_name="System",
                         user_id="system", department="")
        evs = [
            _ev(new_status="Flown", hour=21, agent=sysagent),
            _ev(new_status="Refunded", hour=2, day=16, agent=sysagent),
        ]
        rep = run_pnr_misuse_audit(evs)
        assert rep.flags == ()                               # nothing flagged for system

    def test_clean_pnr_no_flags(self) -> None:
        rep = run_pnr_misuse_audit([_ev(new_status="Issued", hour=11)])
        assert rep.flags == ()


class TestRiskAndTrends:
    def test_corroborated_pnr_ranks_high(self) -> None:
        a = _agent(uid="multi1")
        # One PNR/ticket lit by THREE families: flown->refund (crit), self-refund (high),
        # off-hours (med) — should top the worklist over a single-family PNR.
        hot = [
            _ev(pnr="HOT001", new_status="Issued", agent=a, hour=20),
            _ev(pnr="HOT001", new_status="Flown", agent=a, hour=21, day=15),
            _ev(pnr="HOT001", new_status="Refunded", agent=a, hour=20, day=16),  # 02:00 DAC
        ]
        mild = [_ev(pnr="MILD01", ticket="7792999999999", new_status="Refunded", hour=20)]
        rep = run_pnr_misuse_audit(hot + mild)
        pnr_rows = [r for r in rep.risk_worklist if r.grain == "pnr"]
        assert pnr_rows[0].entity == "HOT001"
        assert len(pnr_rows[0].families) >= 2                # corroboration bonus applied

    def test_agent_activity_counts(self) -> None:
        a = _agent(uid="act1")
        evs = [
            _ev(pnr="P1", new_status="Issued", agent=a),
            _ev(pnr="P2", ticket="7792000000002", new_status="Refunded", agent=a),
            _ev(pnr="P3", ticket="7792000000003", new_status="Voided", agent=a),
        ]
        rep = run_pnr_misuse_audit(evs)
        row = next(r for r in rep.agent_activity if r.agent_user_id == "act1")
        assert row.refunds == 1 and row.voids == 1 and row.issues == 1
        assert row.distinct_pnrs == 3

    def test_report_summary_fields(self) -> None:
        rep = run_pnr_misuse_audit([_ev(new_status="Issued")])
        assert rep.event_count == 1 and rep.pnr_count == 1 and rep.agent_count == 1
        assert rep.date_range[0] is not None


class TestFalsePositiveGuards:
    def test_pnr_fallback_does_not_false_link_across_tickets(self) -> None:
        # Two passengers on one PNR with UNPARSEABLE ticket numbers collapse to a PNR
        # group. Passenger A flew; passenger B refunded. The critical refund_of_flown
        # (and self_refund) must NOT fire — they'd be linking different tickets.
        a, b = _agent(uid="paxA"), _agent(uid="paxB")
        evs = [
            _ev(pnr="SHARED", ticket="", new_status="Issued", agent=a, day=1, hour=9),
            _ev(pnr="SHARED", ticket="", new_status="Flown", agent=a, day=2, hour=8),
            _ev(pnr="SHARED", ticket="", new_status="Issued", agent=b, day=1, hour=9),
            _ev(pnr="SHARED", ticket="", new_status="Refunded", agent=b, day=3, hour=10),
        ]
        rep = run_pnr_misuse_audit(evs)
        dets = {f.detector for f in rep.flags}
        assert "refund_of_flown" not in dets
        assert "self_refund_sod" not in dets
        assert rep.fallback_groups == 1 and rep.real_ticket_groups == 0

    def test_real_ticket_self_refund_still_fires(self) -> None:
        # The same agent issuing AND refunding a REAL ticket must still flag.
        a = _agent(uid="solo1")
        evs = [
            _ev(ticket="7792000000009", new_status="Issued", agent=a),
            _ev(ticket="7792000000009", new_status="Refunded", agent=a, day=16),
        ]
        rep = run_pnr_misuse_audit(evs)
        assert "self_refund_sod" in {f.detector for f in rep.flags}
        assert rep.real_ticket_groups == 1

    def test_excluded_flown_not_attributed_to_human_refunder(self) -> None:
        # A system-set Flown must not make a later human refund a refund_of_flown.
        sysagent = Agent(raw="System (system)", display_name="System",
                         user_id="system", department="")
        human = _agent(uid="human1")
        evs = [
            _ev(ticket="7792000000010", new_status="Issued", agent=human, hour=9),
            _ev(ticket="7792000000010", new_status="Flown", agent=sysagent, hour=21, day=15),
            _ev(ticket="7792000000010", new_status="Refunded", agent=human, hour=10, day=16),
        ]
        rep = run_pnr_misuse_audit(evs)
        assert "refund_of_flown" not in {f.detector for f in rep.flags}

    def test_none_timestamps_do_not_crash(self) -> None:
        import dataclasses
        evs = [
            _ev(new_status="Issued"),
            dataclasses.replace(_ev(new_status="Refunded", hour=2), timestamp=None),
        ]
        rep = run_pnr_misuse_audit(evs)        # must not raise
        assert isinstance(rep.flags, tuple)

    def test_corpus_coverage_counts_flown(self) -> None:
        rep = run_pnr_misuse_audit([
            _ev(ticket="7792000000011", new_status="Flown"),
            _ev(ticket="7792000000011", new_status="Refunded", day=16),
        ])
        assert rep.flown_events == 1

    def test_zero_flag_report_exports_valid_workbook(self, tmp_path) -> None:
        from openpyxl import load_workbook
        from src.excel_io import write_zenith_pnr_misuse_audit
        rep = run_pnr_misuse_audit([_ev(new_status="Issued", hour=11)])
        assert rep.flags == ()
        out = tmp_path / "zero.xlsx"
        write_zenith_pnr_misuse_audit(out, rep)
        wb = load_workbook(out)
        assert wb.sheetnames == ["Cover", "Risk Worklist", "Flags", "Agent Activity"]
        assert wb["Flags"].max_row == 1        # header only, no crash on empty flags


class TestActorType:
    def test_internal_office_codes(self) -> None:
        for dept in ("DAC-02 Customer Service", "BO-3 Revenue Management",
                     "ZYL-2 Sylhet City", "CXB-1 Cox's Bazar", "DAC-17 Uttara USBA-Office"):
            assert classify_actor(_ag("staff123", dept)) == "internal"

    def test_gds_vendors(self) -> None:
        assert classify_actor(_ag("Galileo 1GBS1G", "Galileo 1G 1G")) == "gds"
        assert classify_actor(_ag("AbacusBS1B", "Abacus 1B")) == "gds"
        assert classify_actor(_ag("Sabre 1SBS1S", "Sabre 1S 1S")) == "gds"

    def test_api_and_agency_and_web(self) -> None:
        assert classify_actor(_ag("API_TAKEOFF TRAVELS", "")) == "api"
        assert classify_actor(_ag("api_triplover", "")) == "api"
        assert classify_actor(_ag("Twenty", "")) == "agency"      # blank dept, human login
        assert classify_actor(_ag("Salim@24", "")) == "agency"
        assert classify_actor(_ag("webbot", "WEB")) == "web"

    def test_system_actor(self) -> None:
        sysag = Agent(raw="System (/TTI)", display_name="System", user_id="", department="")
        assert classify_actor(sysag) == "system"

    def test_flags_and_rows_carry_actor_type(self) -> None:
        a = _ag("jahirul3188", "DAC-02 Customer Service")        # internal
        rep = run_pnr_misuse_audit([_ev(new_status="Refunded", hour=20, agent=a)])  # 02:00 DAC
        assert rep.flags and rep.flags[0].actor_type == "internal"
        row = next(r for r in rep.agent_activity if r.agent_user_id == "jahirul3188")
        assert row.actor_type == "internal"
        ag_rows = [r for r in rep.risk_worklist if r.grain == "agent"]
        assert ag_rows and ag_rows[0].actor_type == "internal"

    def test_agency_abuse_still_surfaces(self) -> None:
        # External agency login is NOT excluded — its off-hours void still flags, tagged agency.
        rep = run_pnr_misuse_audit([_ev(new_status="Voided", hour=20, agent=_ag("Twenty", ""))])
        assert any(f.actor_type == "agency" for f in rep.flags)

    def test_web_and_system_actors_excluded(self) -> None:
        assert run_pnr_misuse_audit(
            [_ev(new_status="Refunded", hour=20, agent=_ag("webbot", "WEB"))]).flags == ()

    def test_workbook_has_actor_type_columns(self, tmp_path) -> None:
        from openpyxl import load_workbook
        from src.excel_io import write_zenith_pnr_misuse_audit
        rep = run_pnr_misuse_audit([_ev(new_status="Voided", hour=20, agent=_ag("Twenty", ""))])
        out = tmp_path / "at.xlsx"
        write_zenith_pnr_misuse_audit(out, rep)
        wb = load_workbook(out)
        for sheet in ("Flags", "Risk Worklist", "Agent Activity"):
            assert "Actor Type" in [c.value for c in wb[sheet][1]]
