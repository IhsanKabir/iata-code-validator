"""Tests for the Bulk Mailer split-by-email-column engine.

One main sheet in, one Excel per unique email address out — plus MailRow
conversion with GUI-level CC/BCC stamped on every message. All fixtures are
synthetic workbooks built in tmp_path; no real addresses, no network.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.mailer_split import build_mail_rows, read_headers, split_by_email


HEADERS = ["Agency", "IATA", "Amount", "Agent Email"]


def make_sheet(path: Path, rows: list[list], headers: list[str] = HEADERS,
               sheet_name: str = "Main") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(path)
    return path


def test_read_headers(tmp_path):
    p = make_sheet(tmp_path / "m.xlsx", [["A", "1", 10, "a@x.com"]])
    assert read_headers(p, "Main") == HEADERS
    assert read_headers(p) == HEADERS          # default = active sheet


def test_split_groups_rows_per_email(tmp_path):
    p = make_sheet(tmp_path / "m.xlsx", [
        ["Alpha", "111", 10, "a@x.com"],
        ["Beta", "222", 20, "b@y.com"],
        ["Alpha2", "333", 30, "A@X.COM"],      # same address, different case
    ])
    out = tmp_path / "out"
    res = split_by_email(p, out, email_column="Agent Email", sheet_name="Main")
    assert len(res.groups) == 2
    by_email = {g.email: g for g in res.groups}
    assert by_email["a@x.com"].row_count == 2   # case-insensitive grouping
    assert by_email["b@y.com"].row_count == 1
    # each split file: header + only that recipient's rows
    wb = load_workbook(by_email["a@x.com"].path)
    rows = list(wb.active.iter_rows(values_only=True))
    assert list(rows[0]) == HEADERS
    assert {r[0] for r in rows[1:]} == {"Alpha", "Alpha2"}
    assert not res.unmatched_rows


def test_invalid_and_blank_emails_go_to_unmatched(tmp_path):
    p = make_sheet(tmp_path / "m.xlsx", [
        ["Alpha", "111", 10, "a@x.com"],
        ["NoMail", "222", 20, ""],
        ["BadMail", "333", 30, "not-an-email"],
    ])
    res = split_by_email(p, tmp_path / "out", email_column="Agent Email")
    assert len(res.groups) == 1
    assert len(res.unmatched_rows) == 2         # never silently dropped
    assert res.unmatched_path is not None and Path(res.unmatched_path).is_file()
    wb = load_workbook(res.unmatched_path)
    names = {r[0] for r in wb.active.iter_rows(min_row=2, values_only=True)}
    assert names == {"NoMail", "BadMail"}
    assert res.warnings                          # surfaced to the GUI


def test_multi_address_cell_assigns_row_to_each(tmp_path):
    p = make_sheet(tmp_path / "m.xlsx", [
        ["Shared", "111", 10, "a@x.com; b@y.com"],
        ["Solo", "222", 20, "b@y.com"],
    ])
    res = split_by_email(p, tmp_path / "out", email_column="Agent Email")
    by_email = {g.email: g for g in res.groups}
    assert by_email["a@x.com"].row_count == 1
    assert by_email["b@y.com"].row_count == 2   # shared row lands in both files


def test_missing_email_column_raises(tmp_path):
    p = make_sheet(tmp_path / "m.xlsx", [["A", "1", 10, "a@x.com"]])
    with pytest.raises(ValueError, match="olumn"):
        split_by_email(p, tmp_path / "out", email_column="Nope")


def test_filenames_are_safe(tmp_path):
    # illegal-on-Windows chars in the local part must not break the write
    p = make_sheet(tmp_path / "m.xlsx", [["A", "1", 10, 'we?ird*na<me@x.com']])
    res = split_by_email(p, tmp_path / "out", email_column="Agent Email")
    assert len(res.groups) == 1
    path = Path(res.groups[0].path)
    assert path.is_file()
    assert not any(ch in path.name for ch in '\\/:*?"<>|')


def test_build_mail_rows_stamps_cc_bcc_and_fields(tmp_path):
    p = make_sheet(tmp_path / "m.xlsx", [
        ["Alpha", "111", 10, "a@x.com"],
        ["Alpha2", "333", 30, "a@x.com"],
        ["Beta", "222", 20, "b@y.com"],
    ])
    res = split_by_email(p, tmp_path / "out", email_column="Agent Email")
    rows = build_mail_rows(res, cc="boss@usb.com", bcc="audit@usb.com")
    assert len(rows) == 2
    r = next(x for x in rows if x.email == "a@x.com")
    assert r.is_valid
    assert r.cc == "boss@usb.com" and r.bcc == "audit@usb.com"
    assert len(r.attachments) == 1 and r.attachments[0].is_file()
    assert r.fields["email"] == "a@x.com"
    assert r.fields["rows"] == "2"               # {rows} placeholder for templates
    assert r.fields["file"] == r.attachments[0].name
    assert r.fields["name"]                      # local-part convenience alias


def test_split_empty_sheet_warns(tmp_path):
    p = make_sheet(tmp_path / "m.xlsx", [])
    res = split_by_email(p, tmp_path / "out", email_column="Agent Email")
    assert not res.groups
    assert res.warnings
