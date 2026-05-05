"""Tests for excel_io read/write helpers."""

import openpyxl

from src.excel_io import (
    ResultWriter,
    build_output_path,
    list_columns,
    list_sheet_names,
    read_iata_numbers,
)
from src.parser import LookupResult


def _make_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agents"
    ws.append(["Name", "IATA Number", "Notes"])
    ws.append(["A", 32302491, "ok"])
    ws.append(["B", "12345678", ""])
    ws.append(["C", None, ""])              # blank row in IATA col
    ws.append(["D", "  99999999  ", ""])     # padded
    ws.append(["E", "11-22-33-44", ""])      # hyphenated
    ws2 = wb.create_sheet("Other")
    ws2.append(["X"])
    wb.save(path)


def test_list_sheet_names(tmp_path):
    path = tmp_path / "in.xlsx"
    _make_workbook(path)
    assert list_sheet_names(path) == ["Agents", "Other"]


def test_list_columns(tmp_path):
    path = tmp_path / "in.xlsx"
    _make_workbook(path)
    cols = list_columns(path, "Agents")
    assert cols == ["Name", "IATA Number", "Notes"]


def test_read_iata_numbers_normalizes(tmp_path):
    path = tmp_path / "in.xlsx"
    _make_workbook(path)
    rows = read_iata_numbers(path, "Agents", column_index=1, start_row=2)
    # Row 4 (None) is skipped
    assert rows == [
        (2, "32302491"),
        (3, "12345678"),
        (5, "99999999"),
        (6, "11223344"),
    ]


def test_read_iata_numbers_with_range(tmp_path):
    path = tmp_path / "in.xlsx"
    _make_workbook(path)
    rows = read_iata_numbers(path, "Agents", 1, start_row=3, end_row=5)
    assert rows == [(3, "12345678"), (5, "99999999")]


def test_writer_appends_rows(tmp_path):
    out = tmp_path / "out.xlsx"
    w = ResultWriter(out)
    w.append(LookupResult("11111111", "A", "US", "Y", "VALID", "t", ""))
    w.append(LookupResult("22222222", "", "", "", "INVALID", "t", "n"))
    w.close()

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.max_row == 3  # header + 2 results
    assert ws.cell(row=2, column=1).value == "11111111"
    assert ws.cell(row=2, column=5).value == "VALID"
    assert ws.cell(row=3, column=5).value == "INVALID"


def test_build_output_path(tmp_path):
    p = build_output_path(tmp_path)
    assert p.parent == tmp_path
    assert p.name.startswith("iata_results_")
    assert p.suffix == ".xlsx"
